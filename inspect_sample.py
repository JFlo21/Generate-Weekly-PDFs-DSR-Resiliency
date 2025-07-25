#!/usr/bin/env python3
import openpyxl
import sys

def inspect_excel_file(filepath):
    """Inspect and display the contents of the Excel file."""
    try:
        wb = openpyxl.load_workbook(filepath)
        ws = wb.active
        
        print(f"📊 EXCEL FILE INSPECTION: {filepath}")
        print(f"=" * 80)
        print(f"📋 Worksheet Name: {ws.title}")
        print(f"📏 Dimensions: {ws.max_row} rows × {ws.max_column} columns")
        print(f"=" * 80)
        
        # Show first 20 rows with content
        print("\n🔍 CONTENT PREVIEW (First 20 rows with data):")
        print("-" * 80)
        
        content_rows = 0
        for row_num in range(1, min(ws.max_row + 1, 50)):  # Check first 50 rows
            # Get all cell values in this row
            row_values = []
            has_content = False
            
            for col_num in range(1, min(ws.max_column + 1, 10)):  # Show first 10 columns
                cell = ws.cell(row=row_num, column=col_num)
                value = cell.value
                if value is not None and str(value).strip():
                    has_content = True
                row_values.append(str(value) if value is not None else "")
            
            if has_content and content_rows < 20:  # Only show rows with content
                content_rows += 1
                # Format the row display
                row_display = " | ".join([val[:15].ljust(15) for val in row_values[:8]])
                print(f"Row {row_num:2d}: {row_display}")
        
        # Show some key data cells
        print(f"\n📋 KEY INFORMATION EXTRACTED:")
        print("-" * 40)
        
        # Look for specific patterns
        for row_num in range(1, min(ws.max_row + 1, 30)):
            for col_num in range(1, min(ws.max_column + 1, 6)):
                cell = ws.cell(row=row_num, column=col_num)
                if cell.value:
                    val_str = str(cell.value).strip()
                    if "Foreman:" in val_str:
                        adjacent_cell = ws.cell(row=row_num, column=col_num + 1)
                        print(f"   • Foreman: {adjacent_cell.value}")
                    elif "Work Request" in val_str:
                        adjacent_cell = ws.cell(row=row_num, column=col_num + 1)
                        print(f"   • Work Request: {adjacent_cell.value}")
                    elif "Week Ending" in val_str or "WEEKLY UNITS" in val_str:
                        print(f"   • Found: {val_str}")
                    elif "Total Billed Amount" in val_str:
                        adjacent_cell = ws.cell(row=row_num, column=col_num + 1)
                        print(f"   • Total Billed: {adjacent_cell.value}")
        
        print(f"\n✅ File inspection complete!")
        
    except Exception as e:
        print(f"❌ Error inspecting file: {e}")

if __name__ == "__main__":
    filepath = "generated_docs/WR_89708709_WeekEnding_071325.xlsx"
    inspect_excel_file(filepath)
