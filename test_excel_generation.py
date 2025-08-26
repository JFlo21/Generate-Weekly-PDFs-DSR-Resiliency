#!/usr/bin/env python3
"""
Quick test to verify Excel generation functionality
"""
import os
import datetime
from dateutil import parser
import openpyxl
from openpyxl.styles import Font, numbers, Alignment, PatternFill
from openpyxl.drawing.image import Image
import collections

def parse_price(price_str):
    """Safely converts a price string to a float."""
    if not price_str: return 0.0
    try:
        return float(str(price_str).replace('$', '').replace(',', ''))
    except:
        return 0.0

def test_excel_generation():
    """Test Excel generation with sample data"""
    
    # Sample data mimicking what the script should process
    sample_rows = [
        {
            'Work Request #': '89708709.0',
            'Weekly Reference Logged Date': '2024-06-29',  # Saturday, week ending Sunday 6/30
            'Snapshot Date': '2024-06-24',  # Monday
            'Foreman': 'John Smith',
            'Units Total Price': '$1250.50',
            'Pole #': 'P001',
            'CU': 'CU001',
            'Work Type': 'Installation',
            'CU Description': 'Install new pole',
            'Unit of Measure': 'EA',
            'Quantity': '1',
            'Scope ID': 'SC001',
            'Job #': 'JOB001',
            'Customer Name': 'Test Customer',
            'Work Order #': 'WO001'
        },
        {
            'Work Request #': '89708709.0',
            'Weekly Reference Logged Date': '2024-06-29',  # Same week
            'Snapshot Date': '2024-06-25',  # Tuesday
            'Foreman': 'John Smith',
            'Units Total Price': '$875.25',
            'Pole #': 'P002',
            'CU': 'CU002',
            'Work Type': 'Maintenance',
            'CU Description': 'Repair existing pole',
            'Unit of Measure': 'EA',
            'Quantity': '1',
            'Scope ID': 'SC001',
            'Job #': 'JOB001',
            'Customer Name': 'Test Customer',
            'Work Order #': 'WO001'
        }
    ]
    
    print("🧪 Testing Excel generation with sample data:")
    print(f"Sample data: {len(sample_rows)} rows")
    
    # Process data similar to group_source_rows function
    groups = collections.defaultdict(list)
    
    for r in sample_rows:
        foreman = r.get('Foreman')
        wr = r.get('Work Request #')
        log_date_str = r.get('Weekly Reference Logged Date')

        if not all([foreman, wr, log_date_str]):
            continue

        wr_key = str(wr).split('.')[0]
        
        try:
            # Parse the Weekly Reference Logged Date as the week ending date
            week_ending_date = parser.parse(log_date_str)
            week_end_for_key = week_ending_date.strftime("%m%d%y")
            
            print(f"📅 Processing WR# {wr_key}:")
            print(f"   Weekly Reference Logged Date: {log_date_str}")
            print(f"   Parsed as week ending: {week_ending_date.strftime('%A, %m/%d/%Y')}")
            print(f"   Key format: {week_end_for_key}")
            
            key = f"{week_end_for_key}"
            
            # Add calculated week ending date to the row data
            r['__current_foreman'] = foreman
            r['__week_ending_date'] = week_ending_date
            groups[key].append(r)
            
        except (parser.ParserError, TypeError) as e:
            print(f"❌ Could not parse date '{log_date_str}': {e}")
            continue
    
    print(f"\n📊 Grouped into {len(groups)} groups:")
    for key, rows in groups.items():
        print(f"   Group {key}: {len(rows)} rows")
    
    # Generate Excel for the first group
    if groups:
        group_key, group_rows = list(groups.items())[0]
        print(f"\n📝 Generating Excel for group: {group_key}")
        
        # Test the Excel generation
        generate_test_excel(group_key, group_rows)

def generate_test_excel(group_key, group_rows):
    """Generate a test Excel file"""
    
    first_row = group_rows[0]
    week_end_raw = group_key
    current_foreman = first_row.get('__current_foreman', 'Unknown_Foreman')
    week_ending_date = first_row.get('__week_ending_date')
    
    if week_ending_date:
        week_end_display = week_ending_date.strftime('%m/%d/%y')
    else:
        week_end_display = f"{week_end_raw[:2]}/{week_end_raw[2:4]}/{week_end_raw[4:]}"
    
    wr_numbers = list(set(str(row.get('Work Request #', '')).split('.')[0] for row in group_rows))
    wr_num = wr_numbers[0] if len(wr_numbers) == 1 else f"Multiple_WRs_{len(wr_numbers)}"
    
    output_filename = f"TEST_WR_{wr_num}_WeekEnding_{week_end_raw}.xlsx"
    output_path = os.path.join('generated_docs', output_filename)
    
    print(f"   Filename: {output_filename}")
    print(f"   Week Ending Display: {week_end_display}")
    print(f"   Foreman: {current_foreman}")
    print(f"   Work Request: {wr_num}")
    
    # Create Excel workbook
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = "Work Report"
    
    # Basic formatting
    LINETEC_RED = 'C00000'
    RED_FILL = PatternFill(start_color=LINETEC_RED, end_color=LINETEC_RED, fill_type='solid')
    SUMMARY_HEADER_FONT = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    SUMMARY_LABEL_FONT = Font(name='Calibri', size=10, bold=True)
    SUMMARY_VALUE_FONT = Font(name='Calibri', size=10)
    
    # Title
    current_row = 1
    ws.merge_cells(f'A{current_row}:I{current_row}')
    ws[f'A{current_row}'] = 'LINETEC SERVICES - TEST REPORT'
    ws[f'A{current_row}'].font = Font(name='Calibri', size=16, bold=True)
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
    
    current_row += 2
    
    # Summary section
    ws.merge_cells(f'B{current_row}:D{current_row}')
    ws[f'B{current_row}'] = 'REPORT SUMMARY'
    ws[f'B{current_row}'].font = SUMMARY_HEADER_FONT
    ws[f'B{current_row}'].fill = RED_FILL
    ws[f'B{current_row}'].alignment = Alignment(horizontal='center')
    
    # Calculate total price
    total_price = sum(parse_price(row.get('Units Total Price')) for row in group_rows)
    print(f"   Total Price Calculation:")
    for row in group_rows:
        price_str = row.get('Units Total Price')
        price_val = parse_price(price_str)
        print(f"     Row: '{price_str}' -> ${price_val:.2f}")
    print(f"   Final Total: ${total_price:.2f}")
    
    ws[f'B{current_row+1}'] = 'Total Billed Amount:'
    ws[f'B{current_row+1}'].font = SUMMARY_LABEL_FONT
    ws[f'C{current_row+1}'] = total_price
    ws[f'C{current_row+1}'].font = SUMMARY_VALUE_FONT
    ws[f'C{current_row+1}'].alignment = Alignment(horizontal='right')
    ws[f'C{current_row+1}'].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    
    ws[f'B{current_row+2}'] = 'Total Line Items:'
    ws[f'B{current_row+2}'].font = SUMMARY_LABEL_FONT
    ws[f'C{current_row+2}'] = len(group_rows)
    ws[f'C{current_row+2}'].font = SUMMARY_VALUE_FONT
    ws[f'C{current_row+2}'].alignment = Alignment(horizontal='right')
    
    ws[f'B{current_row+3}'] = 'Week Ending:'
    ws[f'B{current_row+3}'].font = SUMMARY_LABEL_FONT
    ws[f'C{current_row+3}'] = week_end_display
    ws[f'C{current_row+3}'].font = SUMMARY_VALUE_FONT
    ws[f'C{current_row+3}'].alignment = Alignment(horizontal='right')
    
    ws[f'B{current_row+4}'] = 'Foreman:'
    ws[f'B{current_row+4}'].font = SUMMARY_LABEL_FONT
    ws[f'C{current_row+4}'] = current_foreman
    ws[f'C{current_row+4}'].font = SUMMARY_VALUE_FONT
    ws[f'C{current_row+4}'].alignment = Alignment(horizontal='right')
    
    # Data rows
    current_row += 7
    headers = ["Date", "Pole #", "CU", "Work Type", "Description", "UOM", "Qty", "Price"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col_num)
        cell.value = header
        cell.font = SUMMARY_HEADER_FONT
        cell.fill = RED_FILL
        cell.alignment = Alignment(horizontal='center')
    
    for i, row_data in enumerate(group_rows):
        crow = current_row + 1 + i
        price = parse_price(row_data.get('Units Total Price'))
        
        # Extract quantity
        qty_str = str(row_data.get('Quantity', '') or 0)
        try:
            import re
            qty_match = re.search(r'(\d+(?:\.\d+)?)', qty_str)
            quantity = int(float(qty_match.group(1))) if qty_match else 0
        except (ValueError, AttributeError):
            quantity = 0
        
        row_values = [
            row_data.get('Snapshot Date', ''),
            row_data.get('Pole #', ''),
            row_data.get('CU', ''),
            row_data.get('Work Type', ''),
            row_data.get('CU Description', ''),
            row_data.get('Unit of Measure', ''),
            quantity,
            price
        ]
        
        for col_num, value in enumerate(row_values, 1):
            cell = ws.cell(row=crow, column=col_num)
            cell.value = value
            if col_num == 8:  # Price column
                cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    
    # Save the file
    os.makedirs('generated_docs', exist_ok=True)
    workbook.save(output_path)
    print(f"✅ Test Excel file saved: {output_path}")
    
    # Verify the file exists and check file size
    if os.path.exists(output_path):
        file_size = os.path.getsize(output_path)
        print(f"   File size: {file_size:,} bytes")
        if file_size > 5000:  # Reasonable minimum for an Excel file
            print("   ✅ File appears to be properly generated")
        else:
            print("   ⚠️ File seems very small, might be empty")
    else:
        print("   ❌ File was not created")

if __name__ == "__main__":
    test_excel_generation()
