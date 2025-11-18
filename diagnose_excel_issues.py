#!/usr/bin/env python3
"""
Diagnostic script to identify and fix Excel generation issues:
1. Cell merging errors and XML issues
2. Job # field not being populated
"""

import os
import sys
import logging
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import re

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def diagnose_merge_issues():
    """
    Test Excel generation with proper merge cell handling to avoid XML errors.
    """
    print("\n=== DIAGNOSING EXCEL MERGE ISSUES ===\n")
    
    # Create a test workbook
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = "Test Report"
    
    # Track all merges
    merged_ranges = []
    
    def safe_merge_cells(ws, range_str):
        """Safely merge cells, avoiding duplicates."""
        # Check if this range overlaps with any existing merge
        from openpyxl.utils import range_boundaries
        min_col, min_row, max_col, max_row = range_boundaries(range_str)
        
        # Check for existing merges
        for merged in ws.merged_cells.ranges:
            if str(merged) == range_str:
                print(f"  ⚠️  Duplicate merge detected, skipping: {range_str}")
                return False
            # Check for overlapping ranges
            merged_min_col, merged_min_row, merged_max_col, merged_max_row = range_boundaries(str(merged))
            # Check if ranges overlap
            if not (max_col < merged_min_col or min_col > merged_max_col or 
                    max_row < merged_min_row or min_row > merged_max_row):
                print(f"  ⚠️  Overlapping merge detected: {range_str} overlaps with {merged}")
                return False
        
        # Safe to merge
        ws.merge_cells(range_str)
        merged_ranges.append(range_str)
        print(f"  ✓ Successfully merged: {range_str}")
        return True
    
    # Test various merge scenarios
    print("Testing merge operations:")
    
    # Test 1: Logo area merge
    safe_merge_cells(ws, 'A1:C3')
    ws['A1'] = "LINETEC SERVICES"
    
    # Test 2: Title merge
    safe_merge_cells(ws, 'D1:I1')
    ws['D1'] = "WEEKLY UNITS COMPLETED"
    
    # Test 3: Report time merge
    safe_merge_cells(ws, 'D4:I4')
    ws['D4'] = "Report Generated On: 12/01/2024"
    
    # Test 4: Summary header
    safe_merge_cells(ws, 'B7:D7')
    ws['B7'] = "REPORT SUMMARY"
    
    # Test 5: Details header
    safe_merge_cells(ws, 'F7:I7')
    ws['F7'] = "REPORT DETAILS"
    
    # Test 6: Try duplicate merge (should be caught)
    safe_merge_cells(ws, 'F7:I7')  # Duplicate - should be skipped
    
    # Test 7: Detail value merges
    for i in range(8, 14):
        detail_range = f'G{i}:I{i}'
        safe_merge_cells(ws, detail_range)
        ws[f'F{i}'] = f"Label {i}:"
        ws[f'G{i}'] = f"Value {i}"
    
    # Save test file
    test_file = "test_merge_diagnostics.xlsx"
    try:
        workbook.save(test_file)
        print(f"\n✓ Test file saved successfully: {test_file}")
        print(f"  Total merges: {len(merged_ranges)}")
        
        # Try to reopen the file to check for XML errors
        print("\n  Validating Excel file...")
        test_wb = openpyxl.load_workbook(test_file)
        print("  ✓ File opens without XML errors")
        test_wb.close()
        
        # Clean up
        os.remove(test_file)
        print("  ✓ Test file cleaned up")
        
    except Exception as e:
        print(f"\n✗ Error saving or validating Excel file: {e}")
        return False
    
    return True

def diagnose_job_number_issue():
    """
    Diagnose why Job # is not being populated in Excel files.
    """
    print("\n=== DIAGNOSING JOB # POPULATION ISSUE ===\n")
    
    # Check if we have any generated Excel files to examine
    generated_docs = "generated_docs"
    
    if os.path.exists(generated_docs):
        excel_files = [f for f in os.listdir(generated_docs) if f.endswith('.xlsx')]
        
        if excel_files:
            print(f"Found {len(excel_files)} Excel files to examine")
            
            # Examine the first few files
            for filename in excel_files[:3]:
                filepath = os.path.join(generated_docs, filename)
                print(f"\n  Examining: {filename}")
                
                try:
                    workbook = openpyxl.load_workbook(filepath, data_only=True)
                    ws = workbook.active
                    
                    # Look for Job # in the details section
                    found_job = False
                    for row in ws.iter_rows(min_row=1, max_row=30, min_col=1, max_col=10):
                        for cell in row:
                            if cell.value and "Job #:" in str(cell.value):
                                # Found the label, check the value
                                row_num = cell.row
                                value_cell = ws.cell(row=row_num, column=cell.column + 1)
                                job_value = value_cell.value if value_cell else None
                                
                                if not job_value or str(job_value).strip() == "":
                                    print(f"    ⚠️  Job # label found at {cell.coordinate} but value is empty")
                                else:
                                    print(f"    ✓ Job # found: {job_value}")
                                found_job = True
                                break
                        if found_job:
                            break
                    
                    if not found_job:
                        print(f"    ✗ Job # field not found in file")
                    
                    workbook.close()
                    
                except Exception as e:
                    print(f"    ✗ Error reading file: {e}")
        else:
            print("No Excel files found in generated_docs")
    else:
        print("generated_docs folder does not exist")
    
    # Check column name variations
    print("\n=== CHECKING COLUMN NAME VARIATIONS ===\n")
    print("The code looks for 'Job #' in the row data.")
    print("Possible issues:")
    print("  1. Column might be named differently (e.g., 'Job#', 'Job Number', 'JobNumber')")
    print("  2. Data might not be populated in Smartsheet")
    print("  3. Helper rows might have different field structure")
    
    return True

def create_fixed_generate_excel_snippet():
    """
    Create a snippet showing the fixes needed for generate_excel function.
    """
    print("\n=== RECOMMENDED FIXES ===\n")
    
    fixes = """
1. FIX FOR MERGE CELL ERRORS:
   - Add duplicate checking for ALL merge operations
   - Check for overlapping ranges before merging
   
2. FIX FOR JOB # NOT POPULATED:
   - Check multiple column name variations
   - Add fallback logic for missing Job #
   
3. IMPLEMENTATION:

# In generate_weekly_pdfs.py, update the merge operations:

def safe_merge_cells(ws, range_str):
    '''Safely merge cells, avoiding duplicates and overlaps.'''
    # Check if already merged
    for merged in ws.merged_cells.ranges:
        if str(merged) == range_str:
            return False  # Already merged
    # Proceed with merge
    ws.merge_cells(range_str)
    return True

# Update Job # retrieval logic around line 1439:

# Try multiple column name variations for Job #
job_number = (first_row.get('Job #') or 
              first_row.get('Job#') or 
              first_row.get('Job Number') or 
              first_row.get('JobNumber') or 
              first_row.get('Job_Number') or 
              first_row.get('JOB #') or 
              '')

# Add logging if Job # is missing
if not job_number:
    logging.warning(f"Job # not found for WR {wr_num}. Available columns: {list(first_row.keys())[:10]}")
"""
    
    print(fixes)
    return True

def main():
    """Main diagnostic function."""
    print("\n" + "="*60)
    print("EXCEL GENERATION DIAGNOSTICS")
    print("="*60)
    
    # Run diagnostics
    merge_ok = diagnose_merge_issues()
    job_ok = diagnose_job_number_issue()
    
    # Show recommended fixes
    create_fixed_generate_excel_snippet()
    
    print("\n" + "="*60)
    print("DIAGNOSTIC SUMMARY")
    print("="*60)
    print(f"  Merge Cell Handling: {'✓ OK' if merge_ok else '✗ Issues Found'}")
    print(f"  Job # Population: {'⚠️  Needs Investigation' if job_ok else '✗ Issues Found'}")
    print("\nNext Steps:")
    print("  1. Apply the recommended fixes to generate_weekly_pdfs.py")
    print("  2. Add logging to identify missing column names")
    print("  3. Test with actual Smartsheet data")
    
    return 0 if (merge_ok and job_ok) else 1

if __name__ == "__main__":
    sys.exit(main())