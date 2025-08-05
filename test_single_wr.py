#!/usr/bin/env python3
"""
Single Work Request Test - Generate Multiple Week Ending Reports
===============================================================
This script tests the week ending date fix by generating one timesheet
for each week ending date for a single work request.
"""

import os
import datetime
from datetime import timedelta
import logging
from dateutil import parser
import smartsheet
import openpyxl
from openpyxl.styles import Font, numbers, Alignment, PatternFill
from openpyxl.drawing.image import Image
import collections
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()

# --- Configuration ---
API_TOKEN = os.getenv("SMARTSHEET_API_TOKEN")
TARGET_SHEET_ID = 5723337641643908
TARGET_WR_COLUMN_ID = 7941607783092100
LOGO_PATH = "LinetecServices_Logo.png"
OUTPUT_FOLDER = "generated_docs"
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# --- SINGLE WORK REQUEST TEST CONFIGURATION ---
TEST_WORK_REQUEST = "89708709"  # Focus on this specific work request
FORCE_TEST_MODE = True  # Always run in test mode for safety

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

def parse_price(price_str):
    """Safely converts a price string to a float."""
    if not price_str: return 0.0
    try:
        return float(str(price_str).replace('$', '').replace(',', ''))
    except (ValueError, TypeError):
        return 0.0

def is_checked(val):
    """Checks if a value from a checkbox column is considered 'checked'."""
    if isinstance(val, bool):
        return val
    if isinstance(val, int):
        return val == 1
    if isinstance(val, str):
        return val.strip().lower() in ('true', 'checked', 'yes', '1')
    return False

def get_single_wr_data(client):
    """Get data for a single work request from the first available source sheet."""
    # Use the main source sheet
    source_sheet_id = 3239244454645636
    
    # Column mapping for the source sheet
    column_name_mapping = {
        'Foreman': 'Foreman',
        'Work Request #': 'Work Request #',
        'Weekly Reference Logged Date': 'Weekly Referenced Logged Date',
        'Dept #': 'Dept #',
        'Customer Name': 'Customer Name',
        'Work Order #': 'Work Order #',
        'Area': 'Area',
        'Pole #': 'Pole #',
        'CU': 'CU',
        'Work Type': 'Work Type',
        'CU Description': 'CU Description',
        'Unit of Measure': 'Unit of Measure',
        'Quantity': 'Quantity',
        'Units Total Price': 'Redlined Total Price',
        'Snapshot Date': 'Snapshot Date',
        'Scope #': 'Scope ID',
        'Job #': 'Job #',
        'Units Completed?': 'Units Completed',
    }
    
    try:
        # Get the source sheet
        sheet = client.Sheets.get_sheet(source_sheet_id)
        logging.info(f"Processing Sheet: {sheet.name} (ID: {source_sheet_id})")
        
        # Create column mapping by matching column titles
        column_mapping = {}
        for column in sheet.columns:
            if column.title in column_name_mapping:
                column_mapping[column_name_mapping[column.title]] = column.id
        
        # Filter rows for the specific work request
        filtered_rows = []
        for row in sheet.rows:
            cell_map = {c.column_id: c.value for c in row.cells}
            if not any(cell_map.values()):
                continue  # Skip entirely empty rows

            # Create a parsed dictionary of the row's values
            parsed = {key: cell_map.get(col_id) for key, col_id in column_mapping.items()}
            
            # Check if this row belongs to our target work request
            wr_value = parsed.get('Work Request #')
            if not wr_value:
                continue
                
            wr_key = str(wr_value).split('.')[0]
            if wr_key != TEST_WORK_REQUEST:
                continue  # Skip if not our target work request
            
            # Apply filtering criteria
            has_date = parsed.get('Snapshot Date')
            is_complete = is_checked(parsed.get('Units Completed'))
            has_price = parse_price(parsed.get('Redlined Total Price')) > 0

            if not (has_date and is_complete and has_price):
                continue  # If any condition fails, skip this row

            # Add metadata to the row for later use
            parsed['__sheet_id'] = source_sheet_id
            parsed['__row_obj'] = row
            filtered_rows.append(parsed)
        
        logging.info(f"Found {len(filtered_rows)} valid rows for WR# {TEST_WORK_REQUEST}")
        return filtered_rows
        
    except Exception as e:
        logging.error(f"Could not process Sheet ID {source_sheet_id}. Error: {e}")
        return []

def group_wr_rows_by_week_ending(rows):
    """Groups rows by week ending date for the single work request."""
    groups = collections.defaultdict(list)
    
    print(f"\n🔍 Analyzing {len(rows)} rows for WR# {TEST_WORK_REQUEST}")
    print("="*60)
    
    # Track all unique week ending dates found
    week_endings = set()
    
    for r in rows:
        foreman = r.get('Foreman')
        wr = r.get('Work Request #')
        log_date_str = r.get('Weekly Referenced Logged Date')
        snapshot_date_str = r.get('Snapshot Date')

        if not all([foreman, wr, log_date_str, snapshot_date_str]):
            continue  # Skip if key information is missing

        wr_key = str(wr).split('.')[0]
        
        try:
            date_obj = parser.parse(log_date_str)
            
            # Calculate the week ending date (Sunday of that week)
            if date_obj.weekday() == 6:  # If it's already Sunday
                week_ending_date = date_obj
            else:
                days_until_sunday = (6 - date_obj.weekday()) % 7
                week_ending_date = date_obj + timedelta(days=days_until_sunday)
            
            week_end_for_key = week_ending_date.strftime("%m%d%y")
            week_endings.add(week_ending_date.strftime('%m/%d/%Y'))
            
            print(f"📅 Row: Log date {date_obj.strftime('%m/%d/%Y')} ({date_obj.strftime('%A')}) → Week ending {week_ending_date.strftime('%m/%d/%Y')} ({week_ending_date.strftime('%A')})")
            
            key = f"{foreman}_{wr_key}_{week_end_for_key}"
            
            # Add the calculated week ending date to the row data
            r['__week_ending_date'] = week_ending_date
            groups[key].append(r)
            
        except (parser.ParserError, TypeError) as e:
            logging.warning(f"Could not parse date '{log_date_str}' for WR# {wr_key}. Skipping row. Error: {e}")
            continue
    
    print(f"\n📊 Summary for WR# {TEST_WORK_REQUEST}:")
    print(f"   • Unique week ending dates found: {len(week_endings)}")
    for we in sorted(week_endings):
        print(f"     - {we}")
    print(f"   • Total groups to generate: {len(groups)}")
    print("="*60)
    
    return groups

def generate_test_excel(group_key, group_rows):
    """Generates a test Excel report for verification."""
    first_row = group_rows[0]
    foreman, wr_num, week_end_raw = group_key.split('_')
    
    # Get the calculated week ending date from the row data
    week_ending_date = first_row.get('__week_ending_date')
    if week_ending_date:
        week_end_display = week_ending_date.strftime('%m/%d/%y')
        week_end_raw = week_ending_date.strftime('%m%d%y')
    else:
        week_end_display = f"{week_end_raw[:2]}/{week_end_raw[2:4]}/{week_end_raw[4:]}"
    
    scope_id = first_row.get('Scope ID', '')
    job_number = first_row.get('Job #', '')
    output_filename = f"TEST_WR_{wr_num}_WeekEnding_{week_end_raw}.xlsx"
    final_output_path = os.path.join(OUTPUT_FOLDER, output_filename)

    print(f"\n📊 Generating Excel: {output_filename}")
    print(f"   - Work Request: {wr_num}")
    print(f"   - Foreman: {foreman}")
    print(f"   - Week Ending: {week_end_display}")
    if week_ending_date:
        print(f"   - Calculated Week Ending: {week_ending_date.strftime('%A, %m/%d/%Y')}")
    print(f"   - Row Count: {len(group_rows)}")

    # Calculate totals
    total_price = sum(parse_price(row.get('Redlined Total Price')) for row in group_rows)
    
    # Group by snapshot date for daily breakdown
    date_to_rows = collections.defaultdict(list)
    for row in group_rows:
        snap = row.get('Snapshot Date')
        try:
            dt = parser.parse(snap)
            date_to_rows[dt].append(row)
        except (parser.ParserError, TypeError, ValueError):
            continue
    
    print(f"\n   📋 Report Summary:")
    print(f"      • Total Line Items: {len(group_rows)}")
    print(f"      • Total Billed Amount: ${total_price:,.2f}")
    print(f"      • Scope ID: {scope_id}")
    print(f"      • Job #: {job_number}")
    print(f"      • Customer: {first_row.get('Customer Name', '')}")
    print(f"      • Work Order #: {first_row.get('Work Order #', '')}")
    print(f"      • Department #: {first_row.get('Dept #', '')}")
    print(f"      • Area: {first_row.get('Area', '')}")
    
    if date_to_rows:
        print(f"\n   📅 Daily Breakdown:")
        for date_obj in sorted(date_to_rows.keys()):
            day_rows = date_to_rows[date_obj]
            day_total = sum(parse_price(row.get('Redlined Total Price')) for row in day_rows)
            print(f"      • {date_obj.strftime('%A, %m/%d/%Y')}: {len(day_rows)} items, ${day_total:,.2f}")
    
    # Create a simple Excel file for verification
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = "Work Report Test"
    
    # Header information
    ws['A1'] = f"TEST REPORT - WR# {wr_num}"
    ws['A1'].font = Font(size=16, bold=True)
    
    ws['A3'] = f"Week Ending: {week_end_display}"
    ws['A4'] = f"Calculated Week Ending: {week_ending_date.strftime('%A, %m/%d/%Y') if week_ending_date else 'N/A'}"
    ws['A5'] = f"Foreman: {foreman}"
    ws['A6'] = f"Total Amount: ${total_price:,.2f}"
    ws['A7'] = f"Total Items: {len(group_rows)}"
    
    # Column headers
    headers = ['Snapshot Date', 'Point #', 'CU', 'Work Type', 'Description', 'Quantity', 'Price']
    for col, header in enumerate(headers, 1):
        ws.cell(row=9, column=col, value=header).font = Font(bold=True)
    
    # Data rows
    for row_idx, row_data in enumerate(group_rows, 10):
        ws.cell(row=row_idx, column=1, value=row_data.get('Snapshot Date', ''))
        ws.cell(row=row_idx, column=2, value=row_data.get('Pole #', ''))
        ws.cell(row=row_idx, column=3, value=row_data.get('CU', ''))
        ws.cell(row=row_idx, column=4, value=row_data.get('Work Type', ''))
        ws.cell(row=row_idx, column=5, value=row_data.get('CU Description', ''))
        ws.cell(row=row_idx, column=6, value=row_data.get('Quantity', ''))
        ws.cell(row=row_idx, column=7, value=parse_price(row_data.get('Redlined Total Price')))
        ws.cell(row=row_idx, column=7).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
    
    # Auto-adjust column widths
    for col in range(1, 8):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
    
    # Save the file
    workbook.save(final_output_path)
    logging.info(f"📄 Generated test Excel: '{output_filename}'")
    
    return final_output_path, output_filename

def main():
    """Main test execution function."""
    try:
        if not API_TOKEN:
            logging.error("🚨 FATAL: SMARTSHEET_API_TOKEN environment variable not set.")
            return

        client = smartsheet.Smartsheet(API_TOKEN)
        client.errors_as_exceptions(True)

        print(f"\n{'🧪 SINGLE WORK REQUEST TEST 🧪':^80}")
        print(f"{'='*80}")
        print(f"Testing week ending calculation for WR# {TEST_WORK_REQUEST}")
        print(f"Will generate one timesheet for each week ending date found")
        print(f"{'='*80}\n")

        logging.info(f"--- Starting Single WR Test for WR# {TEST_WORK_REQUEST} ---")
        
        # 1. Get data for the specific work request
        wr_rows = get_single_wr_data(client)
        if not wr_rows:
            logging.error(f"No valid rows found for WR# {TEST_WORK_REQUEST}. Exiting.")
            return

        # 2. Group the rows by week ending date
        wr_groups = group_wr_rows_by_week_ending(wr_rows)
        if not wr_groups:
            logging.error(f"No groups created for WR# {TEST_WORK_REQUEST}. Exiting.")
            return

        logging.info(f"Created {len(wr_groups)} week ending groups for WR# {TEST_WORK_REQUEST}")

        # 3. Generate Excel file for each week ending
        print(f"\n🔄 Generating Excel files...")
        print("="*80)
        
        generated_files = []
        for group_key, group_rows in wr_groups.items():
            if not group_rows:
                continue

            excel_path, excel_filename = generate_test_excel(group_key, group_rows)
            generated_files.append((excel_filename, len(group_rows)))

        # 4. Summary
        print(f"\n{'='*80}")
        print(f"🎉 TEST COMPLETE - Week Ending Date Fix Verification")
        print(f"{'='*80}")
        print(f"📈 Results for WR# {TEST_WORK_REQUEST}:")
        print(f"   • Total Excel files generated: {len(generated_files)}")
        print(f"   • Files created in: {OUTPUT_FOLDER}/")
        print(f"\n📄 Generated Files:")
        
        for filename, row_count in generated_files:
            print(f"   • {filename} ({row_count} line items)")
        
        print(f"\n✅ Key Verification Points:")
        print(f"   • Each file shows correct Sunday week ending dates")
        print(f"   • Multiple week endings for same WR# are properly separated")
        print(f"   • Week ending calculation handles all days correctly")
        print(f"   • Files ready for inspection in {OUTPUT_FOLDER} folder")
        print(f"{'='*80}")

    except smartsheet.exceptions.ApiError as e:
        logging.error(f"A Smartsheet API error occurred: {e}")
    except Exception as e:
        logging.error(f"An unexpected error occurred: {e}", exc_info=True)

if __name__ == "__main__":
    main()
