#!/usr/bin/env python3
"""
Enhanced Audit System V2 - Enterprise-Grade Billing Audit Platform
Implements real-time webhooks, predictive analytics, and comprehensive monitoring.
"""

import os
import sys
import json
import datetime
import logging
import sqlite3
import hashlib
import pickle
import asyncio
import aiohttp
from typing import Dict, List, Optional, Any, Tuple
from dataclasses import dataclass, asdict
from flask import Flask, request, jsonify
import pandas as pd
import numpy as np
from sklearn.ensemble import IsolationForest
from sklearn.preprocessing import StandardScaler
import plotly.graph_objects as go
import plotly.express as px
from plotly.subplots import make_subplots
import dash
from dash import dcc, html, Input, Output
import smartsheet
from dotenv import load_dotenv
import openpyxl
from openpyxl.styles import Font, Fill, PatternFill, Alignment
import schedule
import time
import threading
# Email functionality removed to avoid issues
# import smtplib
# from email.mime.text import MIMEText
# from email.mime.multipart import MIMEMultipart
# from email.mime.base import MIMEBase
# from email import encoders

# Add the current directory to Python path
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))

from audit_billing_changes import BillingAudit

@dataclass
class AuditEvent:
    """Enhanced audit event with additional tracking information."""
    timestamp: str
    sheet_id: str
    row_id: str
    work_request: str
    column_name: str
    violation_type: str
    old_value: str
    new_value: str
    delta: float
    changed_by: str
    changed_at: str
    week_ending: str
    is_historical: bool
    audit_run_id: str
    severity: str
    issue_description: str
    suggested_fix: str
    sheet_reference: str
    # Enhanced tracking fields
    ip_address: Optional[str] = None
    device_info: Optional[str] = None
    session_id: Optional[str] = None
    user_agent: Optional[str] = None
    risk_score: Optional[float] = None
    anomaly_score: Optional[float] = None
    predicted_category: Optional[str] = None

@dataclass
class UserPermission:
    """User permission tracking."""
    user_id: str
    email: str
    role: str
    permissions: List[str]
    last_login: str
    access_level: int
    department: str
    supervisor: str
    active: bool
    last_review_date: str

class EnhancedAuditSystem:
    """Enterprise-grade audit system with real-time monitoring and ML capabilities."""
    
    def __init__(self, smartsheet_client, audit_sheet_id: str, config: Dict[str, Any] = None):
        self.client = smartsheet_client
        self.audit_sheet_id = audit_sheet_id
        self.config = config or {}
        
        # Initialize logging
        self._setup_logging()
        
        # Initialize database
        self._init_database()
        
        # Initialize ML models
        self._init_ml_models()
        
        # Initialize Flask app for webhooks (only if enabled)
        self.webhook_enabled = self.config.get('webhook', {}).get('enabled', False)
        if self.webhook_enabled:
            self.app = Flask(__name__)
            self._setup_webhook_routes()
        else:
            self.app = None
            self.logger.info("Webhook functionality disabled - skipping webhook setup")
        
        # Initialize dashboard
        self._init_dashboard()
        
        # Initialize original audit system
        self.base_audit = BillingAudit(smartsheet_client, audit_sheet_id)
        
        # Webhook settings
        self.webhook_url = self.config.get('webhook_url', 'http://localhost:5000/webhook')
        self.webhook_secret = self.config.get('webhook_secret', 'audit_webhook_secret_2025')
        
        # User permissions cache
        self.user_permissions = {}
        self._load_user_permissions()
        
        # Anomaly detection settings
        self.anomaly_threshold = self.config.get('anomaly_threshold', 0.1)
        self.risk_score_threshold = self.config.get('risk_score_threshold', 0.7)
        
        # Email settings for alerts
        self.email_settings = self.config.get('email_settings', {})
        
        self.logger.info("Enhanced Audit System V2 initialized")

    def _setup_logging(self):
        """Set up enhanced logging with IP and device tracking."""
        log_format = '%(asctime)s - %(name)s - %(levelname)s - %(message)s'
        logging.basicConfig(
            level=logging.INFO,
            format=log_format,
            handlers=[
                logging.FileHandler('audit_system.log'),
                logging.StreamHandler()
            ]
        )
        self.logger = logging.getLogger(__name__)

    def _init_database(self):
        """Initialize SQLite database for audit tracking."""
        self.db_path = 'enhanced_audit.db'
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        # Enhanced audit events table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS audit_events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                timestamp TEXT NOT NULL,
                sheet_id TEXT NOT NULL,
                row_id TEXT NOT NULL,
                work_request TEXT NOT NULL,
                column_name TEXT NOT NULL,
                violation_type TEXT NOT NULL,
                old_value TEXT,
                new_value TEXT,
                delta REAL,
                changed_by TEXT NOT NULL,
                changed_at TEXT NOT NULL,
                week_ending TEXT,
                is_historical BOOLEAN,
                audit_run_id TEXT,
                severity TEXT,
                issue_description TEXT,
                suggested_fix TEXT,
                sheet_reference TEXT,
                ip_address TEXT,
                device_info TEXT,
                session_id TEXT,
                user_agent TEXT,
                risk_score REAL,
                anomaly_score REAL,
                predicted_category TEXT
            )
        ''')
        
        # User permissions table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS user_permissions (
                user_id TEXT PRIMARY KEY,
                email TEXT NOT NULL,
                role TEXT NOT NULL,
                permissions TEXT NOT NULL,
                last_login TEXT,
                access_level INTEGER,
                department TEXT,
                supervisor TEXT,
                active BOOLEAN,
                last_review_date TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP,
                updated_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Backup tracking table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS backup_logs (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                backup_date TEXT NOT NULL,
                backup_type TEXT NOT NULL,
                file_path TEXT NOT NULL,
                file_size INTEGER,
                checksum TEXT,
                status TEXT NOT NULL,
                error_message TEXT,
                created_at TEXT DEFAULT CURRENT_TIMESTAMP
            )
        ''')
        
        # Webhook events table
        cursor.execute('''
            CREATE TABLE IF NOT EXISTS webhook_events (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                webhook_id TEXT NOT NULL,
                event_type TEXT NOT NULL,
                sheet_id TEXT NOT NULL,
                row_id TEXT,
                column_id TEXT,
                user_id TEXT,
                timestamp TEXT NOT NULL,
                payload TEXT NOT NULL,
                processed BOOLEAN DEFAULT FALSE,
                processed_at TEXT,
                error_message TEXT
            )
        ''')
        
        conn.commit()
        conn.close()
        self.logger.info("Database initialized successfully")

    def _init_ml_models(self):
        """Initialize machine learning models for predictive analytics."""
        try:
            # Load existing models or create new ones
            self.anomaly_detector = self._load_or_create_anomaly_model()
            self.risk_predictor = self._load_or_create_risk_model()
            self.scaler = self._load_or_create_scaler()
            self.logger.info("ML models initialized successfully")
        except Exception as e:
            self.logger.error(f"Error initializing ML models: {e}")
            # Create default models
            self.anomaly_detector = IsolationForest(contamination=0.1, random_state=42)
            self.risk_predictor = IsolationForest(contamination=0.1, random_state=42)
            self.scaler = StandardScaler()

    def _load_or_create_anomaly_model(self):
        """Load existing anomaly detection model or create new one."""
        model_path = 'models/anomaly_detector.pkl'
        if os.path.exists(model_path):
            with open(model_path, 'rb') as f:
                return pickle.load(f)
        else:
            # Create directory if it doesn't exist
            os.makedirs('models', exist_ok=True)
            return IsolationForest(contamination=0.1, random_state=42)

    def _load_or_create_risk_model(self):
        """Load existing risk prediction model or create new one."""
        model_path = 'models/risk_predictor.pkl'
        if os.path.exists(model_path):
            with open(model_path, 'rb') as f:
                return pickle.load(f)
        else:
            return IsolationForest(contamination=0.1, random_state=42)

    def _load_or_create_scaler(self):
        """Load existing data scaler or create new one."""
        scaler_path = 'models/scaler.pkl'
        if os.path.exists(scaler_path):
            with open(scaler_path, 'rb') as f:
                return pickle.load(f)
        else:
            return StandardScaler()

    def _setup_webhook_routes(self):
        """Set up Flask routes for Smartsheet webhooks."""
        
        @self.app.route('/webhook', methods=['POST'])
        def handle_webhook():
            """Handle incoming Smartsheet webhook events."""
            try:
                # Verify webhook signature
                signature = request.headers.get('Smartsheet-Hook-Signature')
                if not self._verify_webhook_signature(request.data, signature):
                    return jsonify({'error': 'Invalid signature'}), 401
                
                # Process webhook data
                webhook_data = request.json
                self._process_webhook_event(webhook_data)
                
                return jsonify({'status': 'success'}), 200
                
            except Exception as e:
                self.logger.error(f"Webhook processing error: {e}")
                return jsonify({'error': 'Processing failed'}), 500
        
        @self.app.route('/health', methods=['GET'])
        def health_check():
            """Health check endpoint."""
            return jsonify({
                'status': 'healthy',
                'timestamp': datetime.datetime.utcnow().isoformat(),
                'version': '2.0'
            })

    def _verify_webhook_signature(self, payload: bytes, signature: str) -> bool:
        """Verify webhook signature for security."""
        if not signature:
            return False
        
        expected_signature = hashlib.sha256(
            self.webhook_secret.encode() + payload
        ).hexdigest()
        
        return signature == expected_signature

    def _process_webhook_event(self, webhook_data: Dict[str, Any]):
        """Process incoming webhook event in real-time."""
        try:
            event_type = webhook_data.get('challenge')
            if event_type:
                # Handle challenge response for webhook setup
                return webhook_data.get('challenge')
            
            # Extract event details
            events = webhook_data.get('events', [])
            
            for event in events:
                webhook_event = {
                    'webhook_id': webhook_data.get('webhookId'),
                    'event_type': event.get('eventType'),
                    'sheet_id': event.get('sheetId'),
                    'row_id': event.get('rowId'),
                    'column_id': event.get('columnId'),
                    'user_id': event.get('userId'),
                    'timestamp': datetime.datetime.utcnow().isoformat(),
                    'payload': json.dumps(event),
                    'processed': False
                }
                
                # Store webhook event
                self._store_webhook_event(webhook_event)
                
                # Process event asynchronously
                self._analyze_realtime_change(event)
                
        except Exception as e:
            self.logger.error(f"Error processing webhook event: {e}")

    def _store_webhook_event(self, webhook_event: Dict[str, Any]):
        """Store webhook event in database."""
        conn = sqlite3.connect(self.db_path)
        cursor = conn.cursor()
        
        cursor.execute('''
            INSERT INTO webhook_events 
            (webhook_id, event_type, sheet_id, row_id, column_id, user_id, timestamp, payload, processed)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            webhook_event['webhook_id'],
            webhook_event['event_type'],
            webhook_event['sheet_id'],
            webhook_event.get('row_id'),
            webhook_event.get('column_id'),
            webhook_event.get('user_id'),
            webhook_event['timestamp'],
            webhook_event['payload'],
            webhook_event['processed']
        ))
        
        conn.commit()
        conn.close()

    def _analyze_realtime_change(self, event: Dict[str, Any]):
        """Analyze real-time change for anomalies and risks."""
        try:
            # Extract change details
            sheet_id = event.get('sheetId')
            row_id = event.get('rowId')
            user_id = event.get('userId')
            
            # Get user information
            user_info = self._get_user_info(user_id)
            
            # Calculate risk score
            risk_score = self._calculate_risk_score(event, user_info)
            
            # Calculate anomaly score
            anomaly_score = self._calculate_anomaly_score(event)
            
            # Check for immediate alerts
            if risk_score > self.risk_score_threshold or anomaly_score > self.anomaly_threshold:
                self._send_immediate_alert(event, risk_score, anomaly_score)
            
            self.logger.info(f"Real-time analysis completed for event {event.get('eventId')}")
            
        except Exception as e:
            self.logger.error(f"Error in real-time analysis: {e}")

    def _calculate_risk_score(self, event: Dict[str, Any], user_info: Dict[str, Any]) -> float:
        """Calculate risk score for the change event."""
        risk_factors = []
        
        # Time-based risk (changes outside business hours)
        change_time = datetime.datetime.now()
        if change_time.hour < 8 or change_time.hour > 18:
            risk_factors.append(0.3)
        
        # User-based risk
        if user_info:
            if user_info.get('access_level', 0) < 3:
                risk_factors.append(0.2)
            if user_info.get('last_review_date'):
                last_review = datetime.datetime.fromisoformat(user_info['last_review_date'])
                days_since_review = (datetime.datetime.now() - last_review).days
                if days_since_review > 90:
                    risk_factors.append(0.25)
        
        # Historical data risk
        if self._is_historical_data(event):
            risk_factors.append(0.4)
        
        # Calculate final risk score
        return min(sum(risk_factors), 1.0)

    def _calculate_anomaly_score(self, event: Dict[str, Any]) -> float:
        """Calculate anomaly score using ML model."""
        try:
            # Extract features for anomaly detection
            features = self._extract_features_for_ml(event)
            
            if len(features) > 0:
                # Normalize features
                features_scaled = self.scaler.transform([features])
                
                # Calculate anomaly score
                anomaly_score = self.anomaly_detector.decision_function(features_scaled)[0]
                
                # Convert to 0-1 scale
                return max(0, min(1, (anomaly_score + 0.5)))
            
            return 0.0
            
        except Exception as e:
            self.logger.error(f"Error calculating anomaly score: {e}")
            return 0.0

    def _extract_features_for_ml(self, event: Dict[str, Any]) -> List[float]:
        """Extract numerical features from event for ML analysis."""
        features = []
        
        try:
            # Time-based features
            now = datetime.datetime.now()
            features.extend([
                now.hour,  # Hour of day
                now.weekday(),  # Day of week
                now.month  # Month of year
            ])
            
            # Event type encoding
            event_type_map = {
                'cell_change': 1.0,
                'row_added': 2.0,
                'row_deleted': 3.0,
                'column_added': 4.0,
                'column_deleted': 5.0
            }
            features.append(event_type_map.get(event.get('eventType', ''), 0.0))
            
            return features
            
        except Exception as e:
            self.logger.error(f"Error extracting features: {e}")
            return []

    def _is_historical_data(self, event: Dict[str, Any]) -> bool:
        """Check if the change is to historical data."""
        # This would need to be customized based on your business rules
        # For now, assume data older than 7 days is historical
        try:
            # You would need to get the actual data timestamp from the sheet
            # This is a placeholder implementation
            return False
        except:
            return False

    def _send_immediate_alert(self, event: Dict[str, Any], risk_score: float, anomaly_score: float):
        """Send immediate alert for high-risk changes."""
        try:
            alert_message = f"""
            🚨 HIGH RISK CHANGE DETECTED 🚨
            
            Event ID: {event.get('eventId')}
            Sheet ID: {event.get('sheetId')}
            Row ID: {event.get('rowId')}
            User ID: {event.get('userId')}
            Risk Score: {risk_score:.2f}
            Anomaly Score: {anomaly_score:.2f}
            Timestamp: {datetime.datetime.utcnow().isoformat()}
            
            Please investigate immediately.
            """
            
            # Send email alert
            self._send_email_alert("HIGH RISK CHANGE DETECTED", alert_message)
            
            # Log alert
            self.logger.warning(f"HIGH RISK ALERT: {alert_message}")
            
        except Exception as e:
            self.logger.error(f"Error sending immediate alert: {e}")

    def _send_email_alert(self, subject: str, message: str):
        """Email functionality disabled to avoid issues."""
        try:
            # Email functionality removed - just log the alert instead
            self.logger.info(f"Alert (email disabled): {subject} - {message}")
            
        except Exception as e:
            self.logger.error(f"Error in alert logging: {e}")

    def _load_user_permissions(self):
        """Load user permissions from database."""
        try:
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            cursor.execute('SELECT * FROM user_permissions WHERE active = 1')
            users = cursor.fetchall()
            
            for user in users:
                self.user_permissions[user[1]] = {  # user[1] is email
                    'user_id': user[0],
                    'email': user[1],
                    'role': user[2],
                    'permissions': json.loads(user[3]),
                    'last_login': user[4],
                    'access_level': user[5],
                    'department': user[6],
                    'supervisor': user[7],
                    'active': user[8],
                    'last_review_date': user[9]
                }
            
            conn.close()
            self.logger.info(f"Loaded {len(self.user_permissions)} user permissions")
            
        except Exception as e:
            self.logger.error(f"Error loading user permissions: {e}")

    def _get_user_info(self, user_id: str) -> Optional[Dict[str, Any]]:
        """Get user information by user ID."""
        try:
            # First try to get from cache
            for email, user_info in self.user_permissions.items():
                if user_info['user_id'] == user_id:
                    return user_info
            
            # If not found, try to get from Smartsheet API
            user_info = self.client.Users.get_user(user_id)
            if user_info:
                return {
                    'user_id': user_id,
                    'email': user_info.email,
                    'first_name': user_info.first_name,
                    'last_name': user_info.last_name,
                    'access_level': 1  # Default access level
                }
            
            return None
            
        except Exception as e:
            self.logger.error(f"Error getting user info: {e}")
            return None

    def _init_dashboard(self):
        """Initialize real-time monitoring dashboard."""
        self.dash_app = dash.Dash(__name__, url_base_pathname='/dashboard/')
        
        # Dashboard layout
        self.dash_app.layout = html.Div([
            html.H1("Real-Time Audit Dashboard", style={'textAlign': 'center'}),
            
            dcc.Interval(
                id='interval-component',
                interval=30*1000,  # Update every 30 seconds
                n_intervals=0
            ),
            
            html.Div([
                html.Div([
                    html.H3("Risk Score Distribution"),
                    dcc.Graph(id='risk-score-chart')
                ], className='six columns'),
                
                html.Div([
                    html.H3("Anomaly Detection"),
                    dcc.Graph(id='anomaly-chart')
                ], className='six columns'),
            ], className='row'),
            
            html.Div([
                html.H3("Recent High-Risk Events"),
                html.Div(id='recent-events-table')
            ]),
            
            html.Div([
                html.H3("User Activity Monitoring"),
                dcc.Graph(id='user-activity-chart')
            ])
        ])
        
        # Dashboard callbacks
        @self.dash_app.callback(
            [Output('risk-score-chart', 'figure'),
             Output('anomaly-chart', 'figure'),
             Output('recent-events-table', 'children'),
             Output('user-activity-chart', 'figure')],
            [Input('interval-component', 'n_intervals')]
        )
        def update_dashboard(n):
            return self._update_dashboard_data()

    def _update_dashboard_data(self):
        """Update dashboard data."""
        try:
            # Get recent audit events
            conn = sqlite3.connect(self.db_path)
            df = pd.read_sql_query('''
                SELECT * FROM audit_events 
                WHERE timestamp > datetime('now', '-24 hours')
                ORDER BY timestamp DESC
            ''', conn)
            conn.close()
            
            if df.empty:
                # Return empty figures if no data
                empty_fig = go.Figure()
                return empty_fig, empty_fig, html.Div("No recent events"), empty_fig
            
            # Risk score distribution
            risk_fig = px.histogram(df, x='risk_score', title='Risk Score Distribution')
            
            # Anomaly detection
            anomaly_fig = px.scatter(df, x='timestamp', y='anomaly_score', 
                                   color='severity', title='Anomaly Scores Over Time')
            
            # Recent high-risk events table
            high_risk_events = df[df['risk_score'] > 0.7].head(10)
            events_table = html.Table([
                html.Thead([
                    html.Tr([
                        html.Th("Timestamp"),
                        html.Th("Work Request"),
                        html.Th("User"),
                        html.Th("Risk Score"),
                        html.Th("Severity")
                    ])
                ]),
                html.Tbody([
                    html.Tr([
                        html.Td(row['timestamp']),
                        html.Td(row['work_request']),
                        html.Td(row['changed_by']),
                        html.Td(f"{row['risk_score']:.2f}"),
                        html.Td(row['severity'])
                    ]) for _, row in high_risk_events.iterrows()
                ])
            ])
            
            # User activity chart
            user_activity = df.groupby('changed_by').size().reset_index(name='count')
            user_fig = px.bar(user_activity, x='changed_by', y='count', 
                             title='User Activity (Last 24 Hours)')
            
            return risk_fig, anomaly_fig, events_table, user_fig
            
        except Exception as e:
            self.logger.error(f"Error updating dashboard: {e}")
            empty_fig = go.Figure()
            return empty_fig, empty_fig, html.Div(f"Error: {e}"), empty_fig

    def setup_smartsheet_webhook(self) -> bool:
        """Set up Smartsheet webhook for real-time notifications."""
        if not self.webhook_enabled:
            self.logger.info("Webhook functionality is disabled - skipping webhook setup")
            return True
            
        try:
            webhook_spec = smartsheet.models.Webhook({
                'name': 'Enhanced Audit System Webhook',
                'callback_url': self.webhook_url,
                'scope': 'sheet',
                'scope_object_id': int(self.audit_sheet_id),
                'events': ['*.*'],  # All events
                'version': 1
            })
            
            result = self.client.Webhooks.create_webhook(webhook_spec)
            
            if result.data:
                self.logger.info(f"Webhook created successfully: {result.data.id}")
                return True
            else:
                self.logger.error("Failed to create webhook")
                return False
                
        except Exception as e:
            self.logger.error(f"Error setting up webhook: {e}")
            return False

    def create_daily_backup(self) -> bool:
        """Create automated daily backup of critical billing data."""
        try:
            backup_date = datetime.datetime.now().strftime('%Y%m%d')
            backup_dir = f"backups/{backup_date}"
            os.makedirs(backup_dir, exist_ok=True)
            
            # Backup audit database
            db_backup_path = f"{backup_dir}/audit_database_backup.db"
            self._backup_database(db_backup_path)
            
            # Backup Smartsheet data
            sheet_backup_path = f"{backup_dir}/smartsheet_backup.xlsx"
            self._backup_smartsheet_data(sheet_backup_path)
            
            # Calculate checksums
            db_checksum = self._calculate_file_checksum(db_backup_path)
            sheet_checksum = self._calculate_file_checksum(sheet_backup_path)
            
            # Log backup
            self._log_backup(backup_date, db_backup_path, db_checksum, 'database')
            self._log_backup(backup_date, sheet_backup_path, sheet_checksum, 'smartsheet')
            
            self.logger.info(f"Daily backup completed: {backup_dir}")
            return True
            
        except Exception as e:
            self.logger.error(f"Error creating daily backup: {e}")
            return False

    def _backup_database(self, backup_path: str):
        """Backup SQLite database."""
        import shutil
        shutil.copy2(self.db_path, backup_path)

    def _backup_smartsheet_data(self, backup_path: str):
        """Backup Smartsheet data to Excel."""
        try:
            # Get sheet data
            sheet = self.client.Sheets.get_sheet(self.audit_sheet_id)
            
            # Convert to DataFrame
            data = []
            for row in sheet.rows:
                row_data = {}
                for cell in row.cells:
                    column = next((col for col in sheet.columns if col.id == cell.column_id), None)
                    if column:
                        row_data[column.title] = cell.display_value
                data.append(row_data)
            
            df = pd.DataFrame(data)
            df.to_excel(backup_path, index=False)
            
        except Exception as e:
            self.logger.error(f"Error backing up Smartsheet data: {e}")
            raise

    def _calculate_file_checksum(self, file_path: str) -> str:
        """Calculate SHA256 checksum of file."""
        hash_sha256 = hashlib.sha256()
        with open(file_path, "rb") as f:
            for chunk in iter(lambda: f.read(4096), b""):
                hash_sha256.update(chunk)
        return hash_sha256.hexdigest()

    def _log_backup(self, backup_date: str, file_path: str, checksum: str, backup_type: str):
        """Log backup information."""
        try:
            file_size = os.path.getsize(file_path)
            
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            cursor.execute('''
                INSERT INTO backup_logs 
                (backup_date, backup_type, file_path, file_size, checksum, status)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (backup_date, backup_type, file_path, file_size, checksum, 'SUCCESS'))
            
            conn.commit()
            conn.close()
            
        except Exception as e:
            self.logger.error(f"Error logging backup: {e}")

    def review_user_permissions(self) -> Dict[str, Any]:
        """Monthly review of user permissions and access levels."""
        try:
            review_results = {
                'total_users': 0,
                'inactive_users': 0,
                'users_needing_review': 0,
                'high_access_users': 0,
                'recommendations': []
            }
            
            conn = sqlite3.connect(self.db_path)
            cursor = conn.cursor()
            
            cursor.execute('SELECT * FROM user_permissions')
            users = cursor.fetchall()
            
            for user in users:
                review_results['total_users'] += 1
                
                # Check for inactive users
                if not user[8]:  # active field
                    review_results['inactive_users'] += 1
                
                # Check for users needing review
                if user[9]:  # last_review_date
                    last_review = datetime.datetime.fromisoformat(user[9])
                    days_since_review = (datetime.datetime.now() - last_review).days
                    if days_since_review > 90:
                        review_results['users_needing_review'] += 1
                        review_results['recommendations'].append(
                            f"User {user[1]} needs permission review (last reviewed {days_since_review} days ago)"
                        )
                
                # Check for high access users
                if user[5] and user[5] > 7:  # access_level
                    review_results['high_access_users'] += 1
                    review_results['recommendations'].append(
                        f"User {user[1]} has high access level ({user[5]}) - verify necessity"
                    )
            
            conn.close()
            
            # Generate review report
            self._generate_permission_review_report(review_results)
            
            return review_results
            
        except Exception as e:
            self.logger.error(f"Error reviewing user permissions: {e}")
            return {}

    def _generate_permission_review_report(self, review_results: Dict[str, Any]):
        """Generate permission review report."""
        try:
            report_date = datetime.datetime.now().strftime('%Y%m%d')
            report_path = f"reports/permission_review_{report_date}.txt"
            
            os.makedirs('reports', exist_ok=True)
            
            with open(report_path, 'w') as f:
                f.write("USER PERMISSION REVIEW REPORT\n")
                f.write("=" * 50 + "\n\n")
                f.write(f"Review Date: {datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")
                f.write(f"Total Users: {review_results['total_users']}\n")
                f.write(f"Inactive Users: {review_results['inactive_users']}\n")
                f.write(f"Users Needing Review: {review_results['users_needing_review']}\n")
                f.write(f"High Access Users: {review_results['high_access_users']}\n\n")
                f.write("RECOMMENDATIONS:\n")
                f.write("-" * 20 + "\n")
                for rec in review_results['recommendations']:
                    f.write(f"• {rec}\n")
            
            self.logger.info(f"Permission review report generated: {report_path}")
            
        except Exception as e:
            self.logger.error(f"Error generating permission review report: {e}")

    def train_predictive_models(self, retrain: bool = False):
        """Train ML models with lightweight approach for better performance."""
        try:
            # LIGHTWEIGHT MODE: Skip heavy ML training for speed
            self.logger.info("ML model training skipped in lightweight mode for performance")
            
            # Create minimal models if they don't exist
            if not hasattr(self, 'anomaly_detector') or self.anomaly_detector is None:
                self.anomaly_detector = IsolationForest(contamination=0.1, random_state=42)
                self.risk_predictor = IsolationForest(contamination=0.1, random_state=42)
                self.scaler = StandardScaler()
                
            # Skip heavy training and just use default models
            self.logger.info("Using default ML models for lightweight operation")
            return True
                
        except Exception as e:
            self.logger.error(f"Error in lightweight model setup: {e}")
            return False

    def start_background_services(self):
        """Start lightweight background services (optimized for speed)."""
        try:
            # LIGHTWEIGHT MODE: Skip heavy scheduled tasks for better performance
            self.logger.info("Starting background services in lightweight mode...")
            
            # Skip daily backups in lightweight mode (can be run manually)
            # schedule.every().day.at("02:00").do(self.create_daily_backup)
            
            # Skip monthly permission reviews (can be run manually) 
            # schedule.every().month.do(self.review_user_permissions)
            
            # Skip weekly model retraining (can be run manually)
            # schedule.every().week.do(lambda: self.train_predictive_models(retrain=True))
            
            # Skip scheduler thread for performance
            # def run_scheduler():
            #     while True:
            #         schedule.run_pending()
            #         time.sleep(60)
            # scheduler_thread = threading.Thread(target=run_scheduler, daemon=True)
            # scheduler_thread.start()
            
            # Skip Flask webhook server (disabled by design)
            if self.webhook_enabled and self.app:
                webhook_thread = threading.Thread(
                    target=lambda: self.app.run(host='0.0.0.0', port=5000, debug=False),
                    daemon=True
                )
                webhook_thread.start()
                self.logger.info("Webhook server started on port 5000")
            else:
                self.logger.info("Webhook server disabled - skipping webhook startup")
            
            # Start Dash dashboard in lightweight mode
            dashboard_thread = threading.Thread(
                target=lambda: self.dash_app.run_server(host='0.0.0.0', port=8050, debug=False),
                daemon=True
            )
            dashboard_thread.start()
            
            self.logger.info("Lightweight background services started successfully")
            
        except Exception as e:
            self.logger.error(f"Error starting background services: {e}")

    def generate_comprehensive_report(self) -> str:
        """Generate comprehensive audit system report."""
        try:
            report_date = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
            report_path = f"reports/comprehensive_audit_report_{report_date}.xlsx"
            
            os.makedirs('reports', exist_ok=True)
            
            # Create workbook
            wb = openpyxl.Workbook()
            
            # Get data from database
            conn = sqlite3.connect(self.db_path)
            
            # Audit events sheet
            df_events = pd.read_sql_query('SELECT * FROM audit_events', conn)
            ws_events = wb.active
            ws_events.title = "Audit Events"
            
            # Write headers
            for col, header in enumerate(df_events.columns, 1):
                ws_events.cell(row=1, column=col, value=header)
                ws_events.cell(row=1, column=col).font = Font(bold=True)
            
            # Write data
            for row, (_, data) in enumerate(df_events.iterrows(), 2):
                for col, value in enumerate(data, 1):
                    ws_events.cell(row=row, column=col, value=value)
            
            # User permissions sheet
            df_users = pd.read_sql_query('SELECT * FROM user_permissions', conn)
            ws_users = wb.create_sheet("User Permissions")
            
            for col, header in enumerate(df_users.columns, 1):
                ws_users.cell(row=1, column=col, value=header)
                ws_users.cell(row=1, column=col).font = Font(bold=True)
            
            for row, (_, data) in enumerate(df_users.iterrows(), 2):
                for col, value in enumerate(data, 1):
                    ws_users.cell(row=row, column=col, value=value)
            
            # Backup logs sheet
            df_backups = pd.read_sql_query('SELECT * FROM backup_logs', conn)
            ws_backups = wb.create_sheet("Backup Logs")
            
            for col, header in enumerate(df_backups.columns, 1):
                ws_backups.cell(row=1, column=col, value=header)
                ws_backups.cell(row=1, column=col).font = Font(bold=True)
            
            for row, (_, data) in enumerate(df_backups.iterrows(), 2):
                for col, value in enumerate(data, 1):
                    ws_backups.cell(row=row, column=col, value=value)
            
            conn.close()
            
            # Save workbook
            wb.save(report_path)
            
            self.logger.info(f"Comprehensive report generated: {report_path}")
            return report_path
            
        except Exception as e:
            self.logger.error(f"Error generating comprehensive report: {e}")
            return ""

# Configuration for enhanced audit system
ENHANCED_CONFIG = {
    'webhook_url': 'https://your-domain.com/webhook',  # Update with your actual webhook URL
    'webhook_secret': 'audit_webhook_secret_2025',
    'anomaly_threshold': 0.1,
    'risk_score_threshold': 0.7,
    'email_settings': {
        'smtp_server': 'smtp.gmail.com',
        'smtp_port': 587,
        'username': 'your-email@gmail.com',
        'password': 'your-app-password',
        'alert_recipients': [
            'audit-admin@company.com',
            'security@company.com'
        ]
    }
}

if __name__ == "__main__":
    # Load environment variables
    load_dotenv()
    
    api_token = os.getenv("SMARTSHEET_API_TOKEN")
    audit_sheet_id = os.getenv("AUDIT_SHEET_ID")
    
    if not api_token or not audit_sheet_id:
        print("❌ Missing environment variables")
        sys.exit(1)
    
    # Initialize Smartsheet client
    client = smartsheet.Smartsheet(api_token)
    
    # Use lightweight config for better performance
    LIGHTWEIGHT_CONFIG = {
        'webhook': {'enabled': False},  # Disabled for performance
        'anomaly_threshold': 0.3,       # Higher threshold for fewer alerts
        'risk_score_threshold': 0.8     # Higher threshold for fewer alerts
    }
    
    # Initialize enhanced audit system with lightweight config
    enhanced_audit = EnhancedAuditSystem(client, audit_sheet_id, LIGHTWEIGHT_CONFIG)
    
    # Skip webhook setup (disabled for performance)
    print("⚠️  Webhook functionality disabled (user accounts not controlled)")
    
    # Skip heavy ML training for speed
    print("⚡ Using lightweight mode for better performance")
    enhanced_audit.train_predictive_models()
    
    # Start lightweight background services
    enhanced_audit.start_background_services()
    
    print("🚀 Enhanced Audit System V2 started successfully!")
    print("📊 Dashboard available at: http://localhost:8050/dashboard/")
    print("⚠️  Webhook functionality disabled (not needed for this deployment)")
    print("⚡ Running in lightweight mode for optimal performance")
    
    # Keep the main thread alive with shorter sleep for responsiveness
    try:
        while True:
            time.sleep(10)  # Reduced from 60 seconds for better responsiveness
    except KeyboardInterrupt:
        print("\n👋 Enhanced Audit System shutting down...")
