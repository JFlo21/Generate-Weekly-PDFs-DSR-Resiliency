#!/usr/bin/env python3
"""
Analyze specific Excel file: WR_90124277_WeekEnding_092825_120044_a8b10c0045369976.xlsx

This script will read the Excel file and analyze the total calculations to identify
why the expected total of $7,094.58 is not showing correctly.
"""

import openpyxl
import os
from generate_weekly_pdfs import parse_price

def analyze_excel_file(filename="WR_90124277_WeekEnding_092825_120044_a8b10c0045369976.xlsx", expected_total=7094.58):
    """Analyze the specific Excel file for total calculation issues."""
    
    print("📊 EXCEL FILE ANALYSIS")
    print("=" * 80)
    print(f"File: {filename}")
    print(f"Expected Total: ${expected_total:,.2f}")
    
    file_path = f"/workspaces/Generate-Weekly-PDFs-DSR-Resiliency/{filename}"
    
    if not os.path.exists(file_path):
        print(f"❌ File not found: {file_path}")
        return
    
    try:
        # Load the Excel workbook
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        
        print(f"\n📋 WORKBOOK INFO:")
        print(f"   Worksheet name: {ws.title}")
        print(f"   Max row: {ws.max_row}")
        print(f"   Max column: {ws.max_column}")
        
        # Find total billed amount in the summary section
        summary_total = None
        summary_location = None
        
        # Look for "Total Billed Amount" in the worksheet
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value and isinstance(cell.value, str):
                    if "Total Billed Amount" in cell.value:
                        # The total should be in the next column
                        total_cell = ws.cell(row=row, column=col + 1)
                        summary_total = total_cell.value
                        summary_location = f"{total_cell.coordinate}"
                        print(f"\n💰 FOUND SUMMARY TOTAL:")
                        print(f"   Location: {summary_location}")
                        print(f"   Value: {summary_total}")
                        print(f"   Type: {type(summary_total)}")
                        break
        
        # Analyze all pricing data in the worksheet
        print(f"\n🔍 SCANNING ALL PRICING DATA:")
        
        pricing_cells = []
        daily_totals = []
        
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                
                # Look for cells that might contain pricing
                if cell.value is not None:
                    # Check if it's a number that could be a price
                    if isinstance(cell.value, (int, float)) and cell.value > 0:
                        # Check if the cell is formatted as currency or looks like a price
                        if (cell.number_format and ('$' in cell.number_format or 
                            'CURRENCY' in cell.number_format.upper() or
                            '_($' in cell.number_format)):
                            pricing_cells.append((cell.coordinate, cell.value, cell.number_format))
                    
                    # Also check for TOTAL labels and their adjacent values
                    elif isinstance(cell.value, str) and "TOTAL" in cell.value.upper():
                        # Check adjacent cells for totals
                        for offset_col in range(1, 4):  # Check next 3 columns
                            adj_cell = ws.cell(row=row, column=col + offset_col)
                            if isinstance(adj_cell.value, (int, float)) and adj_cell.value > 0:
                                daily_totals.append((adj_cell.coordinate, adj_cell.value, f"Adjacent to {cell.value}"))
        
        print(f"\n💰 PRICING CELLS FOUND ({len(pricing_cells)} cells):")
        total_from_cells = 0.0
        for coord, value, format_str in pricing_cells[:20]:  # Show first 20
            print(f"   {coord}: ${value:,.2f} (format: {format_str})")
            total_from_cells += value
        
        if len(pricing_cells) > 20:
            print(f"   ... and {len(pricing_cells) - 20} more cells")
            for coord, value, format_str in pricing_cells[20:]:
                total_from_cells += value
        
        print(f"\n📊 DAILY TOTALS FOUND ({len(daily_totals)} totals):")
        total_from_daily = 0.0
        for coord, value, description in daily_totals:
            print(f"   {coord}: ${value:,.2f} ({description})")
            total_from_daily += value
        
        print(f"\n🧮 CALCULATION ANALYSIS:")
        print(f"   Summary total (from report): ${summary_total if summary_total else 'NOT FOUND'}")
        print(f"   Sum of all pricing cells: ${total_from_cells:,.2f}")
        print(f"   Sum of daily totals: ${total_from_daily:,.2f}")
        print(f"   Expected total: ${expected_total:,.2f}")
        
        # Compare with expected
        if summary_total:
            diff_summary = abs(float(summary_total) - expected_total)
            print(f"\n🎯 COMPARISON WITH EXPECTED:")
            print(f"   Summary vs Expected: ${diff_summary:.2f} difference")
            
            if diff_summary < 0.01:
                print("   ✅ EXACT MATCH with summary total!")
            elif diff_summary < expected_total * 0.05:
                print("   ⚠️ CLOSE MATCH with summary total (within 5%)")
            else:
                print("   ❌ SIGNIFICANT DIFFERENCE from summary total")
        
        diff_cells = abs(total_from_cells - expected_total)
        print(f"   All cells vs Expected: ${diff_cells:.2f} difference")
        
        # Detailed cell analysis for discrepancies
        if diff_summary and diff_summary > 0.01:
            print(f"\n🔍 DETAILED INVESTIGATION:")
            analyze_cell_details(ws, expected_total)
    
    except Exception as e:
        print(f"❌ Error analyzing Excel file: {e}")
        import traceback
        traceback.print_exc()

def analyze_cell_details(ws, expected_total):
    """Perform detailed analysis of Excel cells to find discrepancies."""
    
    print("Scanning worksheet structure for pricing patterns...")
    
    # Look for specific patterns in the Excel structure
    row_data = []
    
    for row in range(1, min(ws.max_row + 1, 200)):  # Limit to first 200 rows
        row_content = []
        has_price = False
        
        for col in range(1, min(ws.max_column + 1, 10)):  # Check first 10 columns
            cell = ws.cell(row=row, column=col)
            if cell.value is not None:
                row_content.append(str(cell.value))
                
                # Check if this looks like a price cell
                if isinstance(cell.value, (int, float)) and cell.value > 0:
                    if (cell.number_format and ('$' in cell.number_format or 
                        'CURRENCY' in cell.number_format.upper())):
                        has_price = True
        
        if has_price and len(row_content) > 0:
            row_data.append((row, " | ".join(row_content[:8])))  # First 8 columns
    
    print(f"\n📝 ROWS WITH PRICING DATA (showing first 10):")
    for row_num, content in row_data[:10]:
        print(f"   Row {row_num:3d}: {content}")
    
    if len(row_data) > 10:
        print(f"   ... and {len(row_data) - 10} more rows with pricing")

def analyze_work_request_data():
    """Analyze the work request data that should have been used to generate this Excel."""
    
    print(f"\n🔍 SOURCE DATA ANALYSIS:")
    print("Analyzing what data should be in WR 90124277 for week ending 09/28/25...")
    
    # Extract info from filename
    filename = "WR_90124277_WeekEnding_092825_120044_a8b10c0045369976.xlsx"
    parts = filename.replace('.xlsx', '').split('_')
    
    if len(parts) >= 4:
        wr_num = parts[1]
        week_ending = parts[3]
        timestamp = parts[4] if len(parts) > 4 else "Unknown"
        file_hash = parts[5] if len(parts) > 5 else "Unknown"
        
        print(f"   Work Request #: {wr_num}")
        print(f"   Week Ending: {week_ending} (09/28/25)")
        print(f"   Generated at: {timestamp}")
        print(f"   Data Hash: {file_hash}")
        
        print(f"\n💡 RECOMMENDED ACTIONS:")
        print(f"   1. Run diagnostic on source data for WR {wr_num}")
        print(f"   2. Check if all completed rows for week ending 09/28/25 are included")
        print(f"   3. Verify pricing data hasn't changed since file generation")
        print(f"   4. Check for any $0.00 rows that might be excluded")

if __name__ == "__main__":
    # Analyze the specific file
    analyze_excel_file()
    
    # Also analyze what the source data should contain
    analyze_work_request_data()