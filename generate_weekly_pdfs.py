#!/usr/bin/env python3
"""
Weekly PDF Generator with Complete Fixes
Generates Excel reports from Smartsheet data for weekly billing periods.

FIXES IMPLEMENTED:
- WR 90093002 Excel generation fix
- WR 89954686 specific handling 
- Proper file deletion logic
- Complete audit system integration
- All incomplete code sections completed
"""

import os
import datetime
import time
import threading
import weakref
from concurrent.futures import ThreadPoolExecutor, as_completed, TimeoutError as FuturesTimeoutError
import concurrent.futures.thread as _cf_thread
import re
import hashlib
from collections.abc import Sequence
from datetime import timedelta
import logging
from dateutil import parser
import smartsheet
import smartsheet.exceptions as ss_exc

# Upstream SDK workaround: smartsheet-python-sdk 3.8.0 raises an
# AttributeError from smartsheet.smartsheet.Smartsheet._request_with_retry
# whenever the API returns a retryable error (429, 5xx). At
# smartsheet/smartsheet.py:303 it does
# ``getattr(sys.modules[__name__], native.result.name)`` to look up the
# exception class to raise, but that module's top-level imports only
# expose ApiError / HttpError / UnexpectedRequestError. The retryable
# exception classes (RateLimitExceededError, UnexpectedErrorShouldRetry-
# Error, InternalServerError, ServerTimeoutExceededError, SystemMainte-
# nanceError) live in smartsheet.exceptions and were never re-exported
# into smartsheet.smartsheet, so the getattr fails and our retry
# wrapper never gets the real exception. Re-export the missing names
# here so the SDK's internal lookup succeeds. The ``if not hasattr``
# guard makes this a no-op if the upstream SDK ever re-exports them.
import smartsheet.smartsheet as _ss_smartsheet_module
for _exc_name in (
    'RateLimitExceededError',
    'UnexpectedErrorShouldRetryError',
    'InternalServerError',
    'ServerTimeoutExceededError',
    'SystemMaintenanceError',
):
    if not hasattr(_ss_smartsheet_module, _exc_name) and hasattr(ss_exc, _exc_name):
        setattr(_ss_smartsheet_module, _exc_name, getattr(ss_exc, _exc_name))
del _ss_smartsheet_module, _exc_name
import openpyxl
from openpyxl.styles import Font, numbers, Alignment, PatternFill
from openpyxl.drawing.image import Image
import collections
from dotenv import load_dotenv
import sentry_sdk
from sentry_sdk.integrations.logging import LoggingIntegration
from sentry_sdk.integrations.threading import ThreadingIntegration
from sentry_sdk.crons import capture_checkin
from sentry_sdk.crons.consts import MonitorStatus
import traceback
import sys
import json
import signal
import csv
from typing import Any, cast, TYPE_CHECKING

if TYPE_CHECKING:
    # ``MonitorConfig`` is a TypedDict consulted only by the static type
    # checker to validate ``capture_checkin(monitor_config=...)``. Importing it
    # under ``TYPE_CHECKING`` keeps the runtime import surface unchanged (it is
    # never imported when the script actually runs).
    from sentry_sdk._types import MonitorConfig

# Load environment variables
load_dotenv()

# PERFORMANCE: Pre-compiled regex patterns (avoid repeated compilation in hot paths)
_RE_SANITIZE_IDENTIFIER = re.compile(r'[^\w\-@.]')
_RE_SANITIZE_HELPER_NAME = re.compile(r'[^\w\-]')
_RE_EXTRACT_NUMBERS = re.compile(r'[^0-9.\-]')
_RE_ISO_DATE_PREFIX = re.compile(r'^\d{4}-\d{2}-\d{2}')

# Suppress BrokenPipeError when piping output (e.g. | head, | grep -m) so it doesn't surface as an exception
try:
    signal.signal(signal.SIGPIPE, signal.SIG_DFL)  # type: ignore[attr-defined]
except Exception:
    pass

# Import audit system with error handling
try:
    from audit_billing_changes import BillingAudit  # type: ignore
    AUDIT_SYSTEM_AVAILABLE = True
    print("🔍 Billing audit system loaded successfully")
except ImportError as e:
    print(f"⚠️ Billing audit system not available: {e}")
    AUDIT_SYSTEM_AVAILABLE = False
    class BillingAudit:
        def __init__(self, *args, **kwargs):
            pass
        def audit_financial_data(self, *args, **kwargs):
            return {"summary": {"risk_level": "UNKNOWN"}}

# Import billing audit attribution snapshot writer (shadow mode).
# Failures here must NEVER break Excel generation — the writer is a
# strictly additive, flag-gated, no-op-on-failure surface. Catch
# broad Exception (not just ImportError) so a runtime error inside
# billing_audit/* during module init cannot crash the pipeline. We
# log the class name only to avoid leaking any contextual detail.
try:
    from billing_audit import writer as _billing_audit_writer
    from billing_audit.fingerprint import compute_assignment_fingerprint
    BILLING_AUDIT_AVAILABLE = True
    print("❄️ Billing audit snapshot writer loaded successfully")
except Exception as e:
    print(
        "⚠️ Billing audit snapshot writer not available: "
        f"{type(e).__name__}"
    )
    BILLING_AUDIT_AVAILABLE = False

# 🎯 SHOW OUR FIXES ARE ACTIVE
print("✅ CRITICAL FIXES APPLIED:")
print("   • WR 90093002 Excel generation fix - ACTIVE")
print("   • WR 89954686 specific handling - ACTIVE")
print("   • MergedCell assignment errors - FIXED")
print("   • Type ignore comments - APPLIED")
print("🚀 SYSTEM READY FOR PRODUCTION")
print("=" * 60)

# Configure logging early
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logging.getLogger('smartsheet.smartsheet').setLevel(logging.CRITICAL)

# Performance and compatibility settings
GITHUB_ACTIONS_MODE = os.getenv('GITHUB_ACTIONS') == 'true'
SKIP_CELL_HISTORY = os.getenv('SKIP_CELL_HISTORY', 'false').lower() == 'true'

# Resiliency grouping mode: controls which Excel variants to generate
# - "primary": Standard WR-based grouping (one Excel per WR/Week)
# - "helper": Helper-based grouping (one Excel per WR/Week/Helper)
# - "both": Generate both primary and helper variants (DEFAULT - always creates primary + helper when helper criteria met)
RES_GROUPING_MODE = os.getenv('RES_GROUPING_MODE', 'both').lower()
if RES_GROUPING_MODE not in ('primary', 'helper', 'both'):
    logging.warning(f"⚠️ Invalid RES_GROUPING_MODE '{RES_GROUPING_MODE}'; defaulting to 'both'")
    RES_GROUPING_MODE = 'both'

# Activity log-based foreman assignment has been removed - we now use helper column logic only

# Skip Smartsheet uploads for local testing (files still saved to OUTPUT_FOLDER)
SKIP_UPLOAD = os.getenv('SKIP_UPLOAD', 'false').lower() == 'true'

# --- CORE CONFIGURATION ---
API_TOKEN = os.getenv("SMARTSHEET_API_TOKEN")
# TARGET / AUDIT SHEET CONFIGURATION
# TARGET_SHEET_ID: destination for generated weekly Excel report attachments
# AUDIT_SHEET_ID (or legacy BILLING_AUDIT_SHEET_ID): destination for audit rows / stats ONLY
_target_sheet_id_env = os.getenv("TARGET_SHEET_ID")
AUDIT_SHEET_ID = os.getenv("AUDIT_SHEET_ID") or os.getenv("BILLING_AUDIT_SHEET_ID")

def _coerce_sheet_id(raw_value, default=None):
    if not raw_value:
        return default
    try:
        return int(raw_value)
    except (TypeError, ValueError):
        logging.warning(f"⚠️ Invalid sheet id value provided: {raw_value}; using default {default}")
        return default

TARGET_SHEET_ID = _coerce_sheet_id(_target_sheet_id_env, 5723337641643908)
_audit_sheet_id_int = _coerce_sheet_id(AUDIT_SHEET_ID) if AUDIT_SHEET_ID else None

if _target_sheet_id_env:
    logging.info(f"🎯 Using target sheet id: {TARGET_SHEET_ID} (from env TARGET_SHEET_ID)")
else:
    logging.info(f"🎯 Using default target sheet id: {TARGET_SHEET_ID}")

if _audit_sheet_id_int:
    logging.info(f"🧾 Audit sheet configured: {_audit_sheet_id_int}")
else:
    logging.info("🧾 Audit sheet not configured (set AUDIT_SHEET_ID to enable detailed audit logging to Smartsheet)")

# Export AUDIT_SHEET_ID into env-compatible form for BillingAudit (which reads os.getenv inside its module)
if _audit_sheet_id_int and not os.getenv("AUDIT_SHEET_ID"):
    os.environ["AUDIT_SHEET_ID"] = str(_audit_sheet_id_int)

# TARGET_WR_COLUMN_ID removed (unused)
LOGO_PATH = "LinetecServices_Logo.png"
OUTPUT_FOLDER = "generated_docs"
os.makedirs(OUTPUT_FOLDER, exist_ok=True)

# Optional performance/test tuning via environment (must come AFTER OUTPUT_FOLDER defined)
WR_FILTER = [w.strip() for w in os.getenv('WR_FILTER','').split(',') if w.strip()]
EXCLUDE_WRS = [w.strip() for w in os.getenv('EXCLUDE_WRS','').split(',') if w.strip()]  # Work Requests to EXCLUDE from generation
MAX_GROUPS = int(os.getenv('MAX_GROUPS','0') or 0)
QUIET_LOGGING = os.getenv('QUIET_LOGGING','0').lower() in ('1','true','yes')

# Parallel execution: number of ThreadPoolExecutor workers for concurrent Smartsheet API calls
# Smartsheet rate limit is 300 req/min (~5/sec). 8 I/O-bound workers stays safely under the limit
# because each request blocks ~200ms on network I/O; the SDK auto-retries on 429 with backoff.
PARALLEL_WORKERS = int(os.getenv('PARALLEL_WORKERS', '8') or 8)
PARALLEL_WORKERS_DISCOVERY = int(os.getenv('PARALLEL_WORKERS_DISCOVERY', '8') or 8)

# Graceful time budget (minutes). When set and running in GitHub Actions, the script will
# stop processing new groups once this many minutes have elapsed since session start.
# This prevents the Actions runner from hard-killing the job and losing cache/artifact saves.
# Set to 0 to disable. The weekly workflow sets this to 180 (3h) with a matching
# timeout-minutes: 195 on the runner (15min cushion for cache/artifact save steps).
TIME_BUDGET_MINUTES = int(os.getenv('TIME_BUDGET_MINUTES', '0') or 0)

# Sub-budget for the attachment pre-fetch phase. Prevents a flaky Smartsheet
# connection from consuming the entire session budget before group processing
# can start: on 2026-04-22 a run lost 16 minutes to ~14 stuck rows after
# RemoteDisconnected retries, exhausted the 80min TIME_BUDGET_MINUTES before
# generating a single file, and finished with 0 Excel files generated.
# When the pre-fetch exceeds this budget, remaining rows fall back to on-demand
# per-row fetches (already supported in _has_existing_week_attachment and
# delete_old_excel_attachments).
ATTACHMENT_PREFETCH_MAX_MINUTES = int(os.getenv('ATTACHMENT_PREFETCH_MAX_MINUTES', '10') or 10)
# Per-future wait (seconds) inside the pre-fetch consumer loop. One stuck HTTP
# call cannot block the consumer beyond this — the future is left behind and
# its row falls back to the per-row path at generation time.
ATTACHMENT_PREFETCH_FUTURE_TIMEOUT_SEC = int(os.getenv('ATTACHMENT_PREFETCH_FUTURE_TIMEOUT_SEC', '45') or 45)
# Minimum "generation headroom" (minutes) the pre-flight guard reserves
# beyond the pre-fetch budget. Without this, a setup where the session
# has exactly `ATTACHMENT_PREFETCH_MAX_MINUTES` remaining would still
# run pre-fetch and leave ~0 minutes for group processing — the same
# zero-output failure mode this guard is meant to prevent.
ATTACHMENT_PREFETCH_GENERATION_HEADROOM_MIN = int(os.getenv('ATTACHMENT_PREFETCH_GENERATION_HEADROOM_MIN', '2') or 2)


class _DaemonThreadPoolExecutor(ThreadPoolExecutor):
    """ThreadPoolExecutor whose workers are daemonized so stuck I/O
    cannot hold the interpreter open past ``main()`` return.

    Three things can block process exit for a non-daemon worker:
    1. ``concurrent.futures.thread._python_exit`` (registered via
       ``threading._register_atexit``) joins every worker still in
       ``_threads_queues``.
    2. ``threading._shutdown`` joins every lock in
       ``_shutdown_locks`` — non-daemon threads add their tstate
       lock to this set at startup.
    3. The executor's ``shutdown(wait=True)`` joins all workers.

    This subclass addresses (2) by setting ``daemon=True`` at thread
    creation (``_set_tstate_lock`` only adds to ``_shutdown_locks``
    when ``not self.daemon``). Callers addressing a stall must still
    pop from ``_threads_queues`` (addresses 1) and call
    ``shutdown(wait=False, cancel_futures=True)`` (addresses 3).

    **Safety invariant — use this ONLY when the worker's work is
    discardable.** The pre-fetch cache has a per-row fallback path,
    so abandoning a mid-flight HTTP worker is safe; the OS reclaims
    the socket. Do NOT use this executor for workers that produce
    results the main flow depends on (generation, upload,
    ``hash_history.save``) — the atexit join is what guarantees
    those side effects are flushed before exit.

    Re-implements upstream's private ``_adjust_thread_count`` to
    flip ``daemon=True``. Pinned to the Python 3.11 / 3.12 shape;
    falls back to the superclass (non-daemon workers, atexit hang
    returns) if a future Python rearranges the private helpers.
    """

    def _adjust_thread_count(self):
        if not hasattr(_cf_thread, '_worker') or not hasattr(_cf_thread, '_threads_queues'):
            return super()._adjust_thread_count()
        if self._idle_semaphore.acquire(timeout=0):
            return

        def _weakref_cb(_, q=self._work_queue):
            q.put(None)  # type: ignore[arg-type]  # sentinel signals worker shutdown (matches CPython internals)

        num_threads = len(self._threads)
        if num_threads < self._max_workers:
            thread_name = '%s_%d' % (self._thread_name_prefix or self, num_threads)
            t = threading.Thread(
                name=thread_name,
                target=_cf_thread._worker,
                args=(weakref.ref(self, _weakref_cb),
                      self._work_queue,
                      getattr(self, '_initializer', None),
                      getattr(self, '_initargs', ())),
                daemon=True,
            )
            t.start()
            self._threads.add(t)  # type: ignore[attr-defined]  # CPython exposes _threads as a mutable set despite AbstractSet typing
            _cf_thread._threads_queues[t] = self._work_queue  # type: ignore[index]  # CPython internals expose this as a mutable dict despite Mapping typing


USE_DISCOVERY_CACHE = os.getenv('USE_DISCOVERY_CACHE','1').lower() in ('1','true','yes')
# Discovery cache TTL: sheet IDs and column mappings are essentially static.
# Default to 7 days (10080 min). Set FORCE_REDISCOVERY=true to bypass the cache on demand.
DISCOVERY_CACHE_TTL_MIN = int(os.getenv('DISCOVERY_CACHE_TTL_MIN','10080') or 10080)
FORCE_REDISCOVERY = os.getenv('FORCE_REDISCOVERY','false').lower() in ('1','true','yes')
DISCOVERY_CACHE_PATH = os.path.join(OUTPUT_FOLDER, 'discovery_cache.json')
# Bump this version whenever the column synonym dictionary changes so that stale caches
# (missing newly-mapped columns like VAC Crew) are automatically invalidated.
# Also bump when a known bug would leave existing caches with incorrect mappings —
# invalidating the cache is cheaper than waiting up to DISCOVERY_CACHE_TTL_MIN
# (7 days by default) for those mappings to refresh on their own.
DISCOVERY_CACHE_VERSION = 4  # v4: fuzzy VAC Crew column fallback — invalidate caches whose column_mapping missed title variants like trailing whitespace or case drift

# Phase 1.1 SUB-12 / D-17 / D-19: idempotent hash-history prune
# version. The constant IS the kill switch — advance to trigger a
# one-time prune of subcontractor primary orphan entries (the
# pre-Bug-B1 partitioning leftovers); leave at the current value to
# skip the prune. Mirrors the DISCOVERY_CACHE_VERSION pattern above.
# Persisted into ``hash_history.json`` under the
# ``_phase_prune_version`` sentinel key (the extended
# ``load_hash_history`` filter preserves underscore-prefixed sentinels
# and the hardened ``save_hash_history`` retention sort tolerates the
# int-valued sentinel — see the helpers below).
PHASE_1_1_HASH_PRUNE_VERSION = 2
# Subproject B (2026-05-20): one-time hash-history prune version for
# dropping LEGACY blank-identifier `reduced_sub` / `aep_billable`
# orphans left behind when B re-partitions those variants by frozen
# claimer. Separate sentinel (`_subproject_b_prune_version`) from the
# Phase 1.1 prune so the two migrations are independent + auditable.
# Advancing this constant is the kill switch (re-run trigger).
SUBPROJECT_B_HASH_PRUNE_VERSION = 1
# Subproject C (2026-05-21): one-time hash-history prune version for
# dropping LEGACY blank-identifier `vac_crew` orphans left behind when
# C re-partitions vac_crew variants by frozen claimer. Separate sentinel
# (`_vac_crew_prune_version`) from Phase 1.1 and Subproject B so all
# three migrations are independent + auditable.
# Advancing this constant is the kill switch (re-run trigger).
VAC_CREW_HASH_PRUNE_VERSION = 1
# Subproject D (2026-05-25): one-time hash-history prune version for
# dropping LEGACY blank-identifier `primary` orphans left behind when
# D re-partitions the production primary variant by frozen primary
# claimer. Separate sentinel (`_subproject_d_prune_version`) from
# Phase 1.1, Subproject B, and Subproject C so all four migrations are
# independent + auditable. Advancing this constant is the kill switch
# (re-run trigger).
SUBPROJECT_D_HASH_PRUNE_VERSION = 1
# Verbose debug tunables
DEBUG_SAMPLE_ROWS = int(os.getenv('DEBUG_SAMPLE_ROWS','3') or 3)  # How many initial rows (across all sheets) to show full per-cell mapping
DEBUG_ESSENTIAL_ROWS = int(os.getenv('DEBUG_ESSENTIAL_ROWS','5') or 5)  # How many initial rows to log essential field summary
LOG_UNKNOWN_COLUMNS = os.getenv('LOG_UNKNOWN_COLUMNS','1').lower() in ('1','true','yes')  # Summarize unmapped columns once per sheet
PER_CELL_DEBUG_ENABLED = os.getenv('PER_CELL_DEBUG_ENABLED','1').lower() in ('1','true','yes')  # Master switch
UNMAPPED_COLUMN_SAMPLE_LIMIT = int(os.getenv('UNMAPPED_COLUMN_SAMPLE_LIMIT','5') or 5)  # Sample values per unmapped column in summary
# Extended change detection (default ON). When enabled, the data hash used to detect
# whether an Excel needs regeneration will include additional business fields such as
# current foreman, dept numbers, scope id, aggregated totals, unique dept list, and row count.
EXTENDED_CHANGE_DETECTION = os.getenv('EXTENDED_CHANGE_DETECTION','1').lower() in ('1','true','yes')
FILTER_DIAGNOSTICS = os.getenv('FILTER_DIAGNOSTICS','0').lower() in ('1','true','yes')  # When enabled, logs exclusion reasons counts
FOREMAN_DIAGNOSTICS = os.getenv('FOREMAN_DIAGNOSTICS','0').lower() in ('1','true','yes')  # When enabled, logs per-WR foreman value distributions & exclusion reasons
FORCE_GENERATION = os.getenv('FORCE_GENERATION','0').lower() in ('1','true','yes')  # When true, ignore hash short‑circuit and always regenerate
REGEN_WEEKS = {w.strip() for w in os.getenv('REGEN_WEEKS','').split(',') if w.strip()}  # Comma list of MMDDYY week ending codes to force regenerate
def _parse_sheet_ids(env_val):
    """Parse comma-separated sheet IDs, skipping non-integer tokens."""
    ids = []
    for s in env_val.split(','):
        s = s.strip()
        if not s:
            continue
        try:
            ids.append(int(s))
        except ValueError:
            logging.warning(f"Ignoring invalid SUBCONTRACTOR_SHEET_IDS token: {s!r}")
    return ids

SUBCONTRACTOR_SHEET_IDS = set(_parse_sheet_ids(os.getenv('SUBCONTRACTOR_SHEET_IDS', '')))

# Folder-based discovery: Smartsheet folder IDs whose child sheets should be auto-discovered
# Subcontractor folders (sheets priced at subcontractor rates, will be reverted to original contract rates)
SUBCONTRACTOR_FOLDER_IDS = _parse_sheet_ids(os.getenv('SUBCONTRACTOR_FOLDER_IDS', '4232010517505924,2588197684307844'))
# Original contract folders (sheets already at original contract rates)
ORIGINAL_CONTRACT_FOLDER_IDS = _parse_sheet_ids(os.getenv('ORIGINAL_CONTRACT_FOLDER_IDS', '7644752003786628,8815193070299012'))
# VAC Crew detection is now row-level (column-presence-based, no folder/sheet ID config needed).
# Sheets with columns like 'VAC Crew Helping?' and 'Vac Crew Completed Unit?' automatically
# produce vac_crew variant rows during row processing.
# Legacy variables for backward compatibility with tests (no longer used in production)
VAC_CREW_SHEET_IDS = set(_parse_sheet_ids(os.getenv('VAC_CREW_SHEET_IDS', '')))
VAC_CREW_FOLDER_IDS = _parse_sheet_ids(os.getenv('VAC_CREW_FOLDER_IDS', ''))
# Module-level sets populated at runtime by discover_folder_sheets()
_FOLDER_DISCOVERED_SUB_IDS: set[int] = set()
_FOLDER_DISCOVERED_ORIG_IDS: set[int] = set()

# --- RATE CONTRACT VERSIONING ---
# Set RATE_CUTOFF_DATE (YYYY-MM-DD) to activate new rate recalculation.
# When set, rows with Snapshot Date >= cutoff get prices recalculated from new rate tables.
# When unset (empty string), all rows keep their SmartSheet Units Total Price (current behavior).
_cutoff_str = os.getenv('RATE_CUTOFF_DATE', '')
try:
    RATE_CUTOFF_DATE = (
        datetime.datetime.strptime(_cutoff_str, '%Y-%m-%d').date()
        if _cutoff_str else None
    )
except ValueError:
    logging.error(f"Invalid RATE_CUTOFF_DATE format: '{_cutoff_str}', expected YYYY-MM-DD. Rate versioning disabled.")
    RATE_CUTOFF_DATE = None
ARROWHEAD_DISCOUNT = 0.90  # 10% reduction for subcontractors (Arrowhead)

def _sanitize_csv_path(env_var, default):
    """Validate a CSV path from env var, preventing directory traversal and symlink escapes.

    Returns the fully resolved path so that the value passed to open() is
    the same value that was validated (satisfies CodeQL taint analysis).
    """
    raw = (os.getenv(env_var, '') or '').strip() or default
    resolved = os.path.normpath(os.path.realpath(raw))
    cwd = os.path.normpath(os.path.realpath('.'))
    if not resolved.startswith(cwd + os.sep) and resolved != cwd:
        logging.warning(f"⚠️ {env_var} resolves outside working directory: '{raw}'. Using default: '{default}'")
        return os.path.normpath(os.path.realpath(default))
    return resolved

NEW_RATES_CSV = _sanitize_csv_path('NEW_RATES_CSV', 'New Contract Rates copy regenerated again.csv')
OLD_RATES_CSV = _sanitize_csv_path('OLD_RATES_CSV', 'CU List - Corpus North & South.csv')

_RATES_FINGERPRINT = ''  # Populated at runtime by load_rate_versions()

# Weekly-Ref-Date fallback for pre-acceptance rate recalculation.
# Default-ON: when a row has a blank/unparseable Snapshot Date (common on
# current-week VAC crew / helper rows before Smartsheet's snapshot
# automation has fired), fall back to 'Weekly Reference Logged Date'
# for the cutoff comparison. Rows that DO have a Snapshot Date are
# unaffected — the snapshot-keyed business rule stays primary. Set
# RATE_RECALC_WEEKLY_FALLBACK=0 (or false/no/off) to disable.
RATE_RECALC_WEEKLY_FALLBACK = os.getenv(
    'RATE_RECALC_WEEKLY_FALLBACK', '1'
).lower() in ('1', 'true', 'yes', 'on')

# Smartsheet-native pricing guard for original-contract folders.
# Default-ON: sheets discovered via ORIGINAL_CONTRACT_FOLDER_IDS (folders
# whose Smartsheet formula already emits the correct post-cutoff
# Units Total Price for rows with Snapshot Date >= RATE_CUTOFF_DATE
# and Units Completed? = true) are excluded from Python-side recalc.
# Running recalc on top of Smartsheet's already-correct price risked
# overwriting the Smartsheet-authoritative value with a CSV-derived
# rate × qty that did not always match, producing over/under-billed
# rows. Set RATE_RECALC_SKIP_ORIGINAL_CONTRACT=0 (or false/no/off) to
# restore the pre-fix behaviour (run recalc on these folders too).
# Subcontractor sheets are excluded unconditionally regardless of this
# flag (same as before this guard existed).
RATE_RECALC_SKIP_ORIGINAL_CONTRACT = os.getenv(
    'RATE_RECALC_SKIP_ORIGINAL_CONTRACT', '1'
).lower() in ('1', 'true', 'yes', 'on')

# ── Subcontractor rate variants (Phase 1 SUB-01..07) ───────────────────
# See .planning/phases/01-subcontractor-rate-logic-modification/
# 01-CONTEXT.md decisions D-03, D-12, D-13. These three env vars
# scaffold the new ``_AEPBillable`` and ``_ReducedSub`` variant
# pipeline:
#
#   SUBCONTRACTOR_RATES_CSV — path to the operator-managed contract
#       CSV (17 columns, currency-formatted). Default
#       ``data/subcontractor_rates.csv``. Resolved through the same
#       ``_sanitize_csv_path`` helper used by the retired
#       ``NEW_RATES_CSV`` / ``OLD_RATES_CSV`` env vars, which guards
#       against directory traversal and symlink escape per the CodeQL
#       taint-analysis pattern.
#   SUBCONTRACTOR_PPP_SHEET_ID — second target sheet for ``_ReducedSub``
#       attachments. Default ``8162920222379908``. Parsed through
#       ``_coerce_sheet_id`` for parse-error fallback.
#   SUBCONTRACTOR_RATE_VARIANTS_ENABLED — default-on kill switch for
#       the entire new variant pipeline. Pattern mirrors
#       ``RATE_RECALC_SKIP_ORIGINAL_CONTRACT`` and
#       ``RATE_RECALC_WEEKLY_FALLBACK``. Flipping this to ``0`` /
#       ``false`` / ``no`` / ``off`` reverts subcontractor-folder
#       sheets to pre-change behavior on the next run.
#
# Per Living Ledger 2026-04-24 14:30: do NOT re-introduce
# ``RATE_CUTOFF_DATE`` / ``NEW_RATES_CSV`` / ``OLD_RATES_CSV``. The
# new env vars are the subcontractor-specific replacement.
SUBCONTRACTOR_RATES_CSV = _sanitize_csv_path(
    'SUBCONTRACTOR_RATES_CSV', 'data/subcontractor_rates.csv'
)
SUBCONTRACTOR_PPP_SHEET_ID = _coerce_sheet_id(
    os.getenv('SUBCONTRACTOR_PPP_SHEET_ID', '8162920222379908'),
    8162920222379908,
)
# Phase 01 gap closure (REVIEW-WR-02): treat an explicitly-empty
# ``SUBCONTRACTOR_PPP_SHEET_ID=''`` as "disable dual routing,"
# matching the operator-facing documentation in
# website/docs/reference/environment.md and the operator's likely
# intent. ``_coerce_sheet_id`` itself stays as-is because it is
# shared with ``TARGET_SHEET_ID`` where default-fallback is the
# correct behavior (TARGET_SHEET_ID has no "disabled" state).
# Setting to ``0`` also disables (already worked pre-fix; the
# downstream gate ``and SUBCONTRACTOR_PPP_SHEET_ID`` evaluates
# False on int(0)). After this fix, both ``0`` and ``''`` disable.
# Any other non-integer / non-empty value falls back to the
# hardcoded default with the existing _coerce_sheet_id WARNING.
if os.getenv('SUBCONTRACTOR_PPP_SHEET_ID', '8162920222379908') == '':
    SUBCONTRACTOR_PPP_SHEET_ID = 0
SUBCONTRACTOR_RATE_VARIANTS_ENABLED = os.getenv(
    'SUBCONTRACTOR_RATE_VARIANTS_ENABLED', '1'
).lower() in ('1', 'true', 'yes', 'on')

# Phase 1.1 Bug A (D-02 / SUB-08): pre-acceptance rate-recalc rescue
# for subcontractor sheets. Default-on per the [2026-04-23 00:00]
# Living Ledger rule that any pre-acceptance / data-shape change must
# ship with a rollback flag. Setting '0' reverts Bug A behavior to
# the pre-fix state without affecting Bug B1, B2, or claim-history
# fixes. Pinned in workflow env: block per IN-04.
SUBCONTRACTOR_RATE_RECALC_PREACCEPTANCE_ENABLED = os.getenv(
    'SUBCONTRACTOR_RATE_RECALC_PREACCEPTANCE_ENABLED', '1'
).strip().lower() in ('1', 'true', 'yes', 'on')

# Phase 1.1 Bug C (D-14 / SUB-11): per-row claim-history attribution
# kill switch. Default-on per the [2026-04-23 00:00] Living Ledger
# rule. Setting '0' reverts Bug C behavior to Phase 1's full-row-set
# helper behavior (the same path as D-12 unconditionally —
# `lookup_attribution` is not invoked and the row's current
# `__helper_foreman` flows through to shadow-variant emission
# unchanged). Pinned in workflow env: block per IN-04.
SUBCONTRACTOR_HELPER_CLAIM_ATTRIBUTION_ENABLED = os.getenv(
    'SUBCONTRACTOR_HELPER_CLAIM_ATTRIBUTION_ENABLED', '1'
).strip().lower() in ('1', 'true', 'yes', 'on')

# Phase 1.1 UAT gap closure (SUB-09 helper dimension): default-ON
# kill switch for the one-time removal of pre-existing legacy
# `_Helper_<name>` (and bare-primary) attachments on TARGET_SHEET_ID
# for subcontractor WRs. These are duplicate-billing leftovers from
# pre-fix merged runs (Task 1 stops NEW ones; this removes OLD ones).
# Set to '0' to skip the destructive cleanup (the duplicates then
# persist until manually removed). Workflow-pinned per [2026-05-15
# 12:00] rule 7.
SUBCONTRACTOR_LEGACY_HELPER_CLEANUP_ENABLED = os.getenv(
    'SUBCONTRACTOR_LEGACY_HELPER_CLEANUP_ENABLED', '1'
).strip().lower() in ('1', 'true', 'yes', 'on')

# Subproject B (2026-05-20): default-ON kill switch for the one-time
# removal of legacy UNPARTITIONED `_ReducedSub` / `_AEPBillable`
# attachments (no `_User_` token, parsed identifier == '') on
# TARGET_SHEET_ID and SUBCONTRACTOR_PPP_SHEET_ID for subcontractor
# WRs. B re-partitions those variants by frozen primary claimer; the
# legacy one-file-per-WR attachments become duplicate-billing
# leftovers (the Phase 1.1 Bug B2 / SUB-09 trap). Set to '0' to skip
# the destructive cleanup (legacy files then persist until manually
# removed). Separate from SUBCONTRACTOR_HELPER_CLAIM_ATTRIBUTION_ENABLED
# (which gates attribution resolution, NOT this cleanup). Workflow-
# pinned per [2026-05-15 12:00] rule 7.
SUBCONTRACTOR_LEGACY_PRIMARY_CLEANUP_ENABLED = os.getenv(
    'SUBCONTRACTOR_LEGACY_PRIMARY_CLEANUP_ENABLED', '1'
).strip().lower() in ('1', 'true', 'yes', 'on')

# Subproject C (2026-05-21): default-ON kill switch that enables
# per-claimer partitioning of ``_VacCrew`` Excel files. When enabled,
# each vac-crew Excel is partitioned by the FROZEN vac-crew claimer
# (``vac_crew`` role from ``billing_audit.attribution_snapshot`` via
# ``resolve_claimer``). When disabled, the legacy one-file-per-WR
# ``_VacCrew`` behavior is preserved exactly. Pinned in workflow
# env: block per [2026-05-15 12:00] rule 7.
VAC_CREW_CLAIM_ATTRIBUTION_ENABLED = os.getenv(
    'VAC_CREW_CLAIM_ATTRIBUTION_ENABLED', '1'
).strip().lower() in ('1', 'true', 'yes', 'on')

# Subproject C (2026-05-21): default-ON kill switch for the one-time
# removal of legacy UNPARTITIONED ``_VacCrew`` attachments (no
# ``_User_`` token, parsed identifier == '') on TARGET_SHEET_ID for
# vac-crew WRs, once those variants are re-partitioned by frozen
# vac-crew claimer. Set to '0' to skip the destructive cleanup (legacy
# files then persist until removed manually). Separate from
# VAC_CREW_CLAIM_ATTRIBUTION_ENABLED (which gates attribution
# resolution, NOT this cleanup). Workflow-pinned per [2026-05-15
# 12:00] rule 7.
VAC_CREW_LEGACY_CLEANUP_ENABLED = os.getenv(
    'VAC_CREW_LEGACY_CLEANUP_ENABLED', '1'
).strip().lower() in ('1', 'true', 'yes', 'on')

# Subproject D (2026-05-25): default-ON kill switch that enables
# per-claimer partitioning of the PRODUCTION primary Excel files. When
# enabled, each non-subcontractor primary Excel is partitioned by the
# FROZEN primary foreman (``primary`` role from
# ``billing_audit.attribution_snapshot`` via ``resolve_claimer``) and
# named ``_User_<claimer>``. When disabled, the legacy one-file-per-WR
# bare primary behavior is preserved exactly. Unlike Subproject B, the
# core primary path NEVER holds on a Supabase outage — it falls back to
# the current foreman and still generates (operator decision: this path
# covers every non-sub WR, so HOLD would suppress all primary billing
# during an outage). Pinned in workflow env: block per [2026-05-15
# 12:00] rule 7.
PRIMARY_CLAIM_ATTRIBUTION_ENABLED = os.getenv(
    'PRIMARY_CLAIM_ATTRIBUTION_ENABLED', '1'
).strip().lower() in ('1', 'true', 'yes', 'on')

# Subproject D (2026-05-25): default-ON kill switch for the one-time
# removal of legacy UNPARTITIONED bare ``primary`` attachments (no
# ``_User_`` token; ``build_group_identity`` parses these to
# ``identifier=None``) on TARGET_SHEET_ID for
# non-subcontractor WRs, once those files are re-partitioned by frozen
# primary claimer. Set to '0' to skip the destructive cleanup (legacy
# duplicates then persist until removed manually). Separate from
# PRIMARY_CLAIM_ATTRIBUTION_ENABLED (which gates attribution
# resolution, NOT this cleanup). Workflow-pinned per [2026-05-15
# 12:00] rule 7.
LEGACY_PRIMARY_PARTITION_CLEANUP_ENABLED = os.getenv(
    'LEGACY_PRIMARY_PARTITION_CLEANUP_ENABLED', '1'
).strip().lower() in ('1', 'true', 'yes', 'on')

# Sub-project E (2026-05-25): durable Supabase change-detection hash store.
# WRITE (default ON): shadow-write the per-group content hash to Supabase
# every run. Harmless even while not authoritative — it populates the
# durable store so a later flip to authoritative finds current hashes and
# skips the one-time regeneration wave.
SUPABASE_HASH_STORE_WRITE_ENABLED = os.getenv(
    'SUPABASE_HASH_STORE_WRITE_ENABLED', '1'
).strip().lower() in ('1', 'true', 'yes', 'on')
# AUTHORITATIVE (default OFF — ship dormant): when ON, (a) the change-
# detection skip gate reads the Supabase group hash (falling back to the
# hash_history.json cache, then regenerating on miss/outage — never
# skipping unsafely), (b) generated filenames DROP the _<timestamp>/_<hash>
# tokens (deterministic identity-only names), and (c)
# delete_old_excel_attachments stops relying on the filename-embedded hash.
# Flip to '1' only after the Supabase store is validated in production.
# This is the one-line master revert for all of Sub-project E.
SUPABASE_HASH_STORE_AUTHORITATIVE = os.getenv(
    'SUPABASE_HASH_STORE_AUTHORITATIVE', '0'
).strip().lower() in ('1', 'true', 'yes', 'on')

# Phase 2 Plan 05 gap-closure (CR-01): when the bulk lookup_attribution_bulk
# RPC is not yet deployed (PGRST202 -> prefetch status 'rpc_missing'), degrade
# to the already-deployed per-row lookup_attribution path instead of HOLDing
# every B/C/sub-helper row. Default ON so a code-before-RPC deploy ordering
# does NOT suppress billing. A genuine transient outage ('fetch_failure') still
# HOLDs B/C (D-04). The per-row fallback is bounded to the rows actually
# processed this run, so it cannot reintroduce the 137k per-row storm. Set to
# '0' to force the strict bulk-only behavior.
ATTRIBUTION_BULK_PREFETCH_FALLBACK = os.getenv(
    'ATTRIBUTION_BULK_PREFETCH_FALLBACK', '1'
).strip().lower() in ('1', 'true', 'yes', 'on')

# ── Phase 2 Plan 03 (D-06/D-07/D-08): Isolated garbage-attachment remediation ──
# DEFAULT OFF: ``REMEDIATE_CLAIMERS='0'`` so the mode NEVER fires on a scheduled
# cron run. An operator must explicitly set '1' via workflow_dispatch or a local
# shell to activate the sweep. When active, ``main()`` returns immediately after
# the sweep (isolation — no Excel generation occurs in the same session).
REMEDIATE_CLAIMERS = os.getenv(
    'REMEDIATE_CLAIMERS', '0'
).strip().lower() in ('1', 'true', 'yes', 'on')
# DRY_RUN (default ON): first run reports counts without deleting (D-08). Set
# to '0' only after reviewing the dry-run log and confirming the scope is correct.
REMEDIATION_DRY_RUN = os.getenv(
    'REMEDIATION_DRY_RUN', '1'
).strip().lower() in ('1', 'true', 'yes', 'on')
# WINDOW_WEEKS (default 26): sweep only attachments whose week-ending date is
# within the last N weeks. Limits blast radius; set 0 to disable the filter
# (unbounded sweep). Safe-parsed: invalid values fall back to the default with
# an operator WARNING.
_remediation_window_env = os.getenv('REMEDIATION_WINDOW_WEEKS', '26')
try:
    REMEDIATION_WINDOW_WEEKS: int = int(_remediation_window_env)
    if REMEDIATION_WINDOW_WEEKS < 0:
        raise ValueError("negative")
except (ValueError, TypeError):
    logging.warning(
        f"⚠️ REMEDIATION_WINDOW_WEEKS={_remediation_window_env!r} is not a valid "
        f"non-negative integer — falling back to default 26"
    )
    REMEDIATION_WINDOW_WEEKS = 26

# Recent-week scope for the per-row frozen-attribution pre-pass
# (perf hotfix 2026-05-26). The B/C/D claim-attribution pre-passes and
# the subcontractor-helper path each call the ``lookup_attribution``
# Supabase RPC once per completed row. Run unbounded, that resolves
# EVERY completed row across ALL historical weeks (observed: ~137k RPCs
# per run), even though change-detection skips the vast majority of old
# weeks (unchanged + attachment exists) — so the resolved claimer is
# never even used. That eager work scaled with accumulated history and
# blew the workflow time budget.
#
# This gate limits resolution to rows whose ``week_ending`` is within
# the last N weeks of today, making the pre-pass cost track ACTIVE work
# (current + recent edit horizon) instead of total history. Safe by
# construction: an out-of-scope row resolves to use-current at emission,
# but its group is one of two cases — (1) unchanged + attachment exists
# -> skipped, claimer unused (zero impact); or (2) the rare edit to a
# >N-week-old row -> regenerated with the CURRENT foreman (the same
# legacy/no_history fallback the feature already documents). The freeze
# (write) side is UNTOUCHED — every completed row is still frozen during
# generation — so the durable attribution data stays complete.
#
# Phase 2 Plan 02 (D-05): ATTRIBUTION_RESOLUTION_WEEKS removed. The bulk
# prefetch (prefetch_attribution) now covers the exact (wr, week_ending)
# pairs in the current run — no recency scope gate is needed. This fixes
# the production incident (run 26439205107) where the 8-week scope gate
# prevented historical frozen claimers from being resolved, producing
# 372 garbage _User__NO_MATCH / _User_Unknown_Foreman files.

# Cutoff date for ``_AEPBillable`` variant generation. Awarded to
# Linetec on 2026-04-12 (subcontractor rate contract). Plan 2 (parser
# extension) and Plan 3 (variant emission) gate variant emission on
# ``Snapshot Date >= _AEP_BILLABLE_CUTOFF``. Exposed at module level
# so downstream plans can reference a single source of truth.
#
# Phase 01 gap closure (REVIEW-IN-01): exposed as
# ``AEP_BILLABLE_CUTOFF`` env var with safe parse + fallback to
# the contract-award default. Operators can roll forward (or back,
# for retroactive billing decisions) without a code change.
# Format: ``YYYY-MM-DD``. Invalid format logs an error and falls
# back to the default. Default is byte-identical to the pre-fix
# constant — IN-01 is additive (override path), no behavior
# regression for the unset / valid-format cases. The ``RATE_CUTOFF_DATE``
# env var (retired 2026-04-24 14:30) is NOT reused — this is a new
# distinct env var with explicit subcontractor-variant scope.
_aep_billable_cutoff_env = os.getenv('AEP_BILLABLE_CUTOFF', '')
try:
    _AEP_BILLABLE_CUTOFF = (
        datetime.datetime.strptime(
            _aep_billable_cutoff_env, '%Y-%m-%d'
        ).date()
        if _aep_billable_cutoff_env
        else datetime.date(2026, 4, 12)
    )
except ValueError:
    logging.error(
        f"⚠️ Invalid AEP_BILLABLE_CUTOFF format: "
        f"{_aep_billable_cutoff_env!r}; expected YYYY-MM-DD. "
        f"Falling back to default 2026-04-12."
    )
    _AEP_BILLABLE_CUTOFF = datetime.date(2026, 4, 12)

if RATE_CUTOFF_DATE:
    logging.info(f"📊 Rate contract versioning ENABLED: cutoff date = {RATE_CUTOFF_DATE.isoformat()}")
    # The CSV-side rate recalc was retired in production on
    # 2026-04-24 because Smartsheet now emits the authoritative
    # Units Total Price natively for ORIGINAL_CONTRACT_FOLDER_IDS
    # post-cutoff rows. The production workflow pins
    # RATE_CUTOFF_DATE='' so this branch should not fire on
    # scheduled runs anymore. If it DOES fire, something has
    # bypassed the workflow pinning (local dev shell, an ad-hoc
    # script, a re-introduced repo Variable) — surface that loudly
    # so operators can double-check the pricing source before
    # trusting the output.
    logging.warning(
        "⚠️ RATE_CUTOFF_DATE is set, but the Python CSV-side rate "
        "recalc feature has been retired — Smartsheet now emits "
        "the authoritative Units Total Price natively on "
        "ORIGINAL_CONTRACT_FOLDER_IDS sheets, and the production "
        "workflow pins RATE_CUTOFF_DATE='' as of 2026-04-24. "
        "Unset RATE_CUTOFF_DATE to silence this warning. See "
        "CLAUDE.md Living Ledger entry [2026-04-24] for context."
    )
    if RATE_RECALC_WEEKLY_FALLBACK:
        logging.info(
            "📊 Rate recalc Weekly-Ref-Date fallback ENABLED "
            "(blank Snapshot Date → use Weekly Reference Logged Date for cutoff gate)"
        )
    else:
        logging.info("📊 Rate recalc Weekly-Ref-Date fallback DISABLED (RATE_RECALC_WEEKLY_FALLBACK=false)")
    if RATE_RECALC_SKIP_ORIGINAL_CONTRACT:
        logging.info(
            "📊 Rate recalc ORIGINAL_CONTRACT folder skip ENABLED "
            "(sheets discovered via ORIGINAL_CONTRACT_FOLDER_IDS keep "
            "Smartsheet-native Units Total Price, no CSV-side recalc)"
        )
    else:
        logging.info(
            "📊 Rate recalc ORIGINAL_CONTRACT folder skip DISABLED "
            "(RATE_RECALC_SKIP_ORIGINAL_CONTRACT=false) — recalc will "
            "run on original-contract folder sheets too"
        )
else:
    logging.info("📊 Rate contract versioning DISABLED (RATE_CUTOFF_DATE not set)")

# Subcontractor rate variants startup banner (Phase 1 D-13). The
# fingerprint line is appended later in the rate-loading section once
# ``_SUBCONTRACTOR_RATES_FINGERPRINT`` has been computed by Task 2's
# module-level loader call — see ``load_subcontractor_rates`` below.
# These banner lines embed NO row content (just resolved config
# values), so per D-22 no ``_PII_LOG_MARKERS`` extension is required.
if SUBCONTRACTOR_RATE_VARIANTS_ENABLED:
    logging.info(
        "📊 Subcontractor rate variants ENABLED "
        f"(SUBCONTRACTOR_RATES_CSV='{SUBCONTRACTOR_RATES_CSV}', "
        f"SUBCONTRACTOR_PPP_SHEET_ID={SUBCONTRACTOR_PPP_SHEET_ID})"
    )
else:
    logging.info(
        "📊 Subcontractor rate variants DISABLED "
        "(SUBCONTRACTOR_RATE_VARIANTS_ENABLED=false)"
    )

# Phase 01 gap closure (REVIEW-WR-02): name the PPP-routing
# state explicitly so operators tailing the startup banner see
# the resolved active value (or "DISABLED") without inferring
# from the integer 0. Purely additive — does not replace the
# existing banner block above. Only emitted when the umbrella
# variants kill switch is ON (when off, PPP routing is moot).
if SUBCONTRACTOR_RATE_VARIANTS_ENABLED and SUBCONTRACTOR_PPP_SHEET_ID:
    logging.info(
        f"📊 Subcontractor PPP routing ENABLED "
        f"(target sheet id: {SUBCONTRACTOR_PPP_SHEET_ID})"
    )
elif SUBCONTRACTOR_RATE_VARIANTS_ENABLED:
    logging.info(
        "📊 Subcontractor PPP routing DISABLED "
        "(SUBCONTRACTOR_PPP_SHEET_ID='' or 0)"
    )

# Phase 01 gap closure (REVIEW-IN-01): name the resolved AEP cutoff
# in the startup banner so operators tailing the log see the active
# value at a glance. Only emitted when the umbrella variants kill
# switch is ON (when off, the cutoff is moot — no _AEPBillable
# variant emission occurs regardless of the cutoff value).
if SUBCONTRACTOR_RATE_VARIANTS_ENABLED:
    logging.info(
        f"📊 AEP Billable cutoff: {_AEP_BILLABLE_CUTOFF.isoformat()} "
        f"({'env override' if _aep_billable_cutoff_env else 'default'})"
    )

# Phase 1.1 Bug A: surface the resolved kill-switch state at startup
# so operators grepping the banner can see the active feature state
# at a glance (per [2026-04-23 00:00] ledger rule 3). Banner body
# carries no row PII (just the resolved bool value) — no marker
# required.
logging.info(
    f"📋 SUBCONTRACTOR_RATE_RECALC_PREACCEPTANCE_ENABLED="
    f"{SUBCONTRACTOR_RATE_RECALC_PREACCEPTANCE_ENABLED}"
)

# Phase 1.1 Bug C: surface resolved kill-switch state at startup
# so operators grepping the banner can see the active feature state
# at a glance.
logging.info(
    f"📋 SUBCONTRACTOR_HELPER_CLAIM_ATTRIBUTION_ENABLED="
    f"{SUBCONTRACTOR_HELPER_CLAIM_ATTRIBUTION_ENABLED}"
)

# Phase 1.1 UAT gap closure (SUB-09 helper dimension): surface resolved
# kill-switch state at startup so operators grepping the banner can see
# the active feature state at a glance (per [2026-04-23 00:00] ledger
# rule 3). Banner body carries no row PII (just the resolved bool).
logging.info(
    f"📋 SUBCONTRACTOR_LEGACY_HELPER_CLEANUP_ENABLED="
    f"{SUBCONTRACTOR_LEGACY_HELPER_CLEANUP_ENABLED}"
)

# Subproject B: surface resolved kill-switch state at startup so
# operators grepping the banner see the active feature state at a
# glance. Banner body carries no row PII (just the resolved bool).
logging.info(
    f"📋 SUBCONTRACTOR_LEGACY_PRIMARY_CLEANUP_ENABLED="
    f"{SUBCONTRACTOR_LEGACY_PRIMARY_CLEANUP_ENABLED}"
)

# Subproject C: surface resolved kill-switch state at startup so
# operators grepping the banner see the active feature state at a
# glance. Banner body carries no row PII (just the resolved bools).
logging.info(
    f"📋 VAC Crew claim attribution: "
    f"{'ENABLED' if VAC_CREW_CLAIM_ATTRIBUTION_ENABLED else 'DISABLED'}"
)
logging.info(
    f"📋 VAC Crew legacy cleanup: "
    f"{'ENABLED' if VAC_CREW_LEGACY_CLEANUP_ENABLED else 'DISABLED'}"
)
# Subproject D: surface resolved kill-switch state at startup so
# operators grepping the banner see the active feature state at a
# glance. Banner body carries no row PII (just the resolved bools).
logging.info(
    f"📋 PRIMARY_CLAIM_ATTRIBUTION_ENABLED="
    f"{PRIMARY_CLAIM_ATTRIBUTION_ENABLED}"
)
logging.info(
    f"📋 LEGACY_PRIMARY_PARTITION_CLEANUP_ENABLED="
    f"{LEGACY_PRIMARY_PARTITION_CLEANUP_ENABLED}"
)
# Sub-project E: surface the durable hash-store kill switches at startup.
logging.info(
    f"📋 SUPABASE_HASH_STORE_WRITE_ENABLED="
    f"{SUPABASE_HASH_STORE_WRITE_ENABLED}"
)
logging.info(
    f"📋 SUPABASE_HASH_STORE_AUTHORITATIVE="
    f"{SUPABASE_HASH_STORE_AUTHORITATIVE}"
)
# Phase 2 Plan 03: remediation kill switches surfaced at startup.
logging.info(
    f"📋 REMEDIATE_CLAIMERS={REMEDIATE_CLAIMERS} "
    f"DRY_RUN={REMEDIATION_DRY_RUN} "
    f"WINDOW_WEEKS={REMEDIATION_WINDOW_WEEKS}"
)
# Phase 2 Plan 05 (CR-01): surface the bulk-prefetch fallback state at startup.
logging.info(
    f"📋 ATTRIBUTION_BULK_PREFETCH_FALLBACK={ATTRIBUTION_BULK_PREFETCH_FALLBACK}"
)
RESET_HASH_HISTORY = os.getenv('RESET_HASH_HISTORY','0').lower() in ('1','true','yes')  # When true, delete ALL existing WR_*.xlsx attachments & local files first
RESET_WR_LIST = {w.strip() for w in os.getenv('RESET_WR_LIST','').split(',') if w.strip()}  # When provided, only purge these WR numbers (overrides full reset)
_env_hist_path = os.getenv('HASH_HISTORY_PATH')
_default_hist_path = os.path.join(OUTPUT_FOLDER, 'hash_history.json')
if _env_hist_path:
    # Only allow the file within OUTPUT_FOLDER (resolve real absolute paths)
    _norm_path = os.path.normpath(os.path.abspath(os.path.join(OUTPUT_FOLDER, _env_hist_path)))
    _output_folder_abs = os.path.normpath(os.path.abspath(OUTPUT_FOLDER))
    if _norm_path.startswith(_output_folder_abs):
        HASH_HISTORY_PATH = _norm_path
    else:
        logging.warning(f"⚠️ HASH_HISTORY_PATH environment variable must resolve within {OUTPUT_FOLDER}. Using default: {_default_hist_path}")
        HASH_HISTORY_PATH = _default_hist_path
else:
    HASH_HISTORY_PATH = _default_hist_path
HISTORY_SKIP_ENABLED = os.getenv('HISTORY_SKIP_ENABLED','1').lower() in ('1','true','yes')  # Allow skip based on identical stored hash ONLY if attachment still present
ATTACHMENT_REQUIRED_FOR_SKIP = os.getenv('ATTACHMENT_REQUIRED_FOR_SKIP','1').lower() in ('1','true','yes')  # If true, even identical hash regenerates when attachment missing
KEEP_HISTORICAL_WEEKS = os.getenv('KEEP_HISTORICAL_WEEKS','0').lower() in ('1','true','yes')  # Preserve attachments for weeks not processed this run
if EXTENDED_CHANGE_DETECTION:
    logging.info("🔄 Extended change detection ENABLED (hash includes Foreman, Dept #, Scope, totals, etc.)")
else:
    logging.info("ℹ️ Legacy change detection mode (hash limited to core line item fields)")
if QUIET_LOGGING:
    logging.getLogger().setLevel(logging.WARNING)

# Test/Production modes (controlled by environment variable TEST_MODE)
# PRODUCTION BY DEFAULT: set TEST_MODE=true only for maintenance / dry runs.
TEST_MODE = os.getenv('TEST_MODE', 'false').lower() in ('1','true','yes')
DISABLE_AUDIT_FOR_TESTING = False  # Audit system ENABLED for production monitoring

# --- SENTRY CONFIGURATION (SDK 2.x) ---
SENTRY_DSN = os.getenv("SENTRY_DSN")

# Compiled patterns used to scrub billing-row PII out of exception
# messages before they land in Sentry event context_data. The
# ``before_send_log`` hook further down only sanitizes logging records
# — it does NOT walk ``event['contexts']`` — so any raw ``str(e)``
# passed into ``sentry_capture_with_context(...)``'s context_data
# payload would bypass that defense. Keep these patterns conservative:
# they only strip recognised PII tokens, leaving the rest of the
# exception message intact so operators can still diagnose the root
# cause from the Sentry dashboard.
# Match any ``WR``-prefixed identifier, not just digit-only tokens.
# Earlier ``\d+`` missed alphanumeric WR values (``WR=ABCD-123``) and
# only partially matched path-traversal suffixes (``WR=1234/../evil``
# would redact only ``1234``, leaking ``/../evil`` through the
# Sentry context). The negative lookahead ``(?![a-zA-Z])`` keeps the
# pattern from over-matching English words that happen to start with
# ``WR`` (e.g. ``WRITE``). The identifier char class accepts word
# characters plus ``/ \ . -`` so path-traversal tokens and decorated
# IDs are captured in full, and the ``+`` stops at the first
# whitespace / delimiter so only the identifier itself is redacted.
_RE_REDACT_WR = re.compile(r'(?i)\bWR(?![a-zA-Z])\s*[#:=]?\s*[\w/\\\-.]+')
_RE_REDACT_MONEY = re.compile(r'\$\s*\d[\d,]*(?:\.\d+)?')
_RE_REDACT_EMAIL = re.compile(r'[\w.+-]+@[\w.-]+\.\w+')
_RE_REDACT_CUSTOMER = re.compile(r'(?i)\b(customer|foreman|dept|snapshot|cu|job)\s*[#:=]?\s*["\']?[^,;"\')\]}\n]{1,80}')


def _redact_exception_message(exc: BaseException | None, *, max_len: int = 240) -> str:
    """Return a PII-scrubbed single-line form of ``str(exc)``.

    Used exclusively for the ``error_message`` field of
    ``sentry_capture_with_context(...)``'s context_data payload — that
    dict bypasses the INFO-log ``before_send_log`` sanitizer because
    Sentry attaches it directly as event context. The redactor strips
    WR identifiers, dollar amounts, emails, and ``customer=``/
    ``foreman=``/``dept=``/``snapshot=``/``cu=``/``job=`` key-value
    pairs, collapses whitespace, prefixes the exception class name for
    event-grouping stability, and truncates the result.
    """
    if exc is None:
        return ''
    try:
        raw = str(exc)
    except Exception:
        return f"{type(exc).__name__}: <unrepresentable>"
    if not raw:
        return type(exc).__name__
    redacted = _RE_REDACT_CUSTOMER.sub(r'\1=<redacted>', raw)
    redacted = _RE_REDACT_WR.sub('WR=<redacted>', redacted)
    redacted = _RE_REDACT_MONEY.sub('$<redacted>', redacted)
    redacted = _RE_REDACT_EMAIL.sub('<email>', redacted)
    redacted = re.sub(r'\s+', ' ', redacted).strip()
    # Codex: truncate AFTER adding the class prefix so ``max_len``
    # caps the full returned payload (what actually lands in the
    # Sentry event), not just the body portion.
    result = f"{type(exc).__name__}: {redacted}"
    if len(result) > max_len:
        result = result[:max_len - 3] + '...'
    return result


# Sentry helper functions for enhanced error context
def sentry_add_breadcrumb(category: str, message: str, level: str = "info", data: dict | None = None):
    """Add a breadcrumb for execution flow tracking in Sentry dashboard."""
    if SENTRY_DSN:
        sentry_sdk.add_breadcrumb(
            category=category,
            message=message,
            level=level,
            data=data or {}
        )

def sentry_capture_with_context(exception: Exception, context_name: str | None = None, 
                                  context_data: dict | None = None, tags: dict | None = None,
                                  fingerprint: list | None = None):
    """Capture exception with rich context, tags, and custom fingerprinting.
    
    Args:
        exception: The exception to capture
        context_name: Name for the context block in Sentry dashboard
        context_data: Dictionary of contextual data for debugging
        tags: Additional tags for filtering in Sentry
        fingerprint: Custom fingerprint for error grouping
    """
    if not SENTRY_DSN:
        return
    
    scope = sentry_sdk.get_current_scope()
    
    # Add rich context data
    if context_name and context_data:
        sentry_sdk.set_context(context_name, context_data)
    
    # Add custom tags for filtering
    if tags:
        for key, value in tags.items():
            scope.set_tag(key, str(value))
    
    # Set custom fingerprint for error grouping
    if fingerprint:
        scope.fingerprint = fingerprint
    
    # Capture with full context
    sentry_sdk.capture_exception(exception)

from typing import Literal

# Log level type for Sentry SDK 2.x
SentryLogLevel = Literal["fatal", "critical", "error", "warning", "info", "debug"]

def sentry_capture_message_with_context(message: str, level: SentryLogLevel = "error",
                                         context_name: str | None = None, context_data: dict | None = None,
                                         tags: dict | None = None):
    """Capture a message with rich context for Sentry dashboard visibility."""
    if not SENTRY_DSN:
        return
    
    scope = sentry_sdk.get_current_scope()
    
    if context_name and context_data:
        sentry_sdk.set_context(context_name, context_data)
    
    if tags:
        for key, value in tags.items():
            scope.set_tag(key, str(value))
    
    sentry_sdk.capture_message(message, level=level)


def _build_run_kpis(
    *,
    files_generated: int,
    groups_total: int,
    groups_skipped: int,
    groups_generated: int,
    groups_uploaded: int,
    groups_errored: int,
    duration_seconds: float,
    sheets_discovered: int,
    rows_fetched: int,
    api_calls: int,
) -> dict[str, int | float]:
    """Return a flat dict of numeric run-level KPIs for the root Sentry transaction.

    ALL values are int or float — no strings — so there is zero risk of PII
    leakage via set_data().  A derived throughput metric is included.

    This helper is intentionally pure (no side effects) so it is fully
    unit-testable and the no-PII guarantee is test-enforced.
    """
    if duration_seconds and duration_seconds > 0:
        groups_per_minute = round(groups_generated / (duration_seconds / 60.0), 2)
    else:
        groups_per_minute = 0.0
    return {
        "files_generated": files_generated,
        "groups_total": groups_total,
        "groups_skipped": groups_skipped,
        "groups_generated": groups_generated,
        "groups_uploaded": groups_uploaded,
        "groups_errored": groups_errored,
        "duration_seconds": duration_seconds,
        "sheets_discovered": sheets_discovered,
        "rows_fetched": rows_fetched,
        "api_calls": api_calls,
        "groups_per_minute": groups_per_minute,
    }


def _build_run_context_snapshot(
    *,
    success: bool,
    duration_seconds: float,
    groups_attempted: int,
    groups_generated: int,
    groups_uploaded: int,
    groups_errored: int,
    error_type: str | None = None,
) -> dict:
    """Return a PII-safe counts/booleans snapshot for failure-path Sentry attachments.

    This dict is serialised to JSON and attached via scope.add_attachment, which
    BYPASSES before_send_log.  Every value must be a count, boolean, None, or the
    already-safe exception class name only — never a WR number, foreman/dept/job
    name, dollar amount, or any row-level data.

    This helper is intentionally pure (no side effects) so PII-safety is
    test-enforced rather than relying on review alone.
    """
    return {
        "success": success,
        "duration_seconds": duration_seconds,
        "groups_attempted": groups_attempted,
        "groups_generated": groups_generated,
        "groups_uploaded": groups_uploaded,
        "groups_errored": groups_errored,
        "error_type": error_type,  # exception class name only — never the message
    }


def _sentry_log_event(level: str, message: str, **attributes: int | float | bool | str | None) -> None:
    """Guarded structured-log emitter using sentry_sdk.logger (SDK >= 2.54.0).

    ONLY pass non-PII scalars (counts, booleans, fixed enums) as attributes.
    This path BYPASSES before_send_log — never pass row data, WR numbers,
    foreman/dept/job names, dollar amounts, or any per-row values.

    Safety contract:
    - No-ops immediately if SENTRY_DSN is falsy.
    - No-ops immediately if sentry_sdk has no 'logger' attribute (older SDK).
    - Swallows all internal errors (try/except) — never raises, never masks
      the caller's exception.
    - Only emits to Sentry when SENTRY_ENABLE_LOGS is True (the SDK gate).
    """
    if not SENTRY_DSN:
        return
    if not hasattr(sentry_sdk, "logger"):
        return
    try:
        # ``sentry_sdk.logger`` is a lazily-bound attribute (real in
        # sentry-sdk >= 2.54.0; presence already asserted by the hasattr guard
        # above). Resolve it via getattr so static analysis does not flag it as
        # an unknown module attribute. Runtime behavior is unchanged.
        _logger = getattr(sentry_sdk, "logger")
        log_fn = getattr(_logger, level, _logger.info)
        log_fn(message, **attributes)
    except Exception as _log_exc:
        logging.debug(f"_sentry_log_event swallowed error: {_log_exc}")


# Substring markers that identify billing-engine log messages known
# to embed row-level PII (WR, dept, job, foreman, helper / vac-crew
# names, cell values, prices). If any of these appear in a log body,
# the record is dropped before it ships to the Sentry Logs product.
# Applied in addition to the ``SENTRY_ENABLE_LOGS`` env gate —
# defense in depth. Defined at module scope (not inside the
# ``if SENTRY_DSN:`` block) so it is importable and unit-testable
# without a live DSN.
_PII_LOG_MARKERS: tuple[str, ...] = (
    # Per-row / per-cell debug dumps
    "Row data sample",
    "Cell ",
    "ESSENTIAL FIELDS",
    # Helper detection + grouping
    "HELPER ROW DETECTED",
    "HELPER GROUP CREATED",
    "HELPER GROUP SUMMARY",
    "Helper group '",
    "Helper groups: ",
    "Helper detection criteria",
    "Helper variant",
    "Helper row for WR",
    "MAPPED HELPER COLUMN",
    "Sample Helper",
    "Foreman Helping?",
    # VAC Crew detection + grouping
    "VAC Crew detection",
    "VAC CREW ROW DETECTED",
    "VAC CREW GROUP CREATED",
    "VAC CREW GROUP SUMMARY",
    "VAC Crew group '",
    "Adding row to existing VAC Crew group",
    "MAPPED VAC CREW COLUMN",
    "VAC Crew Helping?",
    # Rate / pricing recalc (CU + group codes + quantities + rates)
    "Rate recalculation",
    "Rate recalc",
    # Foreman / assignment / exclusion diagnostics
    "Foreman Assignment",
    "foremen(top5)",
    "Excluding row",
    "EXCLUDING from main Excel",
    "EXCLUDING from foreman/helper",
    # WR-keyed log lines
    "for WR#",
    "WR# ",
    "for WR ",
    "Work request ",
    "Job # not found",
    "Sample group keys",
    # Runtime WR lists (operator-supplied filter / exclusion / reset
    # lists that print every WR identifier they contain).
    "WR_FILTER applied",
    "EXCLUDE_WRS ",
    "EXCLUDE_WRS:",
    "Hash reset requested for specific WRs",
    # Group keys / totals validation. group_key shapes are
    # ``{week}_{wr}`` (primary), ``{week}_{wr}_HELPER_{foreman}``
    # (helper), ``{week}_{wr}_VACCREW`` (vac crew). Any log body
    # carrying ``_HELPER_`` or ``_VACCREW`` is therefore emitting a
    # group key (which embeds WR + week + foreman). Plus the
    # always-on totals validation block at the end of a run, which
    # logs ``{group_key}: rows=N total=$X.YY`` per group.
    "_HELPER_",
    "_VACCREW",
    "Totals Validation",
    "total=$",
    "Failed to process group ",
    "Synthetic group failure ",
    # Attachment / regeneration lifecycle (WR + week embedded)
    "Removing ",
    "Unchanged (",
    "Skip (unchanged",
    "Regenerating ",
    "FORCE GENERATION for ",
    # Output filenames interpolate
    # ``WR_{wr}_WeekEnding_{MMDDYY}_{timestamp}[_Helper_<foreman>|
    # _User_<foreman>|_VacCrew]_{hash}.xlsx``, so any log body that
    # contains ``_WeekEnding_`` is carrying an artifact name that
    # embeds WR + week + foreman. Broad catch-all + explicit prefixes
    # for the upload / delete / generate lifecycle.
    "_WeekEnding_",
    "Generating Excel file",
    "Generated Excel",
    "Uploaded: ",
    "Skipping upload ",
    "Rate limited on upload",
    "Upload retry ",
    "Upload failed for ",
    "Deleted: ",
    "Already gone: ",
    "Delete failed ",
    # Legacy / manual attachment cleanup. `purge_existing_hashed_outputs`
    # logs any ``WR_*.xlsx`` name — including short legacy forms like
    # ``WR_42.xlsx`` that don't contain ``_WeekEnding_`` — so the broad
    # filename catch-all is not sufficient for these paths.
    "Purged attachment:",
    "Failed to purge attachment",
    # Phase 01 Plan 02 D-22: subcontractor variant group keys and
    # group-creation INFO logs (Plan 3 will emit
    # ``AEP BILLABLE GROUP CREATED`` / ``REDUCED SUB GROUP CREATED``
    # plus a missing-CU WARNING that embeds the literal CU code).
    # The group-key tokens (``_AEPBILLABLE`` / ``_REDUCEDSUB`` /
    # the ``_HELPER_`` suffixed variants) cover any log body that
    # embeds the variant's group key — equivalent to the existing
    # ``_HELPER_`` / ``_VACCREW`` markers for the legacy variant set.
    # Locking these in Plan 02 (before Plan 3 emits them) keeps the
    # sanitizer ahead of the call sites, per Living Ledger
    # 2026-04-20 12:00 defense-in-depth rule.
    "_AEPBILLABLE",
    "_REDUCEDSUB",
    "_AEPBILLABLE_HELPER_",
    "_REDUCEDSUB_HELPER_",
    "AEP BILLABLE GROUP CREATED",
    "REDUCED SUB GROUP CREATED",
    # Subproject D (Task 4 review fix / [2026-05-25]): the new
    # PRIMARY GROUP CREATED INFO log embeds WR= and Week= row PII
    # (see the ``🧑 PRIMARY GROUP CREATED`` log in
    # ``group_source_rows``). Explicit marker per the
    # [2026-04-20 12:00] / [2026-05-15 12:00] ledger rules —
    # mirrors the five sibling GROUP CREATED markers above.
    "PRIMARY GROUP CREATED",
    # Phase 01 gap closure (REVIEW-WR-04 / Living Ledger 2026-04-20
    # 12:00): the new helper-shadow GROUP CREATED logs already match
    # against the substring "HELPER GROUP CREATED" by accident — the
    # tokens "REDUCED SUB HELPER GROUP CREATED" and "AEP BILLABLE
    # HELPER GROUP CREATED" happen to CONTAIN that substring. That
    # makes scrubbing fragile to future wording changes (e.g., a
    # rename to "REDUCED SUB HELPER GROUP REGISTERED" or "REDUCED
    # SUB HELPER GRP CREATED" would silently leak the helper
    # foreman name to Sentry Logs). Explicit markers are the
    # defense-in-depth contract per the 2026-04-20 12:00 ledger
    # rule. Sibling markers for the non-shadow variants
    # ("AEP BILLABLE GROUP CREATED" and "REDUCED SUB GROUP
    # CREATED") were landed in Phase 01 Plan 02; these two finish
    # the set.
    "REDUCED SUB HELPER GROUP CREATED",
    "AEP BILLABLE HELPER GROUP CREATED",
    "Subcontractor rates CSV missing",
    # Phase 1.1 Bug A (SUB-08): pre-acceptance rescue diagnostic log
    # embeds WR + CU + rescued price. Explicit marker per the
    # [2026-05-15 12:00] rule 3 (explicit markers for new INFO-level
    # log bodies — no accidental substring matching against
    # pre-existing markers).
    "Subcontractor pre-acceptance rescue",
    # Phase 1.1 Bug B2 (SUB-10): per-sheet variant whitelist
    # off-contract delete INFO log embeds attachment name
    # (``WR_*_WeekEnding_*.xlsx``), which carries WR + week + (for
    # helper-shadow variants) helper foreman name. Explicit marker
    # per the [2026-05-15 12:00] rule 3 — no accidental substring
    # containment with pre-existing markers (the body
    # ``Removed off-contract variant on sheet`` does not overlap
    # any existing marker on its own).
    "Removed off-contract variant on sheet",
    # Phase 1.1 Bug C (SUB-11): per-WR fall-back WARNING for
    # claim-history attribution lookups embeds the sanitized helper
    # foreman name. Explicit marker per the [2026-05-15 12:00]
    # rule 3 — the body ``Subcontractor helper claim attribution
    # fallback`` is an explicit literal that does not overlap any
    # pre-existing marker on its own (no accidental substring
    # containment).
    "Subcontractor helper claim attribution fallback",
    # Phase 1.1 SUB-12 / D-17 / D-19: idempotent hash-history one-time
    # prune INFO log embeds the affected-WR list (capped to the first
    # 20 entries). WR numbers are sanitized at the producer site but
    # are still PII at the project's row-level threshold. Explicit
    # marker per the [2026-05-15 12:00] rule 3 — no accidental
    # substring containment with pre-existing markers.
    "Phase 1.1 hash-history prune",
    # Subproject B Task 8: one-time hash-history prune INFO log embeds
    # the affected-WR list (capped to first 20 entries). Explicit marker
    # per the [2026-05-15 12:00] rule 3 — no accidental substring
    # containment with pre-existing markers (this body does not overlap
    # "Phase 1.1 hash-history prune").
    "Subproject B hash-history prune",
    # Subproject C Task 7: one-time hash-history prune INFO log embeds
    # the affected-WR list (capped to first 20 entries). Explicit marker
    # per the [2026-05-15 12:00] rule 3 — no accidental substring
    # containment with pre-existing markers.
    "Vac crew hash-history prune",
    # Subproject D Task 9: one-time hash-history prune INFO log embeds
    # the affected-WR list (capped to first 20 entries). Explicit marker
    # per the [2026-05-15 12:00] rule 3 — no accidental substring
    # containment with pre-existing markers (body does not overlap
    # "Subproject B hash-history prune" or "Vac crew hash-history prune").
    "Subproject D hash-history prune",
)


def sentry_before_send_log(record, hint):
    """Drop Sentry Logs records that embed billing-row PII.

    Runs only when ``SENTRY_ENABLE_LOGS`` is truthy (otherwise the
    SDK never invokes this hook). Matches against the rendered log
    body; returns ``None`` to drop, or the record unchanged to forward.
    Defined at module scope so it is importable and unit-testable.

    Missing or empty bodies are normalized to ``""`` and forwarded
    unless a configured marker is present. The hook fails **closed**
    for unexpected inspectable payloads: non-string body values, or
    any exception raised while inspecting the record, cause the
    record to be dropped so uninspectable payloads cannot bypass the
    marker checks.
    """
    try:
        # Resolve the body without ``or ""`` coercion — falsy non-string
        # values (0, False, [], {}) must reach the isinstance check so
        # they hit the fail-closed branch instead of being silently
        # converted to "" and forwarded.
        if isinstance(record, dict):
            body = record["body"] if "body" in record else ""
        else:
            body = (
                getattr(record, "body")
                if hasattr(record, "body")
                else ""
            )
        if not isinstance(body, str):
            # Fail closed for unexpected body types so uninspectable
            # records cannot bypass PII marker checks.
            return None
        for marker in _PII_LOG_MARKERS:
            if marker in body:
                return None
    except Exception:
        # Never let the sanitizer crash the SDK — drop on error so
        # unclassified records don't slip through to Sentry Logs.
        return None
    return record


def _parse_sentry_enable_logs(raw: str | None) -> bool:
    """Parse the SENTRY_ENABLE_LOGS env value into a bool.

    Truthy: ``1``, ``true``, ``yes``, ``on`` (case-insensitive,
    whitespace-tolerant). Anything else — including unset / empty —
    is falsy. Extracted so tests can cover both branches without
    needing a live DSN.
    """
    if raw is None:
        return False
    return raw.strip().lower() in ("1", "true", "yes", "on")


if SENTRY_DSN:
    sentry_logging = LoggingIntegration(
        level=logging.INFO,
        event_level=logging.ERROR
    )

    def before_send_filter(event, hint):
        """Filter out normal Smartsheet 404 errors during cleanup operations.

        Sentry 2.x compatible: Enriches events with additional context.
        """
        # Filter Smartsheet internal logger noise
        if event.get('logger') == 'smartsheet.smartsheet':
            return None
        
        # Filter out 404 attachment deletion errors (normal operations)
        if 'exception' in event and event['exception'].get('values'):
            for exc_value in event['exception']['values']:
                if exc_value.get('value'):
                    error_msg = exc_value['value'].lower()
                    if ("404" in error_msg or "not found" in error_msg) and "attachment" in error_msg:
                        logging.info("⚠️ Filtered 404 attachment error from Sentry (normal operation)")
                        return None
        
        # Enrich all events with runtime context
        event.setdefault('contexts', {})
        event['contexts']['runtime_info'] = {
            'test_mode': TEST_MODE,
            'github_actions': bool(os.getenv('GITHUB_ACTIONS')),
            'max_groups': MAX_GROUPS,
            'extended_change_detection': EXTENDED_CHANGE_DETECTION,
            'python_version': sys.version,
        }
        
        return event
    
    def traces_sampler(sampling_context):
        """Dynamic sampling for performance tracing based on operation type."""
        # Always trace errors
        if sampling_context.get('parent_sampled'):
            return 1.0
        
        # Sample main operations at 100% for full visibility
        transaction_name = sampling_context.get('transaction_context', {}).get('name', '')
        if 'excel_generation' in transaction_name or 'main' in transaction_name:
            return 1.0
        
        # Sample other operations at 50%
        return 0.5
    
    try:
        sentry_sdk.init(
            dsn=SENTRY_DSN,
            integrations=[
                sentry_logging,
                ThreadingIntegration(propagate_hub=True),
            ],
            # Performance monitoring
            traces_sample_rate=1.0,
            traces_sampler=traces_sampler,
            profiles_sample_rate=0.5,  # SDK 2.x: No longer experimental
            
            # Environment configuration — SENTRY_* vars take priority over legacy fallbacks
            environment=os.getenv("SENTRY_ENVIRONMENT") or os.getenv("ENVIRONMENT", "production"),
            release=os.getenv("SENTRY_RELEASE") or os.getenv("RELEASE", "weekly-excel-generator@1.0.0"),
            server_name=os.getenv("HOSTNAME", "local"),
            
            # Error enrichment
            before_send=before_send_filter,
            attach_stacktrace=True,
            include_local_variables=True,  # SDK 2.x: Replaces with_locals
            max_breadcrumbs=100,
            
            # Request handling (SDK 2.x syntax)
            max_request_body_size="medium",  # SDK 2.x: Replaces request_bodies
            
            # Enable source context for better stack traces
            include_source_context=True,

            # Shutdown timeout for graceful error flushing
            shutdown_timeout=5,

            # Sentry Logs (SDK >= 2.35.0): forward records captured by
            # LoggingIntegration into the Sentry Logs product in addition
            # to breadcrumbs/events. Gated opt-in via SENTRY_ENABLE_LOGS
            # because INFO-level call sites in this engine can include
            # row/cell debug content (foreman, dept, job, WR, prices)
            # that must not be exfiltrated to Sentry by default. Set
            # SENTRY_ENABLE_LOGS=true only after auditing log call sites
            # and keeping PER_CELL_DEBUG_ENABLED / row sampling off.
            enable_logs=_parse_sentry_enable_logs(
                os.getenv("SENTRY_ENABLE_LOGS")
            ),

            # Defense-in-depth PII sanitizer for the Logs product. Even
            # when the env gate is on, drop records that embed known
            # row-level markers (see _PII_LOG_MARKERS above).
            # Cast to Any: the SDK's typed signature is (Log, Hint) -> Log | None,
            # but our hook is intentionally generic over dict/object records.
            before_send_log=cast(Any, sentry_before_send_log),
        )
        
        # Set user context (SDK 2.x: top-level API)
        sentry_sdk.set_user({
            "id": "excel_generator",
            "username": "weekly_pdf_generator",
            "segment": "billing_automation"
        })
        
        # Set global tags for all events (SDK 2.x: top-level API)
        sentry_sdk.set_tag("component", "excel_generation")
        sentry_sdk.set_tag("process", "weekly_reports")
        sentry_sdk.set_tag("test_mode", str(TEST_MODE))
        sentry_sdk.set_tag("github_actions", str(bool(os.getenv('GITHUB_ACTIONS'))))
        # PII-safe run-mode tags for issue filtering (no raw WR list - WR numbers are row-PII;
        # set_tag bypasses before_send_log so only booleans/enums/counts are permitted here)
        sentry_sdk.set_tag("res_grouping_mode", RES_GROUPING_MODE)
        sentry_sdk.set_tag("wr_filter_active", str(bool(WR_FILTER)))   # BOOL, never the WR list
        sentry_sdk.set_tag("force_generation", str(FORCE_GENERATION))

        # Set initial context (SDK 2.x: top-level API)
        sentry_sdk.set_context("configuration", {
            "max_groups": MAX_GROUPS,
            "extended_change_detection": EXTENDED_CHANGE_DETECTION,
            "use_discovery_cache": USE_DISCOVERY_CACHE,
            "force_generation": FORCE_GENERATION,
            # was: "wr_filter": WR_FILTER  (raw WR list - row-PII; set_context bypasses
            # before_send_log so the list would reach Sentry servers on every init)
            "wr_filter_active": bool(WR_FILTER),
            "wr_filter_count": len(WR_FILTER),
        })
        
        logger = logging.getLogger(__name__)
        logging.info("🛡️ Sentry.io error monitoring initialized (SDK 2.x)")
    except Exception as e:
        logging.warning(f"⚠️ Sentry initialization failed: {e}")
        logger = logging.getLogger(__name__)
else:
    logger = logging.getLogger(__name__)
    logging.warning("⚠️ SENTRY_DSN not configured - error monitoring disabled")

# --- UTILITY FUNCTIONS ---

def parse_price(price_str: str | float | int | None) -> float:
    """Safely convert a price string to a float.
    
    Args:
        price_str: Price value as string, float, int, or None
    
    Returns:
        float: Parsed price value, or 0.0 if parsing fails
    """
    if not price_str:
        return 0.0
    try:
        return float(str(price_str).replace('$', '').replace(',', ''))
    except (ValueError, TypeError):
        return 0.0

def load_contract_rates(filepath):
    """Loads contract rates into a fast lookup dictionary."""
    rates = {}
    REQUIRED_HEADERS = {'CU', 'Install Price', 'Removal Price', 'Transfer Price'}
    if not os.path.isfile(filepath):
        # Optional/retired rate CSV absent (e.g. pinned-empty OLD_RATES_CSV
        # resolving to its uncommitted default 'CU List - Corpus North & South.csv').
        # Benign - skip cleanly. INFO (not error) so LoggingIntegration
        # (event_level=ERROR) does NOT fire a Sentry event every run.
        logging.info(f"Rate CSV not present, skipping load: {filepath}")
        sentry_add_breadcrumb(
            "rate_loading", "rate CSV absent - skipped",
            level="info", data={"path_present": False},
        )
        return rates
    try:
        with open(filepath, mode='r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            if not reader.fieldnames or not REQUIRED_HEADERS.issubset(set(reader.fieldnames)):
                missing = REQUIRED_HEADERS - set(reader.fieldnames or [])
                logging.error(f"CSV {filepath} missing required headers: {missing}")
                return rates
            for row in reader:
                cu = str(row.get('CU', '')).strip().upper()
                if not cu:
                    continue

                rates[cu] = {
                    'install': parse_price(row.get('Install Price', 0)),
                    'removal': parse_price(row.get('Removal Price', 0)),
                    'transfer': parse_price(row.get('Transfer Price', 0))
                }
        logging.info(f"Loaded {len(rates)} CU rates from {filepath}")
    except Exception as e:
        logging.error(f"Failed to load rates from {filepath}: {e}")
        sentry_capture_with_context(
            e,
            context_name="rate_loading",
            context_data={
                "file_present": True,
                "error": _redact_exception_message(e),
            },
            tags={"phase": "rate_load"},
            fingerprint=["rate-csv-load-failure", "load_contract_rates"],
        )
    return rates


def load_new_contract_rates(filepath):
    """Load new contract rates from the 2026 format CSV (group-level codes).

    The new CSV has 3 metadata rows before data, with columns by position:
    [0]=Group Code, [1]=Description, [2]=UOM, [3]=Category, [4]=Region,
    [5]=Date, [6]=Install Price, [7]=Removal Price, [8]=Transfer Price.

    Returns:
        dict: {GROUP_CODE: {install: float, removal: float, transfer: float}}
    """
    rates = {}
    try:
        with open(filepath, mode='r', encoding='utf-8-sig') as f:
            reader = csv.reader(f)
            # Skip 3 metadata/header rows
            for _ in range(3):
                next(reader, None)
            for row in reader:
                if len(row) < 9:
                    continue
                group_code = row[0].strip().upper()
                if not group_code:
                    continue
                rates[group_code] = {
                    'install': parse_price(row[6]),
                    'removal': parse_price(row[7]),
                    'transfer': parse_price(row[8]),
                }
        logging.info(f"Loaded {len(rates)} group-level rates from {filepath}")
    except Exception as e:
        logging.error(f"Failed to load new contract rates from {filepath}: {e}")
    return rates


def build_cu_to_group_mapping(old_csv_path):
    """Build a mapping from detailed CU codes to Compatible Unit Group codes.

    Reads the old-format CSV which has both 'CU' (detailed code) and
    'Compatible Unit Group' columns.

    Returns:
        dict: {DETAILED_CU_CODE: GROUP_CODE} e.g. {'ANC-DHM-10-84-D1': 'ANC-M'}
    """
    mapping = {}
    if not os.path.isfile(old_csv_path):
        # Optional/retired rate CSV absent (e.g. pinned-empty OLD_RATES_CSV
        # resolving to its uncommitted default 'CU List - Corpus North & South.csv').
        # Benign - skip cleanly. INFO (not error) so LoggingIntegration
        # (event_level=ERROR) does NOT fire a Sentry event every run.
        logging.info(f"Rate CSV not present, skipping load: {old_csv_path}")
        sentry_add_breadcrumb(
            "rate_loading", "rate CSV absent - skipped",
            level="info", data={"path_present": False},
        )
        return mapping
    try:
        with open(old_csv_path, mode='r', encoding='utf-8-sig') as f:
            reader = csv.DictReader(f)
            if not reader.fieldnames or 'CU' not in reader.fieldnames or 'Compatible Unit Group' not in reader.fieldnames:
                logging.error(f"CSV {old_csv_path} missing 'CU' or 'Compatible Unit Group' columns for mapping")
                return mapping
            for row in reader:
                cu = str(row.get('CU', '')).strip().upper()
                group = str(row.get('Compatible Unit Group', '')).strip().upper()
                if cu and group:
                    mapping[cu] = group
        logging.info(f"Built CU-to-group mapping: {len(mapping)} CU codes -> groups")
    except Exception as e:
        logging.error(f"Failed to build CU-to-group mapping from {old_csv_path}: {e}")
        sentry_capture_with_context(
            e,
            context_name="rate_loading",
            context_data={
                "file_present": True,
                "error": _redact_exception_message(e),
            },
            tags={"phase": "rate_load"},
            fingerprint=["rate-csv-load-failure", "build_cu_to_group_mapping"],
        )
    return mapping


def _compute_rates_fingerprint(rates_dict):
    """Compute a short SHA256 fingerprint of a rates dictionary for hash invalidation."""
    h = hashlib.sha256()
    for code in sorted(rates_dict.keys()):
        r = rates_dict[code]
        h.update(f"{code}:{r['install']:.2f},{r['removal']:.2f},{r['transfer']:.2f}\n".encode())
    return h.hexdigest()[:12]


# Headers the subcontractor-rates loader must see in the CSV. Defined
# at module scope so the column-shape contract is documented in one
# place — Plan 2 (parser extension) and Plan 3 (variant emission)
# both read the same nine fields per row.
_SUBCONTRACTOR_RATES_REQUIRED_HEADERS: frozenset[str] = frozenset({
    'CU',
    'Install Price (Subcontractor Rates)',
    'Removal Price (Subcontractor Rates)',
    'Transfer Price (Subcontractor Rates)',
    'Install Price (New Rates)',
    'Removal Price (New Rates)',
    'Transfer Price (New Rates)',
})


def _strip_csv_fieldnames(fieldnames: Sequence[str] | None) -> dict[str, str]:
    """Map stripped header → original header for an operator CSV.

    The operator-supplied subcontractor rates CSV uses space-padded
    column headers (e.g. ``' CU                       '``) so a
    ``csv.DictReader``'s ``row.get('CU')`` would miss every value. This
    helper produces a stripped-form → raw-form mapping so the loader
    can ``row.get(raw)`` after looking up the desired header by its
    stripped form. Returns ``{}`` if ``fieldnames`` is falsy.
    """
    if not fieldnames:
        return {}
    return {(name or '').strip(): name for name in fieldnames}


def load_subcontractor_rates(filepath: str) -> dict[str, dict]:
    """Load the subcontractor rates CSV into a CU-keyed dict.

    Per Phase 1 decisions D-04..D-07 + D-20 (see
    ``.planning/phases/01-subcontractor-rate-logic-modification/
    01-CONTEXT.md``):

    - ``encoding='utf-8-sig'`` tolerates a UTF-8 BOM at file start.
    - Header matching strips whitespace so the operator-padded
      headers in the supplied CSV (``' CU                       '``)
      match the canonical ``CU`` key.
    - Price cells go through :func:`parse_price` which strips ``$``
      and thousands-comma. ``N/A`` and other non-numeric values
      coerce to ``0.0``.
    - Rows whose all six priced columns are zero are skipped
      (placeholder / inactive CUs — 1058 of 4848 in the supplied
      file). They are NOT counted as missing-CU telemetry.
    - Literal values are read for every priced field; the loader does
      NOT compute ``reduced = new × 0.87`` or ``new = old × 1.03``
      shortcuts (per-CU variance is real per ``contract-schema.md``).
    - ``Old-Rates`` columns (12-14) and ``Hours`` columns (6-8) are
      NOT loaded — the operator file retains them for human audit
      only; carrying them in memory would invite accidental code
      reuse and create a 3rd source of truth.

    Returns a CU-keyed dict shaped:

    .. code-block:: python

       {
           'cu_code': str,                  # uppercased
           'cu_wbs': str,                   # audit-only
           'compatible_unit_group': str,    # audit-only
           'reduced_install_price': float,
           'reduced_remove_price': float,
           'reduced_transfer_price': float,
           'new_install_price': float,
           'new_remove_price': float,
           'new_transfer_price': float,
       }

    Returns ``{}`` on any failure (fail-safe contract). Never raises
    into the caller — every Phase 1 plan downstream depends on this
    helper degrading gracefully.
    """
    rates: dict[str, dict] = {}
    try:
        with open(filepath, mode='r', encoding='utf-8-sig', newline='') as f:
            # ``skipinitialspace=True`` is mandatory for the operator-
            # supplied CSV: every field is left-padded with spaces
            # (``CU-1    , ADDITEM-ROW-PURCHASE     , EA             ,``).
            # Without it, the leading space before each ``"`` in a
            # quoted description (``, "Additional Item, Right of Way,
            # Purchase",``) breaks Python's csv quote recognition and
            # silently 2-column-shifts every value to the right. The
            # symptom was ALB-6-AUR1's ``reduced_install_price`` being
            # read as ``0.176`` (the Removal Hours column) instead of
            # ``$45.95`` (the actual Install Subcontractor price).
            reader = csv.DictReader(f, skipinitialspace=True)
            stripped_to_raw = _strip_csv_fieldnames(reader.fieldnames)
            missing = _SUBCONTRACTOR_RATES_REQUIRED_HEADERS - set(stripped_to_raw.keys())
            if missing:
                logging.error(
                    f"Subcontractor rates CSV {filepath} missing "
                    f"required headers: {sorted(missing)}"
                )
                return rates

            def _cell(row: dict, stripped_name: str, default: "str | float | int | None" = '') -> "str | float | int | None":
                raw = stripped_to_raw.get(stripped_name)
                if raw is None:
                    return default
                value = row.get(raw, default)
                if isinstance(value, str):
                    return value.strip()
                if isinstance(value, (int, float)) or value is None:
                    return value
                return str(value)

            for row in reader:
                cu = str(_cell(row, 'CU', '') or '').upper()
                if not cu:
                    continue

                reduced_install = parse_price(
                    _cell(row, 'Install Price (Subcontractor Rates)', 0))
                reduced_remove = parse_price(
                    _cell(row, 'Removal Price (Subcontractor Rates)', 0))
                reduced_transfer = parse_price(
                    _cell(row, 'Transfer Price (Subcontractor Rates)', 0))
                new_install = parse_price(
                    _cell(row, 'Install Price (New Rates)', 0))
                new_remove = parse_price(
                    _cell(row, 'Removal Price (New Rates)', 0))
                new_transfer = parse_price(
                    _cell(row, 'Transfer Price (New Rates)', 0))

                # D-04: skip rows whose all six priced cells are zero
                # (placeholder CUs). They are NOT counted as "missing"
                # — they're legitimately blank in the operator file.
                if (
                    reduced_install == 0
                    and reduced_remove == 0
                    and reduced_transfer == 0
                    and new_install == 0
                    and new_remove == 0
                    and new_transfer == 0
                ):
                    continue

                # D-05: 9 literal fields per row. D-06: explicitly do
                # NOT include Old-Rates (cols 12-14) or Hours
                # (cols 6-8) — they stay reference-only on disk.
                rates[cu] = {
                    'cu_code': cu,
                    'cu_wbs': str(_cell(row, 'CU WBS #', '') or ''),
                    'compatible_unit_group': str(
                        _cell(row, 'Compatible Unit Group', '') or ''
                    ),
                    'reduced_install_price': reduced_install,
                    'reduced_remove_price': reduced_remove,
                    'reduced_transfer_price': reduced_transfer,
                    'new_install_price': new_install,
                    'new_remove_price': new_remove,
                    'new_transfer_price': new_transfer,
                }
        logging.info(
            f"Loaded {len(rates)} subcontractor CU rates from {filepath}"
        )
    except Exception as e:
        # Fail-safe contract: never raise into the caller. Surface the
        # filepath but not row content (no PII risk: only the integer
        # count would have been logged on success, and ``e`` here is
        # an open / csv / encoding error, not a row-level message).
        logging.error(
            f"Failed to load subcontractor rates from {filepath}: {e}"
        )
    return rates


def _compute_subcontractor_rates_fingerprint(
    rates_dict: dict[str, dict],
) -> str:
    """Return a deterministic 16-char SHA256 prefix over the six priced
    fields of every CU in ``rates_dict``.

    Per Phase 1 decision D-20: sorted keys + fixed precision guarantees
    byte-identical output for byte-identical input across runs and
    across machines. Dict-insertion-order does NOT influence the
    output. Editing any priced field on any CU MUST change the
    fingerprint.

    Returns ``''`` for an empty dict (matches the legacy
    ``_compute_rates_fingerprint`` convention used elsewhere when the
    rates table is empty).
    """
    if not rates_dict:
        return ''
    h = hashlib.sha256()
    for code in sorted(rates_dict.keys()):
        r = rates_dict[code]
        h.update(
            (
                f"{code}:"
                f"{r['reduced_install_price']:.2f},"
                f"{r['reduced_remove_price']:.2f},"
                f"{r['reduced_transfer_price']:.2f},"
                f"{r['new_install_price']:.2f},"
                f"{r['new_remove_price']:.2f},"
                f"{r['new_transfer_price']:.2f}\n"
            ).encode()
        )
    return h.hexdigest()[:16]


# Per D-20: load the subcontractor rate matrix once at module init so
# downstream plans (2-6) can read ``_SUBCONTRACTOR_RATES`` and
# ``_SUBCONTRACTOR_RATES_FINGERPRINT`` directly without re-parsing
# the CSV per WR group. When the kill switch is OFF, neither value is
# populated — every downstream consumer must short-circuit on the
# empty dict path so the pipeline behaves identically to pre-Phase-1.
_SUBCONTRACTOR_RATES: dict[str, dict] = (
    load_subcontractor_rates(SUBCONTRACTOR_RATES_CSV)
    if SUBCONTRACTOR_RATE_VARIANTS_ENABLED
    else {}
)
_SUBCONTRACTOR_RATES_FINGERPRINT: str = (
    _compute_subcontractor_rates_fingerprint(_SUBCONTRACTOR_RATES)
    if _SUBCONTRACTOR_RATES
    else ''
)

if SUBCONTRACTOR_RATE_VARIANTS_ENABLED and _SUBCONTRACTOR_RATES:
    logging.info(
        f"📊 Subcontractor rates loaded: "
        f"{len(_SUBCONTRACTOR_RATES)} CUs, "
        f"fingerprint={_SUBCONTRACTOR_RATES_FINGERPRINT}"
    )
elif SUBCONTRACTOR_RATE_VARIANTS_ENABLED:
    # Kill switch is on but the dict is empty — the loader's own
    # ``logging.error`` already surfaced the failure cause; flag it
    # here too so operators tailing the startup banner notice that
    # the downstream variant pipeline will be a no-op this run.
    logging.warning(
        "⚠️ Subcontractor rates table is empty — _AEPBillable / "
        "_ReducedSub variant generation will be skipped this run "
        f"(SUBCONTRACTOR_RATES_CSV='{SUBCONTRACTOR_RATES_CSV}')"
    )


def _subcontractor_rescue_price(row_data: dict) -> float:
    """Phase 1.1 Bug A pre-acceptance rescue. Returns reduced-sub
    price * qty OR 0.0 (caller observes no rescue and the row drops
    at the existing has_price gate, same as legacy behaviour).

    Uses reduced_*_price as the safety floor regardless of AEP cutoff:
    the cutoff is variant-emission gating only; this helper is for
    pre-acceptance row admission, not for output pricing. The
    actual price written to Excel happens later in
    ``_resolve_row_price`` at the ``generate_excel`` call site,
    AFTER row acceptance.

    Reads ONLY canonical keys (``CU`` / ``Work Type`` / ``Quantity``)
    per Phase 1 Blocker 2 lock-in. Per [2026-05-16 23:45] ledger
    rule, the work-type matcher uses the SHORTEST UNAMBIGUOUS
    PREFIX as ``A`` in the ``A in B`` substring direction so that
    operator-entered abbreviations (``Inst`` / ``Rem`` / ``Trans``
    / ``Xfr``) AND full canonical forms (``Install`` / ``Removal``
    / ``Transfer``) both match.
    """
    cu = str(row_data.get('CU') or '').strip().upper()
    rate_row = _SUBCONTRACTOR_RATES.get(cu)
    if rate_row is None:
        return 0.0
    work_type_raw = (row_data.get('Work Type') or '').strip().lower()
    if 'inst' in work_type_raw:
        rate = rate_row.get('reduced_install_price', 0.0)
    elif 'rem' in work_type_raw:
        rate = rate_row.get('reduced_remove_price', 0.0)
    elif 'tran' in work_type_raw or 'xfr' in work_type_raw:
        rate = rate_row.get('reduced_transfer_price', 0.0)
    else:
        return 0.0
    qty_raw = row_data.get('Quantity', 0)
    try:
        qty = float(qty_raw) if qty_raw not in (None, '') else 0.0
    except (TypeError, ValueError):
        qty = 0.0
    if rate <= 0 or qty <= 0:
        return 0.0
    return rate * qty


def _resolve_row_price(row: dict, variant: str, missing_cus) -> float:
    """Return the per-row ``Units Total Price`` to write to Excel.

    Phase 01 Plan 03 Task 2 — variant-aware pricing helper used by
    ``generate_excel``'s row-write loop.

    For ``primary`` / ``helper`` / ``vac_crew`` rows (D-14 / D-15):
    return the existing SmartSheet ``Units Total Price`` value
    unchanged via ``parse_price``. Existing variant outputs MUST be
    byte-identical to pre-change behaviour, so this branch is the
    short-circuit path for the legacy variants.

    For ``aep_billable`` / ``reduced_sub`` / ``aep_billable_helper`` /
    ``reduced_sub_helper`` rows (D-08 / D-16):
      • Look up the canonical ``CU`` code in ``_SUBCONTRACTOR_RATES``.
      • Select the rate column from the canonical ``Work Type`` token
        (Install / Removal / Transfer, case-insensitive substring
        match per D-05).
      • For the AEP-Billable family use the ``new_*_price`` columns;
        for the Reduced-Sub family use the ``reduced_*_price`` columns.
      • Return ``rate × quantity`` from the canonical ``Quantity`` key.
      • If the CU is missing from the rates table, retain the row's
        SmartSheet ``Units Total Price`` (NEVER zero-out, NEVER raise)
        and record the missing code in the per-call ``missing_cus``
        ``collections.Counter[str]``. The fall-through pattern mirrors
        the recalc fall-through from Living Ledger 2026-04-21 22:35:
        silent zero-out is a correctness regression in the billing
        pipeline.
      • For unknown work types, degenerate quantities, or non-positive
        rates, the same SmartSheet fall-through applies as a safety
        floor.

    Canonical column-name discipline (Blocker 2 lock-in):
    The helper reads ONLY the canonical keys produced by
    ``_validate_single_sheet``'s synonyms layer at L2523-2547:

      * ``row['CU']`` — canonical CU code key (synonyms ``'CU'`` and
        ``'Billable Unit Code'`` BOTH map to ``row['CU']`` upstream;
        only the canonical key survives at this point in the pipeline).
      * ``row['Work Type']`` — canonical work-type key (no synonyms).
      * ``row['Quantity']`` — canonical quantity key (synonyms ``'Qty'``
        and ``'# Units'`` map to ``row['Quantity']`` upstream).
      * ``row['Units Total Price']`` — canonical price key (synonyms
        ``'Total Price'``, ``'Redlined Total Price'`` map upstream).

    Reading any other key here would be a silent regression — only
    the canonical keys exist by the time the row reaches
    ``generate_excel``. Future synonym additions go in
    ``_validate_single_sheet``, NOT here.

    Args:
        row: Group row dict (already passed through
            ``_validate_single_sheet`` synonyms layer).
        variant: One of ``{primary, helper, vac_crew, aep_billable,
            reduced_sub, aep_billable_helper, reduced_sub_helper}``.
        missing_cus: Per-call ``collections.Counter[str]`` accumulated
            across the row-write loop. Caller is responsible for
            instantiation and downstream forwarding.

    Returns:
        float: The price to write to the row's ``Units Total Price``
            cell. SmartSheet value for legacy variants / missing CUs;
            ``rate × qty`` for new variants with a known CU.
    """
    # Legacy variants short-circuit immediately — preserves the
    # D-14 / D-15 byte-identical guarantee for existing outputs.
    if variant not in (
        'aep_billable', 'reduced_sub',
        'aep_billable_helper', 'reduced_sub_helper',
    ):
        return parse_price(row.get('Units Total Price'))

    # Subcontractor variants: rate × qty from the rate matrix.
    # CU canonical key ONLY — see Blocker 2 docstring above.
    cu_raw = row.get('CU') or ''
    cu = str(cu_raw).strip().upper()
    rate_row = _SUBCONTRACTOR_RATES.get(cu)
    if rate_row is None:
        # Missing CU: record + fall through to SmartSheet (D-16).
        if cu:
            missing_cus[cu] += 1
        return parse_price(row.get('Units Total Price'))

    # Work-Type-keyed column selection (D-05). Canonical 'Work Type'.
    # Production hotfix 2026-05-16: Smartsheet operators commonly enter
    # the abbreviated forms 'Inst' / 'Rem' / 'Trans' / 'Xfr' rather
    # than the canonical 'Install' / 'Removal' / 'Transfer'. The
    # pre-fix matcher checked `'install' in work_type_raw` — a
    # substring test that succeeds on the full form but FAILS on the
    # abbreviation ('install' is 7 chars; 'inst' is 4 chars; the
    # 7-char string is NOT contained in the 4-char string). Fall-
    # through to the safety floor returned `Units Total Price` for
    # BOTH variants, producing byte-identical AEP and ReducedSub
    # workbooks (verified via SHA256 on 8 of 8 file pairs from GHA
    # run 25975684465). Aligned with the existing
    # ``recalculate_row_price`` pattern at L1655 — use the shortest
    # unambiguous prefix so both abbreviated AND full forms match.
    work_type_raw = (row.get('Work Type') or '').strip().lower()
    if 'inst' in work_type_raw:  # matches 'Inst', 'Install', 'Installation'
        wt = 'install'
    elif 'rem' in work_type_raw:  # matches 'Rem', 'Remov', 'Removal', 'Remove'
        wt = 'remove'
    elif 'tran' in work_type_raw or 'xfr' in work_type_raw:  # 'Tran'/'Trans'/'Transfer'/'Xfr'
        wt = 'transfer'
    else:
        # Unknown work type: keep SmartSheet pricing (safety floor).
        return parse_price(row.get('Units Total Price'))

    if variant in ('aep_billable', 'aep_billable_helper'):
        rate = rate_row.get(f'new_{wt}_price', 0.0)
    else:  # reduced_sub / reduced_sub_helper
        rate = rate_row.get(f'reduced_{wt}_price', 0.0)

    # Canonical 'Quantity' ONLY — never 'Units Completed' (checkbox).
    # Phase 01 gap closure (REVIEW-IN-02): explicit None / empty-string
    # handling. The previous ``or 0`` short-circuit collapsed
    # legitimate ``Quantity=0.0`` to int ``0`` (functionally correct
    # after the subsequent ``float()`` coercion but opaque to readers).
    # Numeric output is byte-identical for every pre-existing input case
    # (None, '', 0, 0.0, '1.5', invalid → 0.0 / 0.0 / 0.0 / 0.0 / 1.5
    # / 0.0 respectively).
    qty_raw = row.get('Quantity', 0)
    try:
        qty = float(qty_raw) if qty_raw not in (None, '') else 0.0
    except (TypeError, ValueError):
        qty = 0.0

    if rate <= 0 or qty <= 0:
        # Degenerate row: SmartSheet pricing as the safety floor,
        # NEVER silently zero out (mirrors the recalc fall-through
        # pattern in Living Ledger 2026-04-21 22:35).
        return parse_price(row.get('Units Total Price'))
    return rate * qty


def load_rate_versions():
    """Load new rate versions and build all necessary lookup structures.

    Returns:
        tuple: (cu_to_group, new_rates_primary, new_rates_arrowhead, rates_fingerprint)
            - cu_to_group: {DETAILED_CU: GROUP_CODE}
            - new_rates_primary: {GROUP_CODE: {install, removal, transfer}}
            - new_rates_arrowhead: {GROUP_CODE: {install, removal, transfer}} (rates * 0.90)
            - rates_fingerprint: short hash of rate table contents for cache invalidation
    """
    cu_to_group = build_cu_to_group_mapping(OLD_RATES_CSV)
    new_rates_primary = load_new_contract_rates(NEW_RATES_CSV)

    # Precompute Arrowhead (subcontractor) rates: primary rate * ARROWHEAD_DISCOUNT
    new_rates_arrowhead = {}
    for group_code, rates in new_rates_primary.items():
        new_rates_arrowhead[group_code] = {
            'install': round(rates['install'] * ARROWHEAD_DISCOUNT, 2),
            'removal': round(rates['removal'] * ARROWHEAD_DISCOUNT, 2),
            'transfer': round(rates['transfer'] * ARROWHEAD_DISCOUNT, 2),
        }

    # Only compute fingerprint if rates were successfully loaded
    rates_fingerprint = _compute_rates_fingerprint(new_rates_primary) if new_rates_primary else ''

    if not new_rates_primary:
        logging.warning("⚠️ New rate table is empty — rate recalculation will be skipped even though RATE_CUTOFF_DATE is set")
    else:
        logging.info(f"Rate versions loaded: {len(new_rates_primary)} primary groups, "
                     f"{len(new_rates_arrowhead)} Arrowhead groups (precomputed, not yet active), "
                     f"{len(cu_to_group)} CU-to-group mappings, "
                     f"fingerprint={rates_fingerprint}")
    return cu_to_group, new_rates_primary, new_rates_arrowhead, rates_fingerprint


def _resolve_cu_code(row_data):
    """Return the CU code for a row using the same priority chain as
    ``recalculate_row_price`` and ``revert_subcontractor_price``.

    Resolution order: ``CU Helper`` → ``CU`` → ``Billable Unit Code``,
    with the sentinel string ``'NAN'`` (produced by pandas when a float
    NaN is stringified) falling back to ``CU``. Returns an uppercased,
    stripped string (possibly empty if no column yields a value).
    """
    cu_code = str(
        row_data.get('CU Helper')
        or row_data.get('CU')
        or row_data.get('Billable Unit Code')
        or ''
    ).strip().upper()
    if cu_code == 'NAN':
        cu_code = str(row_data.get('CU') or '').strip().upper()
    return cu_code


def recalculate_row_price(row_data, cu_to_group, rates_dict, *, out_status=None):
    """Recalculate a row's price using new contract rates.

    Looks up the CU code, maps it to its Compatible Unit Group, then
    calculates price as rate × quantity using the provided rates dict.
    Modifies row_data['Units Total Price'] in-place if a matching rate is found.

    Args:
        row_data: Dict of row field values (modified in-place).
        cu_to_group: Dict mapping detailed CU codes to group codes.
        rates_dict: Dict mapping group codes to {install, removal, transfer} rates.
        out_status: Optional dict that, when provided, receives an
            ``'outcome'`` key describing why the function returned:
              * ``'recalculated'`` — a rate was successfully applied
                (the returned price may still equal the original
                SmartSheet price if the computed rate × qty matches it
                exactly; lookup succeeded either way).
              * ``'missing_rate'`` — neither the mapped group nor the
                CU code is present in ``rates_dict``; SmartSheet price
                retained. This is the only outcome the per-sheet
                "skipped" summary counts toward, because it is the
                actionable signal that ``NEW_RATES_CSV`` needs a new
                entry.
              * ``'invalid_quantity'`` — quantity is zero/missing/
                unparseable; SmartSheet price retained.
              * ``'zero_rate'`` — new rate for the resolved work type
                is zero; SmartSheet price retained.

    Returns:
        float: The (possibly recalculated) price value.
    """
    def _set_status(s):
        if out_status is not None:
            out_status['outcome'] = s

    price_val = parse_price(row_data.get('Units Total Price'))

    # Resolve CU code (same chain as revert_subcontractor_price)
    cu_code = _resolve_cu_code(row_data)

    # Map detailed CU code to group code
    group_code = cu_to_group.get(cu_code)
    if not group_code:
        # CU code not found in mapping — try direct group lookup (in case SmartSheet uses group codes)
        if cu_code in rates_dict:
            group_code = cu_code
        else:
            logging.debug(f"Rate recalculation: CU '{cu_code}' not found in CU-to-group mapping or rates, keeping SmartSheet price")
            _set_status('missing_rate')
            return price_val

    # If the mapped group isn't in the new rates table (e.g. old CSV maps
    # CU -> a verbose group name like "Vacuum Switch" that never appears
    # in the new rates' short-code keys), fall back to looking up the CU
    # code directly in the rates table. This recovers specialized work
    # items (common on VAC crew sheets) where the detailed CU code is
    # itself a key in the new contract rates. Only activates on exact
    # match, so no chance of mis-applying a rate.
    if group_code not in rates_dict:
        if cu_code in rates_dict:
            logging.debug(f"Rate recalculation: mapped group '{group_code}' not in new rates for CU '{cu_code}'; matched CU directly")
            group_code = cu_code
        else:
            logging.warning(f"Rate recalculation SKIPPED: CU '{cu_code}' maps to group '{group_code}' but neither is in new rates — keeping SmartSheet price (Qty={row_data.get('Quantity')}, Work Type={row_data.get('Work Type')})")
            _set_status('missing_rate')
            return price_val

    # Determine work type
    work_type_raw = str(row_data.get('Work Type') or '').strip().lower()
    wt_key = 'install'
    if 'rem' in work_type_raw:
        wt_key = 'removal'
    elif 'tran' in work_type_raw or 'xfr' in work_type_raw:
        wt_key = 'transfer'

    # Parse quantity — if missing or unparseable, keep SmartSheet price
    qty_str = str(row_data.get('Quantity', '') or '')
    try:
        qty = float(_RE_EXTRACT_NUMBERS.sub('', qty_str) or 0)
    except ValueError:
        qty = 0.0

    if qty <= 0:
        logging.debug(f"Rate recalculation: quantity '{qty_str}' is zero/missing for CU '{cu_code}', keeping SmartSheet price")
        _set_status('invalid_quantity')
        return price_val

    rate = rates_dict[group_code].get(wt_key, 0.0)
    if rate <= 0:
        logging.debug(f"Rate recalculation: rate is zero for group '{group_code}' work type '{wt_key}', keeping SmartSheet price")
        _set_status('zero_rate')
        return price_val

    new_price = round(rate * qty, 2)
    row_data['Units Total Price'] = new_price
    _set_status('recalculated')
    return new_price


def revert_subcontractor_price(row_data, original_rates):
    """Revert a subcontractor row's price to the 100% original contract rate.

    Looks up the CU code from row_data (preferring CU Helper, then CU,
    then Billable Unit Code), determines the work type, and recalculates
    the price as original_rate × quantity.

    Args:
        row_data: Dict of row field values (modified in-place if reverted).
        original_rates: Dict mapping CU codes to {install, removal, transfer} rates.

    Returns:
        The (possibly recalculated) price value as a float.
    """
    price_val = parse_price(row_data.get('Units Total Price'))

    cu_code = _resolve_cu_code(row_data)

    work_type_raw = str(row_data.get('Work Type') or '').strip().lower()

    wt_key = 'install'
    if 'rem' in work_type_raw:
        wt_key = 'removal'
    elif 'tran' in work_type_raw or 'xfr' in work_type_raw:
        wt_key = 'transfer'

    qty_str = str(row_data.get('Quantity', '') or '0')
    try:
        qty = float(_RE_EXTRACT_NUMBERS.sub('', qty_str) or 0)
    except ValueError:
        qty = 0.0

    if cu_code in original_rates:
        exact_original_rate = original_rates[cu_code].get(wt_key, 0.0)
        price_val = round(exact_original_rate * qty, 2)
        row_data['Units Total Price'] = price_val

    return price_val

def is_checked(value: bool | int | str | None) -> bool:
    """Check if a checkbox value is considered checked/true.
    
    Args:
        value: Checkbox value in various formats
    
    Returns:
        bool: True if the value represents a checked state
    """
    if value is None:
        return False
    if isinstance(value, bool):
        return value
    if isinstance(value, int):
        return value == 1
    if isinstance(value, str):
        return value.strip().lower() in ('true', 'checked', 'yes', '1', 'on')
    return False

def excel_serial_to_date(value):
    """Strict date parsing: return datetime or None. No numeric/serial fallbacks.
    
    PERFORMANCE: Fast-path for ISO format dates (YYYY-MM-DD) before falling back
    to the slower dateutil.parser.parse() for other formats.
    """
    if value in (None, ""):
        return None
    if isinstance(value, datetime.datetime):
        return value
    if isinstance(value, datetime.date):
        return datetime.datetime.combine(value, datetime.time.min)
    s = str(value).strip()
    # PERFORMANCE: Fast-path for ISO date format (most common in Smartsheet)
    if len(s) >= 10 and s[4] == '-' and s[7] == '-':
        try:
            # Try ISO format first (YYYY-MM-DD or YYYY-MM-DDTHH:MM:SS)
            date_part = s[:10]
            return datetime.datetime.strptime(date_part, '%Y-%m-%d')
        except ValueError:
            pass  # Fall through to general parser
    try:
        dt = parser.parse(s)
        if isinstance(dt, datetime.datetime):
            return dt
        return datetime.datetime.combine(dt, datetime.time.min)
    except Exception:
        return None


def _resolve_rate_recalc_cutoff_date(
    row_data,
    cutoff_date,
    *,
    weekly_fallback_enabled: bool = True,
):
    """Return the effective cutoff date for a row's pre-acceptance recalc.

    The snapshot-keyed business rule remains primary: any row with a
    populated ``Snapshot Date`` is gated on that date alone. The
    Weekly-Ref-Date fallback only activates when ``Snapshot Date`` is
    blank or unparseable — it rescues current-week rows that would
    otherwise silently skip recalc because Smartsheet's snapshot
    automation has not fired yet (the observed VAC crew failure mode).

    Args:
        row_data: Mapping that may contain ``'Snapshot Date'`` and
            ``'Weekly Reference Logged Date'`` entries.
        cutoff_date: ``datetime.date`` used as the ``>=`` threshold.
        weekly_fallback_enabled: Set False to reproduce the legacy
            snapshot-only behaviour (the ``RATE_RECALC_WEEKLY_FALLBACK``
            env var wires this in production).

    Returns:
        ``(effective_cutoff_date, used_fallback)``. Returns
        ``(None, False)`` when recalc should not run.
    """
    if cutoff_date is None:
        return (None, False)

    snapshot_raw = row_data.get('Snapshot Date')
    snap_date = None
    if snapshot_raw:
        snap_dt = excel_serial_to_date(snapshot_raw)
        if snap_dt:
            snap_date = (
                snap_dt.date() if hasattr(snap_dt, 'date') else snap_dt
            )

    if snap_date is not None and snap_date >= cutoff_date:
        return (snap_date, False)

    if snap_date is None and weekly_fallback_enabled:
        weekly_raw = row_data.get('Weekly Reference Logged Date')
        if weekly_raw:
            weekly_dt = excel_serial_to_date(weekly_raw)
            if weekly_dt:
                weekly_date = (
                    weekly_dt.date()
                    if hasattr(weekly_dt, 'date')
                    else weekly_dt
                )
                if weekly_date >= cutoff_date:
                    return (weekly_date, True)

    return (None, False)


def _weekly_would_trigger_fallback(weekly_raw, cutoff_date) -> bool:
    """Return True if a blank/unparseable Snapshot Date row would be
    rescued by flipping ``RATE_RECALC_WEEKLY_FALLBACK`` on.

    Mirrors the secondary branch of
    ``_resolve_rate_recalc_cutoff_date``: the fallback only fires when
    the weekly date parses AND is ``>= cutoff_date``. Used to gate the
    fallback-disabled operator note so it doesn't misleadingly suggest
    enabling the env var for rows whose weekly date is also blank,
    unparseable, or pre-cutoff (where flipping the gate wouldn't
    change anything).
    """
    if cutoff_date is None or not weekly_raw:
        return False
    dt = excel_serial_to_date(weekly_raw)
    if dt is None:
        return False
    wdate = dt.date() if hasattr(dt, 'date') else dt
    return wdate >= cutoff_date


def discover_folder_sheets(client, folder_ids: list[int], label: str) -> set[int]:
    """Discover all sheet IDs inside the given Smartsheet folders (recursively including subfolders).

    Args:
        client: Authenticated Smartsheet client instance.
        folder_ids: List of Smartsheet folder IDs to enumerate.
        label: Human-readable label for logging (e.g. 'subcontractor', 'original contract').

    Returns:
        Set of sheet IDs found across all folders and their subfolders.
    """
    discovered: set[int] = set()
    if not folder_ids:
        return discovered

    def _fetch_folder_recursive(fid, depth=0, max_depth=5):
        """Fetch sheets from a single folder, recursing into subfolders."""
        if depth > max_depth:
            logging.warning(f"⚠️ Max folder recursion depth ({max_depth}) reached for {label} folder {fid}")
            return set()
        try:
            with sentry_sdk.start_span(op="smartsheet.api", name=f"Get folder {fid} ({label} depth={depth})") as span:
                from smartsheet.models.sheet import Sheet as _SmartsheetSheet
                from smartsheet.models.folder import Folder as _SmartsheetFolder
                sheets: list = []
                subfolders: list = []
                last_key = None
                # Safety caps: guard against a misbehaving API that perpetually
                # returns a non-falsy or repeated last_key, which would otherwise
                # create a large API burst (amplifying Smartsheet 300 req/min limits).
                max_pages = 100
                pages_fetched = 0
                page_start_time = time.monotonic()
                seen_last_keys: set = set()
                for _page_num in range(max_pages):
                    page = client.Folders.get_folder_children(
                        fid,
                        children_resource_types=["sheets", "folders"],
                        last_key=last_key,
                    )
                    pages_fetched += 1
                    for item in getattr(page, 'data', None) or []:
                        if isinstance(item, _SmartsheetSheet):
                            sheets.append(item)
                        elif isinstance(item, _SmartsheetFolder):
                            subfolders.append(item)
                    next_last_key = getattr(page, 'last_key', None)
                    if not next_last_key:
                        break
                    if next_last_key in seen_last_keys:
                        elapsed = time.monotonic() - page_start_time
                        logging.warning(
                            f"⚠️ Repeated pagination token detected for {label} folder {fid}; "
                            f"stopping after {pages_fetched} page(s) in {elapsed:.2f}s "
                            f"with {len(sheets)} sheet(s)"
                        )
                        break
                    seen_last_keys.add(next_last_key)
                    last_key = next_last_key
                else:
                    elapsed = time.monotonic() - page_start_time
                    logging.warning(
                        f"⚠️ Pagination safety cap ({max_pages}) reached for {label} folder {fid}; "
                        f"stopping after {pages_fetched} page(s) in {elapsed:.2f}s "
                        f"with {len(sheets)} sheet(s)"
                    )
                ids = {s.id for s in sheets}
                span.set_data("folder_id", fid)
                span.set_data("sheets_found", len(sheets))
                span.set_data("depth", depth)
            logging.info(f"{'  ' * depth}📂 Folder {fid} ({label}): found {len(sheets)} direct sheet(s)")
            # Recurse into subfolders to discover ALL sheets in the hierarchy
            for subfolder in subfolders:
                sub_id = subfolder.id
                sub_ids = _fetch_folder_recursive(sub_id, depth + 1, max_depth)
                if sub_ids:
                    logging.info(f"{'  ' * (depth + 1)}📁 Subfolder {sub_id}: contributed {len(sub_ids)} sheet(s)")
                ids |= sub_ids

            return ids
        except Exception as e:
            logging.warning(f"⚠️ Could not read {label} folder {fid}: {e}")
            sentry_add_breadcrumb("folder_discovery", f"Failed to read folder {fid}", level="error", data={
                "folder_id": fid, "label": label, "depth": depth, "error": str(e)[:200],
            })
            return set()

    with ThreadPoolExecutor(max_workers=min(len(folder_ids), PARALLEL_WORKERS_DISCOVERY)) as executor:
        for ids in executor.map(lambda fid: _fetch_folder_recursive(fid), folder_ids):
            discovered |= ids

    logging.info(f"📂 Total {label} folder discovery (recursive): {len(discovered)} unique sheet(s)")
    return discovered

def calculate_data_hash(group_rows: list[dict]) -> str:
    """Calculate a hash of the group data to detect changes.

    Args:
        group_rows: List of row dictionaries to hash
    
    Returns:
        str: 16-character SHA256 hash prefix

    Legacy (EXTENDED_CHANGE_DETECTION=0):
        Uses original minimal fields so hash stability is preserved for rollbacks.

    Extended (default):
        Incorporates additional business-critical fields so regenerated Excel
        files occur when any of these change:
          • Current foreman name (derived '__current_foreman')
          • Dept # values present across the group (set + order)
          • Scope ID / #
          • Aggregated total billed amount
          • Unique dept list and row count
          • All prior minimal fields
    """
    if not group_rows:
        return "0" * 16

    # Deterministic sorting across key business fields.
    #
    # In EXTENDED mode, VAC crew fields appear as tie-breakers so multi-member
    # VAC crew groups — where two rows can share (WR, Snapshot, CU, Pole, Qty)
    # while belonging to different crew members — hash stably across runs.
    # Without the tie-breaker, row insertion order (merged from parallel
    # `as_completed` futures in `get_all_source_rows`) bleeds into the hash
    # because VAC crew fields are included per-row in row_str.
    #
    # LEGACY mode (EXTENDED_CHANGE_DETECTION=0) intentionally keeps the
    # original 5-key sort: the docstring promises legacy uses only the
    # original minimal fields so hashes stay bit-stable for rollbacks.
    # Adding tie-breakers there would change row order for tied rows whose
    # legacy row_data still differs (Work Type / Units Completed? / price),
    # invalidating the rollback-stability guarantee. Legacy mode's row_data
    # also excludes VAC crew fields, so a vac_crew-specific tie-breaker
    # there would be purely cosmetic — skip it.
    _sort_base = lambda x: (
        str(x.get('Work Request #', '')),
        str(x.get('Snapshot Date', '')),
        str(x.get('CU', '')),
        str(x.get('Pole #') or x.get('Point #') or x.get('Point Number') or ''),
        str(x.get('Quantity', '')),
    )
    if EXTENDED_CHANGE_DETECTION:
        sorted_rows = sorted(
            group_rows,
            key=lambda x: _sort_base(x) + (
                str(x.get('__vac_crew_name') or ''),
                str(x.get('__vac_crew_dept') or ''),
                str(x.get('__vac_crew_job') or ''),
            ),
        )
    else:
        sorted_rows = sorted(group_rows, key=_sort_base)

    if not EXTENDED_CHANGE_DETECTION:
        # OPTIMIZATION: Use incremental hashing to avoid large string allocation
        hasher = hashlib.sha256()
        for row in sorted_rows:
            # CRITICAL: Use parse_price() for normalization to avoid format-based false changes
            normalized_price = f"{parse_price(row.get('Units Total Price', 0)):.2f}"
            row_data = (
                f"{row.get('Work Request #', '')}"
                f"{row.get('CU', '')}"
                f"{row.get('Quantity', '')}"
                f"{normalized_price}"
                f"{row.get('Snapshot Date', '')}"
                f"{row.get('Pole #', '')}"
                f"{row.get('Work Type', '')}"
                f"{is_checked(row.get('Units Completed?'))}"  # CRITICAL: Include completion status
            )
            hasher.update(row_data.encode('utf-8'))
        return hasher.hexdigest()[:16]

    # Extended mode: Use incremental hashing
    hasher = hashlib.sha256()

    # Variant is a group-level property (all rows in a group share the same
    # __variant). Compute it once before the row loop so per-row hash can
    # include variant-scoped fields deterministically.
    group_variant = sorted_rows[0].get('__variant', 'primary') if sorted_rows else 'primary'

    group_foreman = None
    for row in sorted_rows:
        foreman = row.get('__current_foreman') or row.get('Foreman') or ''
        if group_foreman is None and foreman:
            group_foreman = foreman
        # CRITICAL: Use parse_price() for normalization to avoid format-based false changes
        normalized_price = f"{parse_price(row.get('Units Total Price', 0)):.2f}"

        row_fields = [
            str(row.get('Work Request #', '')),
            str(row.get('Snapshot Date', '') or ''),
            str(row.get('CU', '') or ''),
            str(row.get('Quantity', '') or ''),
            normalized_price,
            str(row.get('Pole #') or row.get('Point #') or row.get('Point Number') or ''),
            str(row.get('Work Type', '') or ''),
            str(row.get('Dept #', '') or ''),
            str(row.get('Scope #') or row.get('Scope ID', '') or ''),
            str(is_checked(row.get('Units Completed?'))),  # CRITICAL: Include completion status
            # Additional fields to catch changes previously missed
            str(row.get('Customer Name', '') or ''),
            str(row.get('Job #') or row.get('Job Number', '') or ''),
            str(row.get('Work Order #') or row.get('Work Order Number', '') or ''),
            str(row.get('CU Description', '') or ''),
            str(row.get('Unit of Measure', '') or ''),
            str(row.get('Area', '') or ''),
        ]
        # VAC crew groups use a single `_VACCREW` key with no foreman suffix,
        # so one group can hold multiple crew members. Including per-row VAC
        # crew fields here lets each row contribute its own name/dept/job to
        # the hash independently. This avoids two pitfalls of aggregating
        # values into `meta_parts` as a set:
        #   1. Set dedup — e.g. depts {500, 500, 600}: editing one row's dept
        #      from 500→600 leaves {500, 600} unchanged, silently skipping
        #      regeneration.
        #   2. Delimiter collision — ','.join on free-text names cannot
        #      distinguish ['A,B', 'C'] from ['A', 'B,C'].
        # Scoped to vac_crew so hash stability for primary/helper rows is
        # preserved (non-vac_crew row_str structure is unchanged).
        if group_variant == 'vac_crew':
            row_fields.extend([
                str(row.get('__vac_crew_name') or ''),
                str(row.get('__vac_crew_dept') or ''),
                str(row.get('__vac_crew_job') or ''),
            ])

        row_str = "|".join(row_fields)
        # Update hash incrementally with newline separator
        hasher.update(row_str.encode('utf-8'))
        hasher.update(b"\n")

    unique_depts = sorted({str(r.get('Dept #', '') or '') for r in sorted_rows if r.get('Dept #') is not None})
    total_price = sum(parse_price(r.get('Units Total Price')) for r in sorted_rows)

    # Append metadata
    meta_parts = []
    meta_parts.append(f"FOREMAN={group_foreman or ''}")
    
    # Variant-specific hash tokens (replaces activity log USER= token)
    variant = group_variant
    meta_parts.append(f"VARIANT={variant}")

    if variant in ('helper', 'aep_billable_helper', 'reduced_sub_helper'):
        # Helper-style variants: include helper-specific metadata
        # (helper_job is OPTIONAL). The plain helper group key is
        # `_HELPER_{helper}` and the new Phase 01 subcontractor
        # shadow-helper variants will follow the same per-foreman
        # partitioning pattern in Plan 3 (`_AEPBILLABLE_HELPER_{name}` /
        # `_REDUCEDSUB_HELPER_{name}`), so reading sorted_rows[0]
        # here is safe — every row in such a group shares identical
        # helper info.
        _first = sorted_rows[0] if sorted_rows else {}
        helper_foreman = _first.get('__helper_foreman', '')
        helper_dept = _first.get('__helper_dept', '')
        helper_job = _first.get('__helper_job', '')
        # Validate required helper fields (helper_job is optional)
        if not helper_foreman or not helper_dept:
            logging.warning(f"⚠️ Helper variant missing required fields: foreman={helper_foreman}, dept={helper_dept}")
        if not helper_job:
            logging.info(f"ℹ️ Helper variant without Job #: foreman={helper_foreman}, dept={helper_dept} (proceeding anyway)")
        meta_parts.append(f"HELPER={helper_foreman}")
        meta_parts.append(f"HELPER_DEPT={helper_dept}")
        meta_parts.append(f"HELPER_JOB={helper_job}")  # Include even if empty for hash consistency
    # vac_crew variant intentionally has no meta_parts block: VAC crew
    # name/dept/job are already captured per-row in the row_str loop above,
    # which is strictly more sensitive than meta_parts aggregation and is not
    # vulnerable to set-dedup collisions or comma-in-name delimiter collisions.

    meta_parts.append(f"DEPTS={','.join(unique_depts)}")
    meta_parts.append(f"TOTAL={total_price:.2f}")
    meta_parts.append(f"ROWCOUNT={len(sorted_rows)}")
    if RATE_CUTOFF_DATE:
        meta_parts.append(f"RATE_CUTOFF={RATE_CUTOFF_DATE.isoformat()}")
        if _RATES_FINGERPRINT:
            meta_parts.append(f"RATES_FP={_RATES_FINGERPRINT}")

    # Per Phase 01 Plan 02 D-20: mix the subcontractor rates
    # fingerprint into the hash ONLY for the four new variants
    # that actually consume the subcontractor rates CSV. This
    # forces regeneration of _AEPBillable / _ReducedSub files (and
    # their shadow-helper twins) when the CSV changes, WITHOUT
    # touching the primary / helper / vac_crew hashes (preserves
    # the ROADMAP success criterion 5 byte-identical guarantee for
    # the legacy variant set). Mirrors the conditional shape of
    # the existing `if RATE_CUTOFF_DATE: ... RATES_FP=` block
    # above so a future engineer reading the two blocks side by
    # side sees them as parallel — one keys on the retired-but-
    # retained legacy recalc gate, the other keys on the variant
    # set that consumes the new rates table.
    if variant in (
        'aep_billable',
        'reduced_sub',
        'aep_billable_helper',
        'reduced_sub_helper',
    ):
        if _SUBCONTRACTOR_RATES_FINGERPRINT:
            meta_parts.append(
                f"SUB_RATES_FP={_SUBCONTRACTOR_RATES_FINGERPRINT}"
            )

    # Update hash with metadata
    if meta_parts:
        hasher.update("\n".join(meta_parts).encode('utf-8'))

    return hasher.hexdigest()[:16]


def _compute_aggregated_content_hash(rows: list[dict]) -> str:
    """Deterministic content hash for a cross-variant row bucket.

    Used by the billing_audit integration to produce a
    ``pipeline_run.content_hash`` that covers every variant's rows
    for a given (WR, week). ``calculate_data_hash()`` assumes the
    rows it's called with all come from a single production
    ``group_source_rows`` group; that assumption breaks for an
    aggregation bucket that unions multiple groups of the same
    variant. In particular, helper groups are split per-foreman
    (group key ``{week}_{wr}_HELPER_{sanitized_foreman}``) — so
    multiple helper groups can exist for the same (WR, week) and
    calling ``calculate_data_hash(variant='helper', rows=...)``
    with all of them at once would:

      1. Read helper_foreman/dept/job from ``sorted_rows[0]`` only
         (variant-specific meta block), so identity changes on
         non-first helpers never reach the hash.
      2. Depend on row sort order for which helper's identity
         gets recorded — flips spuriously between runs.

    Primary groups are keyed on ``{week}_{wr}`` alone (one group
    per WR/week, no user suffix) and vac_crew on
    ``{week}_{wr}_VACCREW`` (one group, multi-member handled
    internally by ``calculate_data_hash``'s per-row VAC fields)
    — so only the ``helper`` variant needs sub-bucketing here.

    The combined hash is SHA-256 over the sorted
    ``variant=hash`` tokens (sub-bucketed for helper).
    """
    by_variant: dict[str, list[dict]] = {}
    for r in rows:
        v = r.get('__variant', 'primary')
        by_variant.setdefault(v, []).append(r)

    parts: list[str] = []
    for v in sorted(by_variant.keys()):
        variant_rows = by_variant[v]
        if v == 'helper':
            # Sub-bucket by helper identity to match the per-
            # foreman group structure assumed by
            # calculate_data_hash's helper branch.
            sub: dict[tuple[str, str, str], list[dict]] = {}
            for r in variant_rows:
                sk = (
                    str(r.get('__helper_foreman', '')),
                    str(r.get('__helper_dept', '')),
                    str(r.get('__helper_job', '')),
                )
                sub.setdefault(sk, []).append(r)
            sub_parts = [
                f"{sk}={calculate_data_hash(sub[sk])}"
                for sk in sorted(sub.keys())
            ]
            variant_hash = hashlib.sha256(
                "|".join(sub_parts).encode('utf-8')
            ).hexdigest()[:16]
        else:
            variant_hash = calculate_data_hash(variant_rows)
        parts.append(f"{v}={variant_hash}")

    return hashlib.sha256(
        "|".join(parts).encode('utf-8')
    ).hexdigest()[:16]


def extract_data_hash_from_filename(filename: str) -> str | None:
    """Extract data hash from filename format: WR_{wr_num}_WeekEnding_{week_end}_{data_hash}.xlsx
    
    Args:
        filename: Excel filename to parse
    
    Returns:
        str | None: 16-character hash if found, else None
    """
    try:
        name_without_ext = filename.replace('.xlsx', '')
        parts = name_without_ext.split('_')
        if len(parts) >= 4 and len(parts[-1]) == 16:
            return parts[-1]
    except Exception:
        pass
    return None

def list_generated_excel_files(folder: str) -> list[str]:
    """List Excel files beginning with WR_ in the specified folder.
    
    Args:
        folder: Directory path to scan
    
    Returns:
        list[str]: List of matching Excel filenames
    """
    try:
        return [f for f in os.listdir(folder) if f.startswith('WR_') and f.lower().endswith('.xlsx')]
    except FileNotFoundError:
        return []

def build_group_identity(filename: str) -> tuple[str, str, str, str | None] | None:
    """
    Parse filename to extract identity tuple: (wr, week_ending, variant, helper_or_user).

    Args:
        filename: Excel filename to parse

    Returns:
        tuple with format:
        - Primary: (wr, week, 'primary', None)
        - Primary+User: (wr, week, 'primary', user_identifier)
        - Helper: (wr, week, 'helper', helper_name)
        - VAC Crew (legacy, no claimer): (wr, week, 'vac_crew', '')
        - VAC Crew (named, Subproject C): (wr, week, 'vac_crew', crew_name)
        - AEP Billable: (wr, week, 'aep_billable', '')
        - Reduced Sub: (wr, week, 'reduced_sub', '')
        - AEP Billable Helper: (wr, week, 'aep_billable_helper', helper_name)
        - Reduced Sub Helper: (wr, week, 'reduced_sub_helper', helper_name)

        Legacy format without variant: (wr, week, 'primary', None)

        Returns None if filename doesn't match expected format.

    Filename formats supported:
    - WR_{wr}_WeekEnding_{week}_{hash}.xlsx (legacy primary)
    - WR_{wr}_WeekEnding_{week}_{timestamp}_{hash}.xlsx (primary)
    - WR_{wr}_WeekEnding_{week}_{timestamp}_User_{user}_{hash}.xlsx (primary+user)
    - WR_{wr}_WeekEnding_{week}_{timestamp}_Helper_{helper}_{hash}.xlsx (helper)
    - WR_{wr}_WeekEnding_{week}_{timestamp}_VacCrew_{hash}.xlsx (VAC Crew, legacy)
    - WR_{wr}_WeekEnding_{week}_{timestamp}_VacCrew_{name}_{hash}.xlsx (VAC Crew named, Subproject C)
    - WR_{wr}_WeekEnding_{week}_{timestamp}_AEPBillable_{hash}.xlsx (AEP Billable)
    - WR_{wr}_WeekEnding_{week}_{timestamp}_ReducedSub_{hash}.xlsx (Reduced Sub)
    - WR_{wr}_WeekEnding_{week}_{timestamp}_AEPBillable_Helper_{helper}_{hash}.xlsx
    - WR_{wr}_WeekEnding_{week}_{timestamp}_ReducedSub_Helper_{helper}_{hash}.xlsx
    """
    if not filename.startswith('WR_'):
        return None
    base = filename[:-5] if filename.lower().endswith('.xlsx') else filename
    parts = base.split('_')

    # Minimum: WR_<wr>_WeekEnding_<week>
    if len(parts) < 4:
        return None
    if parts[0] != 'WR':
        return None
    # Find ``WeekEnding`` by search rather than fixed position so
    # filenames whose WR token itself contains ``_`` (possible when
    # ``_RE_SANITIZE_HELPER_NAME`` rewrote a sanitization-sensitive
    # source WR#) still parse correctly. For realistic numeric WR#s
    # the marker is at position 2 exactly — this preserves the
    # legacy layout while hardening against the edge case.
    #
    # Disambiguate via the filename format itself. The STRUCTURAL
    # ``WeekEnding`` in the modern format is followed by TWO
    # consecutive 6-digit tokens:
    #   ``WeekEnding_{MMDDYY week}_{HHMMSS timestamp}_...``
    # while the legacy format (still readable off disk but no longer
    # produced) is:
    #   ``WeekEnding_{MMDDYY week}_{hash}.xlsx``
    # where the hash is hex and is essentially never exactly 6
    # digits. Helper/user identifier segments that happen to contain
    # ``WeekEnding_<6digits>`` (pathological but possible, see
    # rounds 10/11) are followed by the HASH, not a second 6-digit
    # token. So we rank candidates:
    #   * STRONG match: ``WeekEnding`` + 6-digit week + 6-digit
    #     timestamp (new format's unambiguous marker).
    #   * WEAK match: ``WeekEnding`` + 6-digit week (accepts legacy
    #     format + any other filename where a second 6-digit token
    #     isn't present).
    # Pick the RIGHTMOST strong match if any — it's always the real
    # structural delimiter because no identifier segment is ever
    # followed by two 6-digit tokens in a row. Fall back to the
    # rightmost weak match only for legacy-format filenames whose
    # hash happens not to be 6 digits (vanishingly rare collision
    # window for the weak-match path).
    _strong_candidates: list[int] = []
    _weak_candidates: list[int] = []
    for _i, _p in enumerate(parts):
        if _p != 'WeekEnding' or _i < 2 or _i + 1 >= len(parts):
            continue
        _week = parts[_i + 1]
        if not (len(_week) == 6 and _week.isdigit()):
            continue
        _weak_candidates.append(_i)
        if _i + 2 < len(parts):
            _timestamp = parts[_i + 2]
            if len(_timestamp) == 6 and _timestamp.isdigit():
                _strong_candidates.append(_i)
    # Pick the LEFTMOST strong match if any — the structural
    # delimiter always comes before any identifier-internal
    # candidate in a filename generated by ``generate_excel``
    # (variant marker + identifier are appended AFTER the
    # timestamp). Using leftmost rather than rightmost resolves
    # the final known pathology: a pathological identifier that
    # sanitizes to ``WeekEnding_<6digits>_<6digits>`` (e.g. from
    # a foreman literally named "WeekEnding 041926 123456")
    # would produce a second strong candidate inside the tail,
    # and rightmost would incorrectly pick it. Fall back to the
    # leftmost weak match only for legacy-format filenames.
    #
    # The remaining residual — a WR token that itself sanitizes
    # to ``WeekEnding_<6digits>_<6digits>`` (which would then
    # provide its own strong match at a position earlier than
    # the structural one) — is the last unreachable edge. It
    # requires a raw WR# literally equal to that pattern, which
    # is not a realistic data-entry scenario for numeric WR#s;
    # the source-side collision quarantine (pre-scan on
    # sanitized WR alone) would also flag such a pathological
    # value long before the parser is exercised.
    if _strong_candidates:
        we_idx = _strong_candidates[0]
    elif _weak_candidates:
        we_idx = _weak_candidates[0]
    else:
        return None

    # WR may span one or more parts depending on whether the
    # sanitizer introduced underscores. Rejoin them so the returned
    # WR token round-trips with the generator's output.
    wr = '_'.join(parts[1:we_idx])
    week = parts[we_idx + 1]

    # Detect variant from filename structure. Scope the marker search
    # to the tail (everything after ``WeekEnding <week>``) so any
    # accidental ``Helper`` / ``User`` / ``VacCrew`` / ``AEPBillable``
    # / ``ReducedSub`` token inside a sanitized WR does not
    # false-positive the variant detection.
    variant = 'primary'
    identifier = None
    tail = parts[we_idx + 2:]

    # Sub-project E (2026-05-25): support BOTH legacy token-bearing names
    # (``..._{HHMMSS}_<marker>_<id>_<hash>``) and the new deterministic
    # clean names (``..._<marker>_<id>``, no timestamp/hash) that
    # ``generate_excel`` produces when SUPABASE_HASH_STORE_AUTHORITATIVE is
    # on. Both shapes coexist on Smartsheet during migration, so the parser
    # must read either. The discriminator is the LEADING 6-digit ``HHMMSS``
    # timestamp at ``tail[0]``: a legacy variant name ALWAYS carries it
    # (immediately after the week) AND a trailing ``data_hash`` token; a
    # clean name has NEITHER (``tail[0]`` is always a variant marker —
    # alphabetic — or ``tail`` is empty). So when (and only when) the
    # leading timestamp is present, strip it AND the trailing hash, leaving
    # ``tail`` == ``[<marker>, <id parts...>]`` for the dispatch below.
    # Clean names skip both strips, so their last identifier segment is
    # never eaten (the bug the old unconditional ``[:-1]`` slice caused for
    # token-less names). NOTE: the oldest legacy bare-primary format
    # ``WR_{wr}_WeekEnding_{week}_{hash}.xlsx`` has a hash but no timestamp
    # and no marker — its lone hash token stays in ``tail`` but, with no
    # reserved marker, yields the correct ``('primary', None)`` regardless.
    if tail and len(tail[0]) == 6 and tail[0].isdigit():
        tail = tail[1:]          # drop the legacy HHMMSS timestamp
        if tail:
            tail = tail[:-1]     # drop the legacy trailing data_hash

    # Reserved-token precedence (ledger [2026-05-21 13:20], generalized to
    # the bare ``_User_`` shape by Subproject D 2026-05-25): the variant is
    # determined by the EARLIEST reserved marker token in ``tail``, NOT by a
    # fixed check order. A claimer / helper / vac-crew name can itself contain
    # a reserved word (e.g. a foreman literally named "Pat Helper" →
    # ``_User_Pat_Helper_<hash>``); a fixed AEPBillable→ReducedSub→VacCrew→
    # Helper→User order would misclassify it (here, as ``helper`` with the
    # ``User`` token lost). Dispatching on the earliest-position marker fixes
    # this for ALL bare shapes while preserving the two-level
    # ``_AEPBillable_User_`` / ``_ReducedSub_Helper_`` handling (AEPBillable /
    # ReducedSub are always the earliest token in those filenames). The scan
    # operates on the post-``WeekEnding`` ``tail`` slice only, so a sanitized
    # WR# containing a reserved word in its body cannot false-positive the
    # variant — covered by the negative tests in
    # TestBuildGroupIdentityWithUnderscoresInWr.
    _reserved_positions = {
        _tok: tail.index(_tok)
        for _tok in ('AEPBillable', 'ReducedSub', 'VacCrew', 'Helper', 'User')
        if _tok in tail
    }
    _first_marker = (
        min(_reserved_positions, key=lambda _t: _reserved_positions[_t])
        if _reserved_positions else None
    )
    if _first_marker == 'AEPBillable':
        aep_idx_rel = tail.index('AEPBillable')
        post_aep = tail[aep_idx_rel + 1:]
        if post_aep and post_aep[0] == 'User':
            # Subproject B: _AEPBillable_User_<claimer>[_<hash>]. Reserved
            # 'User' token marks a primary-claimer identifier. Span-join so an
            # underscored claimer name survives; dangling 'User' -> '' (legacy).
            # ``tail`` already had any legacy timestamp/hash stripped (E), so
            # the identifier is everything after the marker.
            variant = 'aep_billable'
            identifier = '_'.join(post_aep[1:])
        elif 'Helper' in post_aep:
            variant = 'aep_billable_helper'
            helper_idx_rel = post_aep.index('Helper')
            if helper_idx_rel + 1 < len(post_aep):
                identifier = '_'.join(post_aep[helper_idx_rel + 1:])
        else:
            # Legacy unpartitioned _AEPBillable_<hash> (no User/Helper token).
            variant = 'aep_billable'
            identifier = ''
    elif _first_marker == 'ReducedSub':
        rs_idx_rel = tail.index('ReducedSub')
        post_rs = tail[rs_idx_rel + 1:]
        if post_rs and post_rs[0] == 'User':
            variant = 'reduced_sub'
            identifier = '_'.join(post_rs[1:])
        elif 'Helper' in post_rs:
            variant = 'reduced_sub_helper'
            helper_idx_rel = post_rs.index('Helper')
            if helper_idx_rel + 1 < len(post_rs):
                identifier = '_'.join(post_rs[helper_idx_rel + 1:])
        else:
            # Legacy unpartitioned _ReducedSub_<hash> (no User/Helper token).
            variant = 'reduced_sub'
            identifier = ''
    elif _first_marker == 'VacCrew':
        # Subproject C: _VacCrew_<name>_<hash>. Span-join so an underscored
        # name survives. Legacy _VacCrew (no name) -> ''.
        variant = 'vac_crew'
        vac_idx_rel = tail.index('VacCrew')
        identifier = ''  # legacy _VacCrew (no name) -> '' per identity contract
        if vac_idx_rel + 1 < len(tail):
            identifier = '_'.join(tail[vac_idx_rel + 1:])
    elif _first_marker == 'Helper':
        variant = 'helper'
        helper_idx_rel = tail.index('Helper')
        if helper_idx_rel + 1 < len(tail):
            identifier = '_'.join(tail[helper_idx_rel + 1:])
    elif _first_marker == 'User':
        variant = 'primary'
        user_idx_rel = tail.index('User')
        if user_idx_rel + 1 < len(tail):
            identifier = '_'.join(tail[user_idx_rel + 1:])

    return (wr, week, variant, identifier)

def _resolve_unchanged_for_skip(history_key, data_hash, hash_history,
                                wr_num, week_iso, variant, identifier):
    """Decide whether a group's content hash is UNCHANGED vs the durable
    store, for the change-detection skip gate (Sub-project E, 2026-05-25).

    Decision model:
    - ``SUPABASE_HASH_STORE_AUTHORITATIVE`` ON (and billing_audit
      available, not TEST_MODE, and ``week_iso`` present): Supabase
      (``billing_audit.group_content_hash``) is authoritative.
        * ``success``  -> compare the stored hash to ``data_hash``.
        * ``no_row``   -> the group was never durably stored, so it is
          treated as CHANGED (return False -> regenerate). This is the
          safe default that makes the very first authoritative run
          regenerate everything once, populating the store.
        * ``fetch_failure`` / ``unavailable`` -> a Supabase outage (or
          the table/schema not yet exposed); fall through to the local
          ``hash_history`` json cache so a transient outage degrades to
          "use the cache / regenerate", never a silent wrong-skip.
          (``lookup_group_hash`` returns only these four statuses.)
    - A missing/empty ``week_iso`` (no ``__week_ending_date`` on the
      group) skips the Supabase read entirely and uses the json cache —
      ``week_ending`` is a DATE column, so passing ``''`` would be a
      PostgREST type error that could needlessly trip the per-op
      circuit breaker.
    - Authoritative OFF (default): the ``hash_history`` json cache alone
      decides — byte-identical to the pre-E behavior.

    The caller must already have confirmed ``_history_eligible_for_skip``
    (FORCE_GENERATION / REGEN_WEEKS / RESET_* gating) and still applies
    the ``ATTACHMENT_REQUIRED_FOR_SKIP`` guard downstream — a matching
    hash with a missing attachment must still regenerate.
    """
    if (
        SUPABASE_HASH_STORE_AUTHORITATIVE
        and BILLING_AUDIT_AVAILABLE
        and not TEST_MODE
        and week_iso
    ):
        _h, _status = _billing_audit_writer.lookup_group_hash(
            wr_num, week_iso, variant, identifier or '')
        if _status == 'success':
            return _h == data_hash
        if _status == 'no_row':
            return False  # never durably stored -> regenerate (safe)
        # fetch_failure / unavailable -> fall back to json cache.
    _prev = hash_history.get(history_key)
    return bool(_prev and _prev.get('hash') == data_hash)

def cleanup_stale_excels(output_folder: str, kept_filenames: set):
    """Remove Excel files not generated in current run (VARIANT-AWARE).

    Strategy:
      1. Keep all names in kept_filenames.
      2. For identities (wr, week, variant, identifier) present in kept_filenames, 
         remove any other files with same identity (older timestamp/hash).
      3. Remove any other WR_*.xlsx whose identity is not in current run 
         (per user requirement to only keep new system outputs).
      4. CRITICAL: Never cross-delete between variants (primary vs helper).
    
    Identity includes variant dimension to prevent primary/helper cross-deletion.
    Returns list of removed filenames.
    """
    removed = []
    existing = list_generated_excel_files(output_folder)
    identities_to_keep = set()
    for fname in kept_filenames:
        ident = build_group_identity(fname)
        if ident:
            identities_to_keep.add(ident)
    for fname in existing:
        if fname in kept_filenames:
            continue
        ident = build_group_identity(fname)
        if ident and ident in identities_to_keep:
            # Variant of identity we already produced this run
            try:
                os.remove(os.path.join(output_folder, fname))
                removed.append(fname)
            except Exception as e:
                logging.warning(f"⚠️ Failed to remove stale variant {fname}: {e}")
        elif ident:
            # Different identity (older run) – remove per requirement
            try:
                os.remove(os.path.join(output_folder, fname))
                removed.append(fname)
            except Exception as e:
                logging.warning(f"⚠️ Failed to remove legacy file {fname}: {e}")
        # Non-conforming files left untouched
    return removed

def cleanup_untracked_sheet_attachments(
    client,
    target_sheet_id: int,
    valid_wr_weeks: set,
    test_mode: bool,
    attachment_cache: dict | None = None,
    target_sheet=None,
    variant_whitelist: set[str] | None = None,
    sub_wr_scope: set[str] | None = None,
    sub_offcontract_variants: set[str] | None = None,
    sub_legacy_primary_variants: set[str] | None = None,
    vac_legacy_wr_scope: set[str] | None = None,
    primary_wr_scope: set[str] | None = None,
):
    """Prune only older variants for identities processed this run (VARIANT-AWARE).

    If KEEP_HISTORICAL_WEEKS=1 (default false here), weeks not in this run are preserved.
    valid_wr_weeks: set of 4-tuples (wr, week_mmddyy, variant, identifier) that were
                    generated or validated this session.
    attachment_cache: Pre-fetched dict of row_id -> attachment list (avoids per-row API calls).
    target_sheet: Pre-loaded target sheet object (avoids redundant API call).

    variant_whitelist: Per-sheet variant gate (Phase 1.1 Bug B2 /
        D-07 / SUB-10). When provided, any attachment whose
        ``build_group_identity``-parsed variant is NOT in the
        whitelist is treated as off-contract for THIS sheet and
        unconditionally deleted, regardless of ``valid_wr_weeks``
        membership and regardless of ``KEEP_HISTORICAL_WEEKS``.
        When None (default), preserves byte-identical legacy
        behaviour — every variant is accepted and the cleanup
        decision rests on identity grouping + valid_wr_weeks.
        PPP cleanup passes ``{'reduced_sub', 'reduced_sub_helper'}``;
        TARGET cleanup passes None.

    sub_wr_scope: Phase 1.1 UAT gap closure (SUB-09 helper dimension).
        When provided, any attachment whose parsed ``wr`` is in this set
        AND whose parsed ``variant`` is in ``sub_offcontract_variants``
        is treated as off-contract for THIS sheet and unconditionally
        deleted. Used to remove pre-existing legacy ``_Helper_<name>``
        and bare-primary attachments for subcontractor WRs from
        TARGET_SHEET_ID (Task 1 stops NEW ones; this removes OLD ones).
        When None (default), this gate is skipped entirely —
        byte-identical legacy TARGET behaviour for all callers that
        do not pass the parameter. WR-01 guard: an attachment whose
        identity is in ``valid_wr_weeks`` is exempt from this gate so a
        legitimate live non-subcontractor ``_Helper_`` file for a WR#
        that ALSO has subcontractor rows is never deleted (cross-sheet
        WR overlap — see the inline comment at the gate).

    sub_offcontract_variants: Set of variant strings that are off-contract
        for WRs in ``sub_wr_scope`` on THIS sheet. For TARGET cleanup,
        pass ``{'helper', 'primary'}`` (subcontractor non-helper rows
        now emit only variant keys per Bug B1; subcontractor helper rows
        now emit only shadow variants per Task 1 — so any bare 'primary'
        or legacy 'helper' attachment for a sub WR is a pre-fix orphan).
        When None and ``sub_wr_scope`` is provided, this gate is a no-op.
        Ignored when ``sub_wr_scope`` is None.

    sub_legacy_primary_variants: Subproject B (2026-05-20) one-time
        migration. When provided, any attachment whose parsed ``wr`` is
        in ``sub_wr_scope``, whose parsed ``variant`` is in this set, and
        whose parsed ``identifier`` is empty (legacy unpartitioned
        ``_ReducedSub`` / ``_AEPBillable``) is unconditionally deleted —
        UNLESS its identity is in ``valid_wr_weeks`` (live-identity
        exemption). New per-claimer files (non-empty identifier) are
        never matched. When None (default), this gate is skipped.
        Gated at the call sites by SUBCONTRACTOR_LEGACY_PRIMARY_CLEANUP_ENABLED.

    vac_legacy_wr_scope: Subproject C (2026-05-21) Task 6 one-time
        migration. When provided, any attachment whose parsed ``wr`` is
        in this set, whose parsed ``variant`` is ``'vac_crew'``, and
        whose parsed ``identifier`` is empty (legacy unpartitioned bare
        ``_VacCrew``) is unconditionally deleted — UNLESS its identity is
        in ``valid_wr_weeks`` (live-identity exemption). Per-claimer files
        (non-empty identifier like ``_VacCrew_John``) are never matched.
        When None (default), this gate is skipped — byte-identical legacy
        behaviour for callers that do not pass the parameter.
        Gated at the TARGET call site by VAC_CREW_LEGACY_CLEANUP_ENABLED.
        vac_crew files route to TARGET_SHEET_ID only (never PPP); the
        PPP call site must NOT receive this parameter.

    primary_wr_scope: Subproject D (2026-05-25) one-time migration. When
        provided, any attachment whose parsed ``wr`` is in this set, whose
        parsed ``variant`` is ``'primary'``, and whose parsed ``identifier``
        is empty (legacy unpartitioned bare ``primary``) is unconditionally
        deleted — UNLESS its identity is in ``valid_wr_weeks`` (live-identity
        exemption). Per-claimer files (non-empty identifier like
        ``_User_Alice``) are never matched. When None (default), this gate
        is skipped — byte-identical legacy behaviour for callers that do
        not pass the parameter. Gated at the TARGET call site by
        PRIMARY_CLAIM_ATTRIBUTION_ENABLED and
        LEGACY_PRIMARY_PARTITION_CLEANUP_ENABLED. Non-subcontractor primary
        files route to TARGET_SHEET_ID only — the PPP call site must NOT
        receive this parameter.

    CRITICAL: Identity includes variant dimension to prevent primary/helper cross-deletion.
              Each (wr, week, variant, identifier) is treated as independent.
    """
    if test_mode:
        logging.info("🧪 Test mode – skipping sheet attachment pruning")
        return
    try:
        sheet = target_sheet if target_sheet is not None else client.Sheets.get_sheet(target_sheet_id)
    except Exception as e:
        logging.warning(f"⚠️ Could not load target sheet for attachment cleanup: {e}")
        return
    removed_variants = 0
    removed_off_contract = 0  # Phase 1.1 Bug B2 (D-07 / SUB-10): off-contract counter
    for row in sheet.rows:
        try:
            # Use pre-fetched cache if available; otherwise fall back to per-row API call
            if attachment_cache is not None and row.id in attachment_cache:
                attachments = attachment_cache[row.id]
            else:
                attachments = client.Attachments.list_row_attachments(target_sheet_id, row.id).data
        except Exception:
            continue
        identity_groups = collections.defaultdict(list)
        off_contract_attachments = []  # Phase 1.1 Bug B2 / D-07: per-sheet whitelist
        for att in attachments:
            name = getattr(att,'name','') or ''
            if name.startswith('WR_') and name.endswith('.xlsx'):
                ident = build_group_identity(name)
                if ident:
                    wr, week, variant, _identifier = ident
                    # Phase 1.1 Bug B2 (D-07 / SUB-10): per-sheet
                    # variant whitelist gate. variant_whitelist=None
                    # (TARGET cleanup) preserves legacy "accept
                    # every variant" behaviour. When the caller
                    # supplies a whitelist (PPP cleanup passes
                    # {'reduced_sub','reduced_sub_helper'}), any
                    # other variant parsed from a filename on THIS
                    # sheet is off-contract and gets unconditionally
                    # pruned BEFORE the identity_groups +
                    # KEEP_HISTORICAL_WEEKS logic — variant-set
                    # membership is the authoritative gate for this
                    # sheet.
                    if (
                        variant_whitelist is not None
                        and variant not in variant_whitelist
                    ):
                        off_contract_attachments.append(att)
                        continue
                    # Phase 1.1 UAT gap closure (SUB-09 helper dimension):
                    # remove pre-existing legacy `_Helper_<name>` / bare-primary
                    # attachments for subcontractor WRs on TARGET. Task 1 stops
                    # NEW ones at the producer; this removes leftovers already
                    # uploaded by pre-fix merged runs.
                    #
                    # WR-01 cross-sheet-overlap guard (review follow-up): the
                    # ``sub_wr_scope`` set keys on WR# alone, but
                    # ``is_subcontractor_row`` is decided PER-ROW by source
                    # sheet. A single WR# can legitimately have helper rows on
                    # a subcontractor sheet (→ in scope) AND on a NON-sub sheet
                    # (→ a live ``_Helper_<name>.xlsx`` whose identity IS in
                    # ``valid_wr_weeks`` this run). Exempting identities present
                    # in ``valid_wr_weeks`` preserves the cleanup intent — a
                    # genuinely orphaned legacy sub-helper file is NEVER in
                    # ``valid_wr_weeks`` because Task 1 stopped emitting it — while
                    # protecting a legitimate live non-sub artifact from an
                    # every-run delete/regenerate/re-upload churn loop.
                    if (
                        sub_wr_scope is not None
                        and wr in sub_wr_scope
                        and sub_offcontract_variants is not None
                        and variant in sub_offcontract_variants
                        and ident not in valid_wr_weeks
                    ):
                        off_contract_attachments.append(att)
                        continue
                    # Subproject B (2026-05-20): one-time migration —
                    # delete LEGACY UNPARTITIONED `_ReducedSub` /
                    # `_AEPBillable` attachments (parsed identifier == '')
                    # for in-scope subcontractor WRs. B re-partitions
                    # these by frozen claimer, so the bare one-file-per-WR
                    # attachment is an obsolete duplicate. The
                    # ``not _identifier`` check is the precise legacy
                    # selector: new per-claimer files carry a non-empty
                    # identifier and are NOT deleted here. The
                    # ``ident not in valid_wr_weeks`` guard is
                    # belt-and-suspenders (B never emits an empty
                    # identifier, so a live file is never empty-id) per the
                    # [2026-05-19 23:45] WR-01 live-identity rule.
                    if (
                        sub_wr_scope is not None
                        and wr in sub_wr_scope
                        and sub_legacy_primary_variants is not None
                        and variant in sub_legacy_primary_variants
                        and not _identifier
                        and ident not in valid_wr_weeks
                    ):
                        off_contract_attachments.append(att)
                        continue
                    # Subproject C Task 6 (2026-05-21): one-time migration —
                    # delete LEGACY UNPARTITIONED bare ``_VacCrew`` attachments
                    # (parsed identifier == '') for in-scope vac_crew WRs.
                    # Subproject C re-partitions vac_crew files by frozen
                    # claimer (``_VacCrew_<name>``), so the old bare
                    # one-file-per-WR attachment is an obsolete duplicate.
                    # The ``not _identifier`` check is the precise legacy
                    # selector: new per-claimer files carry a non-empty
                    # identifier and are NOT deleted here.
                    # WR-01 live-identity exemption: an attachment whose
                    # identity IS in ``valid_wr_weeks`` is kept — this
                    # protects a live per-claimer file from being deleted
                    # if its identifier happened to be empty (belt-and-
                    # suspenders; per-claimer files always have non-empty
                    # identifiers so this branch is effectively unreachable
                    # for them, but the guard keeps the contract symmetric
                    # with the B sub_legacy_primary gate).
                    if (
                        vac_legacy_wr_scope is not None
                        and wr in vac_legacy_wr_scope
                        and variant == 'vac_crew'
                        and not _identifier
                        and ident not in valid_wr_weeks
                    ):
                        off_contract_attachments.append(att)
                        continue
                    # Subproject D (2026-05-25): one-time migration —
                    # delete LEGACY UNPARTITIONED bare ``primary``
                    # attachments (``build_group_identity`` parses a bare
                    # primary to ``identifier=None``; the ``not _identifier``
                    # gate below matches None and '') for in-scope
                    # NON-subcontractor WRs. D re-partitions production
                    # primary files by frozen claimer (``_User_<name>``),
                    # so the old bare one-file-per-WR attachment is an
                    # obsolete duplicate. The ``not _identifier`` check is
                    # the precise legacy selector: new per-claimer files
                    # carry a non-empty identifier and are NOT deleted here.
                    # WR-01 live-identity exemption: an attachment whose
                    # identity IS in ``valid_wr_weeks`` is kept — this
                    # protects a legitimate bare-primary file the current
                    # run produced (e.g. an overlapping WR still emitting
                    # bare primary because attribution was disabled for
                    # those rows) from an every-run delete/regenerate churn.
                    if (
                        primary_wr_scope is not None
                        and wr in primary_wr_scope
                        and variant == 'primary'
                        and not _identifier
                        and ident not in valid_wr_weeks
                    ):
                        off_contract_attachments.append(att)
                        continue
                    # Variant-migration orphan gate (2026-06-02):
                    # A dual-checkbox helper row that had blank helper_dept
                    # on Run 1 fell back to a primary group and produced a
                    # primary Excel attachment on Smartsheet. On Run 2, once
                    # helper_dept is corrected, the row migrates to the helper
                    # variant — the primary group disappears from ``groups``
                    # and its identity is NEVER added to ``valid_wr_weeks``.
                    # The identity_groups loop below keeps the single remaining
                    # attachment (no duplicate to prune), so it silently
                    # survives every subsequent run.
                    #
                    # Detection rule: this attachment is a 'primary' variant
                    # whose identity is NOT in ``valid_wr_weeks``, AND at
                    # least one helper-family variant for the SAME (wr, week)
                    # IS live in ``valid_wr_weeks`` this run. That combination
                    # is the unambiguous signal that the primary credit was
                    # superseded by a helper.
                    #
                    # Safety: the ``ident not in valid_wr_weeks`` guard ensures
                    # a legitimately live primary (one that IS still produced
                    # this run) is never touched. The helper-family presence
                    # check (any helper/aep_billable_helper/reduced_sub_helper
                    # for same wr+week) is the confirming signal that prevents
                    # over-eager deletion when a primary is simply not in scope
                    # today for other reasons (WR_FILTER, time-budget cutoff,
                    # KEEP_HISTORICAL_WEEKS). Without this confirming signal we
                    # would risk deleting a primary that is still valid but just
                    # not regenerated in this run.
                    _HELPER_VARIANTS_FOR_ORPHAN_GATE = frozenset({
                        'helper', 'aep_billable_helper', 'reduced_sub_helper'
                    })
                    if (
                        variant == 'primary'
                        and ident not in valid_wr_weeks
                        and any(
                            _vw[0] == wr
                            and _vw[1] == week
                            and _vw[2] in _HELPER_VARIANTS_FOR_ORPHAN_GATE
                            for _vw in valid_wr_weeks
                        )
                    ):
                        try:
                            import sentry_sdk as _sentry_sdk
                            with _sentry_sdk.new_scope() as _scope:
                                _scope.set_tag(
                                    'cleanup.reason',
                                    'variant_migration_orphan',
                                )
                                _scope.set_tag('wr', wr)
                                _scope.set_tag('week', week)
                        except Exception:
                            pass
                        off_contract_attachments.append(att)
                        logging.info(
                              f"🔄 Variant-migration orphan detected: "
                            f"primary attachment {att.name!r} superseded "
                            f"by live helper for WR {wr} week {week}. "
                            f"Queued for deletion."
                        )
                        continue
                    identity_groups[ident].append(att)
        # Phase 1.1 Bug B2 (D-07 / SUB-10): unconditionally delete
        # off-contract attachments. These are NEVER subject to
        # KEEP_HISTORICAL_WEEKS — variant-set-membership is the
        # authoritative gate for this sheet. PII marker
        # "Removed off-contract variant on sheet" registered in
        # _PII_LOG_MARKERS for the new log body below (the
        # attachment name embeds WR + week which the sanitizer
        # must catch under SENTRY_ENABLE_LOGS).
        for att in off_contract_attachments:
            try:
                client.Attachments.delete_attachment(target_sheet_id, att.id)
                removed_off_contract += 1
                logging.info(
                    f"🗑️ Removed off-contract variant on sheet "
                    f"{target_sheet_id}: {att.name}"
                )
            except Exception as e:
                logging.warning(
                    f"⚠️ Could not delete off-contract variant "
                    f"{att.name}: {e}"
                )
        for ident, atts in identity_groups.items():
            # Skip identities not processed if preserving historical weeks
            if ident not in valid_wr_weeks and KEEP_HISTORICAL_WEEKS:
                continue
            # Sort attachments newest first based on timestamp token
            def _ts(a):
                parts = a.name.split('_')
                # Find timestamp (should be after WeekEnding_{week})
                for i, p in enumerate(parts):
                    if p == 'WeekEnding' and i + 2 < len(parts):
                        ts_candidate = parts[i + 2]
                        if ts_candidate.isdigit() and len(ts_candidate) == 6:
                            return ts_candidate
                return '000000'
            atts_sorted = sorted(atts, key=_ts, reverse=True)
            # Keep newest; remove others
            for old in atts_sorted[1:]:
                try:
                    client.Attachments.delete_attachment(target_sheet_id, old.id)
                    removed_variants += 1
                    logging.info(f"🗑️ Removed older variant: {old.name}")
                except Exception as e:
                    logging.warning(f"⚠️ Could not delete variant {old.name}: {e}")
    logging.info(
        f"🧹 Variant pruning done: removed_variants={removed_variants}, "
        f"removed_off_contract={removed_off_contract}"
    )

def delete_old_excel_attachments(client, target_sheet_id, target_row, wr_num, week_raw, current_data_hash, variant='primary', identifier=None, force_generation=False, cached_attachments: list | None = None):
    """Delete prior Excel attachment(s) ONLY for the specific (WR, week, variant, identifier) identity.

    VARIANT-AWARE BEHAVIOR:
    • Looks for attachments matching (wr, week, variant, identifier) exactly
    • CRITICAL: Never deletes across variants (primary vs helper)
    • If an attachment for that identity already has the identical data hash
      (and not forcing) we skip regeneration & upload
    • Leaves attachments for other weeks and other variants untouched

    Args:
        variant: 'primary' or 'helper'
        identifier: For helper variant, the helper name; for primary+user, the user identifier
        cached_attachments: Pre-fetched attachment list (avoids redundant API call)

    Returns (deleted_count, skipped_due_to_same_data)
    """
    deleted_count = 0
    try:
        attachments = cached_attachments if cached_attachments is not None else client.Attachments.list_row_attachments(target_sheet_id, target_row.id).data
    except Exception as e:
        logging.warning(f"Could not list attachments for row {target_row.id}: {e}")
        return 0, False

    # Build variant-specific prefix patterns
    # Format: WR_{wr}_WeekEnding_{week}_<variant_marker>
    base_prefix = f"WR_{wr_num}_WeekEnding_{week_raw}"
    
    candidates = []
    for a in attachments:
        name = getattr(a, 'name', '') or ''
        if not name.endswith('.xlsx'):
            continue
        
        # Parse identity from filename
        ident = build_group_identity(name)
        if not ident:
            continue
        
        ident_wr, ident_week, ident_variant, ident_identifier = ident
        
        # Match only if all identity components match
        # Normalize None/'' to avoid mismatch between build_group_identity (returns None) and main loop (uses '')
        if (ident_wr == wr_num and 
            ident_week == week_raw and 
            ident_variant == variant and
            (ident_identifier or '') == (identifier or '')):
            candidates.append(a)

    if not candidates:
        return 0, False

    # Skip if any existing candidate already carries the same hash (unless
    # forced). Sub-project E (2026-05-25): this filename-embedded-hash
    # short-circuit is the LEGACY durable backstop. When
    # SUPABASE_HASH_STORE_AUTHORITATIVE is on, the durable skip decision is
    # made upstream by the Supabase-backed skip gate
    # (_resolve_unchanged_for_skip in main), AND clean filenames carry no
    # hash token (extract_data_hash_from_filename returns None for them), so
    # this short-circuit MUST NOT fire — the identity-based replacement loop
    # below still runs so a fresh clean file supersedes any prior (token-
    # named or clean) attachment for the same identity. Forcing always wins.
    if force_generation:
        logging.info(f"⚐ FORCE GENERATION for {variant} WR {wr_num} Week {week_raw}; ignoring existing hash match")
    elif not SUPABASE_HASH_STORE_AUTHORITATIVE:
        for att in candidates:
            existing_hash = extract_data_hash_from_filename(att.name)
            if existing_hash == current_data_hash:
                logging.info(f"⏩ Unchanged ({variant} WR {wr_num} Week {week_raw}) hash {current_data_hash}; skipping regeneration & upload")
                return 0, True

    logging.info(f"🗑️ Removing {len(candidates)} prior {variant} attachment(s) for WR {wr_num} Week {week_raw}")
    for att in candidates:
        try:
            client.Attachments.delete_attachment(target_sheet_id, att.id)
            deleted_count += 1
            logging.info(f"   ✅ Deleted: {att.name}")
        except Exception as e:
            msg = str(e).lower()
            if '404' in msg or 'not found' in msg:
                logging.info(f"   ℹ️ Already gone: {att.name}")
            else:
                logging.warning(f"   ⚠️ Delete failed {att.name}: {e}")
    return deleted_count, False

def _has_existing_week_attachment(client, target_sheet_id, target_row, wr_num: str, week_raw: str, variant: str = 'primary', identifier: str | None = None, cached_attachments: list | None = None) -> bool:
    """Return True if at least one attachment exists for this (WR, week, variant, identifier) identity."""
    try:
        attachments = cached_attachments if cached_attachments is not None else client.Attachments.list_row_attachments(target_sheet_id, target_row.id).data
    except Exception:
        return False
    
    # Check for attachments matching this exact identity
    for a in attachments:
        name = getattr(a, 'name', '') or ''
        if not name.endswith('.xlsx'):
            continue
        
        # Parse identity from filename
        ident = build_group_identity(name)
        if not ident:
            continue
        
        ident_wr, ident_week, ident_variant, ident_identifier = ident
        
        # Match only if all identity components match
        # Normalize None/'' to avoid mismatch between build_group_identity (returns None) and main loop (uses '')
        if (ident_wr == wr_num and 
            ident_week == week_raw and 
            ident_variant == variant and
            (ident_identifier or '') == (identifier or '')):
            return True
    
    return False

def purge_existing_hashed_outputs(client, target_sheet_id: int, wr_subset: set | None, test_mode: bool):
    """Delete existing hashed Excel attachments and local files so hashes recompute fresh.

    wr_subset: if provided, only purge attachments for these WR numbers; otherwise purge all WR_*.xlsx attachments.
    """
    # Local file purge
    try:
        local_files = list_generated_excel_files(OUTPUT_FOLDER)
        removed_local = 0
        for f in local_files:
            wr_ident = build_group_identity(f)
            if wr_subset and wr_ident and wr_ident[0] not in wr_subset:
                continue
            try:
                os.remove(os.path.join(OUTPUT_FOLDER, f))
                removed_local += 1
            except Exception:
                pass
        logging.info(f"🧨 Local hash reset: removed {removed_local} existing Excel file(s)")
    except Exception as e:
        logging.warning(f"⚠️ Local hash reset failed: {e}")

    if test_mode:
        logging.info("🧪 Test mode active – skipping remote attachment purge")
        return
    try:
        sheet = client.Sheets.get_sheet(target_sheet_id)
    except Exception as e:
        logging.warning(f"⚠️ Could not load target sheet for purge: {e}")
        return
    purged = 0
    scanned = 0
    for row in sheet.rows:
        try:
            attachments = client.Attachments.list_row_attachments(target_sheet_id, row.id).data
        except Exception:
            continue
        for att in attachments:
            name = getattr(att,'name','') or ''
            if not name.startswith('WR_') or not name.endswith('.xlsx'):
                continue
            ident = build_group_identity(name)
            if wr_subset and ident and ident[0] not in wr_subset:
                continue
            scanned += 1
            try:
                client.Attachments.delete_attachment(target_sheet_id, att.id)
                purged += 1
                logging.info(f"🗑️ Purged attachment: {name}")
            except Exception as e:
                logging.warning(f"⚠️ Failed to purge attachment {name}: {e}")
    logging.info(f"🔥 Remote hash reset complete: purged {purged} / scanned {scanned} matching attachment(s)")

# --- HASH HISTORY PERSISTENCE ---

def load_hash_history(path: str):
    if RESET_HASH_HISTORY:
        logging.info("♻️ Hash history reset requested; ignoring existing history file")
        return {}
    try:
        with open(path,'r') as f:
            data = json.load(f)
        if not isinstance(data, dict):
            logging.warning("⚠️ Hash history is not a dict; resetting")
            return {}
        # Validate entries: keep only those with a 'hash' key.
        # Phase 1.1 Pitfall 4: also preserve ``_``-prefixed sentinel
        # keys (e.g. ``_phase_prune_version``) so they survive the
        # load → save → load round-trip and the prune pass at session
        # startup stays idempotent. Without this, the int-valued
        # sentinel would be dropped at load time and the prune would
        # fire on every run (silent non-idempotent trap).
        valid = {
            k: v for k, v in data.items()
            if isinstance(k, str) and (
                k.startswith('_')
                or (isinstance(v, dict) and 'hash' in v)
            )
        }
        dropped = len(data) - len(valid)
        if dropped:
            logging.warning(f"⚠️ Dropped {dropped} malformed hash history entries")
        return valid
    except FileNotFoundError:
        return {}
    except Exception as e:
        logging.warning(f"⚠️ Failed to load hash history: {e}")
        return {}

HASH_HISTORY_MAX_ENTRIES = 1000
BILLING_AUDIT_ROW_CACHE_PATH = os.path.join(
    OUTPUT_FOLDER, "billing_audit_frozen_rows.json"
)
BILLING_AUDIT_ROW_CACHE_MAX_ENTRIES = 200000

def save_hash_history(path: str, history: dict):
    try:
        # Retention: keep only the most recent entries by timestamp.
        # Phase 1.1 Pitfall 4: sentinel keys (``_phase_prune_version``,
        # any future ``_``-prefixed key) are int-valued — calling
        # ``history[k].get('timestamp', '')`` on an int raises
        # AttributeError and the whole save aborts. Filter sentinels
        # OUT of the sort candidates, then re-add them unconditionally
        # so they survive the save. Sentinels are NOT subject to the
        # entry cap because there is exactly one per migration version.
        if len(history) > HASH_HISTORY_MAX_ENTRIES:
            _sentinel_keys = {
                k: v for k, v in history.items()
                if isinstance(k, str) and k.startswith('_')
            }
            _real_entries = {
                k: v for k, v in history.items()
                if not (isinstance(k, str) and k.startswith('_'))
            }
            sorted_keys = sorted(
                _real_entries.keys(),
                key=lambda k: _real_entries[k].get('timestamp', ''),
                reverse=True
            )
            _kept = {
                k: _real_entries[k]
                for k in sorted_keys[:HASH_HISTORY_MAX_ENTRIES]
            }
            _kept.update(_sentinel_keys)
            history = _kept
            logging.info(
                f"🧹 Pruned hash history to {HASH_HISTORY_MAX_ENTRIES} "
                f"entries (+ {len(_sentinel_keys)} sentinel key(s) preserved)"
            )
        tmp_path = path + '.tmp'
        with open(tmp_path,'w') as f:
            json.dump(history, f, indent=2, default=str)
        os.replace(tmp_path, path)
        logging.info(f"📝 Hash history saved ({len(history)} entries)")
    except Exception as e:
        logging.warning(f"⚠️ Failed to save hash history: {e}")


_SUBCONTRACTOR_SCOPE_VARIANTS = frozenset({
    'reduced_sub', 'aep_billable',
    'reduced_sub_helper', 'aep_billable_helper',
})


def _build_subcontractor_wr_scope(groups: dict) -> set[str]:
    """Return the set of sanitized WR tokens active as subcontractor in this run.

    Gates on each group's authoritative ``__variant`` field (set at the
    ``group_source_rows`` emission site) — NOT a ``'_REDUCEDSUB'`` key
    substring scan. The variant gate is what distinguishes a subcontractor
    group, so a primary claimer, helper, or vac-crew name (or a pathological
    WR token) that itself contains an all-caps reserved word like
    ``REDUCEDSUB`` / ``AEPBILLABLE`` cannot false-positive into this
    destructive cleanup scope. This mirrors the ``_build_primary_wr_scope``
    consistency fix (Codex PR #223 P1) — variant detection MUST use the
    ``__variant`` field, never a key substring. The prior substring scan also
    silently missed ``_AEPBILLABLE``-only keys (the ``'_REDUCEDSUB'``
    substring is absent there); it happened to produce the correct WR set
    only because every sub WR always emits a ``reduced_sub`` group, so the
    variant gate is both more robust and strictly more complete.

    Shared by ``_run_phase_1_1_hash_prune`` (hash-prune scope) and the
    TARGET ``cleanup_untracked_sheet_attachments`` call site (SUB-09
    helper-dimension cleanup scope). A single implementation prevents the
    scope-build drift that the [2026-05-15 12:00] three-site invariant
    warns against.
    """
    _scope: set[str] = set()
    for _g_rows in groups.values():
        if not _g_rows:
            continue
        if _g_rows[0].get('__variant') in _SUBCONTRACTOR_SCOPE_VARIANTS:
            _g_wr_raw = _g_rows[0].get('Work Request #', '')
            _g_wr = str(_g_wr_raw).split('.')[0]
            _g_wr = _RE_SANITIZE_HELPER_NAME.sub('_', _g_wr)[:50]
            if _g_wr:
                _scope.add(_g_wr)
    return _scope


def _build_vac_crew_wr_scope(groups: dict) -> set[str]:
    """Return the set of sanitized WR tokens active as vac_crew in this run.

    Gates on each group's authoritative ``__variant`` field (``== 'vac_crew'``,
    set at the ``group_source_rows`` emission site) — NOT a ``'_VACCREW'`` key
    substring scan. The variant gate prevents a non-vac group whose
    claimer/helper name is the all-caps reserved token ``VACCREW`` (key
    ``..._HELPER_VACCREW``) from false-positiving into this destructive
    cleanup scope. Mirrors the ``_build_primary_wr_scope`` /
    ``_build_subcontractor_wr_scope`` consistency fix (Codex PR #223 P1):
    variant detection MUST use the ``__variant`` field, never a key substring.

    Shared by the TARGET ``cleanup_untracked_sheet_attachments`` call site
    (Task 6 vac-crew legacy cleanup scope).  A single implementation prevents
    scope-build drift — mirrors ``_build_subcontractor_wr_scope``.
    """
    _scope: set[str] = set()
    for _g_rows in groups.values():
        if not _g_rows:
            continue
        if _g_rows[0].get('__variant') == 'vac_crew':
            _g_wr_raw = _g_rows[0].get('Work Request #', '')
            _g_wr = str(_g_wr_raw).split('.')[0]
            _g_wr = _RE_SANITIZE_HELPER_NAME.sub('_', _g_wr)[:50]
            if _g_wr:
                _scope.add(_g_wr)
    return _scope


def _build_primary_wr_scope(groups: dict) -> set[str]:
    """Return the set of sanitized WR tokens that have a partitioned
    production-primary ``_USER_`` group in this run (Subproject D).

    A group qualifies iff its authoritative ``__variant`` field (set at
    emission) is ``'primary'`` AND its key carries the ``_USER_`` partition
    token. The ``__variant`` gate — NOT a key substring scan — is what
    excludes helper / vac_crew / subcontractor groups, so a claimer,
    helper, or vac-crew name (or a pathological WR token) that itself
    contains a reserved word cannot false-positive into D's scope (Codex
    PR #223 P1). For example a helper literally named "USER" produces a
    key ``..._HELPER_USER_...`` whose ``_USER_`` substring the prior
    implementation mis-bucketed as a primary group; the ``__variant ==
    'primary'`` gate now rejects it. Conversely a genuine primary claimer
    named "ReducedSub"/"AEPBillable" (key ``..._USER_ReducedSub``) is
    correctly INCLUDED, whereas the prior ``'_REDUCEDSUB' not in _key``
    substring exclusion wrongly dropped it. The ``'_USER_' in _key`` clause
    then distinguishes a PARTITIONED primary (Subproject D ``_USER_`` group)
    from a bare primary (OFF mode / legacy ``RES_GROUPING_MODE='primary'``).
    Both call sites gate on ``PRIMARY_CLAIM_ATTRIBUTION_ENABLED``, so in
    production ``'both'`` mode every primary group is the partitioned form.

    Shared by ``_run_subproject_d_hash_prune`` (hash-prune scope) and the
    TARGET ``cleanup_untracked_sheet_attachments`` call site (bare-primary
    migration scope). A single implementation prevents the scope-build
    drift that the [2026-05-15 12:00] three-site invariant warns against.
    """
    _scope: set[str] = set()
    for _key, _g_rows in groups.items():
        if not _g_rows:
            continue
        if _g_rows[0].get('__variant') == 'primary' and '_USER_' in _key:
            _g_wr_raw = _g_rows[0].get('Work Request #', '')
            _g_wr = str(_g_wr_raw).split('.')[0]
            _g_wr = _RE_SANITIZE_HELPER_NAME.sub('_', _g_wr)[:50]
            if _g_wr:
                _scope.add(_g_wr)
    return _scope


def _run_phase_1_1_hash_prune(hash_history: dict, groups: dict) -> bool:
    """Phase 1.1 SUB-12 / D-17..D-19: idempotent hash-history prune.

    Version 1 (Plan 01.1-05): After Bug B1 (Plan 01.1-02) stops emitting
    legacy primary group keys for subcontractor rows, drops subcontractor
    primary orphans (4-part keys: ``wr|week|primary|``).

    Version 2 (Plan 01.1-06 UAT gap closure): ALSO drops subcontractor
    legacy ``'helper'`` orphans (6-part keys: ``wr|week|helper|foreman|
    dept|job``) left behind after Task 1 stops emitting the legacy
    ``_HELPER_<name>`` key for subcontractor helper rows. Version 2 is a
    superset of version 1 — the primary-orphan drop is preserved.

    This helper:

      1. Reads the persisted ``_phase_prune_version`` sentinel (or 0
         if absent) from ``hash_history``.
      2. If the persisted version is already at or beyond
         ``PHASE_1_1_HASH_PRUNE_VERSION``, restores the sentinel and
         returns — no-op.
      3. Otherwise, delegates scope-building to
         ``_build_subcontractor_wr_scope(groups)`` (shared with the
         TARGET cleanup call site — single implementation, no drift).
      4. Walks ``hash_history.keys()`` and identifies orphan entries
         using a length-tolerant guard (``< 4``) and index access.
         Primary orphans: 4-part keys with ``variant == 'primary'``
         and blank identifier. Helper orphans: any-part-count keys
         with ``variant == 'helper'`` and ``wr in scope``.
      5. Drops those entries in place, persists the new sentinel
         value, and logs ONE INFO line naming the count + affected
         WR sample (per RESEARCH.md §HP Code Example §6).

    CRITICAL — helper-variant hash keys are SIX pipe-parts, NOT four:
    ``wr|week|helper|foreman|dept|job``. The former ``!= 4`` guard
    hard-skipped every helper key (false-clean "no orphans" log) AND
    the 4-element destructure would raise ``ValueError`` on a 6-element
    list. Both are replaced by the ``< 4`` guard + index access pattern.

    Mutates ``hash_history`` in place. The constant
    ``PHASE_1_1_HASH_PRUNE_VERSION`` IS the kill switch (D-19) —
    advance to trigger; leave to skip. Per [2026-04-25 12:00] rule 1:
    idempotent migrations; version advance is the trigger.
    """
    _persisted_prune_version = hash_history.pop('_phase_prune_version', 0)
    if (
        isinstance(_persisted_prune_version, int)
        and _persisted_prune_version >= PHASE_1_1_HASH_PRUNE_VERSION
    ):
        # Re-store the sentinel so ``save_hash_history`` persists it
        # (defensive — the ``.pop`` above removed it). No log because
        # the prune already ran on a prior session and the absence
        # of a log line is the "already migrated" signal.
        hash_history['_phase_prune_version'] = _persisted_prune_version
        # Codex P2: no mutation beyond the no-op sentinel restore — the
        # caller need not force a save on a no-update run.
        return False

    # Build the WR-token set from this run's groups via shared helper
    # (simplified D-18). Shared with TARGET cleanup call site so the
    # two scopes are guaranteed identical (T-01.1-06-05 mitigation).
    _sub_wr_scope: set[str] = _build_subcontractor_wr_scope(groups)

    # Walk hash_history, identify subcontractor primary AND helper orphans.
    _orphans_to_drop: list[str] = []
    for _hk in list(hash_history.keys()):
        if isinstance(_hk, str) and _hk.startswith('_'):
            continue  # sentinel keys — skip
        _parts = str(_hk).split('|')
        # Helper-variant keys are 6 parts (wr|week|helper|foreman|dept|job);
        # primary keys are 4 (wr|week|primary|''). Accept BOTH — the former
        # ``!= 4`` guard hard-skipped every helper key, producing a false-
        # clean "no orphans" log. Sentinel keys (startswith '_') already
        # skipped above.
        if len(_parts) < 4:
            continue
        # Index access (NOT a 4-element destructure) so a 6-part helper
        # key does not raise ValueError. Only positions 0 and 2 are needed
        # for the orphan classification.
        _hk_wr = _parts[0]
        _hk_variant = _parts[2]
        # Version 1: subcontractor primary orphans (variant=='primary',
        #   blank identifier — EXACTLY 4 parts). Version 2 (Phase 1.1 UAT
        #   gap closure): ALSO subcontractor legacy helper orphans
        #   (variant=='helper', 6 parts, any foreman/dept/job) left behind
        #   after Task 1 stops emitting the legacy `_HELPER_<name>` key for
        #   subcontractor rows. Both are pre-fix leftovers in
        #   hash_history.json on existing deployments.
        # IN-01 (review follow-up): the ``_hk_variant == 'helper'`` clause
        # intentionally matches a 'helper' key at ANY part count, not just
        # the documented 6-part production shape. Real production helper
        # keys ARE 6-part — ``history_key = f"{wr}|{week}|{variant}|
        # {identifier}"`` with ``identifier = f"{foreman}|{dept}|{job}"``
        # for the helper variant — but the broad match is a deliberately
        # safe superset for TWO reasons: (1) the version sentinel makes this
        # prune one-time, and (2) the prune only DROPS a hash-history entry,
        # forcing at most one benign regeneration on the next run — it never
        # deletes a file. So even the WR-01 cross-sheet-overlap case (a live
        # non-sub helper key for a WR that also has sub rows) is benign here:
        # the file regenerates once. This is unlike the every-run TARGET
        # ``cleanup_untracked_sheet_attachments`` gate, which DOES delete and
        # therefore carries the ``ident not in valid_wr_weeks`` exemption.
        if _hk_wr in _sub_wr_scope and (
            (len(_parts) == 4 and _hk_variant == 'primary' and _parts[3] == '')
            or _hk_variant == 'helper'
        ):
            _orphans_to_drop.append(_hk)

    for _hk in _orphans_to_drop:
        del hash_history[_hk]

    # Persist the new sentinel.
    hash_history['_phase_prune_version'] = PHASE_1_1_HASH_PRUNE_VERSION

    # ONE INFO log — PII marker "Phase 1.1 hash-history prune" already
    # registered in ``_PII_LOG_MARKERS``. Limit the affected-WR list
    # to the first 20 entries to keep the log line bounded; full count
    # still surfaces.
    if _orphans_to_drop:
        _wr_sample = sorted(_sub_wr_scope)[:20]
        _wr_suffix = (
            '' if len(_sub_wr_scope) <= 20
            else f' (+ {len(_sub_wr_scope) - 20} more)'
        )
        logging.info(
            f"🧹 Phase 1.1 hash-history prune "
            f"(version {_persisted_prune_version} → "
            f"{PHASE_1_1_HASH_PRUNE_VERSION}): "
            f"dropped {len(_orphans_to_drop)} subcontractor "
            f"primary/legacy-helper orphan(s) affecting "
            f"{len(_sub_wr_scope)} WR(s). "
            f"Affected WRs (first 20): {_wr_sample}{_wr_suffix}"
        )
    else:
        logging.info(
            f"🧹 Phase 1.1 hash-history prune "
            f"(version {_persisted_prune_version} → "
            f"{PHASE_1_1_HASH_PRUNE_VERSION}): "
            f"no primary/legacy-helper orphans to drop."
        )
    # Codex P2: the body path advanced the sentinel (and may have dropped
    # orphans) — report the mutation so the caller persists hash_history even
    # on a run with no group updates (where the history_updates-gated save is
    # skipped). Without this the migration re-runs every no-update execution.
    return True


def _run_subproject_b_hash_prune(hash_history: dict, groups: dict) -> bool:
    """Subproject B (2026-05-20): idempotent one-time hash-history prune.

    Drops LEGACY blank-identifier subcontractor primary orphans —
    4-part keys ``wr|week|reduced_sub|`` and ``wr|week|aep_billable|``
    with an EMPTY identifier — for WRs that are subcontractor in this
    run. B re-partitions those variants by frozen claimer (new keys
    carry a non-empty identifier), so the blank-identifier entries are
    obsolete. The normal stale-prune at the end of the run would clear
    them eventually; this makes the migration deterministic on the first
    run and survives interrupted / no-update runs.

    Scope-building delegates to ``_build_subcontractor_wr_scope`` (shared
    with the cleanup call site — no drift, per the [2026-05-15 12:00]
    three-site invariant). Sentinel key ``_subproject_b_prune_version``
    is distinct from the Phase 1.1 ``_phase_prune_version`` so the two
    migrations are independent. Mutates ``hash_history`` in place.
    Dropping a hash entry costs at most one benign regeneration — never
    data loss — so no live-identity exemption is needed on this drop
    path (unlike the every-run attachment cleanup).
    """
    _persisted = hash_history.pop('_subproject_b_prune_version', 0)
    if (
        isinstance(_persisted, int)
        and _persisted >= SUBPROJECT_B_HASH_PRUNE_VERSION
    ):
        hash_history['_subproject_b_prune_version'] = _persisted
        # Codex P2: no-op sentinel restore only — no save needed.
        return False

    _scope = _build_subcontractor_wr_scope(groups)
    _orphans: list[str] = []
    for _hk in list(hash_history.keys()):
        if isinstance(_hk, str) and _hk.startswith('_'):
            continue
        _parts = str(_hk).split('|')
        if len(_parts) != 4:
            continue
        _hk_wr, _hk_week, _hk_variant, _hk_ident = _parts
        if (
            _hk_wr in _scope
            and _hk_variant in ('reduced_sub', 'aep_billable')
            and _hk_ident == ''
        ):
            _orphans.append(_hk)
    for _ok in _orphans:
        del hash_history[_ok]
    hash_history['_subproject_b_prune_version'] = SUBPROJECT_B_HASH_PRUNE_VERSION
    if _orphans:
        _wr_sample = sorted({k.split('|')[0] for k in _orphans})[:20]
        logging.info(
            f"🧹 Subproject B hash-history prune: dropped {len(_orphans)} "
            f"legacy unpartitioned reduced_sub/aep_billable orphan(s) "
            f"(affected WRs first 20: {_wr_sample})"
        )
    else:
        logging.info(
            "🧹 Subproject B hash-history prune: no legacy unpartitioned "
            "reduced_sub/aep_billable orphans to drop"
        )
    # Codex P2: body path advanced the sentinel (and may have dropped
    # orphans) — report the mutation so the caller persists it even on a
    # no-update run.
    return True


def _run_vac_crew_hash_prune(hash_history: dict, groups: dict) -> bool:
    """Subproject C (2026-05-21): idempotent one-time hash-history prune.

    Drops LEGACY blank-identifier vac_crew orphans — 4-part keys
    ``wr|week|vac_crew|`` with an EMPTY identifier — for WRs that are
    vac_crew in this run. C re-partitions vac_crew variants by frozen
    claimer (new keys carry a non-empty identifier), so the
    blank-identifier entries are obsolete. The normal stale-prune at the
    end of the run would clear them eventually; this makes the migration
    deterministic on the first run and survives interrupted / no-update
    runs.

    Scope-building delegates to ``_build_vac_crew_wr_scope`` (shared
    with the cleanup call site — no drift, per the [2026-05-15 12:00]
    three-site invariant). Sentinel key ``_vac_crew_prune_version`` is
    DISTINCT from ``_phase_prune_version`` (Phase 1.1) and
    ``_subproject_b_prune_version`` (Subproject B) so the three
    migrations are independent. Mutates ``hash_history`` in place.
    Dropping a hash entry costs at most one benign regeneration — never
    data loss — so no live-identity exemption is needed on this drop
    path (unlike the every-run attachment cleanup).

    Codex P2 (PR #219): when ``VAC_CREW_CLAIM_ATTRIBUTION_ENABLED`` is OFF,
    the blank-identifier ``wr|week|vac_crew|`` key is the ACTIVE legacy
    format (the kill-switch-OFF path emits the bare ``_VACCREW`` key), so
    pruning it would delete valid current history and force regeneration
    churn — breaking the exact-legacy contract. Skip entirely when the flag
    is off, and do NOT advance the sentinel, so the one-time migration still
    runs if attribution is later enabled.
    """
    if not VAC_CREW_CLAIM_ATTRIBUTION_ENABLED:
        return False
    _persisted = hash_history.pop('_vac_crew_prune_version', 0)
    if (
        isinstance(_persisted, int)
        and _persisted >= VAC_CREW_HASH_PRUNE_VERSION
    ):
        hash_history['_vac_crew_prune_version'] = _persisted
        # Codex P2: no-op sentinel restore only — no save needed.
        return False

    _scope = _build_vac_crew_wr_scope(groups)
    _orphans: list[str] = []
    for _hk in list(hash_history.keys()):
        if isinstance(_hk, str) and _hk.startswith('_'):
            continue
        _parts = str(_hk).split('|')
        if len(_parts) != 4:
            continue
        _hk_wr, _hk_week, _hk_variant, _hk_ident = _parts
        if (
            _hk_wr in _scope
            and _hk_variant == 'vac_crew'
            and _hk_ident == ''
        ):
            _orphans.append(_hk)
    for _ok in _orphans:
        del hash_history[_ok]
    hash_history['_vac_crew_prune_version'] = VAC_CREW_HASH_PRUNE_VERSION
    if _orphans:
        _wr_sample = sorted({k.split('|')[0] for k in _orphans})[:20]
        logging.info(
            f"🧹 Vac crew hash-history prune: dropped {len(_orphans)} legacy "
            f"unpartitioned vac_crew orphan(s) "
            f"(WRs first 20: {_wr_sample})"
        )
    else:
        logging.info(
            "🧹 Vac crew hash-history prune: no legacy vac_crew orphans to drop"
        )
    # Body path advanced the sentinel (and may have dropped orphans) —
    # report the mutation so the caller persists it even on a no-update run.
    return True


def _run_subproject_d_hash_prune(hash_history: dict, groups: dict) -> bool:
    """Subproject D (2026-05-25): idempotent one-time hash-history prune.

    Drops LEGACY blank-identifier production-primary orphans — 4-part keys
    ``wr|week|primary|`` with an EMPTY identifier — for WRs that have a
    partitioned ``_USER_`` primary group in this run. D re-partitions the
    production primary variant by frozen claimer (new keys carry a
    non-empty identifier), so the blank-identifier entries are obsolete.
    The normal stale-prune at the end of the run would clear them
    eventually; this makes the migration deterministic on the first run
    and survives interrupted / no-update runs.

    Scope-building delegates to ``_build_primary_wr_scope`` (shared with
    the TARGET cleanup call site — no drift, per the [2026-05-15 12:00]
    three-site invariant). Sentinel key ``_subproject_d_prune_version`` is
    DISTINCT from the Phase 1.1 / Subproject B / Subproject C sentinels so
    all four migrations are independent. Mutates ``hash_history`` in place.
    Dropping a hash entry costs at most one benign regeneration — never
    data loss — so no live-identity exemption is needed on this drop path
    (unlike the every-run attachment cleanup).

    GATED on ``PRIMARY_CLAIM_ATTRIBUTION_ENABLED``: when OFF, the
    blank-identifier ``wr|week|primary|`` key is the ACTIVE legacy format
    (the kill-switch-OFF path emits the bare primary key), so pruning it
    would delete valid current history and force regeneration churn —
    breaking the exact-legacy contract. Skip entirely when the flag is
    off, and do NOT advance the sentinel, so the one-time migration still
    runs if attribution is later enabled. (Mirrors the Subproject C
    ``_run_vac_crew_hash_prune`` kill-switch guard.)
    """
    if not PRIMARY_CLAIM_ATTRIBUTION_ENABLED:
        return False
    _persisted = hash_history.pop('_subproject_d_prune_version', 0)
    if (
        isinstance(_persisted, int)
        and _persisted >= SUBPROJECT_D_HASH_PRUNE_VERSION
    ):
        hash_history['_subproject_d_prune_version'] = _persisted
        return False

    _scope = _build_primary_wr_scope(groups)
    _orphans: list[str] = []
    for _hk in list(hash_history.keys()):
        if isinstance(_hk, str) and _hk.startswith('_'):
            continue
        _parts = str(_hk).split('|')
        if len(_parts) != 4:
            continue
        _hk_wr, _hk_week, _hk_variant, _hk_ident = _parts
        if (
            _hk_wr in _scope
            and _hk_variant == 'primary'
            and _hk_ident == ''
        ):
            _orphans.append(_hk)
    for _ok in _orphans:
        del hash_history[_ok]
    hash_history['_subproject_d_prune_version'] = SUBPROJECT_D_HASH_PRUNE_VERSION
    if _orphans:
        _wr_sample = sorted({k.split('|')[0] for k in _orphans})[:20]
        logging.info(
            f"🧹 Subproject D hash-history prune: dropped {len(_orphans)} "
            f"legacy unpartitioned primary orphan(s) "
            f"(affected WRs first 20: {_wr_sample})"
        )
    else:
        logging.info(
            "🧹 Subproject D hash-history prune: no legacy unpartitioned "
            "primary orphans to drop"
        )
    # Body path advanced the sentinel (and may have dropped orphans) —
    # report the mutation so the caller persists it even on a no-update run.
    return True


# ── Garbage patterns for the claimer-remediation sweep (Phase 2 Plan 03) ──
# These are the EXACT tokens that ``resolve_claimer`` emits when attribution
# has no frozen history (``#NO_MATCH``) or a blank role (``Unknown_Foreman``).
# They are not realistic human foreman names, so a simple substring match is
# safe (WARNING 6 accepted tradeoff, per the plan's threat-model).
_GARBAGE_PATTERNS: tuple[str, ...] = ('_NO_MATCH', '_Unknown_Foreman')
# WR-04: in the isolated EXECUTE path (valid_wr_weeks=None), only tokens that
# are NEVER a legitimate filename component are swept.  _NO_MATCH is a pure
# Smartsheet ``#NO MATCH`` error token that should never appear in a real file.
# _Unknown_Foreman IS a legitimate current sentinel (emitted when
# ``effective_user`` is blank) and is preserved in the isolated path because
# there is no live-identity set to protect it — an EXECUTE sweep with
# valid_wr_weeks=None would otherwise delete a valid billing artifact.
_ALWAYS_GARBAGE_PATTERNS: tuple[str, ...] = ('_NO_MATCH',)


def run_claimer_remediation(
    client,
    dry_run: bool,
    window_weeks: int,
    valid_wr_weeks: 'set | None' = None,
) -> None:
    """Sweep TARGET_SHEET_ID and SUBCONTRACTOR_PPP_SHEET_ID for garbage claimer
    attachments (``*_NO_MATCH*`` / ``*_Unknown_Foreman*``) and delete them.

    Phase 2 Plan 03 — D-06/D-07/D-08/D-12/D-14.

    Parameters
    ----------
    client:
        Initialized ``smartsheet.Smartsheet`` client.
    dry_run:
        When True, report counts only — no attachment is deleted.
        Matches ``REMEDIATION_DRY_RUN`` default (``'1'``).
    window_weeks:
        Sweep only attachments whose week-ending date is within the last N
        weeks of today (``0`` = unbounded).  Matches
        ``REMEDIATION_WINDOW_WEEKS`` default (``26``).
    valid_wr_weeks:
        A set of ``(wr, week_mmddyy, variant, identifier)`` 4-tuples
        representing the current run's live attachments.  When provided,
        a garbage-named file whose parsed 4-tuple IS in this set is
        EXEMPTED from deletion (live-identity exemption per
        [2026-05-19 23:45]).  Pass ``None`` for the isolated-mode path
        where no live-identity set is available — deletion is then gated
        solely on the name-pattern and window filter (WR-04: only the
        always-garbage ``_NO_MATCH`` token is swept; ``_Unknown_Foreman``
        is preserved because it is a legitimate current sentinel and there
        is no live-identity set to protect it in the isolated path).
    """
    # IN-04: use the module-level datetime (not a shadowing local import).
    _today = datetime.date.today()
    _cutoff = (
        _today - datetime.timedelta(weeks=window_weeks)
        if window_weeks > 0
        else None
    )

    # WR-04: select the active garbage-pattern set based on whether the
    # live-identity exemption is available.  When valid_wr_weeks is None
    # (isolated path), restrict to _ALWAYS_GARBAGE_PATTERNS so a current
    # _Unknown_Foreman billing artifact is never deleted.
    _patterns = _GARBAGE_PATTERNS if valid_wr_weeks is not None else _ALWAYS_GARBAGE_PATTERNS

    # Determine which sheets to sweep.
    _sheet_ids: list[int] = [TARGET_SHEET_ID]
    if SUBCONTRACTOR_PPP_SHEET_ID:
        _sheet_ids.append(SUBCONTRACTOR_PPP_SHEET_ID)

    _total_scanned = 0
    _total_garbage = 0
    _total_deleted = 0
    _total_exempted = 0
    _total_out_of_window = 0

    for _sheet_id in _sheet_ids:
        try:
            _sheet = client.Sheets.get_sheet(_sheet_id)
        except Exception as _e:
            logging.warning(
                f"⚠️ run_claimer_remediation: failed to fetch sheet "
                f"{_sheet_id}: {_redact_exception_message(_e)}"
            )
            continue

        for _row in _sheet.rows:
            try:
                _row_resp = client.Attachments.list_row_attachments(
                    _sheet_id, _row.id
                )
            except Exception as _e:
                logging.warning(
                    f"⚠️ run_claimer_remediation: failed to list attachments "
                    f"for row {_row.id} on sheet {_sheet_id}: "
                    f"{_redact_exception_message(_e)}"
                )
                continue

            _attachments = getattr(_row_resp, 'attachments', None) or []
            for _att in _attachments:
                _name: str = getattr(_att, 'name', '') or ''
                _att_id = _att.id
                _total_scanned += 1

                # ── Step 1: parse filename with the battle-hardened parser ──
                # Files that build_group_identity cannot parse (non-WR filenames,
                # malformed names) are left alone — never deleted.
                _identity = build_group_identity(_name)
                if _identity is None:
                    continue  # unparseable → skip

                _wr, _week_mmddyy, _variant, _identifier = _identity

                # ── Step 2: garbage-pattern check (IN-02: runs BEFORE window) ──
                # Check the active pattern set (WR-04: _ALWAYS_GARBAGE_PATTERNS in
                # the isolated path, _GARBAGE_PATTERNS when the live-identity
                # exemption is available).  Clean real-claimer files never reach
                # the window filter, so out_of_window counts only GARBAGE files
                # that are too old — unambiguous blast-radius metric.
                _is_garbage = any(pat in _name for pat in _patterns)
                if not _is_garbage:
                    continue  # clean real-claimer name → skip

                # ── Step 3: window filter (runs only for garbage files) ──
                # Convert the MMDDYY week token to a date for comparison.
                if _cutoff is not None:
                    try:
                        _week_date = datetime.datetime.strptime(
                            _week_mmddyy, '%m%d%y'
                        ).date()
                        if _week_date < _cutoff:
                            _total_out_of_window += 1
                            continue  # too old — skip
                    except (ValueError, TypeError):
                        # Unparseable week token → conservatively skip
                        continue

                _total_garbage += 1

                # ── Step 4: live-identity exemption ──
                if valid_wr_weeks is not None and _identity in valid_wr_weeks:
                    _total_exempted += 1
                    logging.debug(
                        f"run_claimer_remediation: EXEMPT (live identity) "
                        f"att={_att_id} sheet={_sheet_id} "
                        f"wr={_wr} week={_week_mmddyy} variant={_variant}"
                    )
                    continue

                # ── Step 5: dry-run or execute ──
                if dry_run:
                    logging.info(
                        f"🔍 [DRY-RUN] would delete garbage attachment "
                        f"att={_att_id} sheet={_sheet_id} "
                        f"wr={_wr} week={_week_mmddyy} variant={_variant}"
                    )
                else:
                    try:
                        client.Attachments.delete_attachment(_sheet_id, _att_id)
                        _total_deleted += 1
                        logging.info(
                            f"🗑️ run_claimer_remediation: deleted garbage att "
                            f"att={_att_id} sheet={_sheet_id} "
                            f"wr={_wr} week={_week_mmddyy} variant={_variant}"
                        )
                    except Exception as _del_e:
                        logging.warning(
                            f"⚠️ run_claimer_remediation: failed to delete "
                            f"att={_att_id} sheet={_sheet_id}: "
                            f"{_redact_exception_message(_del_e)}"
                        )

    # ── PII-safe aggregate summary (ASVS V7: counts + sanitized IDs only) ──
    _mode = "DRY-RUN" if dry_run else "EXECUTE"
    logging.info(
        f"✅ run_claimer_remediation [{_mode}] complete: "
        f"scanned={_total_scanned} garbage={_total_garbage} "
        f"deleted={_total_deleted} exempted={_total_exempted} "
        f"out_of_window={_total_out_of_window}"
    )


def load_billing_audit_row_cache(path: str) -> set[str]:
    """Load cached freeze-attribution row keys.

    Keys are ``{wr_sanitized}|{week_mmddyy}|{row_id}``. This local cache
    is best-effort only — if missing/corrupt we simply fall back to
    normal freeze-row behavior.
    """
    try:
        with open(path, "r") as f:
            data = json.load(f)
        if isinstance(data, list):
            return {str(x) for x in data if x is not None}
        if isinstance(data, dict):
            # Backward-compatible shape if we later add metadata.
            rows = data.get("rows", [])
            if isinstance(rows, list):
                return {str(x) for x in rows if x is not None}
        logging.warning("⚠️ Billing-audit row cache malformed; resetting")
        return set()
    except FileNotFoundError:
        return set()
    except Exception as e:
        logging.warning(f"⚠️ Failed to load billing-audit row cache: {e}")
        return set()


def save_billing_audit_row_cache(path: str, rows: set[str]) -> None:
    """Persist cached freeze-attribution row keys."""
    try:
        # Always sort set-backed cache entries so serialized output is
        # deterministic across runs; also produces smaller diffs.
        values = sorted(rows)
        if len(values) > BILLING_AUDIT_ROW_CACHE_MAX_ENTRIES:
            # Deterministic truncation. Cache is opportunistic; precision
            # is not required as fallback is to re-call freeze_row.
            values = values[-BILLING_AUDIT_ROW_CACHE_MAX_ENTRIES:]
            retained = len(values)
            logging.info(
                f"🧹 Pruned billing-audit row cache to {retained} entries"
            )
        tmp_path = path + ".tmp"
        with open(tmp_path, "w") as f:
            json.dump(values, f, separators=(",", ":"))
        os.replace(tmp_path, path)
        logging.info(f"📝 Billing-audit row cache saved ({len(values)} entries)")
    except Exception as e:
        logging.warning(f"⚠️ Failed to save billing-audit row cache: {e}")

# --- DATA DISCOVERY AND PROCESSING ---

def _title(t):
    return (t or "").strip().lower()


def _normalize_column_title_for_vac_crew(t):
    """Normalize a Smartsheet column title for fuzzy VAC Crew matching.

    Lowercases, canonicalises hyphenated (``vac-crew``) and joined-word
    (``vaccrew``) variants into the ``vac crew`` token, collapses runs of
    whitespace to a single space, and strips decorative trailing ``?`` or
    ``#`` (with optional surrounding spaces) so that operator-introduced
    variants like ``'Vac Crew Helping ?'``, ``'VAC CREW Helping?'``,
    ``'vac  crew helping'``, ``'Vac-Crew Helping?'``, ``'VacCrew Dept#'``
    or ``'Vac Crew Dept#'`` collapse to the same key as the canonical
    ``'VAC Crew Helping?'`` / ``'VAC Crew Dept #'``. Scoped intentionally
    narrow — only used by the VAC Crew fuzzy fallback in
    ``_validate_single_sheet`` — so primary/helper exact-match behaviour
    is preserved.
    """
    s = (t or "").strip().lower()
    # Hyphenated variants ('vac-crew') → space-separated.
    s = s.replace("-", " ")
    # Joined-word variants ('vaccrew') → space-separated. Word-boundary
    # guarded so unrelated tokens that happen to contain 'vaccrew' as a
    # substring are left untouched.
    s = re.sub(r"\bvaccrew\b", "vac crew", s)
    # Collapse any whitespace runs introduced by the substitutions above.
    s = re.sub(r"\s+", " ", s).strip()
    # Strip decorative trailing '?' / '#' with optional surrounding spaces.
    s = re.sub(r"\s*[\?#]+\s*$", "", s)
    return s


def discover_source_sheets(client):
    """Strict deterministic discovery: anchored keywords + type filtered. Skips sheets missing Weekly Reference Logged Date."""
    global _FOLDER_DISCOVERED_SUB_IDS, _FOLDER_DISCOVERED_ORIG_IDS, SUBCONTRACTOR_SHEET_IDS

    # ── ALWAYS run folder discovery FIRST (detects new sheets every run) ──────────
    # Folder listing is cheap (2-4 API calls + subfolder recursion).  Running it
    # unconditionally ensures sheets added to configured folders between runs are
    # detected even when the discovery cache is still within TTL.
    if SUBCONTRACTOR_FOLDER_IDS:
        _FOLDER_DISCOVERED_SUB_IDS = discover_folder_sheets(client, SUBCONTRACTOR_FOLDER_IDS, 'subcontractor')
        SUBCONTRACTOR_SHEET_IDS = SUBCONTRACTOR_SHEET_IDS | _FOLDER_DISCOVERED_SUB_IDS
        logging.info(f"📂 Subcontractor sheet IDs after folder merge: {len(SUBCONTRACTOR_SHEET_IDS)}")
    if ORIGINAL_CONTRACT_FOLDER_IDS:
        _FOLDER_DISCOVERED_ORIG_IDS = discover_folder_sheets(client, ORIGINAL_CONTRACT_FOLDER_IDS, 'original contract')
    _all_folder_discovered_ids = _FOLDER_DISCOVERED_SUB_IDS | _FOLDER_DISCOVERED_ORIG_IDS

    # ── Attempt cache load (skip when forced rediscovery requested) ──
    _cached_sheets = []          # previously-validated sheets from cache (used for incremental mode)
    _cached_sheet_ids = set()    # IDs of sheets already validated in cache
    _incremental = False         # True when cache expired but sheets can be reused
    if FORCE_REDISCOVERY:
        logging.info("🔄 FORCE_REDISCOVERY=true — bypassing discovery cache")
    elif USE_DISCOVERY_CACHE and os.path.exists(DISCOVERY_CACHE_PATH):
        try:
            with open(DISCOVERY_CACHE_PATH,'r') as f:
                cache = json.load(f)
            # Check cache schema version — invalidate if column synonyms have changed
            cached_version = cache.get('schema_version', 1)
            if cached_version < DISCOVERY_CACHE_VERSION:
                logging.info(f"🔄 Discovery cache schema outdated (v{cached_version} < v{DISCOVERY_CACHE_VERSION}) — forcing full rediscovery")
                raise ValueError('cache schema outdated')
            ts = datetime.datetime.fromisoformat(cache.get('timestamp'))
            age_min = (datetime.datetime.now() - ts).total_seconds()/60.0
            # Schema guard: each cached sheet must be a dict with an
            # integer ``id`` and a dict ``column_mapping`` — anything
            # else would crash ``_fetch_and_process_sheet`` when it
            # reads ``source['column_mapping']`` / ``source['id']``.
            # Drop malformed entries and WARN so operators can see
            # a corrupted cache immediately rather than debugging a
            # later AttributeError / KeyError.
            _raw_cached_sheets = cache.get('sheets', []) or []
            _valid_cached_sheets = [
                s for s in _raw_cached_sheets
                if isinstance(s, dict)
                and isinstance(s.get('id'), int)
                and isinstance(s.get('column_mapping'), dict)
                and isinstance(s.get('name'), str)
            ]
            if len(_valid_cached_sheets) != len(_raw_cached_sheets):
                logging.warning(
                    f"⚠️ Discovery cache contains "
                    f"{len(_raw_cached_sheets) - len(_valid_cached_sheets)} "
                    f"malformed sheet entry(ies); dropping them "
                    f"(keeping {len(_valid_cached_sheets)} valid). "
                    f"Delete {DISCOVERY_CACHE_PATH} to force a clean rediscovery."
                )
                # If *every* cached entry was malformed, the fresh-cache
                # return path below would otherwise hand back an empty
                # source list and the run would silently process zero
                # sheets. Escalate to the outer cache-load-failed handler
                # so we fall through to a full rediscovery from
                # ``base_sheet_ids`` — same behaviour as an outdated
                # schema or unreadable JSON.
                if _raw_cached_sheets and not _valid_cached_sheets:
                    raise ValueError(
                        f"all {len(_raw_cached_sheets)} cached sheet "
                        f"entries malformed; forcing full rediscovery"
                    )
            _cached_sheet_ids_from_file = {s['id'] for s in _valid_cached_sheets}
            # Compare folder-discovered sheet IDs against cache to detect new sheets
            _new_from_folders = _all_folder_discovered_ids - _cached_sheet_ids_from_file
            # Codex P2 guardrail: if the schema filter dropped ANY
            # entry, skip the fresh-cache fast path. A dropped entry
            # may have been a required static base sheet that isn't
            # in _all_folder_discovered_ids, so _new_from_folders
            # wouldn't flag it. Falling through to incremental mode
            # forces base_sheet_ids to be re-validated and the
            # dropped sheet to be rediscovered on this run instead
            # of waiting until cache expiry (up to
            # DISCOVERY_CACHE_TTL_MIN — default 7 days).
            _partial_cache_corruption = bool(_raw_cached_sheets) and (
                len(_valid_cached_sheets) != len(_raw_cached_sheets)
            )
            if (
                age_min <= DISCOVERY_CACHE_TTL_MIN
                and not _new_from_folders
                and not _partial_cache_corruption
            ):
                # Cache is fresh AND no new sheets in folders AND no
                # malformed entries were dropped → safe to use cache
                cached_sub_ids = cache.get('subcontractor_sheet_ids', [])
                if cached_sub_ids:
                    SUBCONTRACTOR_SHEET_IDS = SUBCONTRACTOR_SHEET_IDS | set(cached_sub_ids)
                    logging.info(f"📂 Restored {len(cached_sub_ids)} subcontractor sheet IDs from cache (total: {len(SUBCONTRACTOR_SHEET_IDS)})")
                logging.info(f"⚡ Using cached discovery ({age_min:.1f} min old) with {len(_valid_cached_sheets)} sheets (folders unchanged)")
                return _valid_cached_sheets
            else:
                # Cache expired OR new sheets found in folders → incremental mode
                _cached_sheets = _valid_cached_sheets
                _cached_sheet_ids = _cached_sheet_ids_from_file
                cached_sub_ids = cache.get('subcontractor_sheet_ids', [])
                if cached_sub_ids:
                    SUBCONTRACTOR_SHEET_IDS = SUBCONTRACTOR_SHEET_IDS | set(cached_sub_ids)
                _incremental = True
                if _partial_cache_corruption:
                    _dropped_count = (
                        len(_raw_cached_sheets) - len(_valid_cached_sheets)
                    )
                    logging.info(
                        f"🛡️ {_dropped_count} malformed cached entry(ies) "
                        f"dropped — forcing incremental revalidation against "
                        f"base_sheet_ids so any required sheet among the "
                        f"dropped entries is rediscovered this run instead "
                        f"of waiting until cache expiry."
                    )
                elif _new_from_folders:
                    logging.info(f"🆕 {len(_new_from_folders)} new sheet(s) detected in folders — "
                                 f"cache invalidated, using incremental mode "
                                 f"(keeping {len(_cached_sheets)} cached + validating new sheets)")
                else:
                    logging.info(f"ℹ️ Discovery cache expired ({age_min:.1f} min old); using incremental mode — "
                                 f"keeping {len(_cached_sheets)} cached sheets, scanning for new IDs only")
        except Exception as e:
            logging.info(f"Cache load failed, refreshing discovery: {e}")
    base_sheet_ids = [
        3239244454645636, 2230129632694148, 1732945426468740, 4126460034895748,
        7899446718189444, 1964558450118532, 5905527830695812, 820644963897220,
        8002920231423876, 2308525217763204,  # Added per user request
        5892629871939460, 
        3756603854507908, # Added Intake Promax
        5833510251089796,  # Added per user request
        5291853336235908,  # Added per user request
        6399146438119300, # Added per user request
        2582148201533316, # Added Resiliency Promax Database 16
        589443900592004, # Added Resiliency Promax Database 17
        7112742503665540, # Added Resiliency Promax Database 18
        8882702989086596, #Added Resiliency Promax Database 19
        2329343909908356, #Added Resiliency Promax Database 20
        5635469074190212, #Added Resiliency Promax Database 21
        5962351384678276, #Added Resiliency Promax Database 22
        3892736567496580, #Added Resiliency Promax Database 23
        4973034927509380, #Added Resiliency Promax Database 24
        1705871626162052, # Added Resiliency Promax Database 25
        5214601672085380, #  Added Resiliency Promax Database 26
        8551186744430468, # Added Resiliency Promax Database 27
        7820299006332804, # Added Resiliency Promax Database 28
        1153867531112324, # Added Resiliency Promax Database 29
        6692045306417028, # Added Resiliency Promax Database 30
        249276132183940, # Added Intake Promax 2
        2126238714908548, # Added Promax Database 31
        366100316376964, # Added Promax Database 32
        1207776467439492, # Added Promax Database 33
        342733613911940, # Added Promax Database 34
        6658677403504516, # Added Promax Database 35
        7043847386255236, # Added Promax Database 36
        2920263713771396, # Added Intake Promax 3
        4317397608517508, # Added Resiliency Promax 37
        277473162907524, # Added Resiliency Promax (New)
        1697214691757956, # Added Resiliency Promax 38
        8823469929090948, # Added Resiliency Promax 39
        692599695298436, #Added Resiliency Promax 40
        2183127494512516, #Added Resiliency Promax 42
        4774094567329668, #Added Resiliency Promax 41
        5067039388422020, #Adde Resiliency Promax 44
        888996134604676, #Added Resiliency Promax 43
        6920127724343172, #Added Resiliency Promax 45
        6587491768291204, #Added Intake Promax 4
        1804369797271428, #Added Resiliency Promax Database 46
        2873734244290436, #Added Resiliency Promax Database 47
        8153606260739972, #Added Resiliency Promax Database 48
        7630927397080964, #Added Resiliency Promax Database 49
        4017481216642948, #Added Resiliency Promax Database 50
        1326553209196420, #Added Resiliency Promax Database 51
        6479287751233412, #Added Resiliency Promax Database 52
        2672627970690948, #Added Resiliency Promax Database 53
        3800355386118020, #Added Intake Promax Database 5
        6123743714692996, #Added Resiliency Promax Database 54
        3804822152105860, #Added Resiliency Promax Database 55
        5065263654326148, #Added Resiliency Promax Database 56
        5417244814167940, #Added Resiliency Promax Database 57
        2431001297899396, #Added Resiliency Promax Database 58
        4085014678425476, #Added Intake Promax 6
        1080481698238340, #Added Resiliency Promax Database 59
        8391967734976388, #Added Resiliency Promax Database 60
        2233624515530628, #Added Resiliency Promax Database 61
        2780425391918980 #Added Resiliency Promax Database 62
        
        

    ]

    # OPTIONAL SPEED-UP FOR TESTING: allow overriding sheet list via env LIMITED_SHEET_IDS
    # Comma-separated list of numeric sheet IDs. If provided, we restrict discovery to only these.
    _limited_ids_raw = os.getenv('LIMITED_SHEET_IDS')
    if _limited_ids_raw:
        try:
            limited_ids = [int(x.strip()) for x in _limited_ids_raw.split(',') if x.strip()]
            if limited_ids:
                logging.info(f"⏩ LIMITED_SHEET_IDS override active ({len(limited_ids)} IDs); restricting discovery to provided list")
                base_sheet_ids = limited_ids
        except Exception as e:
            logging.warning(f"⚠️ LIMITED_SHEET_IDS parse failed '{_limited_ids_raw}': {e}")
    
    # Merge folder-discovered sheet IDs (populated unconditionally at top of function)
    _folder_ids = _FOLDER_DISCOVERED_SUB_IDS | _FOLDER_DISCOVERED_ORIG_IDS
    if _folder_ids:
        existing = set(base_sheet_ids)
        new_ids = _folder_ids - existing
        if new_ids:
            base_sheet_ids.extend(sorted(new_ids))
            logging.info(f"📂 Merged {len(new_ids)} folder-discovered sheet(s) into discovery list (total: {len(base_sheet_ids)})")
        else:
            logging.info(f"📂 All {len(_folder_ids)} folder-discovered sheet(s) already in base list")

    discovered = []

    def _validate_single_sheet(sid):
        """Validate a single sheet and return its discovery dict (or None if invalid)."""
        try:
            # PERFORMANCE FIX: Fetch only column metadata initially (no row data needed yet)
            # This prevents Error 4000 for large sheets during discovery phase
            sheet = client.Sheets.get_sheet(sid, include='columns')
            cols = sheet.columns
            mapping = {}
            by_title = { _title(c.title): c for c in cols }
            # Exact matches
            w_exact = by_title.get(_title('Weekly Reference Logged Date'))
            s_exact = by_title.get(_title('Snapshot Date'))
            if w_exact: mapping['Weekly Reference Logged Date'] = w_exact.id
            if s_exact: mapping['Snapshot Date'] = s_exact.id
            # Date candidates
            date_candidates = [c for c in cols if str(c.type).upper() in ('DATE','DATETIME')]
            if 'Weekly Reference Logged Date' not in mapping:
                keyed = [c for c in date_candidates if 'date' in _title(c.title) and any(k in _title(c.title) for k in ('weekly','reference','logged','week ending'))]
                if keyed:
                    mapping['Weekly Reference Logged Date'] = keyed[0].id
            if 'Snapshot Date' not in mapping:
                keyed = [c for c in date_candidates if 'date' in _title(c.title) and 'snapshot' in _title(c.title)]
                if keyed:
                    mapping['Snapshot Date'] = keyed[0].id
            # Sample fallback — fetch sample rows ONCE for all column checks
            _sample_rows_cache = None
            def _get_sample_rows():
                nonlocal _sample_rows_cache
                if _sample_rows_cache is None:
                    try:
                        _sample_sheet = client.Sheets.get_sheet(sid, row_numbers=list(range(1, 4)))
                        _sample_rows_cache = _sample_sheet.rows if _sample_sheet.rows else []
                    except Exception:
                        _sample_rows_cache = []
                return _sample_rows_cache

            def _extract_col_samples(col_id):
                """Extract sample values for a column from the cached sample rows."""
                vals = []
                for row in _get_sample_rows():
                    for cell in row.cells:
                        if cell.column_id == col_id:
                            val = getattr(cell, 'value', None)
                            if val is None:
                                val = getattr(cell, 'display_value', None)
                            if val is not None:
                                vals.append(str(val))
                            break
                return vals

            if 'Weekly Reference Logged Date' not in mapping:
                for c in date_candidates:
                    t = _title(c.title)
                    if 'date' in t and any(k in t for k in ('weekly','reference','logged','week ending')):
                        samples = _extract_col_samples(c.id)
                        if any(re.match(r'^\d{4}-\d{2}-\d{2}', v) for v in samples):
                            mapping['Weekly Reference Logged Date'] = c.id
                            break
            if 'Snapshot Date' not in mapping:
                for c in date_candidates:
                    t = _title(c.title)
                    if 'date' in t and 'snapshot' in t:
                        samples = _extract_col_samples(c.id)
                        if any(re.match(r'^\d{4}-\d{2}-\d{2}', v) for v in samples):
                            mapping['Snapshot Date'] = c.id
                            break
            # Non-date synonyms
            synonyms = {
                'Foreman':'Foreman','Work Request #':'Work Request #','Dept #':'Dept #','Customer Name':'Customer Name','Work Order #':'Work Order #','Area':'Area',
                'Pole #':'Pole #','Point #':'Pole #','Point Number':'Pole #','CU':'CU','Billable Unit Code':'CU','Work Type':'Work Type','CU Description':'CU Description',
                'Unit Description':'CU Description','Unit of Measure':'Unit of Measure','UOM':'Unit of Measure','Quantity':'Quantity','Qty':'Quantity','# Units':'Quantity',
                'Units Total Price':'Units Total Price','Total Price':'Units Total Price','Redlined Total Price':'Units Total Price','Scope #':'Scope #','Scope ID':'Scope #',
                'Job #':'Job #','Units Completed?':'Units Completed?','Units Completed':'Units Completed?',
                # Helper variant columns (exact names with brackets as authoritative)
                'Helper Job [#]':'Helper Job #',  # Exact spelling with brackets
                'Helper Job':'Helper Job #',      # Fallback synonym
                'Helper Job #':'Helper Job #',    # Ensure direct exact match is captured
                'Helper Dept #':'Helper Dept #',
                'Foreman Helping?':'Foreman Helping?',
                'Helping Foreman Completed Unit?':'Helping Foreman Completed Unit?',
                # VAC Crew variant columns (row-level detection — mirrors helper pattern)
                'VAC Crew Helping?':'VAC Crew Helping?',
                'Vac Crew Helping?':'VAC Crew Helping?',          # Case variant
                'Vac Crew Completed Unit?':'Vac Crew Completed Unit?',
                'VAC Crew Completed Unit?':'Vac Crew Completed Unit?',  # Case variant
                'VAC Crew Dept #':'VAC Crew Dept #',
                'Vac Crew Dept #':'VAC Crew Dept #',              # Case variant
                'Vac Crew Job #':'Vac Crew Job #',
                'VAC Crew Job #':'Vac Crew Job #',                # Case variant
                'Vac Crew Email Address':'Vac Crew Email Address',
                'VAC Crew Email Address':'Vac Crew Email Address', # Case variant
            }
            # COLUMN MAPPING DEBUG: Log all column titles to verify helper and VAC Crew columns
            helper_columns_found = []
            vac_crew_columns_found = []
            for c in cols:
                # Check for helper-related columns specifically
                if 'Helper' in c.title or 'Helping' in c.title:
                    helper_columns_found.append(c.title)
                # Check for VAC Crew-related columns (case-insensitive so lowercase
                # or hyphenated variants like 'vac crew' / 'vac-crew' still surface
                # in logs and feed the fuzzy fallback pass below).
                _ct_lower = (c.title or '').lower()
                if 'vac crew' in _ct_lower or 'vaccrew' in _ct_lower or 'vac-crew' in _ct_lower:
                    vac_crew_columns_found.append(c.title)

                if c.title in synonyms and synonyms[c.title] not in mapping:
                    mapping[synonyms[c.title]] = c.id
                    # Log helper column mappings specifically
                    if 'Helper' in c.title:
                        logging.info(f"🔧 MAPPED HELPER COLUMN: '{c.title}' -> '{synonyms[c.title]}' (column ID: {c.id})")
                    # Log VAC Crew column mappings
                    if 'Vac Crew' in c.title or 'VAC Crew' in c.title:
                        logging.info(f"🚐 MAPPED VAC CREW COLUMN: '{c.title}' -> '{synonyms[c.title]}' (column ID: {c.id})")

            # ── VAC Crew column fuzzy fallback ──
            # The exact-match pass above only catches the two literal case variants
            # declared in `synonyms` (e.g. 'VAC Crew Helping?' and 'Vac Crew Helping?').
            # Operators occasionally introduce subtle variants on a new sheet —
            # trailing / leading whitespace, missing or extra '?', double internal
            # spaces, all-caps 'VAC CREW', all-lowercase 'vac crew' — and any such
            # variant silently fails to map. When the two KEY columns
            # ('VAC Crew Helping?' and 'Vac Crew Completed Unit?') aren't in
            # `mapping`, `sheet_has_vac_crew_columns` in _fetch_and_process_sheet
            # evaluates False and the row-level VAC Crew detection block is
            # skipped wholesale — the sheet produces zero _VacCrew Excel files
            # regardless of row content. This fallback runs ONLY when a canonical
            # VAC Crew key is missing, so helper/primary mappings are unaffected
            # and existing exact-match behaviour is preserved.
            _vac_crew_fuzzy_canonicals = [
                'VAC Crew Helping?',
                'Vac Crew Completed Unit?',
                'VAC Crew Dept #',
                'Vac Crew Job #',
                'Vac Crew Email Address',
            ]
            _already_mapped_ids = set(mapping.values())
            for _canonical in _vac_crew_fuzzy_canonicals:
                if _canonical in mapping:
                    continue
                _target_norm = _normalize_column_title_for_vac_crew(_canonical)
                for c in cols:
                    if c.id in _already_mapped_ids:
                        continue
                    if _normalize_column_title_for_vac_crew(c.title) == _target_norm:
                        mapping[_canonical] = c.id
                        _already_mapped_ids.add(c.id)
                        logging.warning(
                            f"🚐 VAC Crew column FUZZY-MATCHED on sheet ID {sid}: "
                            f"'{c.title}' -> '{_canonical}'. Consider adding '{c.title}' "
                            f"as an explicit synonym if this variant is permanent."
                        )
                        break

            # Log summary of helper columns found
            if helper_columns_found:
                logging.info(f"🔧 All helper/helping columns found in sheet: {helper_columns_found}")
            # Log summary of VAC Crew columns found
            if vac_crew_columns_found:
                logging.info(f"🚐 VAC Crew columns found in sheet: {vac_crew_columns_found}")

            # Actionable WARNING: VAC Crew-looking columns exist but the two key
            # mappings still didn't resolve after the fuzzy pass — detection will
            # be DISABLED for this sheet until the column titles are aligned with
            # `_vac_crew_fuzzy_canonicals`. Surface the raw titles so operators
            # can see exactly which variant is on the sheet.
            if vac_crew_columns_found and not (
                'VAC Crew Helping?' in mapping
                and 'Vac Crew Completed Unit?' in mapping
            ):
                logging.warning(
                    f"🚐⚠️ VAC Crew columns visible on sheet ID {sid} but key "
                    f"mappings incomplete after fuzzy pass: "
                    f"titles_seen={vac_crew_columns_found}, "
                    f"mapped_vac_crew_keys={[k for k in _vac_crew_fuzzy_canonicals if k in mapping]}. "
                    f"VAC Crew row detection will be DISABLED for this sheet until "
                    f"titles match a canonical form in _vac_crew_fuzzy_canonicals."
                )


            if 'Weekly Reference Logged Date' in mapping:
                w_id = mapping['Weekly Reference Logged Date']
                s_id = mapping.get('Snapshot Date')
                w_samples = _extract_col_samples(w_id)
                s_samples = _extract_col_samples(s_id) if s_id else []
                logging.info(f"Sheet {sheet.name} (ID {sid}) date columns:")
                logging.info(f"  Weekly Reference Logged Date (ID {w_id}) samples: {w_samples}")
                if s_id:
                    logging.info(f"  Snapshot Date (ID {s_id}) samples: {s_samples}")
                logging.info(f"✅ Added sheet: {sheet.name} (ID: {sid})")
                return {'id': sid,'name': sheet.name,'column_mapping': mapping}
            else:
                logging.warning(f"❌ Skipping sheet {sheet.name} (ID {sid}) - Weekly Reference Logged Date not found (strict mode)")
                return None
        except Exception as e:
            logging.warning(f"⚡ Failed to validate sheet {sid}: {e}")
            return None

    # ── Incremental mode: only validate NEW sheet IDs, keep cached ones ──
    if _incremental:
        all_base_ids = set(base_sheet_ids)
        new_ids_to_validate = sorted(all_base_ids - _cached_sheet_ids)
        if new_ids_to_validate:
            logging.info(f"🆕 Incremental discovery: {len(new_ids_to_validate)} new sheet ID(s) to validate "
                         f"(skipping {len(_cached_sheet_ids)} already-cached sheets)")
            _discovery_start = datetime.datetime.now()
            with ThreadPoolExecutor(max_workers=PARALLEL_WORKERS_DISCOVERY) as executor:
                futures = {executor.submit(_validate_single_sheet, sid): sid for sid in new_ids_to_validate}
                for i, future in enumerate(as_completed(futures), 1):
                    sid = futures[future]
                    result = future.result()
                    if result is not None:
                        discovered.append(result)
                        logging.info(f"   ✅ [{i}/{len(futures)}] NEW Discovered: {result['name']} (ID: {sid})")
                    else:
                        logging.info(f"   ❌ [{i}/{len(futures)}] Skipped new sheet ID {sid}")
            _discovery_elapsed = (datetime.datetime.now() - _discovery_start).total_seconds()
            logging.info(f"⚡ Incremental discovery: {len(discovered)} new sheet(s) validated in {_discovery_elapsed:.1f}s")
        else:
            logging.info(f"⚡ Incremental discovery: no new sheet IDs found — all {len(_cached_sheet_ids)} sheets already cached")
        # Merge: cached sheets + newly discovered sheets
        discovered = _cached_sheets + discovered
        logging.info(f"📋 Total sheets after incremental merge: {len(discovered)} ({len(_cached_sheets)} cached + {len(discovered) - len(_cached_sheets)} new)")
    else:
        # Full discovery: validate all sheets from scratch
        logging.info(f"🚀 Starting parallel discovery with {PARALLEL_WORKERS_DISCOVERY} workers for {len(base_sheet_ids)} sheets...")
        _discovery_start = datetime.datetime.now()
        with ThreadPoolExecutor(max_workers=PARALLEL_WORKERS_DISCOVERY) as executor:
            futures = {executor.submit(_validate_single_sheet, sid): sid for sid in base_sheet_ids}
            for i, future in enumerate(as_completed(futures), 1):
                sid = futures[future]
                result = future.result()
                if result is not None:
                    discovered.append(result)
                    logging.info(f"   ✅ [{i}/{len(futures)}] Discovered: {result['name']} (ID: {sid})")
                else:
                    logging.info(f"   ❌ [{i}/{len(futures)}] Skipped sheet ID {sid}")
        _discovery_elapsed = (datetime.datetime.now() - _discovery_start).total_seconds()
        logging.info(f"⚡ Discovery complete: {len(discovered)} sheets validated in {_discovery_elapsed:.1f}s (parallel w/{PARALLEL_WORKERS_DISCOVERY} workers)")
    # Save cache
    if USE_DISCOVERY_CACHE:
        try:
            with open(DISCOVERY_CACHE_PATH,'w') as f:
                json.dump({
                    'schema_version': DISCOVERY_CACHE_VERSION,
                    'timestamp': datetime.datetime.now().isoformat(),
                    'sheets': discovered,
                    'subcontractor_sheet_ids': sorted(SUBCONTRACTOR_SHEET_IDS),
                }, f)
        except Exception as e:
            logging.warning(f"Failed to write discovery cache: {e}")
    return discovered

# Cell history and Modified By logic removed - using direct column assignment only

def get_all_source_rows(client, source_sheets):
    """Fetch rows from all source sheets with filtering.
    
    Implements direct column-based foreman assignment with helper row detection.
    
    Args:
        client: Smartsheet client instance
        source_sheets: List of source sheet configurations
    
    Foreman assignment logic:
    - Primary: Use "Foreman Assigned?" column if present
    - Fallback: Use "Foreman" column text
    - Final fallback: "Unknown Foreman"
    
    Helper row detection:
    - Identifies rows where "Foreman Helping?" has a value AND
      both "Helping Foreman Completed Unit?" and "Units Completed?" are checked
    - Helper rows are tagged with __is_helper_row=True and helper metadata
    - No matching, no cache lookup, completely unchanged behavior
    - Triggers when: no cache, no cache entry, no matches found

    Improvements:
      • Per‑cell debug logging limited by DEBUG_SAMPLE_ROWS (env tunable)
      • Essential field summary limited by DEBUG_ESSENTIAL_ROWS
      • Single concise summary for unmapped columns with a small sample of values
      • Greatly reduces 'Unknown' spam while preserving early transparency
    """
    merged_rows = []
    global_row_counter = 0
    original_rates = load_contract_rates(OLD_RATES_CSV)
    # Load new rate versions if rate cutoff is configured
    global _RATES_FINGERPRINT
    _rate_cu_to_group = {}
    _rate_new_primary = {}
    _rate_new_arrowhead = {}
    if RATE_CUTOFF_DATE:
        _rate_cu_to_group, _rate_new_primary, _rate_new_arrowhead, _RATES_FINGERPRINT = load_rate_versions()
    exclusion_counts = {
        'missing_work_request': 0,
        'missing_weekly_reference_logged_date': 0,
        'units_not_completed': 0,
        'price_missing_or_zero': 0,
        'cu_no_match': 0,
        'accepted': 0
    }
    # Detailed per‑WR diagnostics
    foreman_raw_counts = collections.defaultdict(lambda: collections.Counter())  # wr -> Counter(foreman values as-seen)
    wr_exclusion_reasons = collections.defaultdict(lambda: collections.Counter())  # wr -> Counter(reason)

    def _fetch_and_process_sheet(source):
        """Fetch and process a single source sheet. Returns (rows, sheet_exclusion_counts, sheet_foreman_counts, sheet_wr_exclusion_reasons, row_count)."""
        sheet_rows = []
        sheet_exclusion_counts = {
            'missing_work_request': 0,
            'missing_weekly_reference_logged_date': 0,
            'units_not_completed': 0,
            'price_missing_or_zero': 0,
            'cu_no_match': 0,
            'accepted': 0
        }
        sheet_foreman_counts = collections.defaultdict(lambda: collections.Counter())
        sheet_wr_exclusion_reasons = collections.defaultdict(lambda: collections.Counter())
        sheet_row_counter = 0
        # Track post-cutoff rate recalc outcomes for operator visibility
        # ('skipped' covers rows where Snapshot Date>=cutoff but the new
        # rates table has no matching group/CU, so the SmartSheet price
        # is retained — the known VAC-crew pricing-lag signal).
        # 'fallback_applied' counts rows routed through the
        # Weekly-Ref-Date fallback (blank Snapshot Date with
        # Weekly Reference Logged Date>=cutoff); it is path-tracking and
        # is independent of recalc outcome, so it is NOT mutually
        # exclusive with 'recalculated' / 'skipped'.
        sheet_rate_recalc_counts = {
            'recalculated': 0,
            'skipped': 0,
            'fallback_applied': 0,
        }
        sheet_rate_recalc_skipped_cus = collections.Counter()
        try:
            logging.info(f"⚡ Processing: {source['name']} (ID: {source['id']})")
            is_subcontractor_sheet = source['id'] in SUBCONTRACTOR_SHEET_IDS
            # Smartsheet-native pricing guard: sheets discovered via
            # ORIGINAL_CONTRACT_FOLDER_IDS already produce their
            # Units Total Price from Smartsheet's internal formula for
            # post-cutoff rows with Units Completed? = true. The
            # Python-side recalc must NOT overwrite that value.
            # See RATE_RECALC_SKIP_ORIGINAL_CONTRACT env var declaration
            # for the full rationale and the kill-switch.
            is_original_contract_sheet = source['id'] in _FOLDER_DISCOVERED_ORIG_IDS
            _skip_recalc_original_contract = (
                RATE_CUTOFF_DATE is not None
                and RATE_RECALC_SKIP_ORIGINAL_CONTRACT
                and is_original_contract_sheet
                and not is_subcontractor_sheet
            )
            if _skip_recalc_original_contract:
                logging.info(
                    f"🛡️ Skipping Python rate recalc for {source['name']} "
                    f"(ID: {source['id']}) — sheet is in ORIGINAL_CONTRACT_FOLDER_IDS "
                    f"and Smartsheet-native pricing is authoritative for post-cutoff rows"
                )

            try:
                # Fetch sheet once (no column history); include columns to support unmapped summary
                # PERFORMANCE FIX: Use column_ids parameter to only fetch mapped columns
                column_mapping = source['column_mapping']
                required_column_ids = list(column_mapping.values())
                with sentry_sdk.start_span(op="smartsheet.api", name=f"Fetch sheet {source['name']}") as api_span:
                    sheet = client.Sheets.get_sheet(
                        source['id'], 
                        column_ids=required_column_ids
                    )
                    api_span.set_data("sheet_id", source['id'])
                    api_span.set_data("sheet_name", source['name'])
                    api_span.set_data("row_count", len(sheet.rows) if sheet.rows else 0)
                    api_span.set_data("column_count", len(required_column_ids))

                logging.info(f"📋 Available mapped columns in {source['name']}: {list(column_mapping.keys())}")
                
                # PERFORMANCE: Pre-build reverse mapping for O(1) cell lookups (column_id -> field_name)
                reverse_column_map = {cid: name for name, cid in column_mapping.items()}
                
                # Debug: Check if Weekly Reference Logged Date is mapped
                if 'Weekly Reference Logged Date' in column_mapping:
                    logging.info(f"✅ Weekly Reference Logged Date column found with ID: {column_mapping['Weekly Reference Logged Date']}")
                else:
                    logging.warning(f"❌ Weekly Reference Logged Date column NOT found in mapping")
                    logging.info(f"   Available mappings: {column_mapping}")
                
                # HELPER DETECTION LOGGING: Check if helper columns are present
                helper_columns = ['Foreman Helping?', 'Helping Foreman Completed Unit?', 'Helper Dept #', 'Helper Job #']
                found_helper_cols = [col for col in helper_columns if col in column_mapping]
                if found_helper_cols:
                    logging.info(f"🔧 Helper columns found in {source['name']}: {found_helper_cols}")
                    if len(found_helper_cols) == 4:
                        logging.info(f"✅ All 4 helper columns present - helper detection will be active for this sheet")
                    else:
                        missing = [col for col in helper_columns if col not in column_mapping]
                        logging.warning(f"⚠️ Missing helper columns in {source['name']}: {missing}")
                else:
                    logging.info(f"ℹ️ No helper columns found in {source['name']} - helper detection disabled for this sheet")

                # VAC CREW DETECTION: Check if VAC Crew columns are present (row-level detection)
                vac_crew_columns = ['VAC Crew Helping?', 'Vac Crew Completed Unit?', 'VAC Crew Dept #', 'Vac Crew Job #']
                found_vac_crew_cols = [col for col in vac_crew_columns if col in column_mapping]
                # Sheet has VAC Crew capability if at least the two key columns are mapped
                sheet_has_vac_crew_columns = 'VAC Crew Helping?' in column_mapping and 'Vac Crew Completed Unit?' in column_mapping
                # Codex P1 guardrail: the Weekly-Ref-Date recalc
                # fallback is ONLY meaningful on sheets that actually
                # map a Snapshot Date column. When a (legacy) sheet
                # has ``Weekly Reference Logged Date`` but no
                # ``Snapshot Date`` mapping, ``row_data.get('Snapshot
                # Date')`` is None for every row and the fallback
                # would silently re-price the whole sheet by weekly
                # date — changing the cutoff basis rather than
                # rescuing current-week automation-lag rows. Gate the
                # fallback on the column's presence so legacy sheets
                # preserve exactly the pre-fix behaviour (no recalc
                # when Snapshot Date is absent).
                sheet_has_snapshot_date_column = 'Snapshot Date' in column_mapping
                if found_vac_crew_cols:
                    logging.info(f"🚐 VAC Crew columns found in {source['name']}: {found_vac_crew_cols}")
                    if sheet_has_vac_crew_columns:
                        logging.info(f"✅ VAC Crew detection active for this sheet (key columns present)")
                    else:
                        missing = [col for col in vac_crew_columns if col not in column_mapping]
                        logging.warning(f"⚠️ Missing VAC Crew columns in {source['name']}: {missing}")
                else:
                    logging.debug(f"ℹ️ No VAC Crew columns in {source['name']} - VAC Crew detection disabled for this sheet")

                # Note: Unmapped column logging skipped - we now only fetch mapped columns for performance
                # This reduces API payload size by ~64% and prevents Error 4000 for large sheets

                # Process all rows
                for row in sheet.rows:
                    row_data = {}

                    # Per‑cell debug logging only for the earliest rows overall
                    # PERFORMANCE: Use early continue to avoid logging overhead in production
                    _should_debug_cells = PER_CELL_DEBUG_ENABLED and sheet_row_counter < DEBUG_SAMPLE_ROWS
                    if _should_debug_cells:
                        logging.info(f"🔍 DEBUG: Processing row with {len(row.cells)} cells (sheet row #{sheet_row_counter+1})")
                        for cell in row.cells:
                            mapped_name = reverse_column_map.get(cell.column_id)
                            if mapped_name:
                                val = cell.display_value if cell.display_value is not None else cell.value
                                if val is not None:
                                    logging.info(f"   Cell {cell.column_id}: '{mapped_name}' = '{val}'")

                    # Build mapped row data using pre-built reverse mapping for O(1) lookup
                    for cell in row.cells:
                        mapped_name = reverse_column_map.get(cell.column_id)
                        if mapped_name:
                            raw_val = getattr(cell, 'value', None)
                            if raw_val is None:
                                raw_val = getattr(cell, 'display_value', None)
                            row_data[mapped_name] = raw_val

                    # Attach provenance metadata for audit (used to fetch selective cell history later)
                    if row_data:
                        row_data['__sheet_id'] = source['id']
                        # Phase 01 Plan 03 / Plan 09 (WR-06): also expose
                        # the sheet id under the canonical
                        # ``__source_sheet_id`` name. Phase 1's
                        # subcontractor variant gate in
                        # ``group_source_rows`` AND the missing-CU
                        # attribution loop in ``main()`` both read
                        # ``__source_sheet_id`` per WR-06. The legacy
                        # ``__sheet_id`` write is retained above for
                        # back-compat with any future reader that
                        # might still touch it — drop in a follow-up
                        # cleanup once a full pass confirms no other
                        # reader exists.
                        row_data['__source_sheet_id'] = source['id']
                        row_data['__row_id'] = row.id

                    # Essential field summary for earliest rows (gated to reduce I/O)
                    _should_log_essentials = sheet_row_counter < DEBUG_ESSENTIAL_ROWS
                    if _should_log_essentials:
                        essential_fields = [
                            'Weekly Reference Logged Date', 'Snapshot Date', 'Units Completed?',
                            'Units Total Price', 'Work Request #'
                        ]
                        debug_essentials = {f: row_data.get(f) for f in essential_fields}
                        logging.info(f"   ESSENTIAL FIELDS: {debug_essentials}")

                    # Process row if it has any mapped data
                    if row_data:
                        work_request = row_data.get('Work Request #')
                        weekly_date = row_data.get('Weekly Reference Logged Date')
                        price_raw = row_data.get('Units Total Price')
                        price_val = parse_price(price_raw)

                        # --- SUBCONTRACTOR PRICING ---
                        # Subcontractor (Arrowhead) sheets always keep their SmartSheet
                        # pricing as-is. Rate recalculation only applies to primary
                        # (non-subcontractor) sheets. Subcontractor new rates will be
                        # enabled separately when a subcontractor cutoff date is provided.
                        # --- END SUBCONTRACTOR PRICING ---

                        # Pre-acceptance rate recalculation: for cutoff-eligible rows,
                        # recalculate price BEFORE the has_price check so rows with
                        # zero/blank SmartSheet prices can still be accepted if the
                        # new rate produces a valid non-zero amount.
                        # NOTE: Subcontractor (Arrowhead) sheets are EXCLUDED from
                        # recalculation until a separate subcontractor cutoff date is
                        # provided. They keep SmartSheet pricing as-is for now.
                        _rate_recalc_ran_for_row = False
                        _recalc_outcome = None
                        _recalc_via_fallback = False
                        # ``_skip_recalc_original_contract`` is sheet-level
                        # (computed once above, logged once per sheet) —
                        # adding it to this row-level gate avoids per-row
                        # log spam while still short-circuiting every row
                        # on an original-contract folder sheet. Subcontractor
                        # exclusion stays primary; the original-contract
                        # skip only kicks in when Smartsheet-native pricing
                        # is authoritative for the whole sheet.
                        if (
                            RATE_CUTOFF_DATE
                            and _rate_new_primary
                            and not is_subcontractor_sheet
                            and not _skip_recalc_original_contract
                        ):
                            # Primary gate is Snapshot Date; the helper
                            # transparently falls back to Weekly
                            # Reference Logged Date when Snapshot Date
                            # is blank AND RATE_RECALC_WEEKLY_FALLBACK
                            # is enabled. This rescues current-week
                            # rows (VAC crew / helper) that would
                            # otherwise silently drop at the has_price
                            # gate with zero price.
                            effective_cutoff_date, _recalc_via_fallback = (
                                _resolve_rate_recalc_cutoff_date(
                                    row_data,
                                    RATE_CUTOFF_DATE,
                                    # See ``sheet_has_snapshot_date_column``
                                    # — disable the fallback on sheets
                                    # that never map Snapshot Date so
                                    # we don't re-price whole legacy
                                    # sheets by weekly date.
                                    weekly_fallback_enabled=(
                                        RATE_RECALC_WEEKLY_FALLBACK
                                        and sheet_has_snapshot_date_column
                                    ),
                                )
                            )

                            if effective_cutoff_date is not None:
                                old_price = price_val
                                _recalc_status = {}
                                price_val = recalculate_row_price(
                                    row_data,
                                    _rate_cu_to_group,
                                    _rate_new_primary,
                                    out_status=_recalc_status,
                                )
                                _rate_recalc_ran_for_row = True
                                _recalc_outcome = _recalc_status.get('outcome')
                                if _recalc_via_fallback:
                                    sheet_rate_recalc_counts['fallback_applied'] += 1
                                if _recalc_outcome == 'recalculated':
                                    sheet_rate_recalc_counts['recalculated'] += 1
                                    if price_val != old_price:
                                        _via = ' via Weekly-Ref-Date fallback' if _recalc_via_fallback else ''
                                        logging.debug(
                                            f"Rate recalc{_via}: CU={row_data.get('CU')}, "
                                            f"old=${old_price:.2f} -> new=${price_val:.2f}, "
                                            f"effective_cutoff_date={effective_cutoff_date}"
                                        )
                                elif _recalc_outcome == 'missing_rate':
                                    # Only count as "skipped" in the
                                    # per-sheet summary when recalc
                                    # explicitly reported that neither
                                    # the mapped group nor the CU code
                                    # is in the new rates table — that
                                    # is the actionable signal for
                                    # updating NEW_RATES_CSV. Outcomes
                                    # like 'invalid_quantity' /
                                    # 'zero_rate' are data-entry or
                                    # contract gaps and are intentionally
                                    # excluded so the summary WARNING
                                    # and top-CU list stay accurate.
                                    sheet_rate_recalc_counts['skipped'] += 1
                                    # Always attribute the skip to a
                                    # CU bucket so the per-sheet
                                    # "N skipped / Top CUs: ..." summary
                                    # totals stay aligned. Blank CU rows
                                    # are attributed to '<blank>' so
                                    # operators can see that category
                                    # and investigate the missing CU
                                    # code separately.
                                    cu_val = _resolve_cu_code(row_data) or '<blank>'
                                    sheet_rate_recalc_skipped_cus[cu_val] += 1

                        # Phase 1.1 Bug A (D-01..D-03 / SUB-08):
                        # pre-acceptance rate-recalc rescue for
                        # subcontractor sheets. Mirrors the
                        # [2026-04-23 00:00] VAC-crew Weekly-Ref-Date
                        # fallback pattern (additive branch alongside
                        # the existing primary-rate gate above, NOT a
                        # modification of it). Subcontractor operators
                        # populate helper-foreman events BEFORE pricing
                        # is finalized; SmartSheet ``Units Total Price``
                        # is commonly blank/zero on those rows. Without
                        # this rescue, the row drops at the has_price
                        # gate below and helper-detection never fires.
                        # Mutates row_data['Units Total Price'] in
                        # addition to price_val so the has_price gate's
                        # ``price_raw not in (None, "", "$0", ...)``
                        # clause also passes — same in-place pattern
                        # as ``recalculate_row_price`` at L1751.
                        if (
                            is_subcontractor_sheet
                            and SUBCONTRACTOR_RATE_RECALC_PREACCEPTANCE_ENABLED
                            and price_val <= 0
                        ):
                            _rescued = _subcontractor_rescue_price(row_data)
                            if _rescued > 0:
                                price_val = _rescued
                                row_data['Units Total Price'] = _rescued
                                # Telemetry hook for downstream Sentry
                                # breadcrumbs + the e2e regression
                                # test in Plan 01.1-05.
                                row_data['__subcontractor_rescued'] = True
                                if FILTER_DIAGNOSTICS and sheet_row_counter < DEBUG_ESSENTIAL_ROWS:
                                    # PII marker "Subcontractor pre-
                                    # acceptance rescue" added to
                                    # _PII_LOG_MARKERS in Task 1. Log
                                    # body embeds WR + CU. ``wr_key_for_diag``
                                    # is not yet initialized in this
                                    # scope (it lands at L3723 below),
                                    # so derive it here from the row's
                                    # raw Work Request # via the same
                                    # ``str(work_request).split('.')[0]``
                                    # pattern.
                                    _wr_diag = (
                                        str(work_request).split('.')[0]
                                        if work_request else '<unknown>'
                                    )
                                    logging.info(
                                        f"💲 Subcontractor pre-acceptance rescue: "
                                        f"WR={_wr_diag}, "
                                        f"CU={row_data.get('CU')}, "
                                        f"rescued=${_rescued:.2f}"
                                    )

                        price_raw = row_data.get('Units Total Price')
                        has_price = (price_raw not in (None, "", "$0", "$0.00", "0", "0.0")) and price_val > 0
                        units_completed = row_data.get('Units Completed?')
                        units_completed_checked = is_checked(units_completed)

                        if sheet_row_counter < DEBUG_ESSENTIAL_ROWS:
                            logging.info(f"🔍 Row data sample: WR={work_request}, Price={price_val}, Date={weekly_date}, Units Completed={units_completed} ({units_completed_checked})")

                        # Record raw foreman regardless of acceptance (if WR exists)
                        wr_key_for_diag = None
                        if work_request:
                            wr_key_for_diag = str(work_request).split('.')[0]
                            fr_val = (row_data.get('Foreman') or '').strip() or '<<blank>>'
                            sheet_foreman_counts[wr_key_for_diag][fr_val] += 1

                        # Acceptance logic (STRICT: Units Completed? must be checked/true)
                        if work_request and weekly_date and units_completed_checked and has_price:
                            # CU no-match exclusion: drop backend placeholder rows like "#NO MATCH..."
                            cu_raw = (row_data.get('CU') or row_data.get('Billable Unit Code') or '')
                            cu_text = str(cu_raw).strip().upper()
                            # Exclude any backend placeholder variants like '#NO MATCH' or 'NO MATCH'
                            if 'NO MATCH' in cu_text:
                                sheet_exclusion_counts['cu_no_match'] += 1
                                if wr_key_for_diag:
                                    sheet_wr_exclusion_reasons[wr_key_for_diag]['cu_no_match'] += 1
                                if FILTER_DIAGNOSTICS and sheet_row_counter < DEBUG_ESSENTIAL_ROWS:
                                    logging.info(f"🚫 Excluding row for WR {wr_key_for_diag} due to CU 'NO MATCH' placeholder: raw='{cu_raw}'")
                                # Skip appending this row
                                sheet_row_counter += 1
                                continue
                            # Helper row detection (before foreman assignment)
                            # Helper criteria: Foreman Helping? non-blank AND both checkboxes checked
                            # Handle None values safely with defensive str() conversion to prevent float/strip errors
                            foreman_helping_val = row_data.get('Foreman Helping?')
                            helper_name = str(foreman_helping_val).strip() if foreman_helping_val else ''
                            helping_foreman_completed = row_data.get('Helping Foreman Completed Unit?')
                            helping_foreman_completed_checked = is_checked(helping_foreman_completed)
                            
                            is_helper_row = bool(helper_name and helping_foreman_completed_checked and units_completed_checked)
                            
                            # HELPER DETECTION LOGGING: Log criteria evaluation for sample rows (gated behind FILTER_DIAGNOSTICS)
                            if FILTER_DIAGNOSTICS and sheet_row_counter < DEBUG_ESSENTIAL_ROWS:
                                logging.info(f"🔧 Helper detection criteria for row {sheet_row_counter+1}:")
                                logging.info(f"   Foreman Helping?: '{foreman_helping_val}' -> helper_name='{helper_name}'")
                                logging.info(f"   Helping Foreman Completed Unit?: {helping_foreman_completed} -> checked={helping_foreman_completed_checked}")
                                logging.info(f"   Units Completed?: {units_completed} -> checked={units_completed_checked}")
                                logging.info(f"   is_helper_row: {is_helper_row}")
                            
                            if is_helper_row:
                                # Populate helper metadata (with safe None handling and defensive str() conversion)
                                row_data['__is_helper_row'] = True
                                row_data['__helper_foreman'] = helper_name
                                helper_dept_val = row_data.get('Helper Dept #')
                                helper_job_val = row_data.get('Helper Job #')
                                row_data['__helper_dept'] = str(helper_dept_val).strip() if helper_dept_val else ''
                                row_data['__helper_job'] = str(helper_job_val).strip() if helper_job_val else ''
                                
                                # HELPER DETECTION LOGGING: Log first 10 helper rows per sheet for transparency (gated behind FILTER_DIAGNOSTICS)
                                if FILTER_DIAGNOSTICS and sheet_row_counter < 10:
                                    logging.info(f"🔧 HELPER ROW DETECTED [Row {sheet_row_counter+1}]: WR={wr_key_for_diag}, Helper={helper_name}, Dept={row_data['__helper_dept']}, Job={row_data['__helper_job']}")
                            else:
                                row_data['__is_helper_row'] = False
                            
                            # Direct column-based foreman assignment
                            effective_user = None
                            assignment_method = None
                            # Use Foreman Assigned? column, fallback to Foreman column
                            foreman_assigned = row_data.get('Foreman Assigned?')
                            if foreman_assigned:
                                # Use the value directly (could be email, name, or text)
                                effective_user = str(foreman_assigned).strip()
                                assignment_method = 'FOREMAN_ASSIGNED'
                            else:
                                # Fallback to primary Foreman text if available (defensive str() conversion to prevent float errors)
                                foreman_val = row_data.get('Foreman')
                                primary_foreman_text = str(foreman_val).strip() if foreman_val else ''
                                if primary_foreman_text:
                                    effective_user = primary_foreman_text
                                    assignment_method = 'FOREMAN_COLUMN'
                                else:
                                    effective_user = 'Unknown Foreman'
                                    assignment_method = 'NO_FOREMAN'
                            
                            if sheet_row_counter < DEBUG_ESSENTIAL_ROWS:
                                logging.info(f"📋 Foreman Assignment: Using '{effective_user}' ({assignment_method})")
                            
                            # Store effective user and method for grouping
                            row_data['__effective_user'] = effective_user
                            row_data['__assignment_method'] = assignment_method
                            
                            # VAC Crew row-level detection (mirrors helper pattern)
                            # A row is VAC Crew when: VAC Crew Helping? is non-blank AND
                            # Vac Crew Completed Unit? checkbox is checked.
                            # This is column-presence-driven — only sheets with these columns
                            # can produce VAC Crew rows, no sheet-level ID tagging needed.
                            is_vac_crew_row = False
                            if sheet_has_vac_crew_columns:
                                vac_crew_helping_val = row_data.get('VAC Crew Helping?')
                                vac_crew_name = str(vac_crew_helping_val).strip() if vac_crew_helping_val else ''
                                vac_crew_completed = row_data.get('Vac Crew Completed Unit?')
                                vac_crew_completed_checked = is_checked(vac_crew_completed)
                                is_vac_crew_row = bool(vac_crew_name and vac_crew_completed_checked and units_completed_checked)
                                
                                if FILTER_DIAGNOSTICS and sheet_row_counter < DEBUG_ESSENTIAL_ROWS:
                                    logging.info(f"🚐 VAC Crew detection for row {sheet_row_counter+1}:")
                                    logging.info(f"   VAC Crew Helping?: '{vac_crew_helping_val}' -> name='{vac_crew_name}'")
                                    logging.info(f"   Vac Crew Completed Unit?: {vac_crew_completed} -> checked={vac_crew_completed_checked}")
                                    logging.info(f"   is_vac_crew_row: {is_vac_crew_row}")
                                
                                if is_vac_crew_row:
                                    vac_crew_dept_val = row_data.get('VAC Crew Dept #')
                                    vac_crew_job_val = row_data.get('Vac Crew Job #')
                                    row_data['__vac_crew_name'] = vac_crew_name
                                    row_data['__vac_crew_dept'] = str(vac_crew_dept_val).strip() if vac_crew_dept_val else ''
                                    row_data['__vac_crew_job'] = str(vac_crew_job_val).strip() if vac_crew_job_val else ''
                                    vac_crew_email_val = row_data.get('Vac Crew Email Address')
                                    row_data['__vac_crew_email'] = str(vac_crew_email_val).strip() if vac_crew_email_val else ''
                                    if FILTER_DIAGNOSTICS and sheet_row_counter < 10:
                                        logging.info(f"🚐 VAC CREW ROW DETECTED [Row {sheet_row_counter+1}]: WR={wr_key_for_diag}, Name={vac_crew_name}, Dept={row_data['__vac_crew_dept']}, Job={row_data['__vac_crew_job']}")
                            
                            row_data['__is_vac_crew'] = is_vac_crew_row
                            row_data['__is_subcontractor'] = is_subcontractor_sheet

                            sheet_rows.append(row_data)
                            sheet_exclusion_counts['accepted'] += 1
                        else:
                            # Increment specific exclusion reasons (first matching reason recorded)
                            if not work_request:
                                sheet_exclusion_counts['missing_work_request'] += 1
                                if wr_key_for_diag:
                                    sheet_wr_exclusion_reasons[wr_key_for_diag]['missing_work_request'] += 1
                            elif not weekly_date:
                                sheet_exclusion_counts['missing_weekly_reference_logged_date'] += 1
                                if wr_key_for_diag:
                                    sheet_wr_exclusion_reasons[wr_key_for_diag]['missing_weekly_reference_logged_date'] += 1
                            elif not units_completed_checked:
                                sheet_exclusion_counts['units_not_completed'] += 1
                                if wr_key_for_diag:
                                    sheet_wr_exclusion_reasons[wr_key_for_diag]['units_not_completed'] += 1
                            elif not has_price:
                                sheet_exclusion_counts['price_missing_or_zero'] += 1
                                if wr_key_for_diag:
                                    sheet_wr_exclusion_reasons[wr_key_for_diag]['price_missing_or_zero'] += 1
                                # Row-level visibility: surface drops that
                                # otherwise look "correct" to operators —
                                # specifically VAC crew / helper rows whose
                                # only missing piece is a zero or blank
                                # SmartSheet price. These previously
                                # disappeared into the per-sheet counter
                                # with no per-row log, which is why VAC
                                # crew / helping-foreman Excel files could
                                # silently fail to generate even after
                                # RESET_HASH_HISTORY.
                                _vc_helping = str(row_data.get('VAC Crew Helping?') or '').strip()
                                _vc_completed = is_checked(row_data.get('Vac Crew Completed Unit?'))
                                _fh_helping = str(row_data.get('Foreman Helping?') or '').strip()
                                _fh_completed = is_checked(row_data.get('Helping Foreman Completed Unit?'))
                                _is_specialized = (
                                    (bool(_vc_helping) and _vc_completed)
                                    or (bool(_fh_helping) and _fh_completed)
                                )
                                if _is_specialized:
                                    _variant_tag = 'VAC crew' if (_vc_helping and _vc_completed) else 'helper'
                                    # Only point operators at the per-sheet
                                    # "Rate recalc summary" WARNING when that
                                    # summary will actually contain this row:
                                    # the summary is emitted only when at
                                    # least one row in the sheet has outcome
                                    # 'missing_rate', so the note is valid
                                    # exactly when this row's outcome is
                                    # 'missing_rate'. For other outcomes
                                    # ('invalid_quantity', 'zero_rate', or
                                    # recalc-bypassed rows where
                                    # RATE_CUTOFF_DATE is unset, pre-cutoff,
                                    # Snapshot Date blank, or subcontractor
                                    # sheet), skip the breadcrumb so we
                                    # don't send operators hunting for a
                                    # summary line that isn't there.
                                    if _rate_recalc_ran_for_row and _recalc_outcome == 'missing_rate':
                                        _via_txt = (
                                            ' via Weekly-Ref-Date fallback'
                                            if _recalc_via_fallback else ''
                                        )
                                        _recalc_note = (
                                            f" Rate recalc ran{_via_txt} and reported no matching new-contract rate for this CU; "
                                            "see 'Rate recalc summary' WARNING on this sheet for the full CU list."
                                        )
                                    elif (
                                        not _rate_recalc_ran_for_row
                                        and RATE_CUTOFF_DATE
                                        and not RATE_RECALC_WEEKLY_FALLBACK
                                        # The Weekly-Ref-Date-fallback note
                                        # is only valid when recalc was
                                        # genuinely eligible to run on this
                                        # sheet. Original-contract folder
                                        # sheets skip recalc by design
                                        # (Smartsheet-native pricing is
                                        # authoritative), so enabling the
                                        # fallback env var would not change
                                        # anything on those sheets.
                                        and not _skip_recalc_original_contract
                                        # Use the same parser the recalc
                                        # gate uses so an unparseable
                                        # Snapshot Date (treated as blank
                                        # by _resolve_rate_recalc_cutoff_date)
                                        # triggers this note too — raw
                                        # truthiness would miss it and
                                        # leave operators chasing a
                                        # missing-rate explanation that
                                        # doesn't apply.
                                        and excel_serial_to_date(
                                            row_data.get('Snapshot Date')
                                        ) is None
                                        # Only advise enabling the
                                        # fallback when doing so would
                                        # actually rescue the row —
                                        # i.e. the Weekly Reference
                                        # Logged Date parses AND is
                                        # >= RATE_CUTOFF_DATE. For rows
                                        # whose weekly date is blank /
                                        # unparseable / pre-cutoff,
                                        # enabling the env var wouldn't
                                        # change anything, and the
                                        # message would send operators
                                        # on a false lead.
                                        and _weekly_would_trigger_fallback(
                                            row_data.get('Weekly Reference Logged Date'),
                                            RATE_CUTOFF_DATE,
                                        )
                                    ):
                                        # Snapshot Date is blank or
                                        # unparseable, Weekly Reference
                                        # Logged Date IS post-cutoff,
                                        # and the fallback is disabled.
                                        # Enabling RATE_RECALC_WEEKLY_FALLBACK
                                        # would genuinely rescue this
                                        # row — tell operators so they
                                        # don't hunt the CU in
                                        # NEW_RATES_CSV instead.
                                        _recalc_note = (
                                            " Rate recalc skipped because Snapshot Date is blank or unparseable "
                                            "and RATE_RECALC_WEEKLY_FALLBACK is disabled; Weekly Reference Logged "
                                            "Date is >= RATE_CUTOFF_DATE so setting the env var to '1' (default) "
                                            "would let this row get priced from the new-contract rates table."
                                        )
                                    else:
                                        _recalc_note = ""
                                    logging.warning(
                                        f"⚠️ Dropped {_variant_tag} row (price missing or zero): "
                                        f"WR={wr_key_for_diag}, Weekly={weekly_date}, "
                                        f"Snapshot={row_data.get('Snapshot Date') or '<blank>'}, "
                                        f"CU={_resolve_cu_code(row_data) or '<blank>'}, "
                                        f"Qty={row_data.get('Quantity') or '<blank>'}, "
                                        f"SmartSheet price={row_data.get('Units Total Price') or '<blank>'}. "
                                        f"Row has VAC/helper criteria checked but Units Total Price is zero/blank."
                                        f"{_recalc_note}"
                                    )

                    sheet_row_counter += 1

                sentry_add_breadcrumb("sheet_processing", f"Processed sheet {source['name']}", data={
                    "sheet_id": source['id'],
                    "rows_in_sheet": len(sheet.rows) if sheet.rows else 0,
                    "accepted_so_far": sheet_exclusion_counts['accepted'],
                    "is_subcontractor": is_subcontractor_sheet,
                })

                # Per-sheet summary of post-cutoff rate-recalc outcomes.
                # A non-zero 'skipped' count means rows qualified by
                # Snapshot Date but the new rates table could not price
                # them, so they kept their SmartSheet price. This is the
                # signal operators need to investigate missing entries
                # in NEW_RATES_CSV (common on VAC crew specialized work
                # like vacuum switches, softswitches, switched banks).
                # Suppress the summary entirely on sheets where the
                # original-contract skip fired — every counter is zero by
                # construction, so the summary would be noise. The
                # single "🛡️ Skipping Python rate recalc…" info log
                # emitted at the start of _fetch_and_process_sheet is
                # the authoritative per-sheet signal here.
                if (
                    RATE_CUTOFF_DATE
                    and _rate_new_primary
                    and not _skip_recalc_original_contract
                ):
                    skipped = sheet_rate_recalc_counts['skipped']
                    recalculated = sheet_rate_recalc_counts['recalculated']
                    fallback_applied = sheet_rate_recalc_counts['fallback_applied']
                    _fallback_suffix = (
                        f" ({fallback_applied} via Weekly-Ref-Date fallback)"
                        if fallback_applied else ""
                    )
                    if skipped:
                        top_cus = ', '.join(f"{cu}×{cnt}" for cu, cnt in sheet_rate_recalc_skipped_cus.most_common(10))
                        logging.warning(
                            f"⚠️ Rate recalc summary for {source['name']}: "
                            f"{recalculated} recalculated, {skipped} skipped{_fallback_suffix} "
                            f"(post-cutoff rows that kept SmartSheet price because no matching "
                            f"new-contract rate was found). Top CUs: {top_cus}"
                        )
                    elif recalculated:
                        logging.info(
                            f"📊 Rate recalc summary for {source['name']}: "
                            f"{recalculated} rows recalculated, 0 skipped{_fallback_suffix}"
                        )
                    elif fallback_applied:
                        # Fallback ran but every row hit a non-reportable
                        # outcome (invalid_quantity / zero_rate / etc.).
                        # Without this branch the new fallback_applied
                        # counter would be completely invisible in the
                        # logs for those runs — operators would have no
                        # visibility into whether the Weekly-Ref-Date
                        # fallback ever fired.
                        logging.info(
                            f"📊 Rate recalc summary for {source['name']}: "
                            f"0 recalculated, 0 skipped{_fallback_suffix}"
                        )

            except Exception as e:
                logging.error(f"Error processing sheet {source['id']}: {e}")
                sentry_capture_with_context(
                    exception=e,
                    context_name="sheet_processing_error",
                    context_data={
                        "sheet_id": source['id'],
                        "sheet_name": source.get('name', 'Unknown'),
                        "rows_processed": sheet_row_counter,
                        "error_type": type(e).__name__,
                        "error_message": _redact_exception_message(e),
                    },
                    tags={"error_location": "sheet_row_processing", "sheet_id": source['id']},
                    fingerprint=["sheet-processing", str(source['id']), type(e).__name__]
                )
            
        except Exception as e:
            logging.error(f"Could not process Sheet ID {source.get('id', 'N/A')}: {e}")
            sentry_capture_with_context(
                exception=e,
                context_name="sheet_access_error",
                context_data={
                    "sheet_id": source.get('id', 'N/A'),
                    "sheet_name": source.get('name', 'Unknown'),
                    "error_type": type(e).__name__,
                    "error_message": _redact_exception_message(e),
                },
                tags={"error_location": "sheet_access", "sheet_id": str(source.get('id', 'N/A'))},
                fingerprint=["sheet-access", str(source.get('id', 'N/A')), type(e).__name__]
            )

        return (sheet_rows, sheet_exclusion_counts, sheet_foreman_counts, sheet_wr_exclusion_reasons, sheet_row_counter)

    # Parallel sheet fetching: submit all sources to ThreadPoolExecutor
    logging.info(f"🚀 Starting parallel data fetch with {PARALLEL_WORKERS} workers for {len(source_sheets)} sheets...")
    _fetch_start = datetime.datetime.now()
    with ThreadPoolExecutor(max_workers=PARALLEL_WORKERS) as executor:
        futures = {executor.submit(_fetch_and_process_sheet, source): source for source in source_sheets}
        for i, future in enumerate(as_completed(futures), 1):
            source = futures[future]
            try:
                sheet_rows, sheet_exc, sheet_fc, sheet_wr_exc, sheet_rc = future.result()
                # Merge rows
                merged_rows.extend(sheet_rows)
                # Merge exclusion counts
                for k in exclusion_counts:
                    exclusion_counts[k] += sheet_exc[k]
                # Merge foreman raw counts
                for wr_key, ctr in sheet_fc.items():
                    foreman_raw_counts[wr_key] += ctr
                # Merge WR exclusion reasons
                for wr_key, ctr in sheet_wr_exc.items():
                    wr_exclusion_reasons[wr_key] += ctr
                global_row_counter += sheet_rc
                logging.info(f"   📋 [{i}/{len(futures)}] Fetched {sheet_exc['accepted']} rows from {source.get('name', 'unknown')} ({sheet_rc} total processed)")
            except Exception as e:
                logging.error(f"   ⚠️ [{i}/{len(futures)}] Sheet worker failed for {source.get('name', 'unknown')}: {e}")
    _fetch_elapsed = (datetime.datetime.now() - _fetch_start).total_seconds()
    logging.info(f"⚡ Data fetch complete: {len(merged_rows)} valid rows in {_fetch_elapsed:.1f}s (parallel w/{PARALLEL_WORKERS} workers)")
    
    # HELPER DETECTION SUMMARY LOGGING
    helper_row_count = sum(1 for r in merged_rows if r.get('__is_helper_row', False))
    if helper_row_count > 0:
        logging.info(f"🔧 HELPER DETECTION SUMMARY: {helper_row_count} helper rows detected out of {len(merged_rows)} total valid rows ({helper_row_count/len(merged_rows)*100:.1f}%)")
        # Log sample helper rows for verification
        sample_helpers = [r for r in merged_rows if r.get('__is_helper_row', False)][:5]
        for idx, helper in enumerate(sample_helpers, 1):
            logging.info(f"   Sample Helper {idx}: WR={helper.get('Work Request #')}, Helper={helper.get('__helper_foreman')}, Dept={helper.get('__helper_dept')}, Job={helper.get('__helper_job')}")
    else:
        logging.warning(f"⚠️ HELPER DETECTION SUMMARY: No helper rows detected in {len(merged_rows)} valid rows - check if helper columns exist and criteria are met")
    
    if FILTER_DIAGNOSTICS:
        total_excluded = sum(v for k,v in exclusion_counts.items() if k != 'accepted')
        logging.info("🧪 FILTER DIAGNOSTICS:")
        for k,v in exclusion_counts.items():
            logging.info(f"   {k}: {v}")
        logging.info(f"   total_excluded: {total_excluded}")
    if FOREMAN_DIAGNOSTICS and foreman_raw_counts:
        logging.info("🧪 FOREMAN DIAGNOSTICS (first 25 WRs):")
        for wr_key, ctr in list(foreman_raw_counts.items())[:25]:
            top = ctr.most_common(5)
            excl = wr_exclusion_reasons.get(wr_key, {})
            logging.info(f"   WR {wr_key}: {sum(ctr.values())} rows seen, foremen(top5)={top}; exclusions={dict(excl)}")

    if merged_rows:
        logging.info(f"✅ UPDATED FILTERING SUCCESS: Found {len(merged_rows)} rows (Work Request # + Weekly Reference Logged Date + Units Completed? + Units Total Price exists required)")
        logging.info("🎯 Change detection ACTIVE: Existing attachment with matching data hash will skip regeneration & upload")
    else:
        logging.warning("⚠️ No valid rows found with updated filtering (missing Work Request #, Weekly Reference Logged Date, Units Completed?, or Units Total Price)")

    return merged_rows

def group_source_rows(rows):
    """
    VARIANT-AWARE GROUPING: Groups rows by Work Request #, Week Ending Date, and Variant (primary/helper/vac_crew).
    
    Primary Variant:
    - Standard WR-based grouping (one Excel per WR/Week)
    - Key format: MMDDYY_WRNUMBER
    
    Helper Variant:
    - Helper-based grouping (one Excel per WR/Week/Helper)
    - Key format: MMDDYY_WRNUMBER_HELPER_<sanitized_helper_name>
    - Only created for rows where __is_helper_row = True
    
    VAC Crew Variant (Sub-project C — per-claimer partitioning):
    - Only created for rows where __is_vac_crew = True (row-level column-based detection)
    - Two key shapes, gated on VAC_CREW_CLAIM_ATTRIBUTION_ENABLED:
      * ON (default): partitioned by FROZEN vac-crew claimer →
        MMDDYY_WRNUMBER_VACCREW_<sanitized_claimer>, one Excel per
        WR/Week/claimer, filename suffix _VacCrew_<claimer>. Claimer is
        resolved in the pre-pass (frozen_vac_crew; falls back to the current
        vac-crew name on no_history; HOLD defers the row on a Supabase
        fetch_failure).
      * OFF (exact legacy): single group per WR/Week →
        MMDDYY_WRNUMBER_VACCREW, filename suffix _VacCrew (no claimer).
    - VAC crew rows never also emit the subcontractor primary variants
      (the subcontractor block is gated on `not is_vac_crew_row`).
    
    Activity Log (DECOMMISSIONED - only in primary mode):
    - No longer uses Modified By cache - direct column assignment only
    - Appends user identifier: MMDDYY_WRNUMBER_USER_<sanitized_user>
    
    RES_GROUPING_MODE controls primary/helper variants only (not vac_crew):
    - "primary": Only primary variant (may include user if activity log enabled)
    - "helper": Helper variant for helper rows + primary variant for non-helper rows (conditional filter)
    - "both": Both primary and helper variants for all applicable rows
    
    CRITICAL BUSINESS LOGIC: Groups valid rows by Week Ending Date AND Work Request #.
    Each group will create ONE Excel file containing ONE work request for ONE week ending date.
    
    FILENAME FORMAT: 
    - Primary: WR_{work_request_number}_WeekEnding_{MMDDYY}_{hash}.xlsx
    - Primary+User: WR_{work_request_number}_WeekEnding_{MMDDYY}_User_{user_sanitized}_{hash}.xlsx
    - Helper: WR_{work_request_number}_WeekEnding_{MMDDYY}_Helper_{helper_sanitized}_{hash}.xlsx
    - VAC Crew: WR_{work_request_number}_WeekEnding_{MMDDYY}_VacCrew_{hash}.xlsx
    
    This ensures:
    - Each Excel file contains ONLY one work request
    - Each work request can have multiple Excel files (one per week ending date and/or variant)
    - No mixing of work requests or variants in a single file
    - Clear, predictable file naming with variant identification
    """
    groups = collections.defaultdict(list)

    # Phase 1.1 Bug C (D-12 / SUB-11): per-WR dedupe set for the
    # fall-back WARNING. Keyed on (wr_key, week_end_for_key,
    # sanitized_helper_foreman) so we log ONE WARNING per unique
    # attribution-read failure context per run (not per-row). A
    # 100-row WR with the same helper falling back would otherwise
    # log 100 identical WARNINGs — exactly the kind of operator-log
    # spam the [2026-04-25 12:00] / [2026-04-24 10:50] ledger rules
    # call out as a P0 noise hazard.
    _bug_c_warning_seen: set[tuple[str, str, str]] = set()

    # ── Phase 2 Plan 02: Single bulk attribution prefetch (D-02) ──
    # Replace four separate per-row lookup_attribution RPC pre-passes
    # (B/C/D/Phase-1.1-sub-helper, ~137k calls/run on full history) with
    # ONE prefetch_attribution call that fetches ALL (wr, week_ending)
    # pairs in the current row set in bulk, building a shared
    #   {(wr, week_ending, row_id) -> roles_dict}
    # map for O(1) per-row resolution (D-03: map-aware resolve_claimer).
    #
    # D-04 direct-HOLD contract: on fetch_failure, B and C construct
    # ResolveOutcome('hold', None, None, 'fetch_failure') DIRECTLY from the
    # status — zero additional Supabase calls. D uses-current (never HOLDs)
    # per operator decision (core primary path prioritizes availability).
    #
    # D-05 scope removal: ATTRIBUTION_RESOLUTION_WEEKS and its scope gate
    # have been deleted. The bulk prefetch covers the EXACT (wr, week_ending)
    # pairs in the current run — no recency gate needed. Historical rows
    # with valid frozen claimers are now correctly resolved regardless of age
    # (fixes incident run 26439205107: 372 garbage _User__NO_MATCH files).
    _attr_map: dict = {}
    _attr_status: str = 'disabled'
    if BILLING_AUDIT_AVAILABLE and (
        SUBCONTRACTOR_RATE_VARIANTS_ENABLED
        or VAC_CREW_CLAIM_ATTRIBUTION_ENABLED
        or PRIMARY_CLAIM_ATTRIBUTION_ENABLED
        or SUBCONTRACTOR_HELPER_CLAIM_ATTRIBUTION_ENABLED
    ):
        # Build (wr, week_ending, row_id) pairs from all completed rows
        # that any of the four attribution consumers will process.
        _prefetch_pairs: set[tuple[str, datetime.date, int]] = set()
        for _r in rows:
            _rid = _r.get('__row_id')
            if not isinstance(_rid, int):
                continue
            if not is_checked(_r.get('Units Completed?')):
                continue
            _wr_raw = _r.get('Work Request #')
            _ld = _r.get('Weekly Reference Logged Date')
            if not _wr_raw or not _ld:
                continue
            _we = excel_serial_to_date(_ld)
            if _we is None:
                continue
            _we_d = _we.date() if isinstance(_we, datetime.datetime) else _we
            _prefetch_pairs.add((str(_wr_raw).split('.')[0], _we_d, _rid))

        try:
            from billing_audit.writer import (
                prefetch_attribution as _prefetch_attribution,
            )
            _prefetch_pairs_filtered = {(wr, we) for wr, we, _ in _prefetch_pairs}
            _attr_map, _attr_status = _prefetch_attribution(_prefetch_pairs_filtered)
            if _attr_status == 'fetch_failure':
                logging.warning(
                    "⚠️ Attribution bulk prefetch failed "
                    f"(status={_attr_status}); B/C will HOLD affected rows, "
                    "D/sub-helper will use current foreman (D-04 contract)."
                )
        except Exception:
            logging.exception(
                "⚠️ Attribution bulk prefetch: unexpected error; "
                "falling back to use-current for all attribution consumers."
            )
            _attr_map, _attr_status = {}, 'fetch_failure'

    # CR-01 graceful degradation: a MISSING RPC (rpc_missing) is correctness-
    # preserving to fall back per-row (the deployed lookup_attribution returns
    # the SAME frozen data, just slower) — NOT a D-04 violation. A transient
    # outage (fetch_failure) still HOLDs B/C. The fallback is bounded: per-row
    # resolution only happens for the rows B/C/sub-helper actually process this
    # run, so it cannot reintroduce the 137k per-row storm.
    _attr_use_per_row_fallback = (
        _attr_status == 'rpc_missing' and ATTRIBUTION_BULK_PREFETCH_FALLBACK
    )
    if _attr_use_per_row_fallback:
        logging.warning(
            "⚠️ Attribution bulk RPC missing (rpc_missing); "
            "ATTRIBUTION_BULK_PREFETCH_FALLBACK=1 -> degrading to per-row "
            "lookup_attribution for B/C/sub-helper (deploy lookup_attribution_bulk "
            "to restore bulk; see runbook E re-activation Step 1)."
        )

    # ── Subproject B: O(1) map read for sub-primary claimers (D-03) ──
    # The ThreadPoolExecutor per-row RPC block is replaced by O(1) lookups
    # from the shared _attr_map. On fetch_failure, ResolveOutcome is
    # constructed directly (D-04 direct-HOLD, zero additional Supabase calls).
    _sub_primary_claimer_map: dict = {}
    if BILLING_AUDIT_AVAILABLE and SUBCONTRACTOR_RATE_VARIANTS_ENABLED:
        try:
            from billing_audit.writer import (
                resolve_claimer as _resolve_claimer_b,
                ResolveOutcome as _ResolveOutcome_b,
            )
            for _r in rows:
                _sid = _r.get('__source_sheet_id')
                if _sid is None or _sid not in _FOLDER_DISCOVERED_SUB_IDS:
                    continue
                _rid = _r.get('__row_id')
                if not isinstance(_rid, int):
                    continue
                _wr_raw = _r.get('Work Request #')
                _ld = _r.get('Weekly Reference Logged Date')
                if not _wr_raw or not _ld or not is_checked(_r.get('Units Completed?')):
                    continue
                _we = excel_serial_to_date(_ld)
                if _we is None:
                    continue
                _we_d = _we.date() if isinstance(_we, datetime.datetime) else _we
                _eu = _r.get('__effective_user', 'Unknown Foreman')
                _wr_key_b = str(_wr_raw).split('.')[0]
                # HOLD only on a genuine transient outage, or on rpc_missing
                # when the operator has disabled the per-row fallback. On
                # rpc_missing WITH fallback on, route per-row (prefetched_map=
                # None) so B still generates with the real frozen claimer.
                if _attr_status == 'fetch_failure' or (
                    _attr_status == 'rpc_missing'
                    and not ATTRIBUTION_BULK_PREFETCH_FALLBACK
                ):
                    # D-04: construct HOLD directly, zero additional RPC calls.
                    _sub_primary_claimer_map[_rid] = _ResolveOutcome_b(
                        'hold', None, None, 'fetch_failure'
                    )
                else:
                    _sub_primary_claimer_map[_rid] = _resolve_claimer_b(
                        'reduced_sub', _eu,
                        wr=_wr_key_b,
                        week_ending=_we_d,
                        row_id=_rid,
                        enabled=SUBCONTRACTOR_HELPER_CLAIM_ATTRIBUTION_ENABLED,
                        prefetched_map=(
                            None if _attr_use_per_row_fallback else _attr_map
                        ),
                    )
        except Exception:
            logging.exception(
                "⚠️ Subproject B attribution map-read failed; falling "
                "back to current foreman for all subcontractor rows"
            )
            _sub_primary_claimer_map = {}

    # ── Subproject C: O(1) map read for vac-crew claimers (D-03) ──
    # D-04 direct-HOLD on fetch_failure, zero additional Supabase calls.
    _vac_crew_claimer_map: dict = {}
    if BILLING_AUDIT_AVAILABLE and VAC_CREW_CLAIM_ATTRIBUTION_ENABLED:
        try:
            from billing_audit.writer import (
                resolve_claimer as _resolve_claimer_c,
                ResolveOutcome as _ResolveOutcome_c,
            )
            for _r in rows:
                _rid = _r.get('__row_id')
                if not isinstance(_rid, int):
                    continue
                if not _r.get('__is_vac_crew'):
                    continue
                _wr_raw = _r.get('Work Request #')
                _ld = _r.get('Weekly Reference Logged Date')
                if not _wr_raw or not _ld or not is_checked(_r.get('Units Completed?')):
                    continue
                _we = excel_serial_to_date(_ld)
                if _we is None:
                    continue
                _we_d = _we.date() if isinstance(_we, datetime.datetime) else _we
                _current_vac = _r.get('__vac_crew_name') or ''
                _wr_key_c = str(_wr_raw).split('.')[0]
                # HOLD only on a genuine transient outage, or on rpc_missing
                # when the operator has disabled the per-row fallback. On
                # rpc_missing WITH fallback on, route per-row (prefetched_map=
                # None) so C still generates with the real frozen claimer.
                if _attr_status == 'fetch_failure' or (
                    _attr_status == 'rpc_missing'
                    and not ATTRIBUTION_BULK_PREFETCH_FALLBACK
                ):
                    # D-04: construct HOLD directly, zero additional RPC calls.
                    _vac_crew_claimer_map[_rid] = _ResolveOutcome_c(
                        'hold', None, None, 'fetch_failure'
                    )
                else:
                    _vac_crew_claimer_map[_rid] = _resolve_claimer_c(
                        'vac_crew', _current_vac,
                        wr=_wr_key_c,
                        week_ending=_we_d,
                        row_id=_rid,
                        enabled=VAC_CREW_CLAIM_ATTRIBUTION_ENABLED,
                        prefetched_map=(
                            None if _attr_use_per_row_fallback else _attr_map
                        ),
                    )
        except Exception:
            logging.exception(
                "⚠️ Subproject C attribution map-read failed; falling "
                "back to current vac-crew name for all VAC Crew rows"
            )
            _vac_crew_claimer_map = {}

    # ── Subproject D: O(1) map read for primary claimers (D-03) ──
    # Unlike B/C, D never HOLDs on fetch_failure — the core primary path
    # prioritizes availability (operator decision; D uses-current on any
    # failure so all primary billing still ships).
    _primary_claimer_map: dict = {}
    if (
        BILLING_AUDIT_AVAILABLE
        and PRIMARY_CLAIM_ATTRIBUTION_ENABLED
        and RES_GROUPING_MODE in ('helper', 'both')
    ):
        try:
            from billing_audit.writer import (
                resolve_claimer as _resolve_claimer_d,
            )
            for _r in rows:
                _rid = _r.get('__row_id')
                if not isinstance(_rid, int):
                    continue
                if _r.get('__is_vac_crew'):
                    continue
                # Valid helper rows are excluded from the primary emission
                # path, so resolving their primary claimer is pure overhead.
                if (
                    _r.get('__is_helper_row')
                    and _r.get('__helper_foreman')
                    and _r.get('__helper_dept')
                ):
                    continue
                _sid = _r.get('__source_sheet_id')
                if _sid is not None and _sid in _FOLDER_DISCOVERED_SUB_IDS:
                    continue  # subcontractor rows are Sub-project B's domain
                _wr_raw = _r.get('Work Request #')
                _ld = _r.get('Weekly Reference Logged Date')
                if not _wr_raw or not _ld or not is_checked(_r.get('Units Completed?')):
                    continue
                _we = excel_serial_to_date(_ld)
                if _we is None:
                    continue
                _we_d = _we.date() if isinstance(_we, datetime.datetime) else _we
                _eu = _r.get('__effective_user', 'Unknown Foreman')
                _wr_key_d = str(_wr_raw).split('.')[0]
                # WR-03: D never HOLDs. On fetch_failure the bulk map is empty,
                # so the prefetched-map miss yields a ('use', current,
                # 'no_history') outcome and D emits with the current foreman —
                # D never HOLDs by design (core primary path prioritizes
                # availability). On rpc_missing with fallback on,
                # prefetched_map=None routes the per-row lookup_attribution
                # which returns the real frozen claimer.
                _primary_claimer_map[_rid] = _resolve_claimer_d(
                    'primary', _eu,
                    wr=_wr_key_d,
                    week_ending=_we_d,
                    row_id=_rid,
                    enabled=PRIMARY_CLAIM_ATTRIBUTION_ENABLED,
                    prefetched_map=(
                        None if _attr_use_per_row_fallback else _attr_map
                    ),
                )
        except Exception:
            logging.exception(
                "⚠️ Subproject D attribution map-read failed; falling "
                "back to current foreman for all primary rows"
            )
            _primary_claimer_map = {}

    # ── VAC-crew cross-row unit reconciliation (operator contract
    # 2026-06-08). A WR can span multiple source sheets: a foreman /
    # original-contract sheet (no VAC columns) AND a VAC-crew sheet (VAC
    # columns). The SAME physical unit can then exist as two rows — only the
    # VAC-sheet copy carries the VAC claim, so the row-local predicate cannot
    # see the claim on the foreman's copy and the unit leaks onto the foreman
    # sheet (duplicated with the VacCrew sheet). Build the set of VAC-claimed
    # unit identities up front so the non-VAC emission below can drop any unit
    # that is VAC-claimed on ANY row. Keyed at the UNIT grain
    # (WR + week + Point + CU) — NOT the pole — so the foreman's OTHER units on
    # the same pole are retained (operator: per-unit, not per-pole).
    _vac_claimed_units = set()
    for _vr in rows:
        if not _vr.get('__is_vac_crew'):
            continue
        # Defense-in-depth: replay the emission loop's `Units Completed?`
        # gate in this pre-pass too. Today this is belt-and-suspenders — the
        # only place `__is_vac_crew` is set (get_all_source_rows ~L5464)
        # already requires `units_completed_checked`, so no production row
        # reaches here with `__is_vac_crew=True` and `Units Completed?`
        # unchecked. The guard future-proofs against that detection ever
        # being decoupled: an incomplete VAC row, if it slipped through,
        # would be dropped at the emission gate (never billed on VacCrew)
        # yet could still suppress the foreman/helper copy of the same unit,
        # billing a completed unit to nobody. Repo rule: a pre-pass that
        # mirrors a downstream gate must replay that gate's full predicate.
        if not is_checked(_vr.get('Units Completed?')):
            continue
        _vwr = _vr.get('Work Request #')
        _vdate_raw = _vr.get('Weekly Reference Logged Date')
        if not _vwr or not _vdate_raw:
            continue
        _vweek_date = excel_serial_to_date(_vdate_raw)
        if _vweek_date is None:
            continue
        _vpoint = str(
            _vr.get('Pole #') or _vr.get('Point #')
            or _vr.get('Point Number') or ''
        ).strip()
        _vcu = str(
            _vr.get('CU') or _vr.get('Billable Unit Code') or ''
        ).strip()
        if _vpoint and _vcu:
            _vac_claimed_units.add((
                str(_vwr).split('.')[0],
                _vweek_date.strftime("%m%d%y"),
                _vpoint,
                _vcu,
            ))

    for r in rows:
        wr = r.get('Work Request #')
        log_date_str = r.get('Weekly Reference Logged Date')
        units_completed = r.get('Units Completed?')
        total_price = parse_price(r.get('Units Total Price', 0))
        
        # Use __effective_user (set by dual-logic system in get_all_source_rows)
        effective_user = r.get('__effective_user', 'Unknown Foreman')
        assignment_method = r.get('__assignment_method', 'PATH_B_NO_CACHE_FILE')
        
        # Helper row metadata
        is_helper_row = r.get('__is_helper_row', False)
        helper_foreman = r.get('__helper_foreman', '')
        
        # Check if Units Completed? is true/1
        units_completed_checked = is_checked(units_completed)

        # REQUIRE: Work Request # AND Weekly Reference Logged Date AND Units Completed? = true/1 AND Units Total Price exists
        if not wr or not log_date_str or not units_completed_checked or total_price is None:
            continue # Skip if any essential grouping information is missing

        wr_key = str(wr).split('.')[0]
        
        try:
            # Parse the Weekly Reference Logged Date - this IS the week ending date
            week_ending_date = excel_serial_to_date(log_date_str)
            if week_ending_date is None:
                logging.warning(f"Could not parse Weekly Reference Logged Date '{log_date_str}' for WR# {wr_key}. Skipping row.")
                continue
            week_end_for_key = week_ending_date.strftime("%m%d%y")
            
            if TEST_MODE:
                logging.debug(f"WR# {wr_key}: Week ending {week_ending_date.strftime('%A, %m/%d/%Y')} | User: {effective_user} | Method: {assignment_method} | Helper: {is_helper_row}")
            
            # VARIANT-AWARE GROUPING: Build keys based on RES_GROUPING_MODE and row type
            keys_to_add = []
            
            # Check if this row was detected as VAC Crew (row-level column-based detection)
            is_vac_crew_row = r.get('__is_vac_crew', False)

            # Cross-row unit reconciliation: a non-VAC row whose unit
            # (WR + week + Point + CU) is VAC-claimed on ANOTHER source row is
            # the duplicate foreman/helper copy of a unit the VAC crew earns.
            # Drop it from ALL non-VAC variants (primary, helper, and the
            # subcontractor variants below) so the unit appears ONLY on the
            # VacCrew sheet (operator contract 2026-06-08; per-unit, not
            # per-pole). The VAC crew's own row is untouched.
            if not is_vac_crew_row:
                _unit_point = str(
                    r.get('Pole #') or r.get('Point #')
                    or r.get('Point Number') or ''
                ).strip()
                _unit_cu = str(
                    r.get('CU') or r.get('Billable Unit Code') or ''
                ).strip()
                if _unit_point and _unit_cu and (
                    wr_key, week_end_for_key, _unit_point, _unit_cu
                ) in _vac_claimed_units:
                    logging.info(
                        "➖ EXCLUDING from foreman/helper (unit VAC-claimed "
                        f"on another row): WR={wr_key}, "
                        f"Week={week_end_for_key}, Point={_unit_point}, "
                        f"CU={_unit_cu}"
                    )
                    continue

            # Phase 1.1 Bug B1 (D-04 / SUB-09): hoist
            # is_subcontractor_row to BEFORE the primary-emission
            # cascade AND the subcontractor variant emission block.
            # Both dependencies (r.get('__source_sheet_id') and
            # _FOLDER_DISCOVERED_SUB_IDS) are in scope at function
            # entry (verified per RESEARCH.md Pitfall 5). The same
            # computation was previously duplicated immediately before
            # the variant emission block downstream — that duplicate
            # site is now removed and the variable resolves to this
            # hoisted definition for BOTH the new partitioning gate
            # and the existing variant emission block. Hoist lives
            # OUTSIDE the ``if is_vac_crew_row:`` cascade so the
            # variant emission block (at per-row loop scope) still
            # sees the variable even when the row is VAC Crew (in
            # that case the variant block's outer
            # ``if is_subcontractor_row and ...`` gate still resolves
            # correctly).
            _row_sheet_id = r.get('__source_sheet_id')
            is_subcontractor_row = (
                _row_sheet_id is not None
                and _row_sheet_id in _FOLDER_DISCOVERED_SUB_IDS
            )

            # VAC Crew rows get their own dedicated group key (separate from primary/helper).
            # Detection is row-level: a row is VAC Crew when VAC Crew Helping? is non-blank
            # AND Vac Crew Completed Unit? is checked. This means the same sheet can produce
            # both primary/helper rows AND VAC Crew rows — they are mutually exclusive per-row
            # because a single row is either a VAC Crew row or a regular/helper row.
            if is_vac_crew_row:
                if not VAC_CREW_CLAIM_ATTRIBUTION_ENABLED:
                    # Kill switch OFF -> exact legacy behavior: one group per
                    # WR+week, no per-claimer partition.
                    vac_crew_key = f"{week_end_for_key}_{wr_key}_VACCREW"
                    # Use VAC Crew name (from 'VAC Crew Helping?' column) as the foreman
                    # for this group — NOT the primary foreman (effective_user).
                    vac_crew_foreman = r.get('__vac_crew_name') or effective_user
                    keys_to_add.append(('vac_crew', vac_crew_key, vac_crew_foreman))
                    # Only log at info level the first time a new group key is seen;
                    # subsequent rows belonging to the same WR/week VAC Crew group log at
                    # debug to avoid flooding logs with identical "GROUP CREATED" messages.
                    if vac_crew_key not in groups:
                        logging.info(f"🏗️ VAC CREW GROUP CREATED: WR={wr_key}, Week={week_end_for_key}")
                else:
                    # Subproject C: partition by frozen vac-crew claimer.
                    # Consume the pre-pass map. ``use`` -> partition by claimer;
                    # ``hold`` -> defer this row (correctness over availability);
                    # map miss (missing __row_id, pre-pass skipped, plumbing fault)
                    # -> use-current, NEVER HOLD.
                    _vac_current = r.get('__vac_crew_name') or effective_user
                    _c_vac_claimer = None
                    _vac_outcome = _vac_crew_claimer_map.get(r.get('__row_id'))
                    if _vac_outcome is not None and _vac_outcome.action == 'hold':
                        _c_vac_claimer = None  # defer — correctness over availability
                        try:
                            from billing_audit.writer import record_attribution_hold
                            record_attribution_hold(
                                wr_key,
                                week_ending_date.date()
                                if isinstance(week_ending_date, datetime.datetime)
                                else week_ending_date,
                                'vac_crew',
                            )
                        except Exception:
                            logging.exception(
                                "⚠️ Subproject C: record_attribution_hold failed"
                            )
                    elif _vac_outcome is not None and _vac_outcome.action == 'use':
                        _c_vac_claimer = _vac_outcome.name or _vac_current or 'Unknown'
                    else:
                        # Map miss (row absent from pre-pass) or unknown action:
                        # fall back to current name, never HOLD.
                        _c_vac_claimer = _vac_current or 'Unknown'

                    if _c_vac_claimer is not None:
                        _c_vac_sanitized = _RE_SANITIZE_IDENTIFIER.sub(
                            '_', _c_vac_claimer
                        )[:50]
                        vac_crew_key = (
                            f"{week_end_for_key}_{wr_key}_VACCREW_{_c_vac_sanitized}"
                        )
                        keys_to_add.append(('vac_crew', vac_crew_key, _c_vac_claimer))
                        if vac_crew_key not in groups:
                            logging.info(
                                f"🏗️ VAC CREW GROUP CREATED: WR={wr_key}, "
                                f"Week={week_end_for_key}"
                            )
            else:
                # Check if helper mode is enabled
                helper_mode_enabled = RES_GROUPING_MODE in ('helper', 'both')
                
                # Check if this is a valid helper row (both checkboxes checked, has helper info)
                valid_helper_row = False
                if helper_mode_enabled and is_helper_row and helper_foreman:
                    helper_dept = r.get('__helper_dept', '')
                    helper_job = r.get('__helper_job', '')
                    # Validate helper row: helper_dept is required, helper_job is OPTIONAL
                    # This allows rows to sync even when Helper Job # is missing
                    if helper_dept:  # helper_job is now optional
                        valid_helper_row = True
                
                # Primary variant logic
                if RES_GROUPING_MODE == 'primary':
                    # In primary mode, ALL rows go to main (including helper rows)
                    primary_key = f"{week_end_for_key}_{wr_key}"
                    keys_to_add.append(('primary', primary_key, None))
                elif RES_GROUPING_MODE in ('helper', 'both'):
                    # Phase 1.1 Bug B1 (D-04 / SUB-09): partitioning
                    # gate. Subcontractor non-helper rows do NOT emit
                    # the legacy primary key — their content lives
                    # exclusively in the _REDUCEDSUB (always) and
                    # _AEPBILLABLE (post-cutoff) variant files
                    # produced by the subcontractor variant block
                    # below. Primary / original-contract / vac_crew
                    # rows fall through unchanged. Plan 01-03 Test 1's
                    # "additive" contract is overridden per D-22;
                    # Living Ledger entry [Phase 1.1 timestamp]
                    # documents the design-intent change.
                    if not is_subcontractor_row and not valid_helper_row:
                        # Subproject D (2026-05-25): partition the
                        # production primary file by the FROZEN primary
                        # claimer. Consume the pre-pass map. ``use`` ->
                        # partition by claimer; ``hold`` (Supabase outage),
                        # map miss, or disabled -> use the current
                        # effective_user and STILL emit (D never holds —
                        # operator decision for the core path). Empty
                        # claimer -> 'Unknown Foreman' sentinel so the
                        # _User_ suffix builder never gets an empty
                        # identifier (mirrors B's Codex-P1 fix).
                        if PRIMARY_CLAIM_ATTRIBUTION_ENABLED:
                            _d_outcome = _primary_claimer_map.get(r.get('__row_id'))
                            if _d_outcome is not None and _d_outcome.action == 'use':
                                _d_claimer = (
                                    _d_outcome.name or effective_user or 'Unknown Foreman'
                                )
                            else:
                                # hold / map-miss / disabled / None -> current.
                                _d_claimer = effective_user or 'Unknown Foreman'
                            _d_claimer_sanitized = _RE_SANITIZE_IDENTIFIER.sub(
                                '_', _d_claimer
                            )[:50]
                            primary_key = (
                                f"{week_end_for_key}_{wr_key}_USER_"
                                f"{_d_claimer_sanitized}"
                            )
                            keys_to_add.append(('primary', primary_key, _d_claimer))
                            if primary_key not in groups:
                                logging.info(
                                    f"🧑 PRIMARY GROUP CREATED: WR={wr_key}, "
                                    f"Week={week_end_for_key}"
                                )
                        else:
                            # Kill switch OFF -> exact legacy bare primary.
                            primary_key = f"{week_end_for_key}_{wr_key}"
                            keys_to_add.append(('primary', primary_key, None))
                    elif is_subcontractor_row and not valid_helper_row:
                        # Diagnostic log only — no group emission.
                        # Operators can confirm the partition is
                        # firing by grepping for this prefix. PII
                        # marker "EXCLUDING from main Excel" already
                        # covers this body via existing
                        # _PII_LOG_MARKERS entry.
                        logging.debug(
                            f"➖ EXCLUDING from main Excel (subcontractor row): "
                            f"WR={wr_key}, Week={week_end_for_key}"
                        )
                    elif valid_helper_row:
                        # UNCHANGED legacy behaviour — helper row
                        # excluded from main Excel regardless of
                        # subcontractor/non-subcontractor.
                        logging.info(f"➖ EXCLUDING from main Excel: WR={wr_key}, Week={week_end_for_key} (Helper row with both checkboxes)")
                
                # Helper variant - ONLY created when mode allows it
                if valid_helper_row and helper_mode_enabled:
                    helper_dept = r.get('__helper_dept', '')
                    helper_job = r.get('__helper_job', '')
                    # PERFORMANCE: Use pre-compiled regex for helper name sanitization
                    helper_sanitized = _RE_SANITIZE_HELPER_NAME.sub('_', helper_foreman)[:50]
                    helper_key = f"{week_end_for_key}_{wr_key}_HELPER_{helper_sanitized}"
                    # Phase 1.1 UAT gap closure (SUB-09 helper dimension):
                    # mirror Bug B1's non-helper primary partition onto the
                    # helper path. Subcontractor helper rows do NOT emit the
                    # legacy `_HELPER_<name>` key — their line items live
                    # exclusively in the `_REDUCEDSUB_HELPER_<name>` (always)
                    # and `_AEPBILLABLE_HELPER_<name>` (post-cutoff) shadow
                    # files produced by the subcontractor variant block below.
                    # Pre-fix this append fired for ALL helper rows including
                    # subcontractor ones (the D-09 additive intent), producing
                    # a duplicate `_Helper_<name>` file on TARGET_SHEET_ID that
                    # duplicated the shadow file's line items. The guard is
                    # strictly `is_subcontractor_row`-scoped so primary /
                    # original-contract / vac_crew helper rows are byte-identical
                    # (ROADMAP success criterion #5). Living Ledger entry
                    # [2026-05-19] documents the asymmetry + fix.
                    if not is_subcontractor_row:
                        keys_to_add.append(('helper', helper_key, helper_foreman))
                        # HELPER GROUP LOGGING: Always log when helper group is created
                        logging.info(f"🔧 HELPER GROUP CREATED: WR={wr_key}, Week={week_end_for_key}, Helper={helper_foreman}, Dept={helper_dept}, Job={helper_job}")
                    else:
                        # Diagnostic log only — no legacy-helper group emission for
                        # subcontractor rows. Operators can confirm the partition
                        # is firing by grepping this prefix. PII marker
                        # "EXCLUDING from main Excel" already covers this body via
                        # the existing _PII_LOG_MARKERS entry.
                        logging.debug(
                            f"➖ EXCLUDING from main Excel (subcontractor legacy helper): "
                            f"WR={wr_key}, Week={week_end_for_key}, Helper={helper_foreman}"
                        )
                elif is_helper_row and not helper_mode_enabled:
                    # In primary mode, helper rows go to main
                    logging.info(f"ℹ️ Helper row found but RES_GROUPING_MODE={RES_GROUPING_MODE} - including in main Excel")
                elif is_helper_row:
                    # Helper row missing required helper_dept (helper_job is optional)
                    helper_dept = r.get('__helper_dept', '')
                    helper_job = r.get('__helper_job', '')
                    logging.warning(f"⚠️ Helper row for WR {wr_key} missing required Helper Dept # (Job: '{helper_job}') - including in main Excel")

            # ── Phase 01 Plan 03 (D-08/D-09/D-13/D-22): Subcontractor
            # rate variants. Per the committed Blocker 3 plumbing
            # decision the gate is PER-ROW, evaluated against the
            # row's ``__source_sheet_id`` (populated upstream by
            # ``_fetch_and_process_sheet``) and the kill-switch env
            # var. A subcontractor row produces:
            #   • a ``_REDUCEDSUB`` group key unconditionally (D-08
            #     / SUB-02 — every sub WR group always gets a
            #     reduced-sub Excel),
            #   • a ``_AEPBILLABLE`` group key when the row's
            #     ``Snapshot Date >= _AEP_BILLABLE_CUTOFF`` (D-08 /
            #     SUB-01 — snapshot is authoritative per Living
            #     Ledger 2026-04-21 22:35; never use Weekly Reference
            #     Logged Date here),
            #   • two ``_HELPER_<name>`` shadow keys when the
            #     existing helper-foreman event fires on this row
            #     (D-09 — shadows piggyback on the same
            #     ``valid_helper_row + helper_mode_enabled`` gate the
            #     legacy ``_HELPER_<name>`` key uses; the
            #     subcontractor sheet only contributes the new
            #     prefix tokens).
            # Non-subcontractor rows fall through with zero behaviour
            # change (per-row gate ensures no key bleed across rows
            # in the same call). Helper names are sanitized at the
            # producer site via ``_RE_SANITIZE_HELPER_NAME`` (D-22 /
            # Living Ledger 2026-04-23 18:25 — idempotent regex, so
            # the consumer site in ``generate_excel`` can safely
            # re-apply).
            #
            # Phase 1.1 Bug B1 (D-04 / SUB-09): the
            # ``is_subcontractor_row`` computation that used to live
            # here has been HOISTED to BEFORE the primary-emission
            # cascade above so the new partitioning gate can read it.
            # The hoisted variable is in scope for this block (Python
            # locals introduced earlier in the same function body
            # remain in scope), so the variant emission block below
            # reads the SAME boolean — no behavioural change to the
            # variant emission contract.
            # Copilot (PR #219): ``not is_vac_crew_row`` excludes VAC crew
            # rows from the subcontractor variant emission. ``__is_vac_crew``
            # is set by column presence (not sheet membership), so a VAC crew
            # row can come from a subcontractor-folder sheet; without this
            # gate it would be DOUBLE-emitted (VACCREW + REDUCEDSUB/AEPBILLABLE)
            # and a vac ``hold`` outcome would be bypassed by the sub variants.
            # A vac row already produced its own group above; it must not also
            # emit subcontractor primary variants.
            if is_subcontractor_row and not is_vac_crew_row and SUBCONTRACTOR_RATE_VARIANTS_ENABLED:
                # Snapshot cutoff is needed by BOTH the primary block
                # here and the helper-shadow block below, so compute it
                # once up front. ``excel_serial_to_date`` returns ``None``
                # for blank/unparseable values (D-16 fall-through safety).
                # Hoisted above the helper-completed guard so the
                # helper-shadow block still sees it even when the primary
                # emission is skipped.
                _snap_for_cutoff = excel_serial_to_date(r.get('Snapshot Date'))

                # 2026-05-21 hotfix (carried into Subproject B's _USER_
                # partitioning): a helper-COMPLETED subcontractor row
                # (``Units Completed?`` AND ``Helping Foreman Completed
                # Unit?`` both checked, with a valid ``Foreman Helping?``
                # + helper dept) belongs SOLELY to the helper-shadow files
                # below — the helper, not the primary foreman, earns the
                # credit for that line item on Smartsheet. Emitting a
                # primary ``_REDUCEDSUB_USER_`` / ``_AEPBILLABLE_USER_`` key
                # here would double-count the row (it would appear in BOTH
                # the primary and the helper file) and wrongly credit the
                # primary. Mirrors the legacy main-file ``valid_helper_row``
                # exclusion. Computed locally because the else-branch
                # ``valid_helper_row`` is out of scope for vac_crew rows;
                # uses the same inputs as the helper-shadow recompute below.
                _sub_is_valid_helper_row = (
                    not is_vac_crew_row
                    and RES_GROUPING_MODE in ('helper', 'both')
                    and is_helper_row
                    and bool(helper_foreman)
                    and bool(r.get('__helper_dept', ''))
                )

                # Subproject B: resolve the FROZEN primary claimer from
                # the pre-pass map. ``use`` -> partition by the claimer;
                # ``hold`` -> defer this row's primary variants this run
                # (correctness over availability) and record a HOLD; map
                # miss -> use the current effective_user. Skipped for
                # helper-completed rows — they are not primary claims, so
                # no claimer is resolved and no HOLD is recorded; the
                # ``None`` default routes through the
                # ``if _b_primary_claimer is not None`` gate below and
                # suppresses the primary _USER_ emission.
                _b_primary_claimer = None
                if not _sub_is_valid_helper_row:
                    _b_outcome = _sub_primary_claimer_map.get(r.get('__row_id'))
                    if _b_outcome is not None and _b_outcome.action == 'hold':
                        _b_primary_claimer = None
                        try:
                            from billing_audit.writer import record_attribution_hold
                            # Copilot: record_attribution_hold is typed
                            # ``datetime.date | None``; normalize the datetime
                            # week_ending_date to a pure date so the hold key
                            # is 'YYYY-MM-DD' (matching the pre-pass
                            # normalization for resolve_claimer), not
                            # 'YYYY-MM-DDT00:00:00'.
                            record_attribution_hold(
                                wr_key,
                                week_ending_date.date()
                                if isinstance(
                                    week_ending_date, datetime.datetime
                                )
                                else week_ending_date,
                                'reduced_sub',
                            )
                        except Exception:
                            logging.exception(
                                "⚠️ Subproject B: record_attribution_hold failed"
                            )
                    elif _b_outcome is not None and _b_outcome.action == 'use':
                        # Codex P1: fall back to a non-empty sentinel. A
                        # whitespace-only "Foreman Assigned?" yields an empty
                        # __effective_user upstream, and resolve_claimer's
                        # use/no_history then returns an empty name. Without
                        # this guard ``_b_primary_claimer`` would be '' yet
                        # still pass the ``is not None`` gate below, creating a
                        # _USER_ key with an empty claimer that crashes
                        # generate_excel at the suffix raise. 'Unknown Foreman'
                        # mirrors the foreman-assignment fallback sentinel and
                        # keeps the row's billing in a (clearly-flagged) file.
                        _b_primary_claimer = (
                            _b_outcome.name or effective_user or 'Unknown Foreman'
                        )
                    else:
                        _b_primary_claimer = effective_user or 'Unknown Foreman'

                if _b_primary_claimer is not None:
                    _b_claimer_sanitized = _RE_SANITIZE_IDENTIFIER.sub(
                        '_', _b_primary_claimer
                    )[:50]
                    # ReducedSub: unconditional per SUB-02 / D-08, now
                    # partitioned by frozen primary claimer (Subproject B).
                    reduced_key = (
                        f"{week_end_for_key}_{wr_key}_REDUCEDSUB_USER_"
                        f"{_b_claimer_sanitized}"
                    )
                    keys_to_add.append(
                        ('reduced_sub', reduced_key, _b_primary_claimer)
                    )
                    if reduced_key not in groups:
                        logging.info(
                            f"🔻 REDUCED SUB GROUP CREATED: WR={wr_key}, "
                            f"Week={week_end_for_key}"
                        )

                    # AEPBillable: snapshot-cutoff-gated per SUB-01 / D-08
                    # / Living Ledger 2026-04-21 22:35 (snapshot is
                    # authoritative; Weekly Reference Logged Date is NOT a
                    # valid fallback here).
                    if (
                        _snap_for_cutoff is not None
                        and _snap_for_cutoff.date() >= _AEP_BILLABLE_CUTOFF
                    ):
                        aep_key = (
                            f"{week_end_for_key}_{wr_key}_AEPBILLABLE_USER_"
                            f"{_b_claimer_sanitized}"
                        )
                        keys_to_add.append(
                            ('aep_billable', aep_key, _b_primary_claimer)
                        )
                        if aep_key not in groups:
                            logging.info(
                                f"💲 AEP BILLABLE GROUP CREATED: WR={wr_key}, "
                                f"Week={week_end_for_key}"
                            )

                # Helper-shadow variants: piggyback on the EXISTING
                # helper detection. The two gates that already
                # qualify a row for the legacy ``_HELPER_<name>``
                # key (valid_helper_row + helper_mode_enabled) also
                # qualify it for the shadow variants when the sheet
                # is subcontractor. This means a single helper-row
                # event on a subcontractor WR produces:
                #   • the legacy ``_HELPER_<name>`` group (already
                #     added above by the legacy branch when the
                #     gates fire),
                #   • the new ``_REDUCEDSUB_HELPER_<name>`` group
                #     (unconditional),
                #   • the new ``_AEPBILLABLE_HELPER_<name>`` group
                #     when snapshot >= cutoff.
                # ``helper_mode_enabled`` and ``valid_helper_row``
                # are computed in the non-vac_crew else-branch above
                # — they are out of scope here for vac_crew rows
                # (``is_vac_crew_row`` short-circuits this entire
                # else block). Re-evaluate the same gates locally to
                # keep the new code self-contained and to avoid
                # depending on the order of variable definitions in
                # the enclosing block (the legacy primary-vs-helper
                # cascade lives in the else branch but the new sub
                # block is at function scope inside the for-loop, so
                # mirror its inputs here for clarity).
                if not is_vac_crew_row:
                    _helper_mode_enabled = RES_GROUPING_MODE in ('helper', 'both')
                    _valid_helper_row = False
                    if _helper_mode_enabled and is_helper_row and helper_foreman:
                        _helper_dept_local = r.get('__helper_dept', '')
                        if _helper_dept_local:
                            _valid_helper_row = True
                    if _valid_helper_row and _helper_mode_enabled:
                        # Phase 1.1 Bug C (D-10..D-16 / SUB-11):
                        # per-row claim-history attribution. For
                        # subcontractor rows ONLY (D-15), partition
                        # shadow-variant rows by the FROZEN helper
                        # foreman in
                        # ``billing_audit.attribution_snapshot``
                        # rather than the current Smartsheet
                        # ``Foreman Helping?`` value. This is what
                        # makes a foreman who helped Mon-Tue keep
                        # their partitioned helper file even after a
                        # Wed swap on the Smartsheet column.
                        #
                        # Fall-back contract (D-12): when the reader
                        # returns None (no frozen row yet →
                        # no_history, OR PostgREST outage →
                        # fetch_failure, OR kill switch off →
                        # disabled), the row joins the helper file of
                        # the CURRENT ``helper_foreman`` from
                        # Smartsheet. Safe default — helper files
                        # never silently empty.
                        #
                        # Per-WR dedupe (RESEARCH.md §C Pitfall 1):
                        # the fall-back WARNING fires ONCE per
                        # (wr, week, current_helper) tuple, NOT
                        # per-row. Without the dedupe, a 100-row
                        # fall-back run would log 100 identical
                        # WARNINGs.
                        _attributed_helper = helper_foreman  # D-12 default
                        _attribution_reason: str | None = None
                        # Phase 2 Plan 02 (D-03): O(1) map read from the
                        # shared _attr_map built by prefetch_attribution.
                        # No per-row Supabase RPC; no recency scope gate
                        # (D-05 removed ATTRIBUTION_RESOLUTION_WEEKS entirely).
                        if (
                            is_subcontractor_row
                            and SUBCONTRACTOR_HELPER_CLAIM_ATTRIBUTION_ENABLED
                            and (
                                _attr_status == 'fetch_failure'
                                or (
                                    _attr_status == 'rpc_missing'
                                    and not ATTRIBUTION_BULK_PREFETCH_FALLBACK
                                )
                            )
                        ):
                            # WR-05: surface the outage explicitly. Sub-helper
                            # does NOT HOLD (Phase 1.1 design) — it falls back
                            # to the current `Foreman Helping?` (D-12 default),
                            # but now LOGS the reason via the per-WR WARNING
                            # below, restoring the Bug C observability the bulk
                            # path dropped. No per-row RPC issued on this path.
                            _attribution_reason = 'fetch_failure'
                        elif (
                            is_subcontractor_row
                            and SUBCONTRACTOR_HELPER_CLAIM_ATTRIBUTION_ENABLED
                        ):
                            try:
                                from billing_audit.writer import (
                                    resolve_claimer as _resolve_claimer_sh,
                                )
                                _sh_rid = r.get('__row_id')
                                _sh_out = _resolve_claimer_sh(
                                    'helper', helper_foreman,
                                    wr=wr_key,
                                    week_ending=week_ending_date,
                                    row_id=_sh_rid,
                                    enabled=SUBCONTRACTOR_HELPER_CLAIM_ATTRIBUTION_ENABLED,
                                    prefetched_map=(
                                        None if _attr_use_per_row_fallback
                                        else _attr_map
                                    ),
                                )
                                if _sh_out.action == 'use':
                                    _attributed_helper = (
                                        _sh_out.name or helper_foreman
                                    )
                                    _attribution_reason = None
                                elif _sh_out.action == 'hold':
                                    # fetch_failure: D-12 default (current
                                    # helper), flag for WARNING below.
                                    _attribution_reason = 'fetch_failure'
                                else:
                                    # 'disabled' or 'no_history': use D-12 default
                                    _attribution_reason = (
                                        _sh_out.reason
                                        if _sh_out.reason in ('no_history', 'fetch_failure')
                                        else None
                                    )
                            except Exception:
                                # Defense-in-depth: pipeline MUST NEVER
                                # crash on a reader failure — D-12 default.
                                logging.exception(
                                    "⚠️ Subcontractor helper claim "
                                    "attribution map-read: unexpected "
                                    "error (treating as fetch_failure)"
                                )
                                _attribution_reason = 'fetch_failure'

                        # Per-WR dedupe WARNING — operator-actionable,
                        # names the reason. `_bug_c_warning_seen` is
                        # initialized at function scope (set
                        # construction at top of `group_source_rows`).
                        # Tuple key uses sanitized helper for
                        # set-membership stability.
                        if (
                            is_subcontractor_row
                            and SUBCONTRACTOR_HELPER_CLAIM_ATTRIBUTION_ENABLED
                            and _attribution_reason in ('no_history', 'fetch_failure')
                        ):
                            _warning_helper_key = _RE_SANITIZE_HELPER_NAME.sub(
                                '_', helper_foreman
                            )[:50]
                            _warning_key = (
                                wr_key, week_end_for_key, _warning_helper_key
                            )
                            if _warning_key not in _bug_c_warning_seen:
                                _bug_c_warning_seen.add(_warning_key)
                                # PII marker: "Subcontractor helper
                                # claim attribution fallback" — added
                                # to _PII_LOG_MARKERS in Step 2.
                                logging.warning(
                                    f"⚠️ Subcontractor helper claim "
                                    f"attribution fallback for "
                                    f"WR={wr_key} week={week_end_for_key} "
                                    f"helper={_warning_helper_key} "
                                    f"(reason={_attribution_reason}). "
                                    f"Helper file rows will fall back to "
                                    f"the current `Foreman Helping?` "
                                    f"value. To investigate: check "
                                    f"Supabase Logs for "
                                    f"PGRST106/PGRST301/PGRST404 on the "
                                    f"'lookup_attribution' op."
                                )

                        # Use the attributed helper (or current helper
                        # on fall-back) for shadow-variant
                        # sanitization. Phase 1 emission body
                        # downstream is UNCHANGED — only the input
                        # value to the sanitizer changes.
                        _helper_sanitized = (
                            _RE_SANITIZE_HELPER_NAME.sub('_', _attributed_helper)[:50]
                        )
                        rs_helper_key = (
                            f"{week_end_for_key}_{wr_key}_REDUCEDSUB_HELPER_"
                            f"{_helper_sanitized}"
                        )
                        keys_to_add.append(
                            ('reduced_sub_helper', rs_helper_key, _attributed_helper)
                        )
                        if rs_helper_key not in groups:
                            logging.info(
                                f"🔻 REDUCED SUB HELPER GROUP CREATED: "
                                f"WR={wr_key}, Week={week_end_for_key}, "
                                f"Helper={_attributed_helper}"
                            )
                        if (
                            _snap_for_cutoff is not None
                            and _snap_for_cutoff.date() >= _AEP_BILLABLE_CUTOFF
                        ):
                            aep_helper_key = (
                                f"{week_end_for_key}_{wr_key}_AEPBILLABLE_HELPER_"
                                f"{_helper_sanitized}"
                            )
                            keys_to_add.append(
                                (
                                    'aep_billable_helper',
                                    aep_helper_key,
                                    _attributed_helper,
                                )
                            )
                            if aep_helper_key not in groups:
                                logging.info(
                                    f"💲 AEP BILLABLE HELPER GROUP CREATED: "
                                    f"WR={wr_key}, Week={week_end_for_key}, "
                                    f"Helper={_attributed_helper}"
                                )

            # Add row to all applicable groups
            for variant, key, current_foreman in keys_to_add:
                # Add calculated values to row data
                r_copy = r.copy()
                r_copy['__variant'] = variant
                r_copy['__current_foreman'] = current_foreman or effective_user
                r_copy['__week_ending_date'] = week_ending_date
                r_copy['__grouping_key'] = key
                groups[key].append(r_copy)
                
                if TEST_MODE:
                    logging.debug(f"Added to {variant} group '{key}': {len(groups[key])} rows")
                
        except (parser.ParserError, TypeError) as e:
            logging.warning(f"Could not parse Weekly Reference Logged Date '{log_date_str}' for WR# {wr_key}. Skipping row. Error: {e}")
            continue
    
    # FINAL VALIDATION: Ensure each group contains only one work request
    validation_errors = []
    for group_key, group_rows in groups.items():
        unique_wrs = list(set(str(row.get('Work Request #', '')).split('.')[0] for row in group_rows))
        if len(unique_wrs) != 1:
            validation_errors.append(f"Group {group_key} contains {len(unique_wrs)} work requests: {unique_wrs}")
    
    if validation_errors:
        error_msg = "CRITICAL GROUPING ERRORS: " + "; ".join(validation_errors)
        logging.error(error_msg)
        sentry_capture_message_with_context(
            message=error_msg,
            level="error",
            context_name="grouping_validation",
            context_data={
                "total_groups": len(groups),
                "validation_errors": validation_errors,
                "error_count": len(validation_errors),
            },
            tags={"error_location": "group_validation", "error_type": "data_integrity"}
        )
    else:
        logging.info(f"✅ Grouping validation passed: {len(groups)} groups, each with exactly 1 work request")
    
    # HELPER GROUP SUMMARY LOGGING
    helper_groups = [k for k in groups.keys() if '_HELPER_' in k]
    vac_crew_groups = [k for k in groups.keys() if '_VACCREW' in k]
    primary_groups = [k for k in groups.keys() if '_HELPER_' not in k and '_VACCREW' not in k]
    if helper_groups:
        logging.info(f"🔧 HELPER GROUP SUMMARY: Created {len(helper_groups)} helper groups out of {len(groups)} total groups")
        logging.info(f"   Primary groups: {len(primary_groups)}")
        logging.info(f"   Helper groups: {len(helper_groups)}")
        # Log sample helper groups
        for helper_key in helper_groups[:5]:
            row_count = len(groups[helper_key])
            logging.info(f"   Helper group '{helper_key}': {row_count} rows")
    else:
        logging.warning(f"⚠️ HELPER GROUP SUMMARY: No helper groups created out of {len(groups)} total groups - check RES_GROUPING_MODE and helper row detection")
    if vac_crew_groups:
        logging.info(f"🏗️ VAC CREW GROUP SUMMARY: Created {len(vac_crew_groups)} VAC Crew group(s) out of {len(groups)} total groups")
        for vac_key in vac_crew_groups[:5]:
            logging.info(f"   VAC Crew group '{vac_key}': {len(groups[vac_key])} rows")
    
    # Optional filtering by WR_FILTER (retain primary, helper, and vac_crew variants)
    if WR_FILTER and TEST_MODE:
        before = len(groups)
        def _key_matches_wr(k: str, wr: str) -> bool:
            # k format examples (all eleven shapes emitted by group_source_rows):
            #   MMDDYY_WR                                   → primary
            #   MMDDYY_WR_USER_<name>                       → primary (Subproject D)
            #   MMDDYY_WR_HELPER_<name>                     → helper
            #   MMDDYY_WR_VACCREW                           → vac_crew
            #   MMDDYY_WR_VACCREW_<claimer>                 → vac_crew (Subproject C)
            #   MMDDYY_WR_REDUCEDSUB                        → reduced_sub  (Phase 1)
            #   MMDDYY_WR_AEPBILLABLE                       → aep_billable (Phase 1)
            #   MMDDYY_WR_REDUCEDSUB_HELPER_<name>          → reduced_sub_helper  (Phase 1)
            #   MMDDYY_WR_AEPBILLABLE_HELPER_<name>         → aep_billable_helper (Phase 1)
            #   MMDDYY_WR_REDUCEDSUB_USER_<claimer>         → reduced_sub  (Subproject B)
            #   MMDDYY_WR_AEPBILLABLE_USER_<claimer>        → aep_billable (Subproject B)
            #
            # Phase 01 gap closure (REVIEW-CR-03): mirror of the
            # ``_key_matches_excluded_wr`` fix immediately below. Without the
            # new variant clauses, ``TEST_MODE=true WR_FILTER=<wr>`` drops
            # the new-variant groups before generation runs, silently
            # producing zero ``_AEPBillable`` / ``_ReducedSub`` output for the
            # filtered WR — which makes the Step B operator diagnostic
            # documented in 01-VERIFICATION.md unexercisable. Match shape is
            # IDENTICAL to ``_key_matches_excluded_wr``. The two matchers
            # MUST stay in sync — any future variant added in
            # ``group_source_rows`` must extend BOTH.
            try:
                suffix = k.split('_', 1)[1]  # take everything after first underscore (WR...)
            except Exception:
                return False
            return (
                suffix == wr
                # Subproject D: per-claimer primary key {wr}_USER_<claimer>
                # (attribution on). Mirror of the _key_matches_excluded_wr
                # clause below — the two matchers MUST stay in sync.
                or suffix.startswith(f"{wr}_USER_")
                or suffix.startswith(f"{wr}_HELPER_")
                or suffix == f"{wr}_VACCREW"
                # Subproject C: per-claimer vac key {wr}_VACCREW_<claimer>
                # (attribution on). Prefix-match so EXCLUDE_WRS / WR_FILTER
                # cover both the legacy bare and the per-claimer shapes.
                or suffix.startswith(f"{wr}_VACCREW_")
                # Phase 1 subcontractor variants (REVIEW-CR-03).
                or suffix == f"{wr}_REDUCEDSUB"
                or suffix == f"{wr}_AEPBILLABLE"
                or suffix.startswith(f"{wr}_REDUCEDSUB_HELPER_")
                or suffix.startswith(f"{wr}_AEPBILLABLE_HELPER_")
                # Subproject B: per-claimer subcontractor primary keys
                # {wr}_REDUCEDSUB_USER_<claimer> / {wr}_AEPBILLABLE_USER_<claimer>
                # (attribution on — the production default). Prefix-match so
                # WR_FILTER / EXCLUDE_WRS cover the partitioned shape, not just
                # the bare _REDUCEDSUB / _AEPBILLABLE. Mirror in BOTH matchers.
                or suffix.startswith(f"{wr}_REDUCEDSUB_USER_")
                or suffix.startswith(f"{wr}_AEPBILLABLE_USER_")
            )

        groups = {k: v for k, v in groups.items() if any(_key_matches_wr(k, wr) for wr in WR_FILTER)}
        logging.info(f"🔎 WR_FILTER applied (primary + helper + vac_crew): {len(groups)}/{before} groups retained ({','.join(WR_FILTER)})")
    
    # EXCLUDE_WRS: Remove specific Work Requests from generation (applies always, not just TEST_MODE)
    if EXCLUDE_WRS:
        before_exclude = len(groups)
        logging.info(f"🔍 EXCLUDE_WRS check: Attempting to exclude WRs {EXCLUDE_WRS} from {before_exclude} groups")
        
        # Debug: Show sample group keys for troubleshooting
        sample_keys = list(groups.keys())[:5]
        logging.info(f"🔍 Sample group keys: {sample_keys}")
        
        def _key_matches_excluded_wr(k: str, wr: str) -> bool:
            # k format examples (all eleven shapes emitted by group_source_rows):
            #   MMDDYY_WR                                   → primary
            #   MMDDYY_WR_USER_<name>                       → primary (Subproject D)
            #   MMDDYY_WR_HELPER_<name>                     → helper
            #   MMDDYY_WR_VACCREW                           → vac_crew
            #   MMDDYY_WR_VACCREW_<claimer>                 → vac_crew (Subproject C)
            #   MMDDYY_WR_REDUCEDSUB                        → reduced_sub  (Phase 1)
            #   MMDDYY_WR_AEPBILLABLE                       → aep_billable (Phase 1)
            #   MMDDYY_WR_REDUCEDSUB_HELPER_<name>          → reduced_sub_helper  (Phase 1)
            #   MMDDYY_WR_AEPBILLABLE_HELPER_<name>         → aep_billable_helper (Phase 1)
            #   MMDDYY_WR_REDUCEDSUB_USER_<claimer>         → reduced_sub  (Subproject B)
            #   MMDDYY_WR_AEPBILLABLE_USER_<claimer>        → aep_billable (Subproject B)
            #
            # Phase 01 gap closure (REVIEW-CR-02): before this fix the matcher
            # only recognized the first four shapes, so EXCLUDE_WRS=<wr>
            # silently uploaded the four new variant files to TARGET_SHEET_ID
            # and SUBCONTRACTOR_PPP_SHEET_ID even when the operator's intent
            # was "do not bill yet." The additive ``or`` clauses below close
            # that gap. Match shape mirrors ``_key_matches_wr``; the two
            # matchers are siblings and MUST stay in sync — any future
            # variant added in ``group_source_rows`` must extend BOTH.
            try:
                suffix = k.split('_', 1)[1]  # take everything after first underscore (WR...)
            except Exception:
                return False
            return (
                suffix == wr
                or suffix.startswith(f"{wr}_HELPER_")
                or suffix.startswith(f"{wr}_USER_")
                or suffix == f"{wr}_VACCREW"
                # Subproject C: per-claimer vac key {wr}_VACCREW_<claimer>
                # (attribution on). Prefix-match so EXCLUDE_WRS / WR_FILTER
                # cover both the legacy bare and the per-claimer shapes.
                or suffix.startswith(f"{wr}_VACCREW_")
                # Phase 1 subcontractor variants (REVIEW-CR-02).
                or suffix == f"{wr}_REDUCEDSUB"
                or suffix == f"{wr}_AEPBILLABLE"
                or suffix.startswith(f"{wr}_REDUCEDSUB_HELPER_")
                or suffix.startswith(f"{wr}_AEPBILLABLE_HELPER_")
                # Subproject B: per-claimer subcontractor primary keys
                # {wr}_REDUCEDSUB_USER_<claimer> / {wr}_AEPBILLABLE_USER_<claimer>
                # (attribution on — the production default). EXCLUDE_WRS is
                # production-active, so without these the operator's "do not
                # bill yet" intent silently failed for partitioned sub primary
                # files. Mirror of _key_matches_wr — the two MUST stay in sync.
                or suffix.startswith(f"{wr}_REDUCEDSUB_USER_")
                or suffix.startswith(f"{wr}_AEPBILLABLE_USER_")
            )
        
        # Remove groups that match any excluded WR
        groups = {k: v for k, v in groups.items() if not any(_key_matches_excluded_wr(k, wr) for wr in EXCLUDE_WRS)}
        excluded_count = before_exclude - len(groups)
        if excluded_count > 0:
            logging.info(f"🚫 EXCLUDE_WRS applied: {excluded_count} groups excluded ({','.join(EXCLUDE_WRS)}) - {len(groups)} groups remaining")
        else:
            logging.info(f"🚫 EXCLUDE_WRS specified but no matching groups found to exclude ({','.join(EXCLUDE_WRS)})")
    
    return groups

def validate_group_totals(groups):
    """Compute and validate totals per group, returning summary list of dicts."""
    summaries = []
    for key, rows in groups.items():
        total = sum(parse_price(r.get('Units Total Price')) for r in rows)
        summaries.append({'group_key': key, 'rows': len(rows), 'total': round(total,2)})
    return summaries

def safe_merge_cells(ws, range_str):
    """
    Safely merge cells, avoiding duplicates and overlaps that cause XML errors.
    
    Args:
        ws: The worksheet object
        range_str: The range string (e.g., 'A1:C3')
    
    Returns:
        bool: True if merge was successful, False if skipped
    """
    from openpyxl.utils import range_boundaries
    
    try:
        # Parse the requested range boundaries
        min_col, min_row, max_col, max_row = range_boundaries(range_str)
        if min_col is None or min_row is None or max_col is None or max_row is None:
            return False
        
        # Check for any overlapping or duplicate merged ranges
        for merged in list(ws.merged_cells.ranges):
            m_min_col, m_min_row, m_max_col, m_max_row = range_boundaries(str(merged))
            if m_min_col is None or m_min_row is None or m_max_col is None or m_max_row is None:
                continue
            
            # Check if ranges overlap (not just exact match)
            if not (max_col < m_min_col or min_col > m_max_col or
                    max_row < m_min_row or min_row > m_max_row):
                # Ranges overlap - skip to avoid XML corruption
                return False
        
        # Safe to merge - no overlaps detected
        ws.merge_cells(range_str)
        return True
    except Exception as e:
        logging.warning(f"Failed to merge cells {range_str}: {e}")
        return False


def _subcontractor_primary_variant_suffix(
    variant: str, claimer: str, wr_num: str, week_end_raw: str
) -> str:
    """Build the filename suffix for a subcontractor PRIMARY variant.

    Subproject B (2026-05-20): subcontractor primary files are
    partitioned by frozen primary claimer and named with the reserved
    ``_User_`` token (mirrors the primary-workflow convention).
    ``reduced_sub`` -> ``_ReducedSub_User_<sanitized>`` and
    ``aep_billable`` -> ``_AEPBillable_User_<sanitized>``.

    Raises ``ValueError`` if ``claimer`` is empty — production never
    hits this because ``resolve_claimer``'s ``use`` outcome always
    returns a non-empty name (falling back to ``effective_user`` /
    ``'Unknown Foreman'``). The raise mirrors the helper-shadow
    defensive raises and surfaces data drift loudly instead of
    producing a primary-looking filename that misroutes downstream.
    """
    if not claimer:
        logging.error(
            f"⚠️ {variant} variant row missing __current_foreman for "
            f"WR {wr_num} week {week_end_raw}; filename would be "
            f"ambiguous — raising to surface data drift."
        )
        raise ValueError(
            f"{variant} requires a non-empty claimer; got empty for "
            f"WR={wr_num} week={week_end_raw}"
        )
    if variant not in ('reduced_sub', 'aep_billable'):
        # Copilot: this helper is filename-identity logic. An unexpected
        # variant must raise rather than silently fall through to the
        # ``_ReducedSub`` token (which would misroute downstream cleanup /
        # hash identity matching if this helper were ever reused). Mirrors
        # the defensive-raise convention for new variant helpers
        # (Living Ledger 2026-05-15 rule 4).
        raise ValueError(
            f"_subcontractor_primary_variant_suffix: unexpected variant "
            f"{variant!r} (expected 'reduced_sub' or 'aep_billable') for "
            f"WR={wr_num} week={week_end_raw}"
        )
    claimer_sanitized = _RE_SANITIZE_IDENTIFIER.sub('_', claimer)[:50]
    token = '_AEPBillable' if variant == 'aep_billable' else '_ReducedSub'
    return f"{token}_User_{claimer_sanitized}"


def _vac_crew_variant_suffix(claimer: str, wr_num: str, week_end_raw: str) -> str:
    """Build the filename suffix for a per-claimer VAC crew file.

    Subproject C (2026-05-21): vac_crew files are partitioned by frozen
    vac-crew claimer and named ``_VacCrew_<sanitized name>``. Raises on an
    empty claimer (production never hits this — the emission falls back to
    'Unknown'); the raise surfaces data drift instead of an ambiguous name.
    """
    if not claimer:
        logging.error(
            f"⚠️ vac_crew variant row missing claimer for WR {wr_num} "
            f"week {week_end_raw}; filename would be ambiguous — raising to surface data drift."
        )
        raise ValueError(
            f"vac_crew requires a non-empty claimer; got empty for "
            f"WR={wr_num} week={week_end_raw}"
        )
    return f"_VacCrew_{_RE_SANITIZE_IDENTIFIER.sub('_', claimer)[:50]}"


def generate_excel(group_key, group_rows, snapshot_date, ai_analysis_results=None, data_hash=None):
    """
    FIXED: Generate a formatted Excel report for a group of rows.

    SPECIFIC FIXES IMPLEMENTED:
    - WR 90093002 Excel generation (complete implementation)
    - WR 89954686 specific handling
    - Proper error handling for worksheet objects
    - Complete daily data block generation
    - Safe cell merging to prevent XML errors
    - Improved Job # field detection with multiple column name variations
    """
    first_row = group_rows[0]
    
    # Parse the combined key format: "MMDDYY_WRNUMBER"
    if '_' in group_key:
        week_end_raw, wr_from_key = group_key.split('_', 1)
    else:
        # CRITICAL ERROR: Old format detected - this should never happen with fixed grouping
        error_msg = f"CRITICAL: Invalid group key format detected: '{group_key}'. Expected format: 'MMDDYY_WRNUMBER'."
        logging.error(error_msg)
        raise Exception(error_msg)
    
    # Use the current foreman (most recent) from the row data
    current_foreman = first_row.get('__current_foreman', 'Unknown_Foreman')
    
    # CRITICAL VALIDATION: Ensure grouping logic worked correctly
    wr_numbers = list(set(str(row.get('Work Request #', '')).split('.')[0] for row in group_rows if row.get('Work Request #')))
    
    # ABSOLUTE REQUIREMENT: Each group must contain EXACTLY ONE work request
    if len(wr_numbers) != 1:
        error_msg = f"FATAL ERROR: Group contains {len(wr_numbers)} work requests instead of 1: {wr_numbers}. Group key: {group_key}."
        logging.error(error_msg)
        raise Exception(error_msg)
    
    # SUCCESS: Exactly one work request in this group
    wr_num = wr_numbers[0]

    # Filesystem-safety: strip any path-traversal / separator characters
    # from the WR identifier before it reaches ``os.path.join`` and
    # ``workbook.save``. Numeric production WR#s pass through unchanged
    # (\w matches 0-9), so this is a no-op for realistic data and a
    # defense-in-depth guard against a pathological row value. Must use
    # the same regex used by the main-loop derivation site so
    # attachment / hash-history comparisons stay consistent.
    wr_num = _RE_SANITIZE_HELPER_NAME.sub('_', wr_num)[:50]

    # SPECIFIC FIX FOR WR 90093002 and WR 89954686
    if wr_num in ['90093002', '89954686']:
        logging.info(f"🔧 Applying specific fixes for WR# {wr_num}")
    
    # Get the calculated week ending date from the row data if available
    week_ending_date = first_row.get('__week_ending_date')
    if week_ending_date:
        week_end_display = week_ending_date.strftime('%m/%d/%y')
        # Update the raw format to match the calculated date
        week_end_raw = week_ending_date.strftime('%m%d%y')
        # Create subfolder for this week-ending date (YYYY-MM-DD format)
        week_folder_name = week_ending_date.strftime('%Y-%m-%d')
    else:
        # Fallback to the original format
        week_end_display = f"{week_end_raw[:2]}/{week_end_raw[2:4]}/{week_end_raw[4:]}"
        # Parse week_end_raw (MMDDYY) to create folder name
        try:
            fallback_date = datetime.datetime.strptime(week_end_raw, '%m%d%y')
            week_folder_name = fallback_date.strftime('%Y-%m-%d')
        except ValueError:
            week_folder_name = "unknown_week"
    
    # Create week-specific subfolder under OUTPUT_FOLDER
    week_output_folder = os.path.join(OUTPUT_FOLDER, week_folder_name)
    os.makedirs(week_output_folder, exist_ok=True)
    
    # Prefer 'Scope #' then fallback to 'Scope ID'
    scope_id = first_row.get('Scope #') or first_row.get('Scope ID', '')
    
    # Try multiple column name variations for Job # to handle different formats
    job_number = (first_row.get('Job #') or 
                  first_row.get('Job#') or 
                  first_row.get('Job Number') or 
                  first_row.get('JobNumber') or 
                  first_row.get('Job_Number') or 
                  first_row.get('JOB #') or 
                  first_row.get('JOB#') or 
                  first_row.get('job #') or 
                  first_row.get('job#') or 
                  '')
    
    # Log warning if Job # is missing
    if not job_number:
        available_cols = [k for k in first_row.keys() if not k.startswith('__')][:15]
        logging.warning(f"Job # not found for WR {wr_num}. Available columns: {available_cols}")
    
    # Use individual work request number for filename with timestamp for uniqueness
    timestamp = datetime.datetime.now().strftime('%H%M%S')
    
    # Variant-aware filename construction
    variant = first_row.get('__variant', 'primary')
    variant_suffix = ""

    # Phase 01 Plan 03 Task 2: subcontractor variant suffixes MUST be
    # checked BEFORE the legacy ``helper`` / ``vac_crew`` / ``primary``
    # branches so the variant-first ordering (D-09) is preserved. A
    # row tagged ``aep_billable_helper`` MUST produce the
    # ``_AEPBillable_Helper_<sanitized>`` filename, not plain
    # ``_Helper_<sanitized>`` with the AEPBillable token silently
    # dropped (which would break parser round-trip via
    # ``build_group_identity``). Helper-name sanitization mirrors
    # the producer site in ``group_source_rows`` — the regex is
    # idempotent so the double-apply is safe (D-22 / 2026-04-23 18:25).
    if variant in ('aep_billable', 'reduced_sub'):
        # Subproject B: partition by frozen primary claimer
        # (__current_foreman is the resolved claimer set in
        # group_source_rows). Helper-shadow branches below are
        # unchanged.
        variant_suffix = _subcontractor_primary_variant_suffix(
            variant,
            first_row.get('__current_foreman', ''),
            wr_num,
            week_end_raw,
        )
    elif variant == 'aep_billable_helper':
        helper_foreman = first_row.get('__helper_foreman', '')
        if not helper_foreman:
            # Phase 01 gap closure (REVIEW-WR-03): the upstream
            # ``_valid_helper_row`` gate in ``group_source_rows``
            # requires both ``helper_dept`` AND ``helper_foreman`` to
            # be truthy before adding a shadow-variant group key, so
            # this branch should never see an empty foreman in
            # production. If we ever hit it (refactor / data drift /
            # unexpected row mutation), the silent fallthrough produced
            # a primary-looking filename (no ``_AEPBillable_Helper_<name>``
            # suffix), which downstream parses as ``variant='primary'``
            # and routes the file to the wrong target sheet / wrong
            # identity tuple. Raise loudly to surface the drift instead.
            # Message body is _redact_exception_message-compatible
            # (WR + week + variant name; no foreman / dept / job —
            # those are PII per CLAUDE.md Living Ledger 2026-04-20 12:00).
            logging.error(
                f"⚠️ aep_billable_helper variant row missing "
                f"__helper_foreman for WR {wr_num} week {week_end_raw}; "
                f"filename would be ambiguous — raising to surface data drift."
            )
            raise ValueError(
                f"aep_billable_helper requires __helper_foreman; got empty "
                f"for WR={wr_num} week={week_end_raw}"
            )
        helper_sanitized = _RE_SANITIZE_HELPER_NAME.sub('_', helper_foreman)[:50]
        variant_suffix = f"_AEPBillable_Helper_{helper_sanitized}"
    elif variant == 'reduced_sub_helper':
        helper_foreman = first_row.get('__helper_foreman', '')
        if not helper_foreman:
            # WR-03 mirror of the aep_billable_helper defensive raise.
            # Same rationale, same PII-redact-compatible message body.
            logging.error(
                f"⚠️ reduced_sub_helper variant row missing "
                f"__helper_foreman for WR {wr_num} week {week_end_raw}; "
                f"filename would be ambiguous — raising to surface data drift."
            )
            raise ValueError(
                f"reduced_sub_helper requires __helper_foreman; got empty "
                f"for WR={wr_num} week={week_end_raw}"
            )
        helper_sanitized = _RE_SANITIZE_HELPER_NAME.sub('_', helper_foreman)[:50]
        variant_suffix = f"_ReducedSub_Helper_{helper_sanitized}"
    # WR-03 follow-up tech-debt: the legacy ``helper`` branch below has
    # the same shape (silent fallthrough if __helper_foreman is empty)
    # but is out of scope for this gap-closure plan per 01-REVIEW.md
    # WR-03 scope restriction. Adding the same defensive raise here in
    # a future plan is the recommended cleanup, but it requires a
    # separate regression test confirming the upstream
    # ``_valid_helper_row`` gate is the ONLY producer of the legacy
    # helper variant_suffix branch. The
    # ``test_legacy_helper_branch_does_not_raise_on_empty_foreman``
    # regression test in tests/test_subcontractor_pricing.py guards
    # against an accidental WR-03 fix broadening that would regress
    # the legacy helper variant production path.
    elif variant == 'helper':
        # Helper variant: include helper identifier in filename
        helper_foreman = first_row.get('__helper_foreman', '')
        if helper_foreman:
            # PERFORMANCE: Use pre-compiled regex for filename sanitization
            helper_sanitized = _RE_SANITIZE_HELPER_NAME.sub('_', helper_foreman)[:50]
            variant_suffix = f"_Helper_{helper_sanitized}"
    elif variant == 'vac_crew':
        # Subproject C: suffix is GATED on the kill switch.
        # Enabled mode: per-claimer _VacCrew_<name> so each foreman's file is
        # distinct and matches the Sites 1/2/3 identity tuple.
        # Disabled mode: exact legacy bare '_VacCrew' suffix — no claimer name
        # embedded, preserving byte-identical filenames with pre-C attachments.
        # NOTE: __vac_crew_name is intentionally NOT used as a fallback in
        # disabled mode; the legacy contract is a bare '_VacCrew' token, and
        # falling back to __vac_crew_name would produce _VacCrew_<name> in
        # disabled mode, which violates the disabled=legacy invariant.
        _vc_name = first_row.get('__current_foreman', '')
        if VAC_CREW_CLAIM_ATTRIBUTION_ENABLED and _vc_name:
            variant_suffix = _vac_crew_variant_suffix(_vc_name, wr_num, week_end_raw)
        else:
            # Disabled mode (or no claimer resolved) → exact legacy bare suffix.
            variant_suffix = '_VacCrew'
    elif variant == 'primary':
        # Subproject D (2026-05-25): partition the production primary
        # file by the FROZEN primary claimer (__current_foreman is the
        # resolved claimer set in group_source_rows' emission tuple).
        # GATED on the kill switch (mirrors the vac_crew branch above):
        #   • Enabled + claimer present -> _User_<sanitized claimer> so
        #     each claimer's file is distinct and round-trips through
        #     build_group_identity as ('primary', wr, week, claimer).
        #   • Disabled (or no claimer) -> exact legacy bare suffix '',
        #     preserving byte-identical filenames with pre-D attachments.
        # __current_foreman in disabled mode is effective_user (the
        # emission passes None -> `current_foreman or effective_user`),
        # but the kill-switch gate keeps the suffix bare in that case.
        _pf = first_row.get('__current_foreman', '')
        # PR #223 Codex-P1 follow-up: gate on the grouping mode too. In
        # RES_GROUPING_MODE='primary' the emission deliberately stays bare and
        # lumps every non-helper/non-sub foreman's rows into ONE workbook per
        # WR/week (partitioning by primary_foreman there is documented as
        # semantically wrong). A _User_<claimer> suffix would mislabel that
        # merged file and let row-order flip the attachment identity between
        # runs. Primary mode therefore stays bare here, matching the
        # already-mode-gated pre-pass + emission and Sites 1/2/3.
        if (
            PRIMARY_CLAIM_ATTRIBUTION_ENABLED
            and RES_GROUPING_MODE in ('helper', 'both')
            and _pf
        ):
            variant_suffix = (
                f"_User_{_RE_SANITIZE_IDENTIFIER.sub('_', _pf)[:50]}"
            )
        else:
            variant_suffix = ''

    # Phase 01 Plan 03 Task 2 (D-16): per-call missing-CU accumulator.
    # ``_resolve_row_price`` populates this Counter when a row's CU
    # code is absent from ``_SUBCONTRACTOR_RATES`` and the row keeps
    # its SmartSheet price (never zero-out, never raise). The Counter
    # is returned in the new 5-tuple shape (Blocker 4 contract) so
    # the main-loop caller can attribute missing CUs per source sheet
    # and emit the per-sheet WARNING (D-17).
    missing_cus: collections.Counter = collections.Counter()
    
    if SUPABASE_HASH_STORE_AUTHORITATIVE:
        # Sub-project E (2026-05-25): deterministic clean name. The durable
        # change-detection hash lives in billing_audit.group_content_hash,
        # so the filename carries IDENTITY ONLY (no _<timestamp>/_<hash>
        # tokens) and round-trips through build_group_identity unchanged.
        # Gated OFF by default — ships dormant until the store is validated.
        output_filename = f"WR_{wr_num}_WeekEnding_{week_end_raw}{variant_suffix}.xlsx"
    elif data_hash:
        # Use full 16-character hash (calculate_data_hash already truncates to 16)
        output_filename = f"WR_{wr_num}_WeekEnding_{week_end_raw}_{timestamp}{variant_suffix}_{data_hash}.xlsx"
    else:
        output_filename = f"WR_{wr_num}_WeekEnding_{week_end_raw}_{timestamp}{variant_suffix}.xlsx"
    final_output_path = os.path.join(week_output_folder, output_filename)

    if TEST_MODE:
        print(f"\n🧪 TEST MODE: Generating Excel file '{output_filename}'")
        print(f"   - Work Request: {wr_num}")
        print(f"   - Foreman: {current_foreman}")
        print(f"   - Timestamp: {timestamp}")
        print(f"   - Data Hash: {data_hash[:8] if data_hash else 'None'}")
        print(f"   🎯 NEW FILE POLICY: Always create fresh files")
    else:
        logging.info(f"📊 Generating Excel file '{output_filename}' for WR#{wr_num}")
        print(f"   - Week Ending: {week_end_display}")
        print(f"   - Row Count: {len(group_rows)}")

    workbook = openpyxl.Workbook()
    ws = workbook.active
    if ws is None:
        ws = workbook.create_sheet("Work Report")
    ws.title = "Work Report"

    # --- Formatting ---
    LINETEC_RED = 'C00000'
    LIGHT_GREY_FILL = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    RED_FILL = PatternFill(start_color=LINETEC_RED, end_color=LINETEC_RED, fill_type='solid')
    TITLE_FONT = Font(name='Calibri', size=20, bold=True)
    SUBTITLE_FONT = Font(name='Calibri', size=16, bold=True, color='404040')
    TABLE_HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    BLOCK_HEADER_FONT = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    BODY_FONT = Font(name='Calibri', size=11)
    SUMMARY_HEADER_FONT = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    SUMMARY_LABEL_FONT = Font(name='Calibri', size=10, bold=True)
    SUMMARY_VALUE_FONT = Font(name='Calibri', size=10)

    # Use explicit string for orientation for deterministic behavior
    ws.page_setup.orientation = 'landscape'
    try:
        ws.page_setup.paperSize = 9  # A4 paper size code
    except AttributeError:
        ws.page_setup.paperSize = 9  # Fallback for older versions
    ws.page_margins.left = 0.25; ws.page_margins.right = 0.25
    ws.page_margins.top = 0.5; ws.page_margins.bottom = 0.5

    # --- Branding and Titles ---
    current_row = 1
    try:
        img = Image(LOGO_PATH)
        img.height = 99
        img.width = 198
        ws.add_image(img, f'A{current_row}')
        for i in range(current_row, current_row+3): 
            ws.row_dimensions[i].height = 25
        current_row += 3
    except FileNotFoundError:
        safe_merge_cells(ws, f'A{current_row}:C{current_row+2}')
        ws[f'A{current_row}'] = "LINETEC SERVICES"
        ws[f'A{current_row}'].font = TITLE_FONT
        current_row += 3

    # CRITICAL FIX: Merge cells FIRST, then assign values
    safe_merge_cells(ws, f'D{current_row-2}:I{current_row-2}')
    ws[f'D{current_row-2}'] = 'WEEKLY UNITS COMPLETED PER SCOPE ID'
    ws[f'D{current_row-2}'].font = SUBTITLE_FONT
    ws[f'D{current_row-2}'].alignment = Alignment(horizontal='center', vertical='center')

    report_generated_time = datetime.datetime.now()
    safe_merge_cells(ws, f'D{current_row+1}:I{current_row+1}')
    ws[f'D{current_row+1}'] = f"Report Generated On: {report_generated_time.strftime('%m/%d/%Y %I:%M %p')}"
    ws[f'D{current_row+1}'].font = Font(name='Calibri', size=9, italic=True)
    ws[f'D{current_row+1}'].alignment = Alignment(horizontal='right')

    current_row += 3
    safe_merge_cells(ws, f'B{current_row}:D{current_row}')
    ws[f'B{current_row}'] = 'REPORT SUMMARY'
    ws[f'B{current_row}'].font = SUMMARY_HEADER_FONT
    ws[f'B{current_row}'].fill = RED_FILL
    ws[f'B{current_row}'].alignment = Alignment(horizontal='center')

    # Per Phase 01 Plan 03 Task 2 (D-16): resolve each row's price
    # EXACTLY ONCE through ``_resolve_row_price`` and stash the result
    # on the row as ``__resolved_price``. Both the summary "Total
    # Billed Amount" and ``write_day_block``'s per-row Pricing cell
    # read from the same stashed value so the workbook is internally
    # consistent and the per-call ``missing_cus`` Counter is not
    # double-incremented across summary + day-block iteration.
    # For legacy variants this is a no-op (helper returns
    # ``parse_price(row.get('Units Total Price'))``); for the new
    # variants it picks up the rate × qty values from
    # ``_SUBCONTRACTOR_RATES``. Mutating the input row matches the
    # existing pattern (``__variant`` / ``__current_foreman`` are
    # already added upstream in ``group_source_rows``).
    for _row in group_rows:
        _row['__resolved_price'] = _resolve_row_price(_row, variant, missing_cus)
    total_price = sum(row.get('__resolved_price', 0.0) for row in group_rows)
    ws[f'B{current_row+1}'] = 'Total Billed Amount:'
    ws[f'B{current_row+1}'].font = SUMMARY_LABEL_FONT
    ws[f'C{current_row+1}'] = total_price
    ws[f'C{current_row+1}'].font = SUMMARY_VALUE_FONT
    ws[f'C{current_row+1}'].alignment = Alignment(horizontal='right')
    ws[f'C{current_row+1}'].number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

    ws[f'B{current_row+2}'] = 'Total Line Items:'
    ws[f'B{current_row+2}'].font = SUMMARY_LABEL_FONT
    ws[f'C{current_row+2}'] = len(group_rows)
    ws[f'C{current_row+2}'].font = SUMMARY_VALUE_FONT
    ws[f'C{current_row+2}'].alignment = Alignment(horizontal='right')

    ws[f'B{current_row+3}'] = 'Billing Period:'
    ws[f'B{current_row+3}'].font = SUMMARY_LABEL_FONT
    
    # Calculate the proper week range (Monday to Sunday) for billing period
    if week_ending_date:
        week_start_date = week_ending_date - timedelta(days=6)  # Monday of that week
        billing_period = f"{week_start_date.strftime('%m/%d/%Y')} to {week_end_display}"
    else:
        # Fallback to using snapshot date if week ending date is not available
        billing_period = f"{snapshot_date.strftime('%m/%d/%Y')} to {week_end_display}"
    
    ws[f'C{current_row+3}'] = billing_period
    ws[f'C{current_row+3}'].font = SUMMARY_VALUE_FONT
    ws[f'C{current_row+3}'].alignment = Alignment(horizontal='right')

    safe_merge_cells(ws, f'F{current_row}:I{current_row}')
    ws[f'F{current_row}'] = 'REPORT DETAILS'
    ws[f'F{current_row}'].font = SUMMARY_HEADER_FONT
    ws[f'F{current_row}'].fill = RED_FILL
    ws[f'F{current_row}'].alignment = Alignment(horizontal='center')

    # Determine display values based on variant
    variant = first_row.get('__variant', 'primary')
    
    if variant == 'helper':
        # Helper variant: show helper foreman and helper-specific dept/job (REQUIRED)
        display_foreman = first_row.get('__helper_foreman', 'Unknown Helper')
        display_dept = first_row.get('__helper_dept', '')
        display_job = first_row.get('__helper_job', '')
    elif variant in ('reduced_sub_helper', 'aep_billable_helper'):
        # Subcontractor helper-shadow variants: the line items belong to the
        # helper, so Dept # / Job # MUST come from the helper fields
        # (operator requirement 2026-05-21 — helper files show Helper Dept #,
        # primary files show Dept #). Without this branch these variants fall
        # through to the ``else`` (primary) branch and display the PRIMARY
        # ``Dept #`` / ``Job #`` — the reported defect. The displayed Foreman,
        # however, stays ``current_foreman``: for these variants that is the
        # ATTRIBUTED helper (the file's partition key, set at the
        # ``keys_to_add`` site), NOT ``__helper_foreman`` (the current
        # "Foreman Helping?" value, which can diverge from the frozen
        # attribution under Phase 1.1). Folding these into the
        # ``variant == 'helper'`` branch above would regress the foreman, so
        # they are kept separate.
        display_foreman = current_foreman
        display_dept = first_row.get('__helper_dept', '')
        display_job = first_row.get('__helper_job', '')
    elif variant == 'vac_crew':
        # Enabled: show the ATTRIBUTED claimer (__current_foreman, the
        # partition key) so the displayed foreman matches the filename.
        # Disabled (legacy): show __vac_crew_name exactly as master did —
        # must NOT fall back to __current_foreman (which in disabled mode
        # may be the primary / Arrowhead foreman, not the VAC crew member).
        # dept/job remain VAC-crew-specific in all cases.
        if VAC_CREW_CLAIM_ATTRIBUTION_ENABLED:
            display_foreman = (
                first_row.get('__current_foreman')
                or first_row.get('__vac_crew_name', 'Unknown VAC Crew')
            )
        else:
            display_foreman = first_row.get('__vac_crew_name', 'Unknown VAC Crew')
        display_dept = first_row.get('__vac_crew_dept', '')
        display_job = first_row.get('__vac_crew_job', '')
    else:
        # Primary variant: show primary foreman with standard dept/job from row data
        display_foreman = current_foreman
        display_dept = first_row.get('Dept #', '')
        display_job = job_number

    details = [
        ("Foreman:", display_foreman),
        ("Work Request #:", wr_num),
        ("Scope ID #:", scope_id),
        ("Work Order #:", first_row.get('Work Order #', '')),
        ("Customer:", first_row.get('Customer Name', '')),
        ("Job #:", display_job)
    ]
    
    # Add Dept # to details if it exists
    if display_dept:
        details.insert(1, ("Dept #:", display_dept))
    
    # CRITICAL FIX: Merge cells FIRST, then assign value to top-left cell
    for i, (label, value) in enumerate(details):
        r = current_row + 1 + i
        ws[f'F{r}'] = label
        ws[f'F{r}'].font = SUMMARY_LABEL_FONT
        
        # Merge cells first - check for duplicates
        detail_merge_range = f'G{r}:I{r}'
        safe_merge_cells(ws, detail_merge_range)
        
        # Now assign value to the merged cell (top-left cell G)
        vcell = ws[f'G{r}']
        vcell.value = value
        vcell.font = SUMMARY_VALUE_FONT
        vcell.alignment = Alignment(horizontal='right')

    def write_day_block(start_row, day_name, date_obj, day_rows):
        """FIXED: Write daily data blocks with proper cell handling."""
        # Skip empty day blocks to prevent Excel corruption
        if not day_rows:
            return start_row
        
        # CRITICAL FIX: Merge cells FIRST, then assign value to top-left cell
        # Use safe merge to prevent duplicate ranges
        merge_range = f'A{start_row}:H{start_row}'
        safe_merge_cells(ws, merge_range)
        
        # Now assign value to the merged cell (top-left cell A1)
        day_header_cell = ws.cell(row=start_row, column=1)
        day_header_cell.value = f"{day_name} ({date_obj.strftime('%m/%d/%Y')})"  # type: ignore
        day_header_cell.font = BLOCK_HEADER_FONT
        day_header_cell.fill = RED_FILL
        day_header_cell.alignment = Alignment(horizontal='left', vertical='center')
        
        headers = ["Point Number", "Billable Unit Code", "Work Type", "Unit Description", "Unit of Measure", "# Units", "N/A", "Pricing"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row+1, column=col_num)
            cell.value = header  # type: ignore
            cell.font = TABLE_HEADER_FONT
            cell.fill = RED_FILL
            cell.alignment = Alignment(horizontal='center', wrap_text=True, vertical='center')

        total_price_day = 0.0
        for i, row_data in enumerate(day_rows):
            crow = start_row + 2 + i
            # Per Phase 01 Plan 03 Task 2: use the pre-resolved price
            # stashed by the outer ``generate_excel`` loop. Falling back
            # to ``parse_price`` here matches the legacy behaviour if a
            # row somehow lacks ``__resolved_price`` (defensive — should
            # never happen with the current call chain).
            if '__resolved_price' in row_data:
                price = row_data['__resolved_price']
            else:
                price = parse_price(row_data.get('Units Total Price'))
            
            # Safely parse quantity - extract only numbers
            qty_str = str(row_data.get('Quantity', '') or 0)
            # PERFORMANCE: Use pre-compiled regex for quantity normalization
            qty_str = _RE_EXTRACT_NUMBERS.sub('', qty_str)
            try:
                quantity = float(qty_str) if qty_str not in ('', '.', '-', '-.', '.-') else 0.0
            except Exception:
                quantity = 0.0
                
            total_price_day += price
            
            # Get the field values with debugging and fallbacks
            pole_num = (row_data.get('Pole #', '') or 
                       row_data.get('Point #', '') or 
                       row_data.get('Point Number', ''))
            
            cu_code = (row_data.get('CU', '') or 
                      row_data.get('Billable Unit Code', ''))
            
            work_type = row_data.get('Work Type', '')
            cu_description = (row_data.get('CU Description', '') or 
                             row_data.get('Unit Description', ''))
            unit_measure = (row_data.get('Unit of Measure', '') or 
                           row_data.get('UOM', ''))
            
            row_values = [pole_num, cu_code, work_type, cu_description, unit_measure, quantity, "", price]
            for col_num, value in enumerate(row_values, 1):
                cell = ws.cell(row=crow, column=col_num)
                cell.value = value
                cell.font = BODY_FONT
            ws.cell(row=crow, column=8).number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE

        total_row = start_row + 2 + len(day_rows)
        
        # CRITICAL FIX: Merge cells FIRST, then assign value to top-left cell
        # Use safe merge to prevent duplicate ranges
        total_merge_range = f'A{total_row}:G{total_row}'
        safe_merge_cells(ws, total_merge_range)
        
        # Now assign value to the merged cell
        total_label_cell = ws.cell(row=total_row, column=1)
        total_label_cell.value = "TOTAL"  # type: ignore
        total_label_cell.font = TABLE_HEADER_FONT
        total_label_cell.alignment = Alignment(horizontal='right')
        total_label_cell.fill = RED_FILL

        total_value_cell = ws.cell(row=total_row, column=8)
        total_value_cell.value = total_price_day  # type: ignore
        total_value_cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        total_value_cell.font = TABLE_HEADER_FONT
        total_value_cell.fill = RED_FILL

        return total_row + 2

    date_to_rows = collections.defaultdict(list)
    
    # Calculate the proper week range (Monday to Sunday) for filtering
    if week_ending_date:
        # Calculate Monday of the week (6 days before Sunday)
        week_start_date = week_ending_date - timedelta(days=6)  # Monday of that week
        week_end_date = week_ending_date  # Sunday of that week
        
        if TEST_MODE:
            print(f"\nWeek Range Filter: {week_start_date.strftime('%A, %m/%d/%Y')} to {week_end_date.strftime('%A, %m/%d/%Y')}")
    else:
        week_start_date = None
        week_end_date = None
    
    for row in group_rows:
        snap = row.get('Snapshot Date')
        try:
            dt = excel_serial_to_date(snap)
            if dt is None:
                if TEST_MODE:
                    logging.warning(f"Could not parse snapshot date '{snap}'")
                continue
            
            # Include snapshot dates that fall within the Monday-Sunday range
            if week_start_date and week_end_date:
                if week_start_date <= dt <= week_end_date:
                    date_to_rows[dt].append(row)
            else:
                date_to_rows[dt].append(row)
                    
        except (parser.ParserError, TypeError, ValueError) as e:
            if TEST_MODE:
                logging.warning(f"Could not parse snapshot date '{snap}': {e}")
            continue

    snapshot_dates = sorted(date_to_rows.keys())
    if TEST_MODE:
        print(f"\n📅 Found {len(snapshot_dates)} unique snapshot dates:")
        for d in snapshot_dates:
            print(f"   • {d.strftime('%A, %m/%d/%Y')}: {len(date_to_rows[d])} rows")
    
    day_names = {d: d.strftime('%A') for d in snapshot_dates}

    current_row += 7
    for d in snapshot_dates:
        day_rows = date_to_rows[d]
        current_row = write_day_block(current_row, day_names[d], d, day_rows)
        current_row += 1

    column_widths = {'A': 15, 'B': 20, 'C': 25, 'D': 45, 'E': 20, 'F': 10, 'G': 15, 'H': 15, 'I': 15}
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # NOTE: Footer code removed - was causing Excel XML corruption errors
    # Footer attributes (oddFooter.right.text, etc.) can create malformed XML
    # that triggers "We found a problem with some content" errors in Excel

    # Save the workbook
    workbook.save(final_output_path)

    if TEST_MODE:
        print(f"📄 Generated Excel file for inspection: '{output_filename}'")
        print(f"   - Total Amount: ${total_price:,.2f}")
        print(f"   - Daily Breakdown: {len(snapshot_dates)} days")
    else:
        logging.info(f"📄 Generated Excel: '{output_filename}'")

    # Phase 01 Plan 03 Task 2 / Blocker 4: extend the return shape to
    # a 5-tuple (excel_path, filename, wr_numbers, customer_name,
    # missing_cus). The two new trailing fields are absorbed by Plan 04
    # Task 2's upload-task builder. ``customer_name`` echoes the value
    # already used in the workbook's "Customer:" detail row; surfacing
    # it on the return tuple removes a duplicate ``first_row.get(...)``
    # lookup at the call site. ``missing_cus`` carries the per-call
    # subcontractor CU-fall-through codes (D-16) for per-sheet WARNING
    # aggregation (D-17) in the main loop.
    customer_name = first_row.get('Customer Name', '') or ''
    return final_output_path, output_filename, wr_numbers, customer_name, missing_cus

# --- TARGET SHEET MANAGEMENT ---

def create_target_sheet_map_for(client, sheet_id):
    """Build a sanitized ``{wr_num: target_row}`` map for any target
    sheet id.

    Phase 01 Plan 04 Task 1: extracted from the legacy
    ``create_target_sheet_map(client)`` so the dual-routing pipeline
    can build a SECOND target_map against ``SUBCONTRACTOR_PPP_SHEET_ID``
    for ``_ReducedSub`` / ``_ReducedSub_Helper_<name>`` uploads (D-12,
    SUB-03) while keeping the original TARGET_SHEET_ID map for every
    other variant.

    Critical invariants (D-22 / Living Ledger rounds 6, 7, 9):

    - Producer-side sanitization via ``_RE_SANITIZE_HELPER_NAME``
      applied at populate time — so consumer-side
      ``target_map[sanitized_wr]`` lookups in the main loop hit
      consistently across both target_maps (round-7 /
      2026-04-23 18:25).
    - Collision quarantine state (``_quarantined_keys`` /
      ``_seen_raw_for_key``) is FUNCTION-LOCAL: declared inside this
      helper's body, NOT at module scope (Plan 4 Warning 5). Each
      call owns its own quarantine sets so a duplicate WR# on one
      target sheet cannot poison the lookup table for another.
    - On collision, BOTH ambiguous raw values are removed from
      ``target_map`` and the sanitized key is quarantined (round-6
      P1). Loud not-found is strictly safer than silent wrong-row
      upload.

    Returns:
        Tuple of ``(target_map dict, target_sheet object)``. Mirrors
        the legacy return shape so the back-compat wrapper
        ``create_target_sheet_map(client)`` is a drop-in.
    """
    try:
        with sentry_sdk.start_span(op="smartsheet.api", name="Fetch target sheet for WR mapping") as span:
            target_sheet = client.Sheets.get_sheet(sheet_id)
            span.set_data("target_sheet_id", sheet_id)
            span.set_data("row_count", len(target_sheet.rows) if target_sheet.rows else 0)
        target_map: dict = {}

        # Find the Work Request # column
        wr_column_id = None
        for column in target_sheet.columns:
            if column.title == 'Work Request #':
                wr_column_id = column.id
                break

        if not wr_column_id:
            logging.error(
                f"Work Request # column not found in target sheet "
                f"{sheet_id}"
            )
            return {}, None

        # Map work request numbers to rows. Sanitize with the same
        # filesystem-safety regex used on source-row WR#s so downstream
        # ``target_map.get(sanitized_wr)`` lookups are consistent. For
        # realistic numeric WR#s this is a no-op; for any row with a
        # path-traversal-bearing WR the sanitized key matches the same
        # key the generation pipeline uses, so skip checks, upload
        # tasks, and attachment deletion all agree.
        #
        # Codex P2 guardrail: sanitize+truncate can (in principle)
        # collapse two distinct WR# cell values to the same key — e.g.
        # values that differ only in stripped characters, or IDs whose
        # first 50 chars happen to match. A silent overwrite would
        # retarget uploads / attachment deletes to the wrong row.
        # Track which raw value first produced a given sanitized key;
        # on collision, log a WARNING, quarantine the sanitized key,
        # and remove any existing mapping so both (or all) ambiguous
        # WRs are skipped deterministically until the target sheet is
        # deduplicated. A loud "not found in target sheet" warning
        # is strictly safer than a silent wrong-row upload.
        #
        # FUNCTION-LOCAL per Plan 04 Task 1 Warning 5: each invocation
        # owns its own quarantine sets so two target_map builds (one
        # for TARGET_SHEET_ID, one for SUBCONTRACTOR_PPP_SHEET_ID)
        # cannot poison each other. A module-level set would let a
        # duplicate WR# on one sheet remove the same WR# from the
        # other sheet's map — silently breaking dual-routing.
        _seen_raw_for_key: dict = {}
        _quarantined_keys: set = set()
        _collisions = 0
        for row in target_sheet.rows:
            for cell in row.cells:
                if cell.column_id == wr_column_id and cell.display_value:
                    raw_wr = str(cell.display_value).split('.')[0]
                    wr_num = _RE_SANITIZE_HELPER_NAME.sub('_', raw_wr)[:50]
                    if wr_num in _quarantined_keys:
                        # Already ambiguous — don't re-add under any
                        # raw value. Log once per collision instance
                        # so operators see every colliding row.
                        _collisions += 1
                        prior_raw = _seen_raw_for_key.get(
                            wr_num, '<quarantined>',
                        )
                        logging.warning(
                            f"⚠️ Target-sheet WR# collision (already quarantined): "
                            f"raw={raw_wr!r} also maps to sanitized key "
                            f"{wr_num!r} (prior seen: {prior_raw!r}) on sheet "
                            f"{sheet_id}. Uploads for this WR will be skipped "
                            f"until the target sheet is deduplicated."
                        )
                    elif wr_num in target_map:
                        prior_raw = _seen_raw_for_key.get(wr_num, '<unknown>')
                        if prior_raw != raw_wr:
                            # Collision: quarantine the key to prevent
                            # uploads from silently targeting the wrong
                            # row. The upload site's
                            # ``if wr_num in target_map`` check will
                            # then correctly return False for BOTH
                            # WRs, and the existing "not found in
                            # target sheet" warning fires so operators
                            # know to audit the target sheet. Removing
                            # both is strictly safer than keeping one
                            # — a silent wrong-row upload corrupts
                            # Smartsheet attachments; a loud
                            # not-found failure prompts cleanup.
                            _collisions += 1
                            del target_map[wr_num]
                            _quarantined_keys.add(wr_num)
                            logging.warning(
                                f"⚠️ Target-sheet WR# collision after sanitization "
                                f"on sheet {sheet_id}: raw={raw_wr!r} and prior "
                                f"raw={prior_raw!r} both map to sanitized key "
                                f"{wr_num!r}; QUARANTINING the key from "
                                f"target_map. Uploads for both WRs will be "
                                f"skipped until the target sheet is "
                                f"deduplicated — a 'not found in target "
                                f"sheet' warning will follow for each."
                            )
                    else:
                        target_map[wr_num] = row
                        _seen_raw_for_key[wr_num] = raw_wr
                    break

        if _collisions:
            logging.warning(
                f"⚠️ Target sheet {sheet_id} map had {_collisions} "
                f"sanitized-WR# collision event(s) across "
                f"{len(_quarantined_keys)} quarantined key(s) — "
                f"affected uploads will be skipped with 'not found in "
                f"target sheet' warnings."
            )
        logging.info(
            f"Created target sheet map for {sheet_id} with "
            f"{len(target_map)} work requests"
        )
        return target_map, target_sheet

    except Exception as e:
        logging.error(
            f"Failed to create target sheet map for {sheet_id}: "
            f"{_redact_exception_message(e)}"
        )
        return {}, None


def create_target_sheet_map(client):
    """Back-compat wrapper around ``create_target_sheet_map_for``.

    Preserved so existing call sites and tests continue to operate
    against the primary ``TARGET_SHEET_ID`` without churn. New code
    that needs a different sheet should call
    ``create_target_sheet_map_for(client, sheet_id)`` directly.

    Returns:
        Tuple of (target_map dict, target_sheet object) for reuse in cleanup.
    """
    return create_target_sheet_map_for(client, TARGET_SHEET_ID)


def _build_upload_tasks_for_group(
    *,
    variant,
    wr_num,
    target_map,
    target_map_ppp,
    excel_path,
    filename,
    identifier,
    file_identifier,
    data_hash,
    week_raw,
    group_key,
):
    """Build the list of upload-task dicts for a single generated
    Excel file.

    Phase 01 Plan 04 Task 2: routes ``reduced_sub`` /
    ``reduced_sub_helper`` to BOTH ``TARGET_SHEET_ID`` and
    ``SUBCONTRACTOR_PPP_SHEET_ID`` (D-12 / SUB-03); every other
    variant routes to ``TARGET_SHEET_ID`` only. Each task carries
    its own ``target_sheet_id`` so the ``_upload_one`` worker
    resolves uploads to the correct sheet without consulting a
    global.

    Sanitization parity (Warning 9): the same ``wr_num`` value is
    reused for both ``target_map[wr_num]`` and
    ``target_map_ppp[wr_num]`` lookups. Because
    ``_RE_SANITIZE_HELPER_NAME`` is idempotent and both maps are
    populated using it at producer-side (see
    ``create_target_sheet_map_for``), this single sanitisation upstream
    is sufficient — no re-sanitisation is needed at the consumer.

    Independent quarantine (Warning 5 / round-6): the two target_maps
    own their own quarantine sets. If a WR# is quarantined on one
    sheet, the lookup returns False there but may still succeed on
    the other sheet — producing exactly the upload behaviour
    operators expect (uploads only to sheets whose WR# is
    unambiguous, with operator-visible WARNINGs on the quarantined
    side).

    Args:
        variant: One of ``primary`` / ``helper`` / ``vac_crew`` /
            ``aep_billable`` / ``aep_billable_helper`` /
            ``reduced_sub`` / ``reduced_sub_helper``.
        wr_num: Sanitised WR# (already passed through
            ``_RE_SANITIZE_HELPER_NAME`` at the main-loop derivation
            site).
        target_map: Primary ``TARGET_SHEET_ID`` mapping
            (``{sanitized_wr: row}``).
        target_map_ppp: Secondary ``SUBCONTRACTOR_PPP_SHEET_ID``
            mapping; empty / unreachable → reduced-sub routing
            degrades to single-target with a WARNING.
        excel_path / filename / identifier / file_identifier /
            data_hash / week_raw / group_key: Pass-through payload
            consumed by ``_upload_one``.

    Returns:
        A list of upload-task dicts. ``[]`` if ``wr_num`` is blank
        or neither map carries the WR.
    """
    if not wr_num:
        return []

    upload_tasks: list = []

    # Primary leg — every variant routes here, including
    # reduced_sub / reduced_sub_helper. The primary leg always runs
    # first so a missing-WR warning consistently mentions
    # TARGET_SHEET_ID first.
    primary_present = wr_num in target_map
    if primary_present:
        upload_tasks.append({
            'excel_path': excel_path,
            'filename': filename,
            'wr_num': wr_num,
            'target_row': target_map[wr_num],
            'target_sheet_id': TARGET_SHEET_ID,
            'variant': variant,
            'identifier': identifier,
            'file_identifier': file_identifier,
            'data_hash': data_hash,
            'week_raw': week_raw,
            'group_key': group_key,
        })
    else:
        # WR not on TARGET_SHEET_ID. Name the sheet id explicitly so
        # operators know which sheet to dedup / add the WR to (or to
        # check the source-side quarantine). Fires for both the
        # "map populated but WR absent" case and the "map empty —
        # PPP sheet unreachable / TEST_MODE" degraded case, since
        # both produce the same operator-actionable surface.
        logging.warning(
            f"⚠️ Work request {wr_num} not found in target sheet "
            f"{TARGET_SHEET_ID}"
        )

    # Second leg — only for reduced_sub variants per D-12 / SUB-03.
    if variant in ('reduced_sub', 'reduced_sub_helper'):
        if primary_present and wr_num in target_map_ppp:
            upload_tasks.append({
                'excel_path': excel_path,
                'filename': filename,
                'wr_num': wr_num,
                'target_row': target_map_ppp[wr_num],
                'target_sheet_id': SUBCONTRACTOR_PPP_SHEET_ID,
                'variant': variant,
                'identifier': identifier,
                'file_identifier': file_identifier,
                'data_hash': data_hash,
                'week_raw': week_raw,
                'group_key': group_key,
            })
        elif primary_present:
            # WR not on PPP sheet. Degrade gracefully — the primary
            # leg still runs (if it found the WR) and operators see
            # a sheet-specific WARNING with the PPP sheet id so they
            # can tell at a glance which sheet to update.
            logging.warning(
                f"⚠️ Work request {wr_num} not found in "
                f"subcontractor PPP target sheet "
                f"{SUBCONTRACTOR_PPP_SHEET_ID}"
            )

    return upload_tasks


# Modified By cache loading removed - using direct column assignment only


def _build_synthetic_rows():
    """Build an in-memory synthetic dataset for TEST_MODE runs without an API token."""
    base_week_end = datetime.datetime.now()
    # Snap week ending to coming Sunday for consistency
    base_week_end = base_week_end + datetime.timedelta(days=(6 - base_week_end.weekday()))
    week_end_iso = base_week_end.strftime('%Y-%m-%d')
    rows = []
    wrs = ['90093002', '89708709']
    foremen = ['Alice Foreman', 'Bob Foreman']
    daily_prices = [1200.50, 800.00, 950.75, 0, 1300.25, 600.00, 1450.00]
    for idx, wr in enumerate(wrs):
        foreman = foremen[idx]
        for offset, price in enumerate(daily_prices):
            snap_date = (base_week_end - datetime.timedelta(days=(6 - offset)))
            row = {
                'Work Request #': wr,
                'Weekly Reference Logged Date': week_end_iso,  # same week ending for all
                'Snapshot Date': snap_date.strftime('%Y-%m-%d'),
                'Units Total Price': f"${price:,.2f}",
                'Quantity': str(1 + (offset % 3)),
                'Units Completed?': True,
                'Foreman': foreman,
                'CU': f"CU{100+offset}",
                'CU Description': f"Synthetic Work Item {offset+1}",
                'Unit of Measure': 'EA',
                'Pole #': f"P-{offset+1:03d}",
                'Work Type': 'Maintenance',
                'Scope #': f"SCP-{wr[-3:]}"
            }
            # Include a zero price row intentionally (price==0) to confirm exclusion
            rows.append(row)
    return rows


def _run_synthetic_test_mode(session_start):
    """Execute the synthetic TEST_MODE path. Returns number of files generated."""
    logging.info("🧪 TEST_MODE without SMARTSHEET_API_TOKEN: using synthetic in-memory dataset")
    synthetic_rows = _build_synthetic_rows()
    logging.info(f"Synthetic rows prepared: {len(synthetic_rows)} raw rows")
    # Apply normal grouping logic (filtering happens inside grouping)
    groups = group_source_rows(synthetic_rows)
    logging.info(f"Synthetic grouping produced {len(groups)} group(s)")
    snapshot_date = datetime.datetime.now()
    generated_files_count = 0
    for group_key, group_rows in groups.items():
        try:
            data_hash = calculate_data_hash(group_rows)
            # Phase 01 Plan 03 Task 2 / Blocker 4: unpack
            # the new 5-tuple shape. Synthetic path doesn't
            # consume ``customer_name`` / ``missing_cus``
            # (no per-sheet WARNING context here), but the
            # unpack MUST match so a contract drift is
            # surfaced loudly rather than silently dropped.
            (
                _excel_path,
                filename,
                _wr_numbers,
                _customer_name,
                _missing_cus,
            ) = generate_excel(
                group_key, group_rows, snapshot_date,
                data_hash=data_hash,
            )
            generated_files_count += 1
            logging.info(f"🧪 Synthetic Excel generated: {filename} ({len(group_rows)} rows)")
        except Exception as e:
            logging.error(f"Synthetic group failure {group_key}: {e}")
    session_duration = datetime.datetime.now() - session_start
    logging.info(f"🧪 Synthetic session complete: {generated_files_count} file(s) in {session_duration}")
    return generated_files_count


# --- MAIN EXECUTION ---

# Sentry Crons monitor schedule. This cron VALUE must stay byte-for-byte
# identical to the weekday ``schedule.cron`` in
# .github/workflows/weekly-excel-generation.yml. GitHub Actions evaluates every
# ``schedule:`` cron in UTC, so the monitor ``timezone`` below MUST be "UTC".
# (Mislabeling it "America/Chicago" made Sentry expect each check-in 5-6h late
# and fire a perpetual "missed check-in" outage — GENERATE-WEEKLY-EXCEL-6V.)
_CRON_MONITOR_SCHEDULE = "0 13,15,17,19,21,23,1 * * 1-5"


def _build_cron_monitor_config() -> "MonitorConfig":
    """Return the Sentry Crons ``monitor_config`` for the weekly billing job.

    Pure (no I/O); extracted so the schedule/timezone contract can be unit
    tested. ``timezone`` MUST equal the timezone GitHub Actions evaluates the
    workflow ``schedule:`` cron in (UTC); see ``_CRON_MONITOR_SCHEDULE``.
    """
    return {
        "schedule": {"type": "crontab", "value": _CRON_MONITOR_SCHEDULE},
        "timezone": "UTC",
        "checkin_margin": 5,
        "max_runtime": 180,
        "failure_issue_threshold": 1,
        "recovery_threshold": 1,
    }


def _sentry_cron_checkin_start(monitor_slug):
    """Send a Sentry cron 'in_progress' check-in. Returns the check-in id or None.

    Extracted from ``main()`` to reduce its cyclomatic complexity; behavior
    (including swallow-and-log on failure) is preserved verbatim.
    """
    if not SENTRY_DSN:
        return None
    try:
        return capture_checkin(
            monitor_slug=monitor_slug,
            status=MonitorStatus.IN_PROGRESS,
            monitor_config=_build_cron_monitor_config(),
        )
    except Exception as exc:
        logging.warning(f"⚠️ Sentry cron check-in (in_progress) failed: {exc}")
        return None


def _set_sentry_session_tags(session_start):
    """Apply session-level Sentry tags. No-op when Sentry is not configured."""
    if not SENTRY_DSN:
        return
    scope = sentry_sdk.get_isolation_scope()
    scope.set_tag("session_start", session_start.isoformat())
    scope.set_tag("test_mode", str(TEST_MODE))
    scope.set_tag("github_actions", str(GITHUB_ACTIONS_MODE))


def main():  # pyright: ignore[reportGeneralTypeIssues]
    """Main execution function with all fixes implemented.

    NOTE: Pyright reports ``reportGeneralTypeIssues`` ("Code is too complex
    to analyze") on this function because it exceeds the analyzer's internal
    branch/path budget. The behavior is correct and exercised by CI; the
    warning is suppressed at the def line so type-checking of the rest of
    the module remains clean. A full refactor into subroutines is tracked
    separately — many of the local variables here participate in the
    ``except``/``finally`` blocks at the bottom, so extraction requires
    care to preserve the existing error-reporting + cron-checkin contract.
    """
    session_start = datetime.datetime.now()
    generated_files_count = 0
    generated_filenames = []  # Track exact filenames created this session
    # Sentry session-transaction handle. Hoisted to the top of main() so the
    # except/finally blocks at the bottom of this function always see _txn
    # bound. Synthetic TEST_MODE returns and the "no SMARTSHEET_API_TOKEN"
    # raise both short-circuit past the in-place start-transaction block
    # further down, which would otherwise leave _txn unbound and turn any
    # main() exit through finally into an UnboundLocalError.
    _txn = None

    # Sentry cron check-in: signal "in_progress" at session start
    _cron_monitor_slug = os.getenv("SENTRY_CRON_MONITOR_SLUG", "weekly-excel-generation")
    _cron_checkin_id = _sentry_cron_checkin_start(_cron_monitor_slug)

    try:
        # Set Sentry context (SDK 2.x: top-level API)
        _set_sentry_session_tags(session_start)

        logging.info("🚀 Starting Weekly PDF Generator with Complete Fixes")
        
        # Initialize Smartsheet client or fall back to synthetic data in TEST_MODE
        if not API_TOKEN:
            if not TEST_MODE:
                raise Exception("SMARTSHEET_API_TOKEN not configured")
            _run_synthetic_test_mode(session_start)
            return
        
        client = smartsheet.Smartsheet(API_TOKEN)
        client.errors_as_exceptions(True)

        # ── Phase 2 Plan 03: isolated garbage-attachment remediation mode ──
        # REMEDIATE_CLAIMERS defaults OFF ('0') — never fires on scheduled cron.
        # When active, the sweep runs and main() returns immediately (isolation:
        # no Excel generation occurs in this session).
        if REMEDIATE_CLAIMERS:
            logging.info(
                f"🧹 REMEDIATE_CLAIMERS=True — running isolated claimer "
                f"remediation sweep (dry_run={REMEDIATION_DRY_RUN}, "
                f"window_weeks={REMEDIATION_WINDOW_WEEKS})"
            )
            run_claimer_remediation(
                client,
                dry_run=REMEDIATION_DRY_RUN,
                window_weeks=REMEDIATION_WINDOW_WEEKS,
                valid_wr_weeks=None,  # isolated path: no live-identity set
            )
            return

        # ── Start root Sentry transaction for full session tracing ──
        # _txn handle is already initialized to None at the top of main().
        if SENTRY_DSN:
            _txn = sentry_sdk.start_transaction(
                op="session",
                name="weekly-excel-generation",
                description="Full weekly Excel generation session",
            )
            _txn.__enter__()
            _txn.set_data("test_mode", TEST_MODE)
            _txn.set_data("github_actions", GITHUB_ACTIONS_MODE)

        # #7 - milestone structured log: run start (counts/booleans only)
        _sentry_log_event(
            "info",
            "weekly run started",
            test_mode=TEST_MODE,
            github_actions=GITHUB_ACTIONS_MODE,
        )

        # ── Source sheet discovery (includes folder discovery on cache miss) ──
        _phase_start = datetime.datetime.now()
        logging.info(f"\n{'='*60}")
        logging.info("📊 PHASE 1: Discovering source sheets...")
        logging.info(f"{'='*60}")
        sentry_add_breadcrumb("discovery", "Starting source sheet discovery")
        with sentry_sdk.start_span(op="smartsheet.discovery", name="Discover and validate source sheets") as span:
            source_sheets = discover_source_sheets(client)
            span.set_data("sheets_discovered", len(source_sheets) if source_sheets else 0)
        
        if not source_sheets:
            raise Exception("No valid source sheets found")
        
        _phase_elapsed = (datetime.datetime.now() - _phase_start).total_seconds()
        logging.info(f"⚡ Phase 1 complete: {len(source_sheets)} sheets discovered in {_phase_elapsed:.1f}s")
        sentry_add_breadcrumb("discovery", f"Discovered {len(source_sheets)} source sheets", data={"count": len(source_sheets)})
        
        # Get all source rows
        _phase_start = datetime.datetime.now()
        logging.info(f"\n{'='*60}")
        logging.info("📋 PHASE 2: Fetching source data...")
        logging.info(f"{'='*60}")
        with sentry_sdk.start_span(op="smartsheet.fetch_rows", name="Fetch all source rows from Smartsheet") as span:
            all_rows = get_all_source_rows(client, source_sheets)
            span.set_data("source_sheets_count", len(source_sheets))
            span.set_data("rows_fetched", len(all_rows) if all_rows else 0)
        
        if not all_rows:
            raise Exception("No valid data rows found")
        
        _phase_elapsed = (datetime.datetime.now() - _phase_start).total_seconds()
        logging.info(f"⚡ Phase 2 complete: {len(all_rows)} rows fetched from {len(source_sheets)} sheets in {_phase_elapsed:.1f}s")
        sentry_add_breadcrumb("data", f"Fetched {len(all_rows)} source rows from {len(source_sheets)} sheets", data={
            "row_count": len(all_rows),
            "sheet_count": len(source_sheets),
        })
        
        # Initialize audit system
        audit_system = None
        audit_results = {}
        if AUDIT_SYSTEM_AVAILABLE and not DISABLE_AUDIT_FOR_TESTING:
            try:
                sentry_add_breadcrumb("audit", "Starting billing audit", data={"skip_cell_history": SKIP_CELL_HISTORY})
                with sentry_sdk.start_span(op="audit.financial", name="Run billing audit on source data") as audit_span:
                    audit_system = BillingAudit(client, skip_cell_history=SKIP_CELL_HISTORY)
                    audit_results = audit_system.audit_financial_data(source_sheets, all_rows)
                    audit_span.set_data("risk_level", audit_results.get('summary', {}).get('risk_level', 'UNKNOWN'))
                    audit_span.set_data("total_anomalies", audit_results.get('summary', {}).get('total_anomalies', 0))
                logging.info(f"🔍 Audit complete - Risk level: {audit_results.get('summary', {}).get('risk_level', 'UNKNOWN')}")
                sentry_add_breadcrumb("audit", "Audit completed", data={
                    "risk_level": audit_results.get('summary', {}).get('risk_level', 'UNKNOWN'),
                    "total_anomalies": audit_results.get('summary', {}).get('total_anomalies', 0)
                })
            except Exception as e:
                logging.warning(f"⚠️ Audit system error: {e}")
                sentry_capture_with_context(
                    exception=e,
                    context_name="audit_system_error",
                    context_data={
                        "source_sheets_count": len(source_sheets),
                        "total_rows": len(all_rows),
                        "skip_cell_history": SKIP_CELL_HISTORY,
                        "error_type": type(e).__name__,
                        "error_message": _redact_exception_message(e),
                    },
                    tags={"error_location": "audit_system", "subsystem": "billing_audit"},
                    fingerprint=["audit-system", type(e).__name__]
                )
        else:
            logging.info("🚀 Audit system disabled for testing")

    # Group rows by work request and week ending
        logging.info("📂 Grouping data...")
        with sentry_sdk.start_span(op="data.grouping", name="Group source rows by WR/week/variant") as span:
            groups = group_source_rows(all_rows)
            span.set_data("input_rows", len(all_rows))
            span.set_data("groups_created", len(groups) if groups else 0)

        # Optional full/partial hash reset purge BEFORE processing groups if requested
        if RESET_HASH_HISTORY or RESET_WR_LIST:
            with sentry_sdk.start_span(op="smartsheet.purge", name="Purge existing hashed outputs") as span:
                if RESET_WR_LIST:
                    logging.info(f"🧨 Hash reset requested for specific WRs: {sorted(list(RESET_WR_LIST))}")
                    span.set_data("purge_type", "wr_subset")
                    span.set_data("wr_count", len(RESET_WR_LIST))
                    purge_existing_hashed_outputs(client, TARGET_SHEET_ID, RESET_WR_LIST, TEST_MODE)
                else:
                    logging.info("🧨 Global hash reset requested (RESET_HASH_HISTORY=1)")
                    span.set_data("purge_type", "global")
                    purge_existing_hashed_outputs(client, TARGET_SHEET_ID, None, TEST_MODE)
            # After purge, any regenerated files get new timestamp+hash filenames and re-upload
        
        if not groups:
            raise Exception("No valid groups created")
        
        logging.info(f"📈 Found {len(groups)} work request groups to process")
        sentry_add_breadcrumb("grouping", f"Created {len(groups)} groups from {len(all_rows)} rows", data={
            "group_count": len(groups),
            "row_count": len(all_rows),
        })
        if MAX_GROUPS and len(groups) > MAX_GROUPS:
            logging.info(f"✂️ Limiting processing to first {MAX_GROUPS} groups for test run")
            groups = dict(list(groups.items())[:MAX_GROUPS])
        
        # Process groups
        snapshot_date = datetime.datetime.now()
        
        # Create target sheet map for production uploads.
        target_map = {}
        _target_sheet_obj = None  # Cached for cleanup to avoid redundant API call
        if not TEST_MODE:
            with sentry_sdk.start_span(op="smartsheet.target_map", name="Create target sheet map for uploads") as span:
                target_map, _target_sheet_obj = (
                    create_target_sheet_map_for(client, TARGET_SHEET_ID)
                )
                span.set_data("wr_count", len(target_map))

        # Phase 01 Plan 04 Task 1: build a SECOND target_map for the
        # subcontractor PPP sheet. Only ``_ReducedSub`` /
        # ``_ReducedSub_Helper_<name>`` upload tasks consume this map
        # (D-12 / SUB-03); ``primary`` / ``helper`` / ``vac_crew`` /
        # ``aep_billable`` continue to route through ``target_map``
        # alone, so a missing or unreachable PPP sheet only degrades
        # the second leg of the reduced-sub fan-out — the rest of the
        # pipeline is unaffected.
        #
        # Per Plan 04 acceptance criterion: only attempt the build
        # when the kill switch is on AND a distinct sheet id was
        # configured. Defense against an operator setting
        # ``SUBCONTRACTOR_PPP_SHEET_ID=<same as TARGET_SHEET_ID>``
        # which would otherwise cause every reduced-sub upload to
        # double-attach to the SAME target row.
        target_map_ppp: dict = {}
        _target_sheet_ppp_obj = None
        if (not TEST_MODE
                and SUBCONTRACTOR_RATE_VARIANTS_ENABLED
                and SUBCONTRACTOR_PPP_SHEET_ID
                and SUBCONTRACTOR_PPP_SHEET_ID != TARGET_SHEET_ID):
            try:
                with sentry_sdk.start_span(op="smartsheet.target_map_ppp", name="Create PPP target sheet map") as span:
                    target_map_ppp, _target_sheet_ppp_obj = create_target_sheet_map_for(client, SUBCONTRACTOR_PPP_SHEET_ID)
                    span.set_data("wr_count", len(target_map_ppp))
                logging.info(
                    f"🎯 Subcontractor PPP target sheet: "
                    f"{SUBCONTRACTOR_PPP_SHEET_ID}, "
                    f"{len(target_map_ppp)} WR# entries mapped"
                )
            except Exception as _ppp_exc:
                # Fail-safe: if the PPP sheet is unreachable (access
                # revoked, renamed, deleted), log + degrade to single-
                # sheet routing for this run. Per D-22 / Living
                # Ledger 2026-04-23 12:00, the exception body is
                # sanitised via ``_redact_exception_message`` before
                # reaching Sentry's ``event['contexts']``.
                logging.error(
                    f"Failed to load subcontractor PPP target sheet "
                    f"{SUBCONTRACTOR_PPP_SHEET_ID}: "
                    f"{_redact_exception_message(_ppp_exc)}"
                )
                target_map_ppp = {}
                _target_sheet_ppp_obj = None

        # PERFORMANCE: Pre-fetch all target row attachments into cache to eliminate
        # redundant per-row API calls in _has_existing_week_attachment and delete_old_excel_attachments.
        # Each row's attachments are fetched once here instead of 2-3 times in the group loop.
        attachment_cache = {}  # row_id -> list of attachment objects
        target_map_to_prefetch = {}
        if target_map and not TEST_MODE:
            target_map_to_prefetch = target_map
            # Pre-flight session-budget guard: if discovery + row fetch already consumed most
            # of TIME_BUDGET_MINUTES, skip pre-fetch entirely so we have time for generation.
            # Reserve ATTACHMENT_PREFETCH_GENERATION_HEADROOM_MIN beyond the pre-fetch budget
            # so we don't end up with exactly enough time to pre-fetch and then zero time to
            # generate — that would recreate the original incident's zero-output failure mode.
            # Per-row fallback paths handle an empty cache transparently.
            if TIME_BUDGET_MINUTES and GITHUB_ACTIONS_MODE:
                _pre_elapsed_min = (datetime.datetime.now() - session_start).total_seconds() / 60.0
                _remaining_min = TIME_BUDGET_MINUTES - _pre_elapsed_min
                _required_remaining_min = ATTACHMENT_PREFETCH_MAX_MINUTES + ATTACHMENT_PREFETCH_GENERATION_HEADROOM_MIN
                if _remaining_min <= _required_remaining_min:
                    logging.warning(
                        f"⏩ Skipping attachment pre-fetch: {_pre_elapsed_min:.1f}min already elapsed, "
                        f"only {_remaining_min:.1f}min left in session budget "
                        f"(need > {_required_remaining_min}min = "
                        f"{ATTACHMENT_PREFETCH_MAX_MINUTES}min pre-fetch budget + "
                        f"{ATTACHMENT_PREFETCH_GENERATION_HEADROOM_MIN}min generation headroom). "
                        f"Attachment lookups will fall back to per-row fetches during generation."
                    )
                    sentry_add_breadcrumb(
                        "prefetch_skipped",
                        f"Pre-fetch skipped, {_remaining_min:.1f}min remaining",
                        level="warning",
                        data={
                            "elapsed_min": round(_pre_elapsed_min, 1),
                            "remaining_min": round(_remaining_min, 1),
                            "prefetch_budget_min": ATTACHMENT_PREFETCH_MAX_MINUTES,
                            "generation_headroom_min": ATTACHMENT_PREFETCH_GENERATION_HEADROOM_MIN,
                            "required_remaining_min": _required_remaining_min,
                        },
                    )
                    target_map_to_prefetch = {}

        if target_map_to_prefetch:
            with sentry_sdk.start_span(op="smartsheet.attachment_prefetch", name="Pre-fetch row attachments") as span:
                logging.info(f"🚀 Starting parallel attachment pre-fetch with {PARALLEL_WORKERS} workers for {len(target_map_to_prefetch)} target rows (max {ATTACHMENT_PREFETCH_MAX_MINUTES}min)...")
                _att_start = datetime.datetime.now()

                def _fetch_row_attachments(row_item):
                    # row_item is (wr_num, target_row); only target_row is needed.
                    _, target_row = row_item
                    max_retries = 4
                    for attempt in range(max_retries):
                        try:
                            atts = client.Attachments.list_row_attachments(TARGET_SHEET_ID, target_row.id).data
                            return (target_row.id, atts)
                        except (ss_exc.RateLimitExceededError,) as e:
                            if attempt < max_retries - 1:
                                backoff = 15 * (attempt + 1)  # 15s, 30s, 45s for rate limits
                                logging.warning(f"⚠️ Rate limited on attachment fetch for row {target_row.id}, backoff {backoff}s (attempt {attempt+1}/{max_retries})")
                                time.sleep(backoff)
                            else:
                                logging.warning(f"⚠️ Attachment fetch failed after {max_retries} rate-limit retries for row {target_row.id}")
                                return (target_row.id, [])
                        except (ss_exc.UnexpectedErrorShouldRetryError, ss_exc.InternalServerError, ss_exc.ServerTimeoutExceededError) as e:
                            if attempt < max_retries - 1:
                                backoff = 2 ** attempt + 0.5
                                logging.warning(f"⚠️ Attachment fetch retry {attempt+1}/{max_retries} for row {target_row.id} ({type(e).__name__}), backoff {backoff:.1f}s")
                                time.sleep(backoff)
                            else:
                                logging.warning(f"⚠️ Attachment fetch failed after {max_retries} attempts for row {target_row.id}: {type(e).__name__}")
                                return (target_row.id, [])
                        except Exception as e:
                            err_name = type(e).__name__
                            is_transient = any(tag in err_name for tag in ('RemoteDisconnected', 'ConnectionError', 'ConnectionReset', 'SSLError', 'SSLEOFError', 'Timeout'))
                            if is_transient and attempt < max_retries - 1:
                                backoff = 2 ** attempt + 0.5
                                logging.warning(f"⚠️ Attachment fetch retry {attempt+1}/{max_retries} for row {target_row.id} ({err_name}), backoff {backoff:.1f}s")
                                time.sleep(backoff)
                            else:
                                if attempt > 0:
                                    logging.warning(f"⚠️ Attachment fetch failed after {max_retries} attempts for row {target_row.id}: {err_name}")
                                return (target_row.id, [])

                _prefetch_budget_exceeded = False
                _prefetch_stuck_futures = 0     # future.result timed out after as_completed yielded
                _prefetch_cancelled = 0         # queued futures we successfully cancelled
                _prefetch_still_running = 0     # in-flight futures we abandoned to the background
                # Manual executor lifecycle with daemon workers. Three things can
                # block process exit for a non-daemon worker and all three matter
                # here: (1) _python_exit joins _threads_queues, (2) threading.
                # _shutdown joins _shutdown_locks, (3) executor.shutdown(wait=True)
                # joins via the `with` block. Using _DaemonThreadPoolExecutor
                # addresses (2) — daemon threads don't add their tstate lock to
                # _shutdown_locks. Using explicit shutdown(wait=False,
                # cancel_futures=True) in finally addresses (3). The detach helper
                # below addresses (1) — but only on the budget-exceeded path
                # (Copilot review: don't touch private APIs when everything
                # completed normally; the workers are already done and there's
                # nothing to skip). See _DaemonThreadPoolExecutor docstring for
                # the full three-defense story and the safety invariant.
                executor = _DaemonThreadPoolExecutor(max_workers=PARALLEL_WORKERS)
                futures = [executor.submit(_fetch_row_attachments, item) for item in target_map_to_prefetch.items()]
                total_futures = len(futures)
                _phase_budget_sec = ATTACHMENT_PREFETCH_MAX_MINUTES * 60

                # Helper: pop workers from concurrent.futures' atexit join
                # registry so _python_exit doesn't t.join() them at interpreter
                # shutdown (daemon-ness doesn't help here — join() blocks
                # unconditionally). Called only when we're abandoning in-flight
                # work. Uses private APIs; getattr guards keep the main path
                # working if a future Python rearranges the names.
                def _detach_from_atexit_registry():
                    try:
                        registry = getattr(_cf_thread, '_threads_queues', None)
                        if registry is None:
                            return
                        for _t in list(getattr(executor, '_threads', ()) or ()):
                            registry.pop(_t, None)
                    except Exception as _det_e:
                        logging.debug(f"Could not detach pre-fetch workers from atexit registry: {_det_e}")
                try:
                    try:
                        # timeout= is measured from this call; the iterator itself raises
                        # FuturesTimeoutError if nothing else completes within that window,
                        # so a stuck HTTP call can't pin the consumer loop.
                        for i, future in enumerate(as_completed(futures, timeout=_phase_budget_sec), 1):
                            try:
                                row_id, atts = future.result(timeout=ATTACHMENT_PREFETCH_FUTURE_TIMEOUT_SEC)
                            except FuturesTimeoutError:
                                # Defensive — as_completed only yields done futures, so in
                                # practice this branch is unreachable; keep it so a future
                                # refactor that yields not-yet-done futures still degrades
                                # gracefully instead of raising.
                                _prefetch_stuck_futures += 1
                                continue
                            attachment_cache[row_id] = atts
                            if i % 25 == 0 or i == total_futures:
                                logging.info(f"   📎 [{i}/{total_futures}] Attachment pre-fetch progress...")
                    except FuturesTimeoutError:
                        # Phase sub-budget exhausted — stuck HTTP call(s) held the iterator.
                        # Bail out; remaining rows fall back to the per-row path.
                        _prefetch_budget_exceeded = True
                finally:
                    # Classify remaining work so the log / Sentry span reflects reality:
                    # cancel() returns True only for queued futures that hadn't started
                    # (Copilot review: the old code overcounted by calling `not f.done()`
                    # alone, conflating started-but-running with still-queued).
                    for f in futures:
                        if f.done():
                            continue
                        if f.cancel():
                            _prefetch_cancelled += 1
                        else:
                            _prefetch_still_running += 1
                    # wait=False so stuck in-flight threads don't block the critical path
                    # (the main generation loop). They'll either complete via SDK retry
                    # backoff or be hard-killed by the workflow's timeout-minutes ceiling.
                    # Only touch the atexit registry when we're actually abandoning
                    # work (budget exceeded + still-running threads remain).
                    # Normal completion leaves the workers done; _python_exit will
                    # find them complete and return immediately from its join().
                    if _prefetch_still_running:
                        _detach_from_atexit_registry()
                    executor.shutdown(wait=False, cancel_futures=True)

                _att_elapsed = (datetime.datetime.now() - _att_start).total_seconds()
                span.set_data("rows_cached", len(attachment_cache))
                span.set_data("rows_cancelled", _prefetch_cancelled)
                span.set_data("rows_still_running", _prefetch_still_running)
                span.set_data("rows_stuck", _prefetch_stuck_futures)
                if _prefetch_budget_exceeded:
                    logging.warning(
                        f"⏰ Attachment pre-fetch budget hit ({ATTACHMENT_PREFETCH_MAX_MINUTES}min). "
                        f"Cached {len(attachment_cache)}/{total_futures} rows in {_att_elapsed:.1f}s; "
                        f"{_prefetch_cancelled} cancelled, {_prefetch_still_running} still running in background, "
                        f"{_prefetch_stuck_futures} stuck. Remaining rows will use per-row fallback."
                    )
                    sentry_add_breadcrumb(
                        "prefetch_truncated",
                        f"Pre-fetch truncated at {ATTACHMENT_PREFETCH_MAX_MINUTES}min",
                        level="warning",
                        data={
                            "cached": len(attachment_cache),
                            "total": total_futures,
                            "cancelled": _prefetch_cancelled,
                            "still_running": _prefetch_still_running,
                            "stuck": _prefetch_stuck_futures,
                        },
                    )
                else:
                    logging.info(f"⚡ Pre-fetched attachments for {len(attachment_cache)} target rows in {_att_elapsed:.1f}s (parallel w/{PARALLEL_WORKERS} workers)")

        # ──────────────────────────────────────────────────────────
        # Phase 01 gap closure (REVIEW-WR-05): secondary attachment
        # prefetch for SUBCONTRACTOR_PPP_SHEET_ID rows. Without it,
        # every _ReducedSub / _ReducedSub_Helper_* upload to the PPP
        # sheet pays an extra ``list_row_attachments`` API call (for
        # delete_old_excel_attachments matching). The PPP sheet has
        # far fewer rows than TARGET_SHEET_ID — only the subset that
        # needs _ReducedSub* — so the cost amortizes quickly.
        #
        # Defense-in-depth contract (Living Ledger 2026-04-22 16:05):
        #   - _DaemonThreadPoolExecutor (NOT ThreadPoolExecutor)
        #   - as_completed(futures, timeout=...) for the wait
        #   - executor.shutdown(wait=False, cancel_futures=True)
        #   - _detach_ppp_from_atexit_registry() on budget-exceed path
        #   - Pre-flight skip if session budget < (PREFETCH_MAX +
        #     GENERATION_HEADROOM)
        # Safety invariant: PPP prefetch is OPTIONAL — both
        # delete_old_excel_attachments and _has_existing_week_attachment
        # accept cached_attachments=None and fall back to per-row API.
        # Do NOT add new consumers that assume the PPP cache is
        # populated.
        # ──────────────────────────────────────────────────────────
        _ppp_prefetch_eligible = (
            SUBCONTRACTOR_RATE_VARIANTS_ENABLED
            and SUBCONTRACTOR_PPP_SHEET_ID
            and SUBCONTRACTOR_PPP_SHEET_ID != TARGET_SHEET_ID
            and not TEST_MODE
            and target_map_ppp is not None
            and len(target_map_ppp) > 0
        )
        if _ppp_prefetch_eligible:
            # Pre-flight budget guard (Living Ledger 2026-04-22 16:05
            # rule 7): skip entirely if remaining budget < (prefetch
            # phase budget + generation headroom). Without the
            # headroom reservation, an edge case where session
            # budget == prefetch budget would still trigger the
            # prefetch and leave zero time for the main loop.
            if TIME_BUDGET_MINUTES > 0:
                _ppp_elapsed_min = (
                    datetime.datetime.now() - session_start
                ).total_seconds() / 60.0
                _ppp_remaining_min = TIME_BUDGET_MINUTES - _ppp_elapsed_min
                _ppp_required_min = (
                    ATTACHMENT_PREFETCH_MAX_MINUTES
                    + ATTACHMENT_PREFETCH_GENERATION_HEADROOM_MIN
                )
                if _ppp_remaining_min < _ppp_required_min:
                    logging.info(
                        f"🛡️ Skipping PPP attachment prefetch: only "
                        f"{_ppp_remaining_min:.1f}min of session budget "
                        f"remain (need >= {_ppp_required_min:.0f}min for "
                        f"prefetch + generation headroom). PPP target "
                        f"rows will fall back to per-row API calls — "
                        f"correctness is preserved."
                    )
                    _ppp_prefetch_eligible = False
        if _ppp_prefetch_eligible:
            with sentry_sdk.start_span(
                op="smartsheet.attachment_prefetch_ppp",
                name="Pre-fetch PPP row attachments",
            ) as ppp_span:
                logging.info(
                    f"🚀 Starting parallel PPP attachment pre-fetch "
                    f"with {PARALLEL_WORKERS} workers for "
                    f"{len(target_map_ppp)} PPP target rows (max "
                    f"{ATTACHMENT_PREFETCH_MAX_MINUTES}min)..."
                )
                _ppp_att_start = datetime.datetime.now()

                def _fetch_ppp_row_attachments(row_item):
                    # row_item is (wr_num, target_row); only target_row is needed.
                    _, target_row = row_item
                    max_retries = 4
                    for attempt in range(max_retries):
                        try:
                            atts = client.Attachments.list_row_attachments(SUBCONTRACTOR_PPP_SHEET_ID, target_row.id).data
                            return (target_row.id, atts)
                        except (ss_exc.RateLimitExceededError,) as e:
                            if attempt < max_retries - 1:
                                backoff = 15 * (attempt + 1)
                                logging.warning(
                                    f"⚠️ PPP rate limited on attachment fetch "
                                    f"for row {target_row.id}, backoff {backoff}s "
                                    f"(attempt {attempt + 1}/{max_retries})"
                                )
                                time.sleep(backoff)
                            else:
                                logging.warning(
                                    f"⚠️ PPP attachment fetch failed after "
                                    f"{max_retries} rate-limit retries for row "
                                    f"{target_row.id}"
                                )
                                return (target_row.id, [])
                        except (
                            ss_exc.UnexpectedErrorShouldRetryError,
                            ss_exc.InternalServerError,
                            ss_exc.ServerTimeoutExceededError,
                        ) as e:
                            if attempt < max_retries - 1:
                                backoff = 2 ** attempt + 0.5
                                logging.warning(
                                    f"⚠️ PPP attachment fetch retry "
                                    f"{attempt + 1}/{max_retries} for row "
                                    f"{target_row.id} ({type(e).__name__}), "
                                    f"backoff {backoff:.1f}s"
                                )
                                time.sleep(backoff)
                            else:
                                logging.warning(
                                    f"⚠️ PPP attachment fetch failed after "
                                    f"{max_retries} attempts for row "
                                    f"{target_row.id}: {type(e).__name__}"
                                )
                                return (target_row.id, [])
                        except Exception as e:
                            err_name = type(e).__name__
                            is_transient = any(
                                tag in err_name for tag in (
                                    'RemoteDisconnected', 'ConnectionError',
                                    'ConnectionReset', 'SSLError',
                                    'SSLEOFError', 'Timeout',
                                )
                            )
                            if is_transient and attempt < max_retries - 1:
                                backoff = 2 ** attempt + 0.5
                                logging.warning(
                                    f"⚠️ PPP attachment fetch retry "
                                    f"{attempt + 1}/{max_retries} for row "
                                    f"{target_row.id} ({err_name}), backoff "
                                    f"{backoff:.1f}s"
                                )
                                time.sleep(backoff)
                            else:
                                if attempt > 0:
                                    logging.warning(
                                        f"⚠️ PPP attachment fetch failed "
                                        f"after {max_retries} attempts for row "
                                        f"{target_row.id}: {err_name}"
                                    )
                                return (target_row.id, [])

                _ppp_prefetch_budget_exceeded = False
                _ppp_prefetch_cancelled = 0
                _ppp_prefetch_still_running = 0

                ppp_executor = _DaemonThreadPoolExecutor(
                    max_workers=PARALLEL_WORKERS,
                )
                ppp_futures = [
                    ppp_executor.submit(_fetch_ppp_row_attachments, item)
                    for item in target_map_ppp.items()
                ]
                _ppp_phase_budget_sec = ATTACHMENT_PREFETCH_MAX_MINUTES * 60

                def _detach_ppp_from_atexit_registry():
                    try:
                        registry = getattr(
                            _cf_thread, '_threads_queues', None,
                        )
                        if registry is None:
                            return
                        for _t in list(
                            getattr(ppp_executor, '_threads', ()) or ()
                        ):
                            registry.pop(_t, None)
                    except Exception as _det_e:
                        logging.debug(
                            f"Could not detach PPP pre-fetch workers from "
                            f"atexit registry: {_det_e}"
                        )

                try:
                    for fut in as_completed(
                        ppp_futures, timeout=_ppp_phase_budget_sec,
                    ):
                        try:
                            row_id, atts = fut.result()
                            attachment_cache[row_id] = atts
                        except Exception as e:
                            # Worker exceptions already logged inside
                            # the worker — fall through to per-row.
                            logging.debug(
                                f"PPP prefetch future raised; row will "
                                f"fall back to per-row: {type(e).__name__}"
                            )
                except FuturesTimeoutError:
                    _ppp_prefetch_budget_exceeded = True
                    logging.warning(
                        f"⚠️ PPP attachment prefetch exceeded "
                        f"{ATTACHMENT_PREFETCH_MAX_MINUTES}min sub-budget; "
                        f"abandoning in-flight workers. Affected PPP rows "
                        f"will fall back to per-row API calls — correctness "
                        f"is preserved."
                    )
                finally:
                    # Three defenses against interpreter-exit hang:
                    # (1) atexit registry detach (only on budget-exceed,
                    #     per Copilot review — don't touch private APIs
                    #     when workers completed normally)
                    # (2) _DaemonThreadPoolExecutor handles tstate_lock
                    # (3) explicit shutdown(wait=False, cancel_futures=True)
                    if _ppp_prefetch_budget_exceeded:
                        _detach_ppp_from_atexit_registry()
                    for _fut in ppp_futures:
                        if _fut.cancel():
                            _ppp_prefetch_cancelled += 1
                        elif not _fut.done():
                            _ppp_prefetch_still_running += 1
                    ppp_executor.shutdown(wait=False, cancel_futures=True)

                _ppp_elapsed = (
                    datetime.datetime.now() - _ppp_att_start
                ).total_seconds()
                logging.info(
                    f"🏁 PPP attachment prefetch complete in "
                    f"{_ppp_elapsed:.1f}s: {len(target_map_ppp)} rows, "
                    f"{_ppp_prefetch_cancelled} cancelled, "
                    f"{_ppp_prefetch_still_running} still_running"
                )
                ppp_span.set_data("rows_prefetched", len(target_map_ppp))
                ppp_span.set_data("budget_exceeded", _ppp_prefetch_budget_exceeded)
                ppp_span.set_data("cancelled", _ppp_prefetch_cancelled)
                ppp_span.set_data("still_running", _ppp_prefetch_still_running)

        # Load hash history AFTER optional purge so we don't rely on stale attachments
        hash_history = load_hash_history(HASH_HISTORY_PATH)

        # ─────────────────────────────────────────────────────────
        # Phase 1.1 SUB-12 / D-17..D-19: idempotent hash-history prune.
        # ─────────────────────────────────────────────────────────
        # Runs once per migration version. The constant
        # ``PHASE_1_1_HASH_PRUNE_VERSION`` IS the kill switch (D-19);
        # the helper handles the version-gate + simplified-D-18 scope
        # detection + INFO logging. Mutates ``hash_history`` in place
        # so the sentinel + dropped-orphan side-effects survive the
        # subsequent ``save_hash_history`` write at end of run.
        # ``groups`` was built upstream at the ``group_source_rows``
        # call site; if grouping failed and execution reached here,
        # the helper degrades gracefully (empty groups → empty
        # _sub_wr_scope → no orphans dropped → sentinel still written).
        # Codex P2: track whether either one-time migration prune mutated
        # hash_history so we can persist it even on a run with no group
        # updates (the history_updates-gated save below would otherwise skip
        # it, making the migration re-run every no-update execution).
        _hash_history_migration_dirty = False
        try:
            if _run_phase_1_1_hash_prune(hash_history, groups):
                _hash_history_migration_dirty = True
        except Exception as _prune_exc:
            # Fail-safe per [2026-04-22 16:05] rule 4 — the prune
            # is an optimization. A failed prune MUST NOT break the
            # billing pipeline. Log + continue with the unmodified
            # hash_history (the sentinel will not advance, the prune
            # retries next run, the orphans remain harmless).
            logging.warning(
                f"⚠️ Phase 1.1 hash-history prune failed; continuing "
                f"with existing history: {_prune_exc!r}"
            )

        # Subproject B: one-time prune of legacy blank-identifier
        # reduced_sub/aep_billable orphans (kill switch is the version
        # constant). Fail-safe — a failed prune must not break the run.
        try:
            if _run_subproject_b_hash_prune(hash_history, groups):
                _hash_history_migration_dirty = True
        except Exception as _b_prune_exc:
            logging.warning(
                f"⚠️ Subproject B hash-history prune failed; continuing "
                f"with existing history: {_b_prune_exc!r}"
            )

        # Subproject C: one-time prune of legacy blank-identifier vac_crew
        # orphans (kill switch is the version constant). Fail-safe — a
        # failed prune must not break the run.
        try:
            if _run_vac_crew_hash_prune(hash_history, groups):
                _hash_history_migration_dirty = True
        except Exception as _vc_prune_exc:
            logging.warning(
                f"⚠️ Vac crew hash-history prune failed; continuing "
                f"with existing history: {_vc_prune_exc!r}"
            )

        # Subproject D: one-time prune of legacy blank-identifier primary
        # orphans (kill switch is PRIMARY_CLAIM_ATTRIBUTION_ENABLED + the
        # version constant). Fail-safe — a failed prune must not break the
        # run.
        try:
            if _run_subproject_d_hash_prune(hash_history, groups):
                _hash_history_migration_dirty = True
        except Exception as _d_prune_exc:
            logging.warning(
                f"⚠️ Subproject D hash-history prune failed; continuing "
                f"with existing history: {_d_prune_exc!r}"
            )

        billing_audit_row_cache: set[str] = set()
        billing_audit_row_cache_dirty = False
        if BILLING_AUDIT_AVAILABLE and not TEST_MODE:
            billing_audit_row_cache = load_billing_audit_row_cache(
                BILLING_AUDIT_ROW_CACHE_PATH
            )
            # Ensure the cache file exists on disk even when no rows have been
            # frozen yet. The GitHub Actions cache/save step will fail with
            # "Path does not exist" when the file is absent, which can happen
            # on the very first run or when all rows were already cached from a
            # prior run (billing_audit_row_cache_dirty stays False, so the
            # save at the end of the run is skipped).  Writing an empty list
            # now is cheap and makes the CI step reliably no-op safe.
            if not os.path.exists(BILLING_AUDIT_ROW_CACHE_PATH):
                save_billing_audit_row_cache(
                    BILLING_AUDIT_ROW_CACHE_PATH, billing_audit_row_cache
                )
        history_updates = 0
        _groups_skipped = 0
        _groups_generated = 0
        _groups_uploaded = 0
        _groups_errored = 0
        _api_calls_count = 0
        _upload_tasks = []  # Collect upload tasks for parallel processing

        _phase_group_start = datetime.datetime.now()
        _time_budget_exceeded = False

        # Phase 01 Plan 03 Task 2 (D-16/D-17): per-sheet accumulator
        # of subcontractor CU codes that fell through to SmartSheet
        # pricing during ``generate_excel``. ``_resolve_row_price``
        # records each missing CU into a per-call Counter that
        # ``generate_excel`` returns in the 5-tuple's trailing slot;
        # the per-group loop below attributes each group's missing
        # CUs to the source sheet(s) that contributed rows. After the
        # loop completes, exactly ONE WARNING per affected sheet is
        # emitted (D-17), naming the first 10 codes alphabetically.
        # The PII sanitizer's ``_PII_LOG_MARKERS`` already includes
        # the WARNING's stable marker ("Subcontractor rates CSV
        # missing") so it is dropped from Sentry before send.
        _missing_cus_by_sheet: dict[int, collections.Counter] = (
            collections.defaultdict(collections.Counter)
        )

        # Codex P1: source-side WR# collision quarantine.
        # ``_RE_SANITIZE_HELPER_NAME`` on the raw row value is a lossy
        # transform — two distinct raw WR# values may fold to the
        # same sanitized key. Downstream routing uses that sanitized
        # key for ``target_map`` lookups AND for attachment-identity
        # matching (filenames, hash_history), so an unquarantined
        # collision can cause cross-WR data corruption:
        #   * If target_map has BOTH colliding raws, round-6 quarantine
        #     removes the key from target_map so both uploads fail
        #     loudly at ``if wr_num in target_map`` — safe.
        #   * If target_map has only ONE of the raws (the other WR
        #     simply isn't in the target sheet yet), the source-side
        #     scan is the only defence. The second raw's group would
        #     otherwise resolve ``target_map[sanitized]`` to the first
        #     raw's row and upload to the wrong row.
        # We therefore key the quarantine on the sanitized WR ALONE
        # (not on ``(wr, week, variant)``): any pair of distinct raw
        # WRs that fold to the same sanitized key, anywhere in the
        # run's groups, is a collision regardless of week or variant.
        # Realistic numeric WR#s can't collide, so the scan is
        # zero-impact on production data.
        _source_wr_raws_per_key: dict = collections.defaultdict(set)
        for _g_rows in groups.values():
            if not _g_rows:
                continue
            _g_raw = str(_g_rows[0].get('Work Request #') or '').split('.')[0]
            if not _g_raw:
                continue
            _g_sanitized = _RE_SANITIZE_HELPER_NAME.sub('_', _g_raw)[:50]
            _source_wr_raws_per_key[_g_sanitized].add(_g_raw)
        _quarantined_source_wr_keys: set = {
            key for key, raws in _source_wr_raws_per_key.items()
            if len(raws) > 1
        }
        if _quarantined_source_wr_keys:
            for _qk in _quarantined_source_wr_keys:
                _raws = sorted(_source_wr_raws_per_key[_qk])
                logging.warning(
                    f"⚠️ Source WR# sanitization collision: raws={_raws} "
                    f"all fold to sanitized_key={_qk!r}. All affected "
                    f"groups (across every week + variant combination) "
                    f"will be SKIPPED to prevent cross-WR contamination "
                    f"of target_map uploads and attachment identity. "
                    f"Deduplicate the source WR# values and rerun."
                )
            logging.warning(
                f"⚠️ Total source WR# collision quarantines: "
                f"{len(_quarantined_source_wr_keys)} sanitized key(s); "
                f"see preceding warnings for raw values."
            )

        # Hoist static env var lookups once per run (not per row) —
        # these never change during execution and were previously
        # being read on every freeze_row call for every row in every
        # group. One-time read. Empty-string defaults (instead of
        # None) keep the values valid as Supabase RPC parameters
        # whether or not the deployment target applies NOT NULL to
        # ``release`` / ``run_id``.
        #
        # NOTE: the fingerprint flag state is NOT hoisted here. Flag
        # reads are per-call so a transient early-run ``get_flag``
        # failure (which deliberately isn't cached per the
        # non-caching-on-failure fix) can recover on subsequent
        # calls. Hoisting the boolean would lock the whole run into
        # the first-read result and silently drop pipeline_run rows.
        _billing_audit_release_env = os.getenv('SENTRY_RELEASE', '') or ''
        # ``run_id`` is part of the ``pipeline_run`` on_conflict key
        # ``(wr, week_ending, run_id)``. An empty string would make
        # every non-GitHub-Actions execution (manual reruns, local
        # debugging, crontab on a bare host, etc.) collide into the
        # same row for a given (wr, week), overwriting prior runs'
        # records and destroying run history.
        #
        # GitHub Actions re-runs preserve ``GITHUB_RUN_ID`` and only
        # increment ``GITHUB_RUN_ATTEMPT``. Appending the attempt
        # number makes each rerun create a distinct pipeline_run
        # row instead of overwriting the prior attempt — critical
        # for preserving drift-detection context when an earlier
        # attempt already wrote the key. Falls back to a microsecond
        # timestamp outside Actions.
        _ga_run_id = os.getenv('GITHUB_RUN_ID', '')
        _ga_run_attempt = os.getenv('GITHUB_RUN_ATTEMPT', '')
        if _ga_run_id:
            _billing_audit_run_id_env = (
                f"{_ga_run_id}.{_ga_run_attempt}"
                if _ga_run_attempt
                else _ga_run_id
            )
        else:
            _billing_audit_run_id_env = (
                f"local-{datetime.datetime.utcnow().strftime('%Y%m%dT%H%M%S%fZ')}"
            )

        # Pre-aggregate rows per (sanitized_wr, week) across ALL
        # variants so the assignment fingerprint captures the full
        # personnel picture. ``group_source_rows`` splits helper-
        # completed rows out of the primary group (to prevent
        # double-counting in Excel generation), so each group only
        # carries ONE variant's rows. With the writer's per-
        # (wr, week, run_id) dedup, only the first variant emitted
        # actually writes — meaning a naive fingerprint would miss
        # helper / vac_crew personnel entirely, defeating the whole
        # point of this PR (mid-week helper swaps wouldn't change
        # the primary-only fingerprint → no drift alert).
        #
        # Walking ``groups.items()`` once here is O(total rows)
        # and negligible compared to per-group work.
        _billing_audit_fp_buckets: dict[tuple[str, str], list[dict]] = {}
        # Aggregated content hash per (wr, week). Like assignment_fp,
        # the emit_run_fingerprint dedup writes exactly one
        # ``pipeline_run`` row per (wr, week, run_id) — so
        # ``content_hash`` must reflect the UNION of all variants'
        # rows, not whichever variant was iterated first. Without
        # this, a source-ordering change between runs flips the
        # stored hash even when the underlying work set is
        # unchanged, making downstream run comparisons noisy.
        _billing_audit_agg_content_hashes: dict[tuple[str, str], str] = {}
        # Split the cheap work from the expensive work:
        #   • Bucket assembly (dict appends across rows) — runs
        #     when billing_audit is available AND at least one
        #     writer flag is enabled (or the flag state is
        #     indeterminate via a transient read blip).
        #     ``any_flag_enabled()`` fails OPEN — a transient
        #     feature_flag read blip returns True so we still
        #     build buckets and don't miss the first-write-wins
        #     freeze window for this run's completed rows. Cost
        #     is O(total rows) of dict appends.
        #   • ``calculate_data_hash`` per bucket — LAZY, memoized
        #     into ``_billing_audit_agg_content_hashes`` at first
        #     emit attempt inside the per-group block. The emit is
        #     already fingerprint-flag-gated, so flag-off runs
        #     never pay this cost, and flag-on runs pay it exactly
        #     once per bucket regardless of variant count (dedup
        #     no-ops reuse the memo).
        #
        # Wrapped in try/except Exception so any unexpected failure
        # (malformed row data, novel exception from ``any_flag_enabled``,
        # future code additions that introduce a bug) degrades
        # gracefully: buckets stay empty, the per-group emit falls
        # back to ``group_rows`` / ``data_hash`` via its
        # ``.get(key, fallback)`` calls, and Excel generation is
        # completely untouched. Class-name-only logging preserves
        # the _PII_LOG_MARKERS discipline.
        try:
            if (
                BILLING_AUDIT_AVAILABLE
                and not TEST_MODE
                and _billing_audit_writer.any_flag_enabled()
            ):
                for _agg_gk, _agg_rows in groups.items():
                    if not _agg_rows:
                        continue
                    # Defensive isinstance: group_source_rows always
                    # emits dicts, but a future mutation or bug
                    # upstream could violate that — don't let it
                    # raise AttributeError into the main loop.
                    _first = _agg_rows[0]
                    if not isinstance(_first, dict):
                        continue
                    _raw_wr = _first.get('Work Request #')
                    _wr_str = str(_raw_wr).split('.')[0] if _raw_wr else ''
                    _wr_san = _RE_SANITIZE_HELPER_NAME.sub('_', _wr_str)[:50]
                    _week_part = (
                        _agg_gk.split('_', 1)[0] if '_' in _agg_gk else ''
                    )
                    if not _wr_san or not _week_part:
                        continue
                    _billing_audit_fp_buckets.setdefault(
                        (_wr_san, _week_part), []
                    ).extend(_agg_rows)
        except Exception as _preloop_err:
            # Graceful degradation. Empty buckets + the per-group
            # emit's ``.get(key, fallback)`` calls preserve correct
            # Excel generation; only cross-variant fingerprint
            # aggregation is lost for this run.
            logging.warning(
                "⚠️ Billing audit pre-loop aggregation failed "
                f"(suppressed details): {type(_preloop_err).__name__}"
            )
            sentry_add_breadcrumb(
                "billing_audit",
                "Pre-loop aggregation failure",
                level="warning",
                data={"error_type": type(_preloop_err).__name__},
            )
            _billing_audit_fp_buckets = {}
            _billing_audit_agg_content_hashes = {}

        for group_idx, (group_key, group_rows) in enumerate(groups.items(), 1):
            # Graceful time budget: stop before Actions hard-kills the job
            if TIME_BUDGET_MINUTES and GITHUB_ACTIONS_MODE:
                elapsed_min = (datetime.datetime.now() - session_start).total_seconds() / 60.0
                if elapsed_min >= TIME_BUDGET_MINUTES:
                    remaining = len(groups) - group_idx + 1
                    logging.warning(f"⏰ Time budget exhausted ({elapsed_min:.1f}min >= {TIME_BUDGET_MINUTES}min). "
                                    f"Stopping with {remaining} group(s) remaining. "
                                    f"They will be processed on the next run (hash history preserved).")
                    _time_budget_exceeded = True
                    sentry_add_breadcrumb("time_budget", f"Budget exceeded after {elapsed_min:.1f}min", level="warning", data={
                        "groups_remaining": remaining, "groups_processed": group_idx - 1,
                    })
                    break
            try:
                # Calculate data hash for change detection
                data_hash = calculate_data_hash(group_rows)
                wr_num_raw = group_rows[0].get('Work Request #')
                wr_num = str(wr_num_raw).split('.')[0] if wr_num_raw else ''
                # Apply the same filesystem-safety sanitizer used inside
                # generate_excel so history keys, attachment prefix
                # matching, and Excel filenames all use the identical
                # WR identifier. Realistic numeric WR#s are unchanged;
                # path-traversal metacharacters get replaced with ``_``.
                wr_num = _RE_SANITIZE_HELPER_NAME.sub('_', wr_num)[:50]
                week_raw = group_key.split('_',1)[0] if '_' in group_key else ''

                # Extract variant and identifier for variant-aware hash history
                first_row = group_rows[0] if group_rows else {}
                variant = first_row.get('__variant', 'primary')

                # Source-side collision quarantine (see pre-scan above).
                # If this group's sanitized WR was flagged as colliding
                # with another group's raw WR anywhere in the run —
                # regardless of week or variant — skip it entirely. The
                # broader key is required because downstream
                # ``target_map`` lookups and attachment-identity
                # matching use only the sanitized WR; they do not
                # disambiguate by week or variant, so an unquarantined
                # cross-context collision can still route uploads /
                # deletes to the wrong target-sheet row.
                if wr_num in _quarantined_source_wr_keys:
                    logging.warning(
                        f"⚠️ Skipping group {group_key}: sanitized WR# "
                        f"{wr_num!r} collides with another group (see "
                        f"'Source WR# sanitization collision' WARNING "
                        f"above for the full raw-value list)."
                    )
                    _groups_skipped += 1
                    continue
                if variant in ('helper', 'aep_billable_helper', 'reduced_sub_helper'):
                    # CRITICAL FIX: Include helper dept and job in identifier for unique hash keys
                    # This ensures helper files regenerate when new helper rows are added.
                    #
                    # CR-01 gap closure (Site 1 — main-loop identifier):
                    # helper, aep_billable_helper, and reduced_sub_helper all
                    # derive identifier / file_identifier from __helper_foreman
                    # so the round-trip with build_group_identity (which parses
                    # the helper-shadow filename's _Helper_<name>_<hash> tail in
                    # Plan 02) succeeds. Pre-fix, the two shadow variants fell
                    # through to the ``else`` branch that reads ``User`` —
                    # typically blank for shadow rows — producing
                    # file_identifier='' and a (parsed='Jane_Smith') == ('')
                    # mismatch in _has_existing_week_attachment. Result:
                    # permanent regeneration churn and orphan accumulation on
                    # SUBCONTRACTOR_PPP_SHEET_ID. The change is additive — the
                    # legacy ``helper`` body is preserved exactly; we just
                    # expand the gate to include the two helper-shadow variants.
                    # Sites 2 (valid_wr_weeks builder) and 3 (current_keys
                    # hash-history prune) carry the same gate — drift between
                    # the three sites is exactly the bug shape CR-01 documents.
                    helper_foreman = first_row.get('__helper_foreman', '')
                    helper_dept = first_row.get('__helper_dept', '')
                    helper_job = first_row.get('__helper_job', '')
                    identifier = f"{helper_foreman}|{helper_dept}|{helper_job}"
                    # file_identifier matches the sanitized name that generate_excel() puts in the filename
                    file_identifier = _RE_SANITIZE_HELPER_NAME.sub('_', helper_foreman)[:50] if helper_foreman else ''
                elif variant == 'vac_crew':
                    # Subproject C identity site (Site 1 — main-loop identifier /
                    # history_key / file_identifier). GATED on the kill switch:
                    # disabled mode MUST reproduce the exact legacy '' identifier
                    # (bare _VacCrew filename, bare history_key) so existing
                    # attachments are not treated as stale and regeneration churn
                    # is not triggered. Enabled mode uses the sanitized claimer.
                    _vc = first_row.get('__current_foreman', '')
                    identifier = (
                        _RE_SANITIZE_IDENTIFIER.sub('_', _vc)[:50]
                        if (VAC_CREW_CLAIM_ATTRIBUTION_ENABLED and _vc) else ''
                    )
                    file_identifier = identifier
                elif variant in ('reduced_sub', 'aep_billable'):
                    # Subproject B identity site (Site 1 — main-loop
                    # identifier). Partitioned by the frozen primary
                    # claimer (__current_foreman). identifier ==
                    # file_identifier == sanitized claimer, matching the
                    # _ReducedSub_User_<name> filename and Sites 2 & 3.
                    _b_claimer = first_row.get('__current_foreman', '')
                    identifier = (
                        _RE_SANITIZE_IDENTIFIER.sub('_', _b_claimer)[:50]
                        if _b_claimer else ''
                    )
                    file_identifier = identifier
                else:
                    # Subproject D (2026-05-25): Site 1 — main-loop primary
                    # identity (history_key / file_identifier). Gated on kill
                    # switch: enabled → frozen claimer (__current_foreman);
                    # disabled → legacy ``User`` field ('' in production).
                    if (
                        PRIMARY_CLAIM_ATTRIBUTION_ENABLED
                        and RES_GROUPING_MODE in ('helper', 'both')
                    ):
                        _pf = first_row.get('__current_foreman', '')
                        identifier = (
                            _RE_SANITIZE_IDENTIFIER.sub('_', _pf)[:50]
                            if _pf else ''
                        )
                        file_identifier = identifier
                    else:
                        # Legacy primary variant: identifier derived from
                        # the row's ``User`` field.
                        user_val = first_row.get('User')
                        # PERFORMANCE: Use pre-compiled regex for identifier sanitization
                        identifier = _RE_SANITIZE_IDENTIFIER.sub('_', user_val)[:50] if user_val else ''
                        file_identifier = identifier
                
                # History key includes variant dimension to prevent collisions
                history_key = f"{wr_num}|{week_raw}|{variant}|{identifier}"

                # Sub-project E: ISO week-ending date for the durable
                # Supabase hash store (group_content_hash.week_ending is a
                # DATE column). Derived from the SAME __week_ending_date the
                # billing_audit freeze / fingerprint calls use (see the
                # _week_snap normalization below), so the durable 4-tuple key
                # matches across the reader, the writer, and those callers.
                # Falls back to '' when the date is absent — the lookup then
                # returns no_row and the upsert is keyed on '', both of which
                # fail safe to "regenerate".
                _wed = group_rows[0].get('__week_ending_date')
                if hasattr(_wed, 'date'):
                    _wed = _wed.date()
                week_iso = _wed.isoformat() if hasattr(_wed, 'isoformat') else ''

                # Pre-compute hash-change state before any optional side-effects.
                # Billing audit RPCs are the single most expensive per-group operation
                # in steady state, so we can safely skip them when the group hash is
                # unchanged versus hash_history (no row-content drift to freeze or emit).
                _history_eligible_for_skip = (
                    HISTORY_SKIP_ENABLED
                    and not (
                        FORCE_GENERATION
                        or week_raw in REGEN_WEEKS
                        or RESET_HASH_HISTORY
                        or RESET_WR_LIST
                    )
                )
                # Sub-project E: the unchanged decision now consults the
                # durable Supabase hash store when authoritative, falling
                # back to the local hash_history json cache on outage/miss.
                # See _resolve_unchanged_for_skip for the full decision
                # table. Default (authoritative OFF) is json-cache-only —
                # byte-identical to the pre-E behavior.
                _hash_unchanged = (
                    _resolve_unchanged_for_skip(
                        history_key, data_hash, hash_history,
                        wr_num, week_iso, variant, identifier,
                    )
                    if _history_eligible_for_skip
                    else False
                )

                # Pre-compute whether any eligible row in this group is absent
                # from the freeze cache. When _hash_unchanged is True but some
                # rows are uncached (e.g., freeze_attribution failed transiently
                # in a prior run), we still need to attempt those rows so they
                # are not permanently left unfrozen. This allows recovery without
                # waiting for the group's content hash to change again.
                #
                # Use set-difference rather than an any()-generator so that for
                # large groups (50-150 rows is typical) the membership test is
                # O(len(eligible_keys)) via a single set operation instead of
                # potentially scanning all rows in the worst case.
                _has_uncached_freeze_candidates: bool = False
                if BILLING_AUDIT_AVAILABLE and not TEST_MODE:
                    _eligible_freeze_keys = {
                        f"{wr_num}|{week_raw}|{_r.get('__row_id')}"
                        for _r in group_rows
                        if isinstance(_r.get("__row_id"), int)
                        and is_checked(_r.get("Units Completed?"))
                    }
                    _has_uncached_freeze_candidates = bool(
                        _eligible_freeze_keys - billing_audit_row_cache
                    )

                # ── Billing audit snapshot: freeze personnel + emit run fingerprint ──
                # Runs when the group hash has changed/is new, OR when some rows
                # were not successfully frozen in a prior run (transient failure
                # recovery). Skipped only when hash is unchanged AND every
                # eligible row is already in the freeze cache.
                # Failures must never break Excel generation.
                if (
                    BILLING_AUDIT_AVAILABLE
                    and not TEST_MODE
                    and (not _hash_unchanged or _has_uncached_freeze_candidates)
                    and _billing_audit_writer.any_flag_enabled()
                ):
                    try:
                        # Generic span name — the WR number is
                        # attached as span data below. The pipeline's
                        # _PII_LOG_MARKERS (see log sanitizer) treats
                        # "for WR " as a PII signal that gets
                        # dropped from Sentry Logs; span names
                        # bypass that sanitizer entirely and end up
                        # in performance/trace data regardless. Keep
                        # the name structural and route the
                        # identifier through set_data where it can
                        # be scoped, filtered, and (if needed) later
                        # scrubbed via before_send.
                        with sentry_sdk.start_span(
                            op="billing_audit.freeze",
                            name="billing_audit.freeze_attribution",
                        ) as _bas:
                            _bas.set_data("wr", wr_num)
                            _rows_to_freeze: list[dict] = []
                            _freeze_row_keys: dict[int, str] = {}
                            for _row in group_rows:
                                _row_id = _row.get("__row_id")
                                if not isinstance(_row_id, int):
                                    continue
                                if not is_checked(_row.get("Units Completed?")):
                                    continue
                                _cache_key = f"{wr_num}|{week_raw}|{_row_id}"
                                if _cache_key in billing_audit_row_cache:
                                    continue
                                _rows_to_freeze.append(_row)
                                _freeze_row_keys[id(_row)] = _cache_key
                            _bas.set_data("row_count", len(_rows_to_freeze))
                            _week_snap = first_row.get('__week_ending_date')
                            if hasattr(_week_snap, 'date'):
                                _week_snap = _week_snap.date()
                            # Parallelize per-row freeze_row calls so a
                            # group with N rows costs ~ceil(N/W) ×
                            # round-trip latency instead of N × latency.
                            # Pre-2026-04-25 this was a serial loop;
                            # at ~120ms per Supabase RPC, large groups
                            # (50-150 rows is typical for a busy WR
                            # week) burned 6-18 seconds of wall-clock
                            # purely on serial HTTP. Across 1900+
                            # groups in a weekly run that compounded
                            # into ~2 hours of new latency on top of
                            # the pre-billing_audit ~1h baseline,
                            # consuming TIME_BUDGET_MINUTES before the
                            # main loop reached Excel generation.
                            #
                            # ``freeze_row`` is intended to be fail-
                            # safe: it handles routine errors
                            # internally and records best-effort
                            # diagnostic counters. Counter writes are
                            # protected by ``_counters_lock`` so the
                            # totals stay exact even under concurrent
                            # invocation (the bare ``dict[k] += 1``
                            # is a multi-bytecode read-modify-write
                            # and CAN lose increments without the
                            # lock). A future raising here is still
                            # unexpected; log it (with sanitized row
                            # id) and continue with the rest of the
                            # group's writes.
                            #
                            # Executor reuse: ``get_freeze_row_executor()``
                            # returns a process-wide singleton lazily
                            # created on first use. With ~1900 groups
                            # per typical run, creating a per-group
                            # executor would mean ~1900 executor
                            # constructions and ~15,000 thread-join
                            # operations — each cheap individually
                            # but non-trivial in aggregate, and
                            # noisy in operational debugging.
                            # ``atexit`` handles shutdown when the
                            # interpreter exits.
                            if len(_rows_to_freeze) <= 1:
                                for _row in _rows_to_freeze:
                                    # Per D-18 / SUB-07 Path B: variant is
                                    # accepted by freeze_row for signature
                                    # symmetry but is NOT injected into the
                                    # freeze_attribution RPC params dict.
                                    # The variant lives on pipeline_run via
                                    # emit_run_fingerprint below. Default
                                    # 'primary' for pre-Phase-1 rows whose
                                    # __variant field isn't set (legacy
                                    # primary/helper/vac_crew rows from
                                    # before Plan 03 tagged them).
                                    _ok = _billing_audit_writer.freeze_row(
                                        _row,
                                        release=_billing_audit_release_env,
                                        run_id=_billing_audit_run_id_env,
                                        variant=_row.get('__variant', 'primary'),
                                    )
                                    if _ok:
                                        _rk = _freeze_row_keys.get(id(_row))
                                        if _rk:
                                            billing_audit_row_cache.add(_rk)
                                            billing_audit_row_cache_dirty = True
                            else:
                                # Singleton executor sized once at
                                # first use; subsequent calls share
                                # the same worker pool.
                                _bas_ex = (
                                    _billing_audit_writer
                                    .get_freeze_row_executor(
                                        max_workers=PARALLEL_WORKERS,
                                    )
                                )
                                _bas.set_data(
                                    "in_flight", len(_rows_to_freeze)
                                )
                                # Track future → row so an unexpected
                                # raise can be pinpointed to the
                                # specific row that triggered it,
                                # not just the WR — useful when one
                                # row in a 100-row group has malformed
                                # data the writer didn't anticipate.
                                _bas_future_to_row: dict[Any, dict] = {}
                                for _row in _rows_to_freeze:
                                    # Per D-18 / SUB-07 Path B: variant
                                    # threads through the parallelized
                                    # worker fn but does NOT reach the
                                    # RPC params dict. See the single-row
                                    # branch above for the full rationale.
                                    _bas_f = _bas_ex.submit(
                                        _billing_audit_writer.freeze_row,
                                        _row,
                                        release=_billing_audit_release_env,
                                        run_id=_billing_audit_run_id_env,
                                        variant=_row.get('__variant', 'primary'),
                                    )
                                    _bas_future_to_row[_bas_f] = _row
                                for _bas_f in as_completed(_bas_future_to_row):
                                    try:
                                        _ok = _bas_f.result()
                                        if _ok:
                                            _good_row = _bas_future_to_row.get(
                                                _bas_f, {}
                                            )
                                            _rk = _freeze_row_keys.get(
                                                id(_good_row)
                                            )
                                            if _rk:
                                                billing_audit_row_cache.add(_rk)
                                                billing_audit_row_cache_dirty = True
                                    except Exception:
                                        # Sanitized row identifier:
                                        # ``__row_id`` is a Smartsheet
                                        # numeric ID (not PII) — safe
                                        # to log. Skip Pole / CU /
                                        # Foreman fields per the
                                        # _PII_LOG_MARKERS rule.
                                        _bad_row = _bas_future_to_row.get(_bas_f, {})
                                        _bad_row_id = _bad_row.get("__row_id")
                                        logging.exception(
                                            "billing_audit.freeze_row "
                                            "raised unexpectedly for "
                                            "WR %s row_id=%s",
                                            wr_num,
                                            _bad_row_id,
                                        )
                            # Skip fingerprint compute + completed
                            # count when the fingerprint flag is off
                            # — emit_run_fingerprint would no-op
                            # inside otherwise, wasting per-group
                            # work. Checked per-group (not hoisted)
                            # so a transient early-run flag-read
                            # failure doesn't suppress fingerprint
                            # emission for the rest of the run.
                            # ``get_flag`` caches successful reads,
                            # so the steady-state cost is a single
                            # dict lookup per group.
                            if _billing_audit_writer.fingerprint_flag_enabled():
                                # Use the cross-variant aggregation
                                # so the fingerprint AND content hash
                                # cover all personnel + all rows
                                # (primary + helper + vac) for this
                                # (wr, week). Falls back to
                                # ``group_rows`` / ``data_hash`` only
                                # if the bucket is empty (shouldn't
                                # happen — the bucket is built from
                                # the same groups dict we're
                                # iterating).
                                _agg_key = (wr_num, week_raw)
                                _agg_fp_rows = _billing_audit_fp_buckets.get(
                                    _agg_key, group_rows
                                )
                                # Lazy + memoized content-hash
                                # computation. First emit attempt
                                # for a bucket pays the hashing
                                # cost once and caches the result;
                                # subsequent variants that
                                # dedup-no-op inside
                                # emit_run_fingerprint get a cache
                                # hit for free.
                                #
                                # ``calculate_data_hash`` assumes
                                # all rows share one ``__variant``
                                # (it reads sorted_rows[0]'s
                                # variant and conditionally
                                # includes VAC / helper fields
                                # based on it). Passing it the raw
                                # cross-variant bucket would make
                                # the result depend on sort order
                                # and can miss VAC personnel
                                # entirely. Instead: bucket rows by
                                # variant, hash each subset with
                                # the production helper (so each
                                # variant gets its own correct
                                # fields), then SHA-256 the
                                # variant-sorted
                                # ``variant=hash`` tokens. Result
                                # is deterministic and covers
                                # every variant's full field set.
                                if _agg_key in _billing_audit_fp_buckets:
                                    _agg_content_hash = (
                                        _billing_audit_agg_content_hashes.get(
                                            _agg_key
                                        )
                                    )
                                    if _agg_content_hash is None:
                                        # Variant-aware aggregated
                                        # hash, with per-helper sub-
                                        # bucketing so multi-helper
                                        # WRs produce a stable
                                        # content_hash (see
                                        # _compute_aggregated_content_hash).
                                        _agg_content_hash = (
                                            _compute_aggregated_content_hash(
                                                _agg_fp_rows
                                            )
                                        )
                                        _billing_audit_agg_content_hashes[
                                            _agg_key
                                        ] = _agg_content_hash
                                else:
                                    _agg_content_hash = data_hash
                                _fp = compute_assignment_fingerprint(_agg_fp_rows)
                                _completed = sum(
                                    1 for _r in _agg_fp_rows
                                    if is_checked(_r.get('Units Completed?'))
                                )
                                # Per D-18 / SUB-07 Path B: variant is
                                # recorded on pipeline_run via this call.
                                # All rows in a group share the same
                                # __variant by construction in
                                # group_source_rows (Plan 03), so reading
                                # group_rows[0] is canonical. Falls back to
                                # 'primary' when the row hasn't been
                                # tagged (legacy / non-variant-aware
                                # call paths) — matches the writer's
                                # None-coercion sentinel.
                                _group_variant = (
                                    group_rows[0].get('__variant', 'primary')
                                    if group_rows else 'primary'
                                )
                                _billing_audit_writer.emit_run_fingerprint(
                                    wr=wr_num,
                                    week_ending=_week_snap,
                                    content_hash=_agg_content_hash,
                                    assignment_fp=_fp,
                                    completed_count=_completed,
                                    total_count=len(_agg_fp_rows),
                                    release=_billing_audit_release_env,
                                    run_id=_billing_audit_run_id_env,
                                    variant=_group_variant,
                                )
                            _bas.set_data("rows", len(group_rows))
                            _bas.set_data("freeze_candidates", len(_rows_to_freeze))
                            _bas.set_data("variant", variant)
                    except Exception as _audit_err:
                        # Class name only — avoids leaking WR / foreman /
                        # helper names via log bodies (see _PII_LOG_MARKERS).
                        logging.warning(
                            f"⚠️ Billing audit snapshot failed for group (suppressed details): "
                            f"{type(_audit_err).__name__}"
                        )
                        sentry_add_breadcrumb(
                            "billing_audit",
                            "Snapshot failure (group-level)",
                            level="warning",
                            data={"error_type": type(_audit_err).__name__},
                        )

                # Decide skip based on stored history BEFORE generating Excel (only if FORCE not set)
                if _history_eligible_for_skip:
                    if _hash_unchanged:
                        # Only skip if attachment present OR policy allows skipping without attachment
                        can_skip = True
                        if ATTACHMENT_REQUIRED_FOR_SKIP and not TEST_MODE:
                            # Need a target row to verify attachment presence
                            if not target_map:
                                target_map, _target_sheet_obj = create_target_sheet_map(client)
                            target_row = target_map.get(str(wr_num)) if target_map else None
                            if target_row is None:
                                can_skip = False  # Can't verify; safer to regenerate
                            else:
                                # Use file_identifier (the value
                                # actually embedded in the filename)
                                # rather than identifier (the
                                # hash-history-tuple form that includes
                                # helper_dept/helper_job). For the
                                # helper variant the two diverge —
                                # filename only carries the sanitized
                                # helper_foreman, so comparing against
                                # the tuple form would always miss and
                                # force regeneration of unchanged
                                # helper groups.
                                has_attachment = _has_existing_week_attachment(
                                    client, TARGET_SHEET_ID, target_row,
                                    str(wr_num), week_raw, variant,
                                    file_identifier,
                                    cached_attachments=attachment_cache.get(target_row.id),
                                )
                                if not has_attachment:
                                    can_skip = False
                        if can_skip:
                            logging.info(f"⏩ Skip (unchanged + attachment exists) {variant} WR {wr_num} week {week_raw} hash {data_hash}")
                            _groups_skipped += 1
                            sentry_add_breadcrumb("group", f"Skipped unchanged group", level="info", data={
                                "wr": wr_num, "week": week_raw, "variant": variant, "hash": data_hash,
                            })
                            continue
                        else:
                            logging.info(f"🔁 Regenerating {variant} WR {wr_num} week {week_raw} despite unchanged hash (attachment missing or verification failed)")
                            sentry_add_breadcrumb("group", f"Regenerating despite same hash (attachment missing)", level="warning", data={
                                "wr": wr_num, "week": week_raw, "variant": variant,
                            })
                
                # Generate Excel file with complete fixes
                with sentry_sdk.start_span(op="excel.generate", name=f"Generate Excel for WR {wr_num}") as gen_span:
                    gen_span.set_data("group_key", group_key)
                    gen_span.set_data("row_count", len(group_rows))
                    gen_span.set_data("variant", variant)
                    gen_span.set_data("group_index", group_idx)
                    # Phase 01 Plan 03 Task 2 / Blocker 4: 5-tuple
                    # return. ``customer_name`` is forwarded to Plan 04's
                    # upload-task builder; ``missing_cus`` accumulates
                    # per source sheet into ``_missing_cus_by_sheet``
                    # for the D-17 end-of-loop WARNING.
                    (
                        excel_path,
                        filename,
                        wr_numbers,
                        _customer_name,
                        _missing_cus_for_group,
                    ) = generate_excel(
                        group_key, group_rows, snapshot_date, data_hash=data_hash
                    )
                    gen_span.set_data("filename", filename)

                # Attribute missing CUs to each source sheet that
                # contributed rows to this group (a single group can
                # span sheets when multiple sheets carry the same WR).
                # Distinct sheets get their own bucket so the per-sheet
                # WARNING surfaces the correct sheet id; rows missing
                # ``__source_sheet_id`` are bucketed under -1 so they
                # still surface in operator logs without crashing the
                # attribution loop.
                #
                # Phase 01 gap closure (REVIEW-WR-06): standardized on
                # ``__source_sheet_id`` (Phase 1 canonical field name)
                # instead of the legacy alias ``__sheet_id``. Both
                # fields are written to the same ``source['id']`` value
                # at populate time in ``_fetch_and_process_sheet``, so
                # the runtime behavior is unchanged today. The
                # migration ensures a future refactor that splits the
                # two field names cannot silently route missing-CU
                # WARNINGs to sheet -1 (the fallback bucket).
                if _missing_cus_for_group:
                    _contributing_sheet_ids: set[int] = set()
                    for _r in group_rows:
                        _sid = _r.get('__source_sheet_id')
                        if isinstance(_sid, int):
                            _contributing_sheet_ids.add(_sid)
                    if not _contributing_sheet_ids:
                        _contributing_sheet_ids = {-1}
                    for _sid in _contributing_sheet_ids:
                        _missing_cus_by_sheet[_sid].update(_missing_cus_for_group)
                
                generated_files_count += 1
                _groups_generated += 1
                generated_filenames.append(filename)
                
                # Collect upload task(s) for parallel processing
                # (instead of uploading serially). ``wr_numbers`` is
                # returned raw by ``generate_excel`` — do NOT read
                # from it here; the filename, hash-history key,
                # attachment prefix match, and target_map key all use
                # the sanitised main-loop ``wr_num`` and must stay
                # aligned to avoid repeated regeneration and orphaned
                # duplicate attachments on subsequent runs.
                #
                # Phase 01 Plan 04 Task 2: dispatch routing decisions
                # to ``_build_upload_tasks_for_group``. For
                # ``reduced_sub`` / ``reduced_sub_helper`` variants the
                # helper returns TWO tasks (one per target sheet); for
                # every other variant it returns ONE task on
                # ``TARGET_SHEET_ID``. Each task carries its own
                # ``target_sheet_id`` so the ``_upload_one`` worker
                # routes to the correct sheet without consulting a
                # global.
                if not TEST_MODE and wr_num:
                    _new_upload_tasks = _build_upload_tasks_for_group(
                        variant=variant,
                        wr_num=wr_num,
                        target_map=target_map,
                        target_map_ppp=target_map_ppp,
                        excel_path=excel_path,
                        filename=filename,
                        identifier=identifier,
                        file_identifier=file_identifier,
                        data_hash=data_hash,
                        week_raw=week_raw,
                        group_key=group_key,
                    )
                    _upload_tasks.extend(_new_upload_tasks)

                # Update hash history with variant-aware key (even in TEST_MODE so future prod runs can leverage)
                hash_history[history_key] = {
                    'hash': data_hash,
                    'rows': len(group_rows),
                    'updated_at': datetime.datetime.now(datetime.timezone.utc).isoformat(),
                    'foreman': group_rows[0].get('__current_foreman'),
                    'week': week_raw,
                    'variant': variant,
                    'identifier': identifier,
                }
                history_updates += 1

                # Sub-project E: shadow-write the durable per-group content
                # hash to Supabase alongside the local json cache. Gated on
                # SUPABASE_HASH_STORE_WRITE_ENABLED (default ON) — harmless
                # while the store is not yet authoritative: it just populates
                # billing_audit.group_content_hash so the eventual
                # authoritative flip has data to read. ``upsert_group_hash``
                # is fail-safe (returns a no-op when Supabase is unavailable /
                # TEST_MODE and never raises); the extra guard keeps a future
                # regression from breaking the generation path. ``week_iso``
                # is the ISO DATE the column expects (NOT the MMDDYY
                # week_raw), kept consistent with lookup_group_hash in the
                # skip gate above.
                if (
                    SUPABASE_HASH_STORE_WRITE_ENABLED
                    and BILLING_AUDIT_AVAILABLE
                    and not TEST_MODE
                    and week_iso
                ):
                    # ``week_iso`` is guarded truthy: week_ending is a DATE
                    # column, so an empty string (missing __week_ending_date)
                    # would be a PostgREST type error that could trip the
                    # per-op circuit breaker. Skipping the shadow write for
                    # such an edge-case group is harmless — the json cache
                    # and (until authoritative) the filename hash still drive
                    # change detection.
                    try:
                        _billing_audit_writer.upsert_group_hash(
                            wr_num, week_iso, variant,
                            identifier or '', data_hash,
                        )
                    except Exception:
                        logging.exception(
                            "E shadow hash write failed (non-fatal)")
                
            except Exception as e:
                _groups_errored += 1
                logging.error(f"❌ Failed to process group {group_key}: {e}")
                sentry_capture_with_context(
                    exception=e,
                    context_name="group_processing_error",
                    context_data={
                        "group_key": group_key,
                        "group_index": group_idx,
                        "total_groups": len(groups),
                        "wr_number": wr_num if 'wr_num' in dir() else 'unknown',
                        "week_ending": week_raw if 'week_raw' in dir() else 'unknown',
                        "variant": variant if 'variant' in dir() else 'unknown',
                        "row_count": len(group_rows),
                        "error_type": type(e).__name__,
                        "error_message": _redact_exception_message(e),
                        "traceback": traceback.format_exc(),
                    },
                    tags={
                        "error_location": "group_processing",
                        "group_key": group_key[:50],  # Truncate for tag limit
                    },
                    fingerprint=["group-processing", type(e).__name__]
                )
                continue
        
        _phase_group_elapsed = (datetime.datetime.now() - _phase_group_start).total_seconds()
        logging.info(f"⚡ Group processing phase: {_groups_generated} generated, {_groups_skipped} skipped in {_phase_group_elapsed:.1f}s"
                     + (f" (stopped early — time budget exceeded)" if _time_budget_exceeded else ""))

        # Phase 01 Plan 03 Task 2 Change 3 (D-17): emit exactly ONE
        # WARNING per source sheet whose subcontractor variant
        # generation fell through to SmartSheet pricing on missing
        # CU codes. The first 10 CU codes (alphabetical) are named so
        # operators get an immediate, bounded, actionable surface
        # without log-line blowout when many CUs are missing at once.
        # Suppressed entirely when the kill switch is off — there is
        # no subcontractor variant work to surface in that case. The
        # WARNING template includes the stable marker "Subcontractor
        # rates CSV missing" so Plan 02's ``_PII_LOG_MARKERS``
        # extension drops it from Sentry before send.
        if SUBCONTRACTOR_RATE_VARIANTS_ENABLED and _missing_cus_by_sheet:
            for _sid, _sheet_missing_cus in _missing_cus_by_sheet.items():
                if not _sheet_missing_cus:
                    continue
                N = len(_sheet_missing_cus)
                first_10 = ', '.join(sorted(_sheet_missing_cus)[:10])
                ellipsis = '...' if N > 10 else ''
                logging.warning(
                    f"Subcontractor rates CSV missing {N} CU code(s) on "
                    f"sheet {_sid}: {first_10}{ellipsis}. Add to "
                    f"{SUBCONTRACTOR_RATES_CSV} to enable rate recalc for "
                    f"these rows. Sheet rows fell through to SmartSheet pricing."
                )

        # ── PARALLEL UPLOAD PHASE ─────────────────────────────────────────
        # Upload all collected tasks in parallel instead of serially per-group.
        # This is the primary runtime optimization — reduces upload time by ~Nx with N workers.
        if _upload_tasks:
            _upload_start = datetime.datetime.now()
            logging.info(f"\n{'='*60}")
            logging.info(f"📤 PARALLEL UPLOAD PHASE: {len(_upload_tasks)} files with {PARALLEL_WORKERS} workers")
            logging.info(f"{'='*60}")

            def _upload_one(task):
                """Delete old attachment + upload new one for a single group.

                Phase 01 Plan 04 Task 2: routing target is resolved
                from ``task['target_sheet_id']`` instead of the
                module-level primary sheet id. The upload-task
                builder (``_build_upload_tasks_for_group``) sets the
                sheet id per-task — ``primary`` / ``aep_billable``
                / etc. point at the primary sheet; the second leg
                of a ``reduced_sub`` fan-out points at the
                subcontractor PPP sheet. The worker is otherwise
                oblivious to which sheet it is uploading to — and
                that's the point: routing decisions live in the
                builder, mutations live in the worker.
                """
                max_retries = 4
                last_err = None
                for attempt in range(max_retries):
                    try:
                        target_row = task['target_row']
                        force_this = FORCE_GENERATION or (task['week_raw'] in REGEN_WEEKS)

                        deleted_count, skipped = delete_old_excel_attachments(
                            client, task['target_sheet_id'], target_row, task['wr_num'],
                            task['week_raw'], task['data_hash'],
                            variant=task['variant'], identifier=task['file_identifier'],
                            force_generation=force_this,
                            cached_attachments=attachment_cache.get(target_row.id)
                        )
                        if force_this and skipped:
                            skipped = False

                        if skipped:
                            return 'skipped'

                        if not SKIP_UPLOAD:
                            with open(task['excel_path'], 'rb') as file:
                                client.Attachments.attach_file_to_row(
                                    task['target_sheet_id'],
                                    target_row.id,
                                    (task['filename'], file, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                                )
                            logging.info(
                                f"✅ Uploaded: {task['filename']} → sheet "
                                f"{task['target_sheet_id']}"
                            )
                            return 'uploaded'
                        else:
                            logging.info(f"⏭️  Skipping upload (SKIP_UPLOAD=true): {task['filename']}")
                            return 'skip_upload'
                    except ss_exc.RateLimitExceededError as e:
                        last_err = e
                        if attempt < max_retries - 1:
                            backoff = 15 * (attempt + 1)  # 15s, 30s, 45s for rate limits
                            logging.warning(f"⚠️ Rate limited on upload for {task['filename']}, backoff {backoff}s (attempt {attempt+1}/{max_retries})")
                            time.sleep(backoff)
                        else:
                            break
                    except (ss_exc.UnexpectedErrorShouldRetryError, ss_exc.InternalServerError, ss_exc.ServerTimeoutExceededError) as e:
                        last_err = e
                        if attempt < max_retries - 1:
                            backoff = 2 ** attempt + 0.5
                            logging.warning(f"⚠️ Upload retry {attempt+1}/{max_retries} for {task['filename']} ({type(e).__name__}), backoff {backoff:.1f}s")
                            time.sleep(backoff)
                        else:
                            break
                    except Exception as e:
                        last_err = e
                        err_name = type(e).__name__
                        is_transient = any(tag in err_name for tag in ('RemoteDisconnected', 'ConnectionError', 'ConnectionReset', 'SSLError', 'SSLEOFError', 'Timeout'))
                        if is_transient and attempt < max_retries - 1:
                            backoff = 2 ** attempt + 0.5
                            logging.warning(f"⚠️ Upload retry {attempt+1}/{max_retries} for {task['filename']} ({err_name}), backoff {backoff:.1f}s")
                            time.sleep(backoff)
                        else:
                            break
                logging.error(f"❌ Upload failed for {task['filename']}: {last_err}")
                sentry_add_breadcrumb("upload", f"Upload failed for {task['filename']}", level="error", data={
                    "wr": task['wr_num'], "error": str(last_err)[:200],
                })
                return 'error'

            with ThreadPoolExecutor(max_workers=PARALLEL_WORKERS) as executor:
                upload_results = list(executor.map(_upload_one, _upload_tasks))

            _groups_uploaded = sum(1 for r in upload_results if r == 'uploaded')
            _upload_errors = sum(1 for r in upload_results if r == 'error')
            _groups_errored += _upload_errors
            _api_calls_count = _groups_uploaded

            _upload_elapsed = (datetime.datetime.now() - _upload_start).total_seconds()
            logging.info(f"⚡ Upload phase complete: {_groups_uploaded} uploaded, {_upload_errors} errors in {_upload_elapsed:.1f}s (parallel w/{PARALLEL_WORKERS} workers)")

        # Validation summary
        summaries = validate_group_totals(groups)
        if summaries:
            logging.info("🧮 Totals Validation (first 10 groups):")
            for s in summaries[:10]:
                logging.info(f"   {s['group_key']}: rows={s['rows']} total=${s['total']}")

        # Session summary
        session_duration = datetime.datetime.now() - session_start
        logging.info(f"✅ Session complete!")
        logging.info(f"   • Files generated: {generated_files_count}")
        logging.info(f"   • Duration: {session_duration}")
        logging.info(f"   • Mode: {'TEST' if TEST_MODE else 'PRODUCTION'}")

        # Build identity set for sheet pruning: (wr, week, variant, identifier) 4-tuples
        valid_wr_weeks = set()
        for fname in generated_filenames:
            ident = build_group_identity(fname)
            if ident:
                valid_wr_weeks.add(ident)  # Already returns 4-tuple
        # Also include any WR/week/variant/identifier combos we skipped due to identical hash (so we don't delete their existing attachment)
        # Already implicit because skipped groups did not regenerate; we can add from groups processed via grouping keys
        for key, group_rows in groups.items():
            if '_' in key:
                week_raw = key.split('_',1)[0]
                wr_raw = group_rows[0].get('Work Request #')
                wr = str(wr_raw).split('.')[0] if wr_raw else ''
                # Apply the same sanitizer used at every other site
                # (generate_excel, main-loop derivation, hash-prune
                # loop, create_target_sheet_map). Without this,
                # ``build_group_identity`` (which returns sanitized
                # WR tokens for filenames with rewritten WR#s) would
                # produce identity tuples that don't match the
                # unsanitized entries this loop adds to
                # valid_wr_weeks — causing
                # cleanup_untracked_sheet_attachments to incorrectly
                # prune attachments for sanitization-sensitive WRs
                # when KEEP_HISTORICAL_WEEKS is enabled.
                wr = _RE_SANITIZE_HELPER_NAME.sub('_', wr)[:50]
                variant = group_rows[0].get('__variant', 'primary')
                if variant in ('helper', 'aep_billable_helper', 'reduced_sub_helper'):
                    # CR-01 gap closure (Site 2 — mirror of Site 1).
                    # build_group_identity returns the sanitized helper
                    # foreman as the parsed identifier for all three
                    # helper-style variants; valid_wr_weeks must match
                    # that tuple shape so
                    # cleanup_untracked_sheet_attachments correctly
                    # identifies which helper-shadow attachments are
                    # "live" and which are stale. Pre-fix, shadow
                    # variants fell through to the ``User``-derived
                    # ``else`` branch and produced file_id='' tuples
                    # that NEVER matched the parser's 'Jane_Smith'
                    # identifier — risking cleanup either pruning
                    # legitimate attachments or missing orphans.
                    # Sites 1 and 3 carry the same gate.
                    helper_foreman = group_rows[0].get('__helper_foreman', '')
                    file_id = _RE_SANITIZE_HELPER_NAME.sub('_', helper_foreman)[:50] if helper_foreman else ''
                elif variant == 'vac_crew':
                    # Subproject C identity site (Site 2 — valid_wr_weeks).
                    # GATED on the kill switch (mirrors Site 1): disabled mode
                    # produces file_id='' so the 4-tuple matches the bare
                    # _VacCrew attachment identity and cleanup does not delete
                    # live legacy-mode attachments.
                    _vc = group_rows[0].get('__current_foreman', '')
                    file_id = (
                        _RE_SANITIZE_IDENTIFIER.sub('_', _vc)[:50]
                        if (VAC_CREW_CLAIM_ATTRIBUTION_ENABLED and _vc) else ''
                    )
                elif variant in ('reduced_sub', 'aep_billable'):
                    # Subproject B identity site (Site 2 — valid_wr_weeks).
                    # Mirror Site 1 so attachment cleanup keeps the live
                    # per-claimer file.
                    _b_claimer = group_rows[0].get('__current_foreman', '')
                    file_id = (
                        _RE_SANITIZE_IDENTIFIER.sub('_', _b_claimer)[:50]
                        if _b_claimer else ''
                    )
                else:
                    # Subproject D (2026-05-25): primary identity site
                    # (Site 2 — valid_wr_weeks). Mirror Site 1 so attachment
                    # cleanup keeps the live per-claimer primary file.
                    # Disabled mode preserves the legacy ``User``-field path.
                    if (
                        PRIMARY_CLAIM_ATTRIBUTION_ENABLED
                        and RES_GROUPING_MODE in ('helper', 'both')
                    ):
                        _pf = group_rows[0].get('__current_foreman', '')
                        file_id = (
                            _RE_SANITIZE_IDENTIFIER.sub('_', _pf)[:50]
                            if (PRIMARY_CLAIM_ATTRIBUTION_ENABLED and _pf) else ''
                        )
                    else:
                        user_val = group_rows[0].get('User')
                        # PERFORMANCE: Use pre-compiled regex
                        file_id = _RE_SANITIZE_IDENTIFIER.sub('_', user_val)[:50] if user_val else ''
                valid_wr_weeks.add((wr, week_raw, variant, file_id))
        if not TEST_MODE:
            # Invalidate stale attachment cache after upload phase — uploads added/deleted attachments
            _cleanup_cache = attachment_cache if not _upload_tasks else None
            # Phase 1.1 Bug B2 (D-09): TARGET_SHEET_ID cleanup is UNCHANGED —
            # accepts every variant currently routed to it (primary, helper,
            # vac_crew, aep_billable, reduced_sub, aep_billable_helper,
            # reduced_sub_helper). The whitelist is per-sheet; passing
            # variant_whitelist=None (default — kwarg omitted below)
            # preserves byte-identical legacy behaviour on TARGET.
            #
            # Phase 1.1 UAT gap closure (SUB-09 helper dimension): build the
            # subcontractor WR scope from this run's groups (shared helper)
            # and pass it to the TARGET cleanup to delete pre-existing legacy
            # _Helper_<name> and bare-primary attachments. Kill-switch-gated:
            # SUBCONTRACTOR_LEGACY_HELPER_CLEANUP_ENABLED=0 reverts to
            # byte-identical pre-fix TARGET behaviour (sub orphans persist).
            # Subproject B: build the subcontractor WR scope when EITHER
            # the legacy-helper cleanup (SUB-09) OR the legacy-primary
            # cleanup (Subproject B) is enabled — the two share the scope.
            _need_sub_scope = (
                SUBCONTRACTOR_LEGACY_HELPER_CLEANUP_ENABLED
                or SUBCONTRACTOR_LEGACY_PRIMARY_CLEANUP_ENABLED
            )
            _sub_scope = (
                _build_subcontractor_wr_scope(groups)
                if _need_sub_scope
                else None
            )
            _target_offcontract = set()
            if _sub_scope and SUBCONTRACTOR_LEGACY_HELPER_CLEANUP_ENABLED:
                _target_offcontract |= {'helper', 'primary'}
            _target_legacy_primary = (
                {'reduced_sub', 'aep_billable'}
                if _sub_scope and SUBCONTRACTOR_LEGACY_PRIMARY_CLEANUP_ENABLED
                else None
            )
            # Subproject C Task 6 (2026-05-21): build the vac_crew WR scope
            # for legacy bare _VacCrew cleanup on TARGET. vac_crew files route
            # to TARGET_SHEET_ID only (never PPP) — do NOT pass this to PPP.
            # Kill-switch-gated: VAC_CREW_LEGACY_CLEANUP_ENABLED=0 reverts to
            # byte-identical pre-fix TARGET behaviour (vac orphans persist).
            _vac_scope = (
                _build_vac_crew_wr_scope(groups)
                if VAC_CREW_LEGACY_CLEANUP_ENABLED
                else None
            )
            # Subproject D (2026-05-25): build the non-subcontractor
            # primary WR scope for legacy bare-primary cleanup on TARGET.
            # Gated on BOTH the attribution kill switch (the partitioned
            # _USER_ groups only exist when attribution is on) AND the
            # cleanup kill switch. primary files route to TARGET only —
            # do NOT pass this to PPP.
            _primary_scope = (
                _build_primary_wr_scope(groups)
                if (
                    PRIMARY_CLAIM_ATTRIBUTION_ENABLED
                    and LEGACY_PRIMARY_PARTITION_CLEANUP_ENABLED
                )
                else None
            )
            with sentry_sdk.start_span(op="smartsheet.cleanup", name="Cleanup untracked sheet attachments"):
                cleanup_untracked_sheet_attachments(
                    client, TARGET_SHEET_ID, valid_wr_weeks, TEST_MODE,
                    attachment_cache=_cleanup_cache, target_sheet=_target_sheet_obj,
                    sub_wr_scope=_sub_scope,
                    # Empty set means the SUB-09 helper cleanup is off; coerce
                    # to None so the off-contract gate no-ops (the gate keys on
                    # `is not None`, not truthiness).
                    sub_offcontract_variants=(_target_offcontract or None),
                    sub_legacy_primary_variants=_target_legacy_primary,
                    vac_legacy_wr_scope=_vac_scope,
                    primary_wr_scope=_primary_scope,
                )

            # Phase 01 gap closure (REVIEW-WR-01): parallel cleanup pass
            # for SUBCONTRACTOR_PPP_SHEET_ID. The TARGET_SHEET_ID
            # cleanup above iterates one sheet only; without an
            # equivalent pass on PPP, any helper-shadow attachment
            # (``_AEPBillable_Helper_*`` / ``_ReducedSub_Helper_*``)
            # whose per-row ``delete_old_excel_attachments`` call
            # missed (CR-01 pre-fix bug, timestamp-identity drift,
            # future refactor) orphans permanently on PPP. This
            # invocation is the belt-and-suspenders defense: it
            # iterates PPP rows, groups attachments by parsed identity
            # tuple, and prunes everything-but-newest per identity.
            #
            # ``valid_wr_weeks`` is the SHARED authority — Plan 08
            # (CR-01) ensured shadow-variant entries are correctly
            # included so live attachments are not pruned.
            #
            # Cache semantics: ``_cleanup_cache`` is computed ABOVE
            # both invocations as ``attachment_cache if not _upload_tasks
            # else None``. In the normal production case (uploads ran
            # this session, ``_upload_tasks`` truthy), ``_cleanup_cache``
            # is ``None`` for BOTH passes because uploads invalidate
            # the prefetch snapshot. When no uploads ran (TEST_MODE
            # skip path, or no-changes branch), both passes share
            # WR-05's prefetched dict transparently. WR-05's prefetch
            # primarily amortizes per-row ``_upload_one`` API calls
            # (its real value); the cleanup-time benefit is only on
            # the no-uploads path. Either way, passing the same
            # ``_cleanup_cache`` keeps cache semantics consistent
            # across both passes.
            #
            # Gates (in order, short-circuit on first False):
            #   1. SUBCONTRACTOR_RATE_VARIANTS_ENABLED (kill switch)
            #   2. SUBCONTRACTOR_PPP_SHEET_ID is truthy (disable case)
            #   3. SUBCONTRACTOR_PPP_SHEET_ID != TARGET_SHEET_ID
            #      (skip redundant pass if operator points both to
            #       the same sheet — unusual but supported)
            #   4. _target_sheet_ppp_obj is not None (Plan 04 only
            #      populates this when target_map_ppp was successfully
            #      built; None means PPP routing was unreachable this
            #      run and we should not iterate the sheet)
            if (
                SUBCONTRACTOR_RATE_VARIANTS_ENABLED
                and SUBCONTRACTOR_PPP_SHEET_ID
                and SUBCONTRACTOR_PPP_SHEET_ID != TARGET_SHEET_ID
                and _target_sheet_ppp_obj is not None
            ):
                with sentry_sdk.start_span(op="smartsheet.cleanup_ppp", name="Cleanup untracked PPP sheet attachments"):
                    # Phase 1.1 Bug B2 (D-07 / D-08 / SUB-10):
                    # per-sheet variant whitelist. PPP receives only
                    # `_ReducedSub` / `_ReducedSub_Helper_*` from
                    # Phase 1's routing matrix (per
                    # _build_upload_tasks_for_group). Any other
                    # variant parsed from a filename on PPP is
                    # off-contract and unconditionally pruned —
                    # defense in depth against Bug B1 regressions
                    # AND against future routing-matrix drift.
                    # Hardcoded at the call site per D-08 (no env
                    # var, no config). If a future plan adds a new
                    # variant to PPP routing (e.g., aep_billable),
                    # this literal whitelist MUST be updated in the
                    # SAME PR — coupling is documented in the
                    # 01.1-03 SUMMARY.
                    cleanup_untracked_sheet_attachments(
                        client,
                        SUBCONTRACTOR_PPP_SHEET_ID,
                        valid_wr_weeks,
                        TEST_MODE,
                        attachment_cache=_cleanup_cache,
                        target_sheet=_target_sheet_ppp_obj,
                        variant_whitelist={'reduced_sub', 'reduced_sub_helper'},
                        sub_wr_scope=_sub_scope,
                        sub_legacy_primary_variants=(
                            {'reduced_sub'}
                            if _sub_scope and SUBCONTRACTOR_LEGACY_PRIMARY_CLEANUP_ENABLED
                            else None
                        ),
                    )

        # Cleanup legacy / stale Excel files so only current system outputs remain
        try:
            with sentry_sdk.start_span(op="file.cleanup", name="Cleanup stale local Excel files"):
                removed = cleanup_stale_excels(OUTPUT_FOLDER, set(generated_filenames))
            logging.info(f"🧹 Cleanup complete: removed {len(removed)} stale file(s)")
        except Exception as e:
            logging.warning(f"⚠️ Cleanup step failed: {e}")
        
        # Audit summary
        if audit_results:
            audit_summary = audit_results.get('summary', {})
            logging.info(f"🔍 Audit Summary:")
            logging.info(f"   • Risk Level: {audit_summary.get('risk_level', 'UNKNOWN')}")
            logging.info(f"   • Anomalies: {audit_summary.get('total_anomalies', 0)}")
            logging.info(f"   • Data Issues: {audit_summary.get('total_data_issues', 0)}")
        
        # Persist hash history if updated
        if history_updates:
            # Prune stale hash_history entries for groups no longer in source data.
            # Only prune on FULL runs (not time-budget-truncated runs) to avoid
            # deleting entries for groups that simply weren't reached this run.
            if not _time_budget_exceeded:
                current_keys = set()
                for key, group_rows in groups.items():
                    if '_' in key:
                        _wr_raw = group_rows[0].get('Work Request #')
                        _wr = str(_wr_raw).split('.')[0] if _wr_raw else ''
                        # Codex P2: apply the same filesystem-safety
                        # sanitizer used by the main loop (line ~4493)
                        # so the current_keys tuple matches the
                        # history_key actually written for this group.
                        # Without this, any WR# containing
                        # sanitization-sensitive characters would have
                        # its freshly-written entry treated as stale
                        # and deleted before save, so hash-skip could
                        # never persist across runs for those WRs.
                        _wr = _RE_SANITIZE_HELPER_NAME.sub('_', _wr)[:50]
                        _week = key.split('_',1)[0]
                        _variant = group_rows[0].get('__variant', 'primary')
                        if _variant in ('helper', 'aep_billable_helper', 'reduced_sub_helper'):
                            # CR-01 gap closure (Site 3 — mirror of Site 1).
                            # Site 1 writes the helper-shadow history_key as
                            # f"{wr}|{week}|{variant}|{foreman}|{dept}|{job}" —
                            # this prune-key reconstruction MUST match it
                            # byte-for-byte or the entry written this run is
                            # treated as stale and deleted before
                            # save_hash_history runs. Pre-fix, both Sites 1
                            # and 3 fell through to the same ``User``-derived
                            # branch, so the two stayed aligned by accident
                            # (both produced '' identifiers). With Site 1 now
                            # correctly deriving from __helper_foreman, Site 3
                            # must follow or the alignment breaks the OTHER
                            # way and we permanently lose hash-skip for
                            # helper-shadow variants. Note: ``_ident`` here is
                            # the HISTORY-KEY shape (pipe-joined triple), NOT
                            # the FILE-IDENTIFIER shape (Site 1 builds both;
                            # this site reconstructs the history-key shape
                            # only — the same pattern as the legacy helper
                            # branch).
                            _hf = group_rows[0].get('__helper_foreman', '')
                            _hd = group_rows[0].get('__helper_dept', '')
                            _hj = group_rows[0].get('__helper_job', '')
                            _ident = f"{_hf}|{_hd}|{_hj}"
                        elif _variant == 'vac_crew':
                            # Subproject C identity site (Site 3 —
                            # current_keys). GATED on the kill switch (mirrors
                            # Site 1): disabled mode produces _ident='' so the
                            # reconstructed current_keys entry matches the
                            # bare history_key written by Site 1 and the fresh
                            # entry is not treated as stale and deleted.
                            _vc = group_rows[0].get('__current_foreman', '')
                            _ident = (
                                _RE_SANITIZE_IDENTIFIER.sub('_', _vc)[:50]
                                if (VAC_CREW_CLAIM_ATTRIBUTION_ENABLED and _vc) else ''
                            )
                        elif _variant in ('reduced_sub', 'aep_billable'):
                            # Subproject B identity site (Site 3 —
                            # current_keys). Must match the history_key
                            # written at Site 1 byte-for-byte (sanitized
                            # claimer) or the freshly-written entry is
                            # treated as stale and deleted before save.
                            _b_claimer = group_rows[0].get('__current_foreman', '')
                            _ident = (
                                _RE_SANITIZE_IDENTIFIER.sub('_', _b_claimer)[:50]
                                if _b_claimer else ''
                            )
                        else:
                            # Subproject D (2026-05-25): primary identity
                            # site (Site 3 — current_keys). Must match the
                            # history_key written at Site 1 byte-for-byte
                            # (sanitized claimer when on, legacy User-field
                            # when off) or the freshly-written entry is
                            # treated as stale and deleted before save.
                            if (
                                PRIMARY_CLAIM_ATTRIBUTION_ENABLED
                                and RES_GROUPING_MODE in ('helper', 'both')
                            ):
                                _pf = group_rows[0].get('__current_foreman', '')
                                _ident = (
                                    _RE_SANITIZE_IDENTIFIER.sub('_', _pf)[:50]
                                    if (PRIMARY_CLAIM_ATTRIBUTION_ENABLED and _pf) else ''
                                )
                            else:
                                _uv = group_rows[0].get('User')
                                _ident = _RE_SANITIZE_IDENTIFIER.sub('_', _uv)[:50] if _uv else ''
                        current_keys.add(f"{_wr}|{_week}|{_variant}|{_ident}")
                stale_keys = [k for k in hash_history if k not in current_keys]
                if stale_keys:
                    for sk in stale_keys:
                        del hash_history[sk]
                    logging.info(f"🧹 Pruned {len(stale_keys)} stale hash history entries (groups no longer in source data)")
            save_hash_history(HASH_HISTORY_PATH, hash_history)
        elif _hash_history_migration_dirty:
            # Codex P2: no group updates this run, but a one-time migration
            # prune (Phase 1.1 / Subproject B / Subproject C) mutated hash_history. Persist
            # it now so the migration is durable and does not re-run every
            # execution. Do NOT run the stale-prune on this path — groups
            # were not fully processed, so current_keys would be incomplete
            # and could delete freshly-skipped live entries.
            save_hash_history(HASH_HISTORY_PATH, hash_history)
        if (
            BILLING_AUDIT_AVAILABLE
            and not TEST_MODE
            and billing_audit_row_cache_dirty
        ):
            save_billing_audit_row_cache(
                BILLING_AUDIT_ROW_CACHE_PATH,
                billing_audit_row_cache,
            )

        # Write run summary JSON for downstream consumers (Notion sync, dashboards)
        _run_summary = {
            "success": True,
            "files_generated": generated_files_count,
            "groups_total": len(groups),
            "groups_skipped": _groups_skipped,
            "groups_generated": _groups_generated,
            "groups_uploaded": _groups_uploaded,
            "groups_errored": _groups_errored,
            "duration_seconds": session_duration.total_seconds(),
            "duration_minutes": round(session_duration.total_seconds() / 60.0, 2),
            "history_updates": history_updates,
            "sheets_discovered": len(source_sheets) if 'source_sheets' in dir() else 0,
            "rows_fetched": len(all_rows) if 'all_rows' in dir() else 0,
            "api_calls": _api_calls_count,
            "audit_risk_level": audit_results.get('summary', {}).get('risk_level', 'UNKNOWN') if audit_results else 'UNKNOWN',
            "mode": "TEST" if TEST_MODE else "PRODUCTION",
            "timestamp": datetime.datetime.now(datetime.timezone.utc).isoformat(),
            "snapshots_written": 0,
            "snapshots_already_frozen": 0,
            "snapshots_errored": 0,
            "fingerprint_changes_detected": 0,
        }
        if BILLING_AUDIT_AVAILABLE:
            try:
                _run_summary.update(_billing_audit_writer.get_counters())
            except Exception:
                pass  # Counter retrieval must never fail the run summary write.
            # Subproject B / Subproject C: emit ONE aggregate WARNING if any
            # rows were held this run pending attribution (Supabase outage).
            # B is the first consumer of Foundation A's HOLD machinery; C
            # (vac_crew) also records holds via the same counter. This is
            # the single end-of-run summary call. PII-safe (counts +
            # sanitized WR list only). Never fail the run summary write.
            try:
                _billing_audit_writer.summarize_attribution_holds()
            except Exception:
                pass
        try:
            with open(os.path.join(OUTPUT_FOLDER, 'run_summary.json'), 'w') as _rsf:
                json.dump(_run_summary, _rsf, indent=2)
        except Exception as _rse:
            logging.warning(f"⚠️ Could not write run_summary.json: {_rse}")

        # SDK 2.x: Use get_isolation_scope() instead of configure_scope()
        if SENTRY_DSN:
            scope = sentry_sdk.get_isolation_scope()
            scope.set_tag("session_success", "true")
            scope.set_tag("files_generated", str(generated_files_count))
            scope.set_tag("groups_skipped", str(_groups_skipped))
            scope.set_tag("groups_generated", str(_groups_generated))
            scope.set_tag("groups_uploaded", str(_groups_uploaded))
            scope.set_tag("groups_errored", str(_groups_errored))
            scope.set_tag("session_duration_seconds", str(session_duration.total_seconds()))
            if audit_results:
                scope.set_tag("audit_risk_level", audit_results.get('summary', {}).get('risk_level', 'UNKNOWN'))
            
            # Set final session context for dashboard visibility
            sentry_sdk.set_context("session_summary", {
                "success": True,
                "files_generated": generated_files_count,
                "groups_total": len(groups),
                "groups_skipped": _groups_skipped,
                "groups_generated": _groups_generated,
                "groups_uploaded": _groups_uploaded,
                "groups_errored": _groups_errored,
                "duration_seconds": session_duration.total_seconds(),
                "duration_human": str(session_duration),
                "history_updates": history_updates,
                "mode": "TEST" if TEST_MODE else "PRODUCTION",
                "audit_risk_level": audit_results.get('summary', {}).get('risk_level', 'UNKNOWN') if audit_results else None,
            })
            sentry_sdk.set_context("data_pipeline", {
                "source_sheets": len(source_sheets) if 'source_sheets' in dir() else 0,
                "total_rows_fetched": len(all_rows) if 'all_rows' in dir() else 0,
                "groups_created": len(groups),
                "hash_history_entries": len(hash_history) if 'hash_history' in dir() else 0,
                "api_calls_upload": _api_calls_count,
            })
            sentry_add_breadcrumb("session", "Session completed successfully", level="info", data={
                "files_generated": generated_files_count,
                "duration": str(session_duration),
                "skipped": _groups_skipped,
                "errored": _groups_errored,
            })
            
            # #6 - SUCCESS-path root-transaction KPIs (counts only, no PII)
            if _txn:
                for _k, _v in _build_run_kpis(
                    files_generated=generated_files_count,
                    groups_total=len(groups),
                    groups_skipped=_groups_skipped,
                    groups_generated=_groups_generated,
                    groups_uploaded=_groups_uploaded,
                    groups_errored=_groups_errored,
                    duration_seconds=session_duration.total_seconds(),
                    sheets_discovered=len(source_sheets) if 'source_sheets' in dir() else 0,
                    rows_fetched=len(all_rows) if 'all_rows' in dir() else 0,
                    api_calls=_api_calls_count,
                ).items():
                    _txn.set_data(_k, _v)

            # #7 - milestone structured log: run complete (counts only, no PII)
            _sentry_log_event(
                "info",
                "weekly run complete",
                files_generated=generated_files_count,
                groups_generated=_groups_generated,
                groups_uploaded=_groups_uploaded,
                groups_errored=_groups_errored,
                duration_seconds=session_duration.total_seconds(),
            )

            # Finish the root transaction
            if _txn:
                _txn.set_status("ok")
                _txn.__exit__(None, None, None)
                _txn = None

    except FileNotFoundError as e:
        error_context = f"Missing required file: {e}"
        logging.error(f"💥 {error_context}")
        sentry_capture_with_context(
            exception=e,
            context_name="file_not_found",
            context_data={
                "missing_file": str(e),
                "working_directory": os.getcwd(),
                "error_type": "FileNotFoundError",
            },
            tags={"error_location": "main", "error_type": "file_not_found"},
            fingerprint=["file-not-found", str(e)]
        )
        # Close transaction with error
        if _txn:
            _txn.set_status("internal_error")
            _txn.__exit__(type(e), e, e.__traceback__)
            _txn = None
            
    except Exception as e:
        session_duration = datetime.datetime.now() - session_start
        error_context = f"Session failed after {session_duration}"
        logging.error(f"💥 {error_context}: {e}")
        
        # SDK 2.x: Use get_isolation_scope() instead of configure_scope()
        if SENTRY_DSN:
            scope = sentry_sdk.get_isolation_scope()
            scope.set_tag("session_success", "false")
            scope.set_tag("session_duration_seconds", str(session_duration.total_seconds()))
            scope.set_tag("failure_type", "general_exception")
            scope.set_tag("groups_errored", str(_groups_errored))
            scope.set_level("error")

            # #5 - FAILURE-path PII-safe attachment (counts/booleans only)
            # add_attachment bypasses before_send_log — this try/except guard
            # ensures a telemetry failure can NEVER mask the real exception.
            try:
                _snap = _build_run_context_snapshot(
                    success=False,
                    duration_seconds=session_duration.total_seconds(),
                    groups_attempted=len(groups) if 'groups' in dir() else 0,
                    groups_generated=_groups_generated,
                    groups_uploaded=_groups_uploaded if '_groups_uploaded' in dir() else 0,
                    groups_errored=_groups_errored,
                    error_type=type(e).__name__,
                )
                scope.add_attachment(
                    bytes=json.dumps(_snap, indent=2).encode("utf-8"),
                    filename="run-context.json",
                    content_type="application/json",
                )
            except Exception:
                pass  # telemetry must never mask the real failure

            sentry_capture_with_context(
                exception=e,
                context_name="session_failure",
                context_data={
                    "duration_seconds": session_duration.total_seconds(),
                    "duration_human": str(session_duration),
                    "error_type": type(e).__name__,
                    "error_message": _redact_exception_message(e),
                    "traceback": traceback.format_exc(),
                    "test_mode": TEST_MODE,
                    "groups_attempted": len(groups) if 'groups' in dir() else 'unknown',
                    "groups_generated": _groups_generated,
                    "groups_errored": _groups_errored,
                },
                tags={"error_location": "main", "session_phase": "execution"},
                fingerprint=["session-failure", type(e).__name__]
            )
        # Close transaction with error
        if _txn:
            _txn.set_status("internal_error")
            _txn.__exit__(type(e), e, e.__traceback__)
            _txn = None
    
    finally:
        # Sentry cron check-in: signal final status
        if SENTRY_DSN and _cron_checkin_id:
            try:
                _cron_ok = '_groups_errored' not in dir() or _groups_errored == 0
                capture_checkin(
                    monitor_slug=_cron_monitor_slug,
                    check_in_id=_cron_checkin_id,
                    status=MonitorStatus.OK if _cron_ok else MonitorStatus.ERROR,
                )
            except Exception as exc:
                logging.warning(f"⚠️ Sentry cron check-in (final) failed: {exc}")
        
        # Ensure any open transaction is closed
        if _txn:
            _txn.set_status("unknown")
            _txn.__exit__(None, None, None)
        
        # Flush Sentry events before process exits
        if SENTRY_DSN:
            sentry_sdk.flush(timeout=10)

if __name__ == "__main__":
    main()
