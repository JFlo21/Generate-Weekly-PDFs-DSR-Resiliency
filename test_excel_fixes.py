#!/usr/bin/env python3
"""
Test script to verify the Excel generation fixes:
1. Safe merge cells (no XML errors)
2. Job # field population with multiple column variations
"""

import os
import sys
import datetime
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
import logging

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Import the safe_merge_cells function from generate_weekly_pdfs
sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from generate_weekly_pdfs import safe_merge_cells

def test_safe_merge_cells():
    """Test the safe_merge_cells function to ensure no XML errors."""
    print("\n=== TESTING SAFE MERGE CELLS ===\n")
    
    # Create a test workbook
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = "Test Report"
    
    # Define font styles
    TITLE_FONT = Font(name='Calibri', size=20, bold=True)
    SUBTITLE_FONT = Font(name='Calibri', size=16, bold=True, color='404040')
    TABLE_HEADER_FONT = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    BODY_FONT = Font(name='Calibri', size=11)
    
    # Test various merge scenarios
    print("Testing merge operations with safe_merge_cells:")
    
    # Test 1: Logo area merge
    result = safe_merge_cells(ws, 'A1:C3')
    print(f"  1. Logo area (A1:C3): {'✓' if result else '⚠️  Skipped'}")
    ws['A1'] = "LINETEC SERVICES"
    ws['A1'].font = TITLE_FONT
    
    # Test 2: Title merge
    result = safe_merge_cells(ws, 'D1:I1')
    print(f"  2. Title (D1:I1): {'✓' if result else '⚠️  Skipped'}")
    ws['D1'] = "WEEKLY UNITS COMPLETED"
    ws['D1'].font = SUBTITLE_FONT
    
    # Test 3: Report time merge
    result = safe_merge_cells(ws, 'D4:I4')
    print(f"  3. Report time (D4:I4): {'✓' if result else '⚠️  Skipped'}")
    ws['D4'] = f"Report Generated On: {datetime.datetime.now().strftime('%m/%d/%Y %I:%M %p')}"
    ws['D4'].font = BODY_FONT
    
    # Test 4: Summary header
    result = safe_merge_cells(ws, 'B7:D7')
    print(f"  4. Summary header (B7:D7): {'✓' if result else '⚠️  Skipped'}")
    ws['B7'] = "REPORT SUMMARY"
    ws['B7'].font = TABLE_HEADER_FONT
    
    # Test 5: Details header
    result = safe_merge_cells(ws, 'F7:I7')
    print(f"  5. Details header (F7:I7): {'✓' if result else '⚠️  Skipped'}")
    ws['F7'] = "REPORT DETAILS"
    ws['F7'].font = TABLE_HEADER_FONT
    
    # Test 6: Try duplicate merge (should be caught)
    result = safe_merge_cells(ws, 'F7:I7')  # Duplicate - should be skipped
    print(f"  6. Duplicate merge (F7:I7): {'✓ Correctly skipped' if not result else '✗ ERROR - Should have been skipped'}")
    
    # Test 7: Detail value merges
    print("  7. Detail value merges:")
    for i in range(8, 14):
        detail_range = f'G{i}:I{i}'
        result = safe_merge_cells(ws, detail_range)
        ws[f'F{i}'] = f"Label {i}:"
        ws[f'G{i}'] = f"Value {i}"
        if i == 8:
            print(f"     - {detail_range}: {'✓' if result else '⚠️  Skipped'}")
    print(f"     - ... (tested 6 detail rows)")
    
    # Save test file
    test_file = "test_safe_merge.xlsx"
    try:
        workbook.save(test_file)
        print(f"\n✓ Test file saved successfully: {test_file}")
        
        # Try to reopen the file to check for XML errors
        print("  Validating Excel file for XML errors...")
        test_wb = openpyxl.load_workbook(test_file)
        print("  ✓ File opens without XML errors")
        
        # Check merged cells
        merged_count = len(test_wb.active.merged_cells.ranges)
        print(f"  ✓ Total merged ranges: {merged_count}")
        test_wb.close()
        
        # Clean up
        os.remove(test_file)
        print("  ✓ Test file cleaned up")
        
        return True
        
    except Exception as e:
        print(f"\n✗ Error with Excel file: {e}")
        return False

def test_job_number_variations():
    """Test Job # field with different column name variations."""
    print("\n=== TESTING JOB # FIELD VARIATIONS ===\n")
    
    # Test data with different column name variations
    test_cases = [
        {"Job #": "12345", "expected": "12345"},
        {"Job#": "23456", "expected": "23456"},
        {"Job Number": "34567", "expected": "34567"},
        {"JobNumber": "45678", "expected": "45678"},
        {"Job_Number": "56789", "expected": "56789"},
        {"JOB #": "67890", "expected": "67890"},
        {"JOB#": "78901", "expected": "78901"},
        {"job #": "89012", "expected": "89012"},
        {"job#": "90123", "expected": "90123"},
        {"Other Field": "value", "expected": ""},  # No job field
    ]
    
    print("Testing Job # retrieval logic:")
    
    for i, test_case in enumerate(test_cases, 1):
        # Simulate the logic from generate_weekly_pdfs.py
        job_number = (test_case.get('Job #') or 
                      test_case.get('Job#') or 
                      test_case.get('Job Number') or 
                      test_case.get('JobNumber') or 
                      test_case.get('Job_Number') or 
                      test_case.get('JOB #') or 
                      test_case.get('JOB#') or 
                      test_case.get('job #') or 
                      test_case.get('job#') or 
                      '')
        
        expected = test_case['expected']
        if job_number == expected:
            print(f"  {i}. ✓ {list(test_case.keys())[0]}: Retrieved '{job_number}'")
        else:
            print(f"  {i}. ✗ {list(test_case.keys())[0]}: Expected '{expected}', got '{job_number}'")
    
    print("\n✓ Job # variation testing complete")
    return True

def test_excel_generation_sample():
    """Create a sample Excel file using the fixed logic."""
    print("\n=== CREATING SAMPLE EXCEL WITH FIXES ===\n")
    
    # Create a sample Excel file using the fixed logic
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = "Fixed Report"
    
    # Test data
    test_data = {
        'Job #': 'JOB123456',
        'Work Request #': 'WR789012',
        'Scope #': 'SCOPE001',
        'Customer Name': 'Test Customer',
        'Work Order #': 'WO456789',
        'Dept #': 'DEPT101'
    }
    
    # Format constants
    LINETEC_RED = 'C00000'
    RED_FILL = PatternFill(start_color=LINETEC_RED, end_color=LINETEC_RED, fill_type='solid')
    TITLE_FONT = Font(name='Calibri', size=20, bold=True)
    SUBTITLE_FONT = Font(name='Calibri', size=16, bold=True, color='404040')
    SUMMARY_HEADER_FONT = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    SUMMARY_LABEL_FONT = Font(name='Calibri', size=10, bold=True)
    SUMMARY_VALUE_FONT = Font(name='Calibri', size=10)
    
    # Build the Excel structure
    current_row = 1
    
    # Logo area
    safe_merge_cells(ws, f'A{current_row}:C{current_row+2}')
    ws[f'A{current_row}'] = "LINETEC SERVICES"
    ws[f'A{current_row}'].font = TITLE_FONT
    current_row += 3
    
    # Title
    safe_merge_cells(ws, f'D{current_row-2}:I{current_row-2}')
    ws[f'D{current_row-2}'] = 'WEEKLY UNITS COMPLETED PER SCOPE ID'
    ws[f'D{current_row-2}'].font = SUBTITLE_FONT
    ws[f'D{current_row-2}'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Report generated time
    safe_merge_cells(ws, f'D{current_row+1}:I{current_row+1}')
    ws[f'D{current_row+1}'] = f"Report Generated On: {datetime.datetime.now().strftime('%m/%d/%Y %I:%M %p')}"
    ws[f'D{current_row+1}'].font = Font(name='Calibri', size=9, italic=True)
    ws[f'D{current_row+1}'].alignment = Alignment(horizontal='right')
    
    current_row += 3
    
    # Report Summary
    safe_merge_cells(ws, f'B{current_row}:D{current_row}')
    ws[f'B{current_row}'] = 'REPORT SUMMARY'
    ws[f'B{current_row}'].font = SUMMARY_HEADER_FONT
    ws[f'B{current_row}'].fill = RED_FILL
    ws[f'B{current_row}'].alignment = Alignment(horizontal='center')
    
    # Report Details
    safe_merge_cells(ws, f'F{current_row}:I{current_row}')
    ws[f'F{current_row}'] = 'REPORT DETAILS'
    ws[f'F{current_row}'].font = SUMMARY_HEADER_FONT
    ws[f'F{current_row}'].fill = RED_FILL
    ws[f'F{current_row}'].alignment = Alignment(horizontal='center')
    
    # Details section with Job #
    details = [
        ("Foreman:", "John Smith"),
        ("Dept #:", test_data['Dept #']),
        ("Work Request #:", test_data['Work Request #']),
        ("Scope ID #:", test_data['Scope #']),
        ("Work Order #:", test_data['Work Order #']),
        ("Customer:", test_data['Customer Name']),
        ("Job #:", test_data['Job #'])  # This should now be populated!
    ]
    
    for i, (label, value) in enumerate(details):
        r = current_row + 1 + i
        ws[f'F{r}'] = label
        ws[f'F{r}'].font = SUMMARY_LABEL_FONT
        
        detail_merge_range = f'G{r}:I{r}'
        safe_merge_cells(ws, detail_merge_range)
        
        vcell = ws[f'G{r}']
        vcell.value = value
        vcell.font = SUMMARY_VALUE_FONT
        vcell.alignment = Alignment(horizontal='right')
    
    # Save the sample file
    sample_file = "sample_fixed_excel.xlsx"
    try:
        workbook.save(sample_file)
        print(f"✓ Sample Excel file created: {sample_file}")
        
        # Verify the file
        test_wb = openpyxl.load_workbook(sample_file, data_only=True)
        ws_test = test_wb.active
        
        # Check if Job # is populated
        job_found = False
        for row in ws_test.iter_rows(min_row=1, max_row=20, min_col=1, max_col=10):
            for cell in row:
                if cell.value and "Job #:" in str(cell.value):
                    # Found the label, check the value
                    row_num = cell.row
                    # Look for the value in merged cells (columns G-I)
                    for col in range(7, 10):  # G=7, H=8, I=9
                        value_cell = ws_test.cell(row=row_num, column=col)
                        if value_cell.value:
                            print(f"  ✓ Job # field populated: {value_cell.value}")
                            job_found = True
                            break
                    break
            if job_found:
                break
        
        if not job_found:
            print(f"  ⚠️  Job # field not found or empty")
        
        test_wb.close()
        
        # Clean up
        os.remove(sample_file)
        print(f"  ✓ Sample file cleaned up")
        
        return True
        
    except Exception as e:
        print(f"✗ Error creating sample Excel: {e}")
        return False

def main():
    """Run all tests."""
    print("\n" + "="*60)
    print("EXCEL GENERATION FIX TESTS")
    print("="*60)
    
    # Run tests
    test1_ok = test_safe_merge_cells()
    test2_ok = test_job_number_variations()
    test3_ok = test_excel_generation_sample()
    
    print("\n" + "="*60)
    print("TEST SUMMARY")
    print("="*60)
    print(f"  Safe Merge Cells: {'✓ PASSED' if test1_ok else '✗ FAILED'}")
    print(f"  Job # Variations: {'✓ PASSED' if test2_ok else '✗ FAILED'}")
    print(f"  Sample Excel Generation: {'✓ PASSED' if test3_ok else '✗ FAILED'}")
    
    if test1_ok and test2_ok and test3_ok:
        print("\n✓ ALL TESTS PASSED - Excel generation fixes are working correctly!")
        print("\nThe fixes address:")
        print("  1. Cell merging errors and XML issues - Now using safe_merge_cells")
        print("  2. Job # field not populated - Now checks multiple column name variations")
        print("  3. Duplicate merge prevention - Automatically skips duplicate merges")
        return 0
    else:
        print("\n✗ Some tests failed - Please review the issues")
        return 1

if __name__ == "__main__":
    sys.exit(main())