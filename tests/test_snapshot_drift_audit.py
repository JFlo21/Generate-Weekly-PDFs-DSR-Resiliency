"""Tests for the snapshot-date drift audit (quick task 260812-jqx).

RED-first suite covering the 9 research cases: detection with zero
extra Smartsheet API calls, first-sight seeding, classification via
targeted cell-history lookups (capped/paced/budget-aware), the
hold-prior-week override (both fields), and the audit risk-level
wiring. Mirrors ``tests/test_rate_sanity_audit.py`` (no-network
fixture pattern) and ``tests/test_billing_audit_shadow.py`` (Supabase
writer mocks).

Task 1 covers detection only: no classifier and no hold override
exist yet at this point in the suite's history, so drift candidates
are asserted generically (not pinned to a specific classification
value) -- Task 2 adds the classifier and Task 3 adds the hold
override in later sections of this file.
"""

from __future__ import annotations

import datetime
import types
import unittest
from unittest import mock

from pipeline.snapshot_drift import apply_snapshot_drift_holds


def _row(
    sheet_id: int = 111,
    row_id: int = 222,
    wr: str = "90001",
    cu: str = "ABC-123",
    week: str = "2026-08-09",
    snapshot: str = "2026-08-05",
) -> dict:
    return {
        "__source_sheet_id": sheet_id,
        "__row_id": row_id,
        "Work Request #": wr,
        "CU": cu,
        "Weekly Reference Logged Date": week,
        "Snapshot Date": snapshot,
    }


def _baseline(
    sheet_id: int = 111,
    row_id: int = 222,
    billed_week: str = "2026-08-09",
    snapshot_date: str = "2026-08-05",
    wr: str = "90001",
    cu: str = "ABC-123",
    first_seen_at: str = "2026-07-01T00:00:00+00:00",
) -> dict:
    return {
        "sheet_id": sheet_id,
        "row_id": row_id,
        "wr": wr,
        "cu": cu,
        "snapshot_date": snapshot_date,
        "billed_week": billed_week,
        "run_id": "prior-run",
        "first_seen_at": first_seen_at,
        "last_seen_at": "2026-08-05T00:00:00+00:00",
    }


class SnapshotDriftDetectionTestBase(unittest.TestCase):
    """Shared patches for the Supabase snapshot_store surface."""

    def setUp(self) -> None:
        self.session_start = datetime.datetime.now()
        self.client = mock.MagicMock()

        self._fetch_patch = mock.patch(
            "billing_audit.snapshot_store.fetch_snapshot_provenance"
        )
        self._upsert_patch = mock.patch(
            "billing_audit.snapshot_store.upsert_snapshot_provenance"
        )
        self._insert_patch = mock.patch(
            "billing_audit.snapshot_store.insert_snapshot_drift_events"
        )
        self.mock_fetch = self._fetch_patch.start()
        self.mock_upsert = self._upsert_patch.start()
        self.mock_insert = self._insert_patch.start()
        self.addCleanup(self._fetch_patch.stop)
        self.addCleanup(self._upsert_patch.stop)
        self.addCleanup(self._insert_patch.stop)


class TestTask1NoBaselineSeeds(SnapshotDriftDetectionTestBase):
    """No provenance baseline -> silent seed, zero drift, zero API (D-09)."""

    def test_no_baseline_seeds_silently(self) -> None:
        self.mock_fetch.return_value = ({}, "no_row")
        row = _row()

        summary = apply_snapshot_drift_holds(
            [row], [], self.client, self.session_start
        )

        self.assertTrue(summary["enabled"])
        self.assertEqual(summary["candidates"], 0)
        self.assertEqual(summary["seeded"], 1)
        self.client.Cells.get_cell_history.assert_not_called()
        self.mock_upsert.assert_called_once()
        records = self.mock_upsert.call_args[0][0]
        self.assertEqual(len(records), 1)
        self.assertEqual(records[0]["sheet_id"], 111)
        self.assertEqual(records[0]["row_id"], 222)
        # No baseline -> no drift -> no event written.
        self.mock_insert.assert_called_once_with([])


class TestTask1WeekUnchangedCostsNothing(SnapshotDriftDetectionTestBase):
    """Computed week == baseline billed_week -> zero drift, zero API (D-04)."""

    def test_unchanged_week_costs_zero_api_calls(self) -> None:
        self.mock_fetch.return_value = (
            {(111, 222): _baseline(billed_week="2026-08-09")},
            "success",
        )
        row = _row(week="2026-08-09")

        summary = apply_snapshot_drift_holds(
            [row], [], self.client, self.session_start
        )

        self.assertEqual(summary["candidates"], 0)
        self.assertEqual(summary["unchanged"], 1)
        self.client.Cells.get_cell_history.assert_not_called()
        self.mock_insert.assert_called_once_with([])


class TestTask1WeekDriftIsACandidate(SnapshotDriftDetectionTestBase):
    """Computed week != baseline billed_week -> candidate, not held,
    row fields left untouched at this point in the audit (Task 1)."""

    def test_drifted_week_emits_candidate_without_mutation(self) -> None:
        self.mock_fetch.return_value = (
            {(111, 222): _baseline(
                billed_week="2026-08-02", snapshot_date="2026-07-29"
            )},
            "success",
        )
        row = _row(week="2026-08-09", snapshot="2026-08-05")

        summary = apply_snapshot_drift_holds(
            [row], [], self.client, self.session_start
        )

        self.assertEqual(summary["candidates"], 1)
        self.assertEqual(summary["automation_self_fire_holds"], 0)
        # Task 1 never mutates rows.
        self.assertEqual(row["Weekly Reference Logged Date"], "2026-08-09")
        self.assertEqual(row["Snapshot Date"], "2026-08-05")

        self.mock_insert.assert_called_once()
        events = self.mock_insert.call_args[0][0]
        self.assertEqual(len(events), 1)
        event = events[0]
        self.assertEqual(event["sheet_id"], 111)
        self.assertEqual(event["row_id"], 222)
        self.assertEqual(event["prior_billed_week"], "2026-08-02")
        self.assertEqual(event["new_week"], "2026-08-09")
        self.assertFalse(event["held"])
        self.assertTrue(event["classification"])  # non-empty placeholder/real


class TestTask1KillSwitchDisabled(SnapshotDriftDetectionTestBase):
    """SNAPSHOT_DRIFT_AUDIT_ENABLED=false -> zeroed summary, no
    Supabase call, rows byte-identical (D-08)."""

    def test_disabled_is_a_pure_noop(self) -> None:
        row = _row()
        original = dict(row)

        with mock.patch.dict(
            "os.environ", {"SNAPSHOT_DRIFT_AUDIT_ENABLED": "false"}
        ):
            summary = apply_snapshot_drift_holds(
                [row], [], self.client, self.session_start
            )

        self.assertFalse(summary["enabled"])
        self.assertEqual(summary["candidates"], 0)
        self.mock_fetch.assert_not_called()
        self.mock_upsert.assert_not_called()
        self.mock_insert.assert_not_called()
        self.assertEqual(row, original)


class TestTask1ClientUnavailable(SnapshotDriftDetectionTestBase):
    """get_client() unavailable -> flagged unavailable, no exception."""

    def test_unavailable_client_flags_summary_and_does_not_raise(self) -> None:
        self.mock_fetch.return_value = ({}, "unavailable")
        row = _row()

        summary = apply_snapshot_drift_holds(
            [row], [], self.client, self.session_start
        )

        self.assertFalse(summary["available"])
        self.assertEqual(summary["candidates"], 0)
        self.client.Cells.get_cell_history.assert_not_called()


class TestTask1FetchRaisesDegradesToSeed(SnapshotDriftDetectionTestBase):
    """A raised exception from the bulk read is swallowed and
    degrades to the no-baseline (seed) path."""

    def test_fetch_exception_is_swallowed(self) -> None:
        self.mock_fetch.side_effect = Exception("boom")
        row = _row()

        summary = apply_snapshot_drift_holds(
            [row], [], self.client, self.session_start
        )

        self.assertFalse(summary["available"])
        self.assertEqual(summary["seeded"], 1)
        self.assertEqual(summary["candidates"], 0)
        # CR-01 regression: a fetch failure must never poison existing
        # baselines by upserting the degraded (empty-baseline) seed
        # records.
        self.mock_upsert.assert_not_called()


class TestTask1FetchFailureStatusSkipsUpsert(SnapshotDriftDetectionTestBase):
    """CR-01 regression: an explicit ``fetch_failure`` status (bulk
    read's own retry-exhaustion / circuit-breaker path, not a raised
    exception) must ALSO skip the provenance upsert -- otherwise a
    transient read failure rebases every existing baseline's
    ``billed_week`` to the current week and silently launders any
    in-flight drift."""

    def test_fetch_failure_status_skips_upsert(self) -> None:
        self.mock_fetch.return_value = ({}, "fetch_failure")
        row = _row()

        summary = apply_snapshot_drift_holds(
            [row], [], self.client, self.session_start
        )

        self.assertFalse(summary["available"])
        self.assertEqual(summary["seeded"], 1)
        self.mock_upsert.assert_not_called()


class TestTask1BatchedProvenanceUpsert(SnapshotDriftDetectionTestBase):
    """Provenance upsert is invoked at most once per run with a
    batched payload -- never once per row (RESEARCH caveat 6)."""

    def test_upsert_called_once_with_batched_payload(self) -> None:
        self.mock_fetch.return_value = (
            {
                (111, 222): _baseline(sheet_id=111, row_id=222, billed_week="2026-08-09"),
            },
            "success",
        )
        rows = [
            _row(sheet_id=111, row_id=222, wr="90001", week="2026-08-09"),
            _row(sheet_id=111, row_id=333, wr="90002", week="2026-08-09"),
            _row(sheet_id=222, row_id=444, wr="90003", week="2026-08-09"),
        ]

        summary = apply_snapshot_drift_holds(
            rows, [], self.client, self.session_start
        )

        self.assertEqual(self.mock_upsert.call_count, 1)
        records = self.mock_upsert.call_args[0][0]
        self.assertEqual(len(records), 3)
        self.assertEqual(summary["seeded"] + summary["unchanged"], 3)


SNAPSHOT_COL = 5001
UNITS_COL = 5002
AUTOMATION_EMAIL = "automation@smartsheet.com"


def _source_sheets(sheet_id: int = 111, snapshot_col=SNAPSHOT_COL,
                    units_col=UNITS_COL) -> list:
    mapping = {}
    if snapshot_col is not None:
        mapping["Snapshot Date"] = snapshot_col
    if units_col is not None:
        mapping["Units Completed?"] = units_col
    return [{"id": sheet_id, "name": "Sheet A", "column_mapping": mapping}]


def _entry(modified_at: str, modified_by):
    return types.SimpleNamespace(modified_at=modified_at, modified_by=modified_by)


def _history(entries: list):
    return types.SimpleNamespace(data=list(entries))


def _drift_row_and_baseline(sheet_id=111, row_id=222, wr="90001"):
    row = _row(sheet_id=sheet_id, row_id=row_id, wr=wr,
               week="2026-08-09", snapshot="2026-08-05")
    baseline = {
        (sheet_id, row_id): _baseline(
            sheet_id=sheet_id, row_id=row_id, wr=wr,
            billed_week="2026-08-02", snapshot_date="2026-07-29",
        )
    }
    return row, baseline


class SnapshotDriftClassifierTestBase(SnapshotDriftDetectionTestBase):
    """Adds a configurable Cells.get_cell_history side_effect keyed by
    column id, so each test only needs to describe the two histories."""

    def _wire_history(self, snapshot_entries=None, units_entries=None):
        snapshot_entries = snapshot_entries or []
        units_entries = units_entries or []

        def _side_effect(sheet_id, row_id, column_id, include_all=True):
            if column_id == SNAPSHOT_COL:
                return _history(snapshot_entries)
            if column_id == UNITS_COL:
                return _history(units_entries)
            raise AssertionError(f"unexpected column_id {column_id}")

        self.client.Cells.get_cell_history.side_effect = _side_effect


class TestTask2AutomationSelfFire(SnapshotDriftClassifierTestBase):
    def test_automation_write_no_nearby_units_change_is_self_fire(self) -> None:
        row, baseline = _drift_row_and_baseline()
        self.mock_fetch.return_value = (baseline, "success")
        self._wire_history(
            snapshot_entries=[
                _entry("2026-08-12T10:00:00Z",
                       types.SimpleNamespace(email=AUTOMATION_EMAIL)),
            ],
            units_entries=[],
        )

        summary = apply_snapshot_drift_holds(
            [row], _source_sheets(), self.client, self.session_start
        )

        self.assertEqual(summary["automation_self_fire"], 1)
        self.assertEqual(summary["manual"], 0)
        self.assertEqual(summary["unclassified"], 0)
        self.assertEqual(self.client.Cells.get_cell_history.call_count, 2)


class TestTask2AutomationWriteWithNearbyUnitsChangeIsManual(
    SnapshotDriftClassifierTestBase
):
    def test_nearby_units_change_within_window_is_manual(self) -> None:
        row, baseline = _drift_row_and_baseline()
        self.mock_fetch.return_value = (baseline, "success")
        self._wire_history(
            snapshot_entries=[
                _entry("2026-08-12T10:00:00Z",
                       types.SimpleNamespace(email=AUTOMATION_EMAIL)),
            ],
            units_entries=[
                _entry("2026-08-12T09:59:30Z",
                       types.SimpleNamespace(email="foreman@example.com")),
            ],
        )

        summary = apply_snapshot_drift_holds(
            [row], _source_sheets(), self.client, self.session_start
        )

        self.assertEqual(summary["manual"], 1)
        self.assertEqual(summary["automation_self_fire"], 0)


class TestTask2HumanEmailIsManualNoSecondCall(SnapshotDriftClassifierTestBase):
    def test_human_email_is_manual_zero_second_call(self) -> None:
        row, baseline = _drift_row_and_baseline()
        self.mock_fetch.return_value = (baseline, "success")
        self._wire_history(
            snapshot_entries=[
                _entry("2026-08-12T10:00:00Z",
                       types.SimpleNamespace(email="foreman@example.com")),
            ],
        )

        summary = apply_snapshot_drift_holds(
            [row], _source_sheets(), self.client, self.session_start
        )

        self.assertEqual(summary["manual"], 1)
        # Snapshot Date history alone already rules out automation --
        # skip the Units Completed? call and save the quota.
        self.assertEqual(self.client.Cells.get_cell_history.call_count, 1)


class TestTask2ApiErrorIsUnclassified(SnapshotDriftClassifierTestBase):
    def test_get_cell_history_raises_is_unclassified(self) -> None:
        row, baseline = _drift_row_and_baseline()
        self.mock_fetch.return_value = (baseline, "success")
        self.client.Cells.get_cell_history.side_effect = Exception("429")

        summary = apply_snapshot_drift_holds(
            [row], _source_sheets(), self.client, self.session_start
        )

        self.assertEqual(summary["unclassified"], 1)
        self.assertEqual(summary["automation_self_fire"], 0)
        self.assertEqual(summary["manual"], 0)


class TestTask2MissingColumnIdIsUnclassified(SnapshotDriftClassifierTestBase):
    def test_missing_column_id_unclassified_zero_calls(self) -> None:
        row, baseline = _drift_row_and_baseline()
        self.mock_fetch.return_value = (baseline, "success")

        summary = apply_snapshot_drift_holds(
            [row], _source_sheets(units_col=None), self.client,
            self.session_start,
        )

        self.assertEqual(summary["unclassified"], 1)
        self.client.Cells.get_cell_history.assert_not_called()


class TestTask2PerRunCap(SnapshotDriftClassifierTestBase):
    def test_cap_of_one_classifies_first_leaves_second_unclassified(self) -> None:
        row1, baseline1 = _drift_row_and_baseline(row_id=100, wr="90001")
        row2, baseline2 = _drift_row_and_baseline(row_id=200, wr="90002")
        merged_baseline = {**baseline1, **baseline2}
        self.mock_fetch.return_value = (merged_baseline, "success")
        # row1 (lower row id, classified first in deterministic order)
        # requires BOTH calls: automation write with no nearby units
        # change.
        self._wire_history(
            snapshot_entries=[
                _entry("2026-08-12T10:00:00Z",
                       types.SimpleNamespace(email=AUTOMATION_EMAIL)),
            ],
            units_entries=[],
        )

        with mock.patch.dict(
            "os.environ", {"SNAPSHOT_DRIFT_MAX_ROWS": "1"}
        ):
            summary = apply_snapshot_drift_holds(
                [row1, row2], _source_sheets(), self.client,
                self.session_start,
            )

        self.assertEqual(
            summary["automation_self_fire"] + summary["manual"], 1
        )
        self.assertEqual(summary["unclassified"], 1)
        self.assertEqual(self.client.Cells.get_cell_history.call_count, 2)


class TestTask2BudgetGuardSkipsAll(SnapshotDriftClassifierTestBase):
    def test_low_remaining_budget_skips_classification_entirely(self) -> None:
        row, baseline = _drift_row_and_baseline()
        self.mock_fetch.return_value = (baseline, "success")
        self._wire_history(
            snapshot_entries=[
                _entry("2026-08-12T10:00:00Z",
                       types.SimpleNamespace(email=AUTOMATION_EMAIL)),
            ],
        )
        # session_start far enough in the past that remaining budget
        # (TIME_BUDGET_MINUTES - elapsed) is comfortably below
        # SNAPSHOT_DRIFT_MAX_MINUTES (5) -- margin wide enough that
        # normal test-execution latency cannot flip the outcome
        # (a boundary-exact 160/165 margin was observed flaky under
        # the full suite's slower wall-clock scheduling).
        stale_session_start = (
            datetime.datetime.now() - datetime.timedelta(minutes=163)
        )

        with mock.patch.dict(
            "os.environ",
            {"GITHUB_ACTIONS": "true", "TIME_BUDGET_MINUTES": "165"},
        ):
            summary = apply_snapshot_drift_holds(
                [row], _source_sheets(), self.client, stale_session_start
            )

        self.assertEqual(summary["unclassified"], 1)
        self.assertIsNotNone(summary["skip_reason"])
        self.client.Cells.get_cell_history.assert_not_called()


class TestTask2HistoryOrderingIndependence(SnapshotDriftClassifierTestBase):
    def test_newest_first_and_oldest_first_agree(self) -> None:
        older = _entry("2026-08-12T09:00:00Z",
                        types.SimpleNamespace(email="foreman@example.com"))
        newer = _entry("2026-08-12T10:00:00Z",
                        types.SimpleNamespace(email=AUTOMATION_EMAIL))

        results = []
        for order in ([older, newer], [newer, older]):
            row, baseline = _drift_row_and_baseline()
            self.mock_fetch.return_value = (baseline, "success")
            self._wire_history(snapshot_entries=order, units_entries=[])
            summary = apply_snapshot_drift_holds(
                [row], _source_sheets(), self.client, self.session_start
            )
            if summary["automation_self_fire"]:
                results.append("automation_self_fire")
            elif summary["manual"]:
                results.append("manual")
            else:
                results.append("unclassified")

        self.assertEqual(results[0], results[1])
        self.assertEqual(results[0], "automation_self_fire")


class TestTask2ModifiedByShapes(SnapshotDriftClassifierTestBase):
    def test_modified_by_object_string_and_none_never_raise(self) -> None:
        shapes = [
            types.SimpleNamespace(email=AUTOMATION_EMAIL),
            AUTOMATION_EMAIL,
            None,
        ]
        for shape in shapes:
            row, baseline = _drift_row_and_baseline()
            self.mock_fetch.return_value = (baseline, "success")
            self._wire_history(
                snapshot_entries=[_entry("2026-08-12T10:00:00Z", shape)],
                units_entries=[],
            )
            # Must not raise.
            summary = apply_snapshot_drift_holds(
                [row], _source_sheets(), self.client, self.session_start
            )
            self.assertIn(
                summary["automation_self_fire"] + summary["manual"]
                + summary["unclassified"],
                [1],
            )


class TestTask2PacingBetweenCalls(SnapshotDriftClassifierTestBase):
    def test_sleep_between_calls_not_before_first(self) -> None:
        row1, baseline1 = _drift_row_and_baseline(row_id=100, wr="90001")
        row2, baseline2 = _drift_row_and_baseline(row_id=200, wr="90002")
        merged_baseline = {**baseline1, **baseline2}
        self.mock_fetch.return_value = (merged_baseline, "success")
        self._wire_history(
            snapshot_entries=[
                _entry("2026-08-12T10:00:00Z",
                       types.SimpleNamespace(email=AUTOMATION_EMAIL)),
            ],
            units_entries=[],
        )

        with mock.patch("pipeline.snapshot_drift.time.sleep") as mock_sleep:
            apply_snapshot_drift_holds(
                [row1, row2], _source_sheets(), self.client,
                self.session_start,
            )

        # 2 candidates x 2 calls each = 4 calls -> 3 sleeps, never
        # before the first call.
        self.assertEqual(self.client.Cells.get_cell_history.call_count, 4)
        self.assertEqual(mock_sleep.call_count, 3)
        for call in mock_sleep.call_args_list:
            self.assertAlmostEqual(call.args[0], 2.0)


# ── Task 3: hold-prior-week override + audit risk wiring ────────────

def _billable_row(sheet_id=111, row_id=222, wr="90001", cu="ABC-123",
                   week="2026-08-09", snapshot="2026-08-05",
                   price="$777.00"):
    row = _row(sheet_id=sheet_id, row_id=row_id, wr=wr, cu=cu,
               week=week, snapshot=snapshot)
    row.update({
        "Units Total Price": price,
        "Quantity": "2",
        "Customer Name": "TestCustomer",
        "Dept #": "500",
        "Job #": "J-1",
        "Work Type": "Install",
        "Pole #": "P-1",
        "Foreman": "TestForeman",
        "__variant": "primary",
        "__current_foreman": "TestForeman",
        "__effective_user": "TestForeman",
        "Units Completed?": True,
    })
    return row


class TestTask3HoldRewritesBothFields(SnapshotDriftClassifierTestBase):
    def test_hold_rewrites_both_fields_and_preserves_originals(self) -> None:
        row, baseline = _drift_row_and_baseline()
        self.mock_fetch.return_value = (baseline, "success")
        self._wire_history(
            snapshot_entries=[
                _entry("2026-08-12T10:00:00Z",
                       types.SimpleNamespace(email=AUTOMATION_EMAIL)),
            ],
            units_entries=[],
        )

        with mock.patch.dict(
            "os.environ", {"SNAPSHOT_DRIFT_HOLD_ENABLED": "true"}
        ):
            summary = apply_snapshot_drift_holds(
                [row], _source_sheets(), self.client, self.session_start
            )

        self.assertEqual(summary["automation_self_fire_holds"], 1)
        self.assertEqual(row["Weekly Reference Logged Date"], "2026-08-02")
        self.assertEqual(row["Snapshot Date"], "2026-07-29")
        self.assertEqual(
            row["__drifted_weekly_reference_logged_date"], "2026-08-09"
        )
        self.assertEqual(row["__drifted_snapshot_date"], "2026-08-05")
        self.assertEqual(
            row["__snapshot_drift_classification"], "automation_self_fire"
        )


class TestTask3HeldRowSurvivesExcelFilter(SnapshotDriftClassifierTestBase):
    def test_held_row_appears_in_generated_workbook(self) -> None:
        import tempfile

        import openpyxl

        import generate_weekly_pdfs as _gwp
        from pipeline.excel import generate_excel

        row = _billable_row()
        baseline = {
            (111, 222): _baseline(
                billed_week="2026-08-02", snapshot_date="2026-07-29"
            )
        }
        self.mock_fetch.return_value = (baseline, "success")
        self._wire_history(
            snapshot_entries=[
                _entry("2026-08-12T10:00:00Z",
                       types.SimpleNamespace(email=AUTOMATION_EMAIL)),
            ],
            units_entries=[],
        )

        with mock.patch.dict(
            "os.environ", {"SNAPSHOT_DRIFT_HOLD_ENABLED": "true"}
        ):
            summary = apply_snapshot_drift_holds(
                [row], _source_sheets(), self.client, self.session_start
            )
        self.assertEqual(summary["automation_self_fire_holds"], 1)

        # After the hold, grouping.py would compute the row's week from
        # the (restored) Weekly Reference Logged Date -- reproduce that
        # here since this test calls generate_excel directly.
        row["__week_ending_date"] = datetime.datetime(2026, 8, 2)

        tmp = tempfile.TemporaryDirectory()
        self.addCleanup(tmp.cleanup)
        orig_output = _gwp.OUTPUT_FOLDER
        _gwp.OUTPUT_FOLDER = tmp.name
        self.addCleanup(lambda: setattr(_gwp, "OUTPUT_FOLDER", orig_output))

        result = generate_excel(
            "080226_90001", [row], datetime.datetime(2026, 8, 2),
            data_hash="deadbeefcafe0002",
        )
        excel_path = result[0]

        wb = openpyxl.load_workbook(excel_path)
        ws = wb.active
        found = any(
            ws.cell(row=r, column=c).value == 777.0
            for r in range(1, ws.max_row + 1)
            for c in range(1, ws.max_column + 1)
        )
        self.assertTrue(
            found,
            "Held row's price must appear in a day-table -- if it does "
            "not, the row was silently dropped by the Monday-Sunday "
            "Snapshot Date filter (RESEARCH pitfall 1)."
        )


class TestTask3HoldSkippedWhenBaselineSnapshotDateIsNull(
    SnapshotDriftClassifierTestBase
):
    """CR-02 regression: a hold whose baseline ``snapshot_date`` is
    NULL must NOT write ``None`` into ``Snapshot Date`` -- that value
    is silently dropped by the Excel Monday-Sunday day-block filter,
    which is strictly worse than the drift itself. The candidate must
    be skipped (row left untouched, not counted as a hold)."""

    def test_null_prior_snapshot_date_skips_hold(self) -> None:
        row = _row(week="2026-08-09", snapshot="2026-08-05")
        original = dict(row)
        baseline = {
            (111, 222): _baseline(
                billed_week="2026-08-02", snapshot_date=None,
            )
        }
        self.mock_fetch.return_value = (baseline, "success")
        self._wire_history(
            snapshot_entries=[
                _entry("2026-08-12T10:00:00Z",
                       types.SimpleNamespace(email=AUTOMATION_EMAIL)),
            ],
            units_entries=[],
        )

        with mock.patch.dict(
            "os.environ", {"SNAPSHOT_DRIFT_HOLD_ENABLED": "true"}
        ):
            summary = apply_snapshot_drift_holds(
                [row], _source_sheets(), self.client, self.session_start
            )

        self.assertEqual(summary["automation_self_fire"], 1)
        self.assertEqual(summary["automation_self_fire_holds"], 0)
        self.assertEqual(row, original)


class TestTask3PriorWeekHashStability(SnapshotDriftClassifierTestBase):
    def test_hash_stable_across_drift_and_hold_both_modes(self) -> None:
        import generate_weekly_pdfs as _gwp
        from pipeline.change_detection import calculate_data_hash

        orig_extended = _gwp.EXTENDED_CHANGE_DETECTION
        self.addCleanup(
            lambda: setattr(_gwp, "EXTENDED_CHANGE_DETECTION", orig_extended)
        )

        for extended in (False, True):
            with self.subTest(extended=extended):
                _gwp.EXTENDED_CHANGE_DETECTION = extended
                row = _billable_row(snapshot="2026-07-29", week="2026-08-02")
                hash_before = calculate_data_hash([row])

                # Simulate the automation drift: both fields re-stamped.
                row["Snapshot Date"] = "2026-08-05"
                row["Weekly Reference Logged Date"] = "2026-08-09"

                baseline = {
                    (111, 222): _baseline(
                        billed_week="2026-08-02", snapshot_date="2026-07-29"
                    )
                }
                self.mock_fetch.return_value = (baseline, "success")
                self._wire_history(
                    snapshot_entries=[
                        _entry("2026-08-12T10:00:00Z",
                               types.SimpleNamespace(email=AUTOMATION_EMAIL)),
                    ],
                    units_entries=[],
                )
                with mock.patch.dict(
                    "os.environ", {"SNAPSHOT_DRIFT_HOLD_ENABLED": "true"}
                ):
                    apply_snapshot_drift_holds(
                        [row], _source_sheets(), self.client,
                        self.session_start,
                    )

                hash_after = calculate_data_hash([row])
                self.assertEqual(hash_before, hash_after)


class TestTask3ManualNeverMutated(SnapshotDriftClassifierTestBase):
    def test_manual_candidate_never_mutated_even_with_hold_enabled(self) -> None:
        row, baseline = _drift_row_and_baseline()
        original = dict(row)
        self.mock_fetch.return_value = (baseline, "success")
        self._wire_history(
            snapshot_entries=[
                _entry("2026-08-12T10:00:00Z",
                       types.SimpleNamespace(email="human@example.com")),
            ],
        )

        with mock.patch.dict(
            "os.environ", {"SNAPSHOT_DRIFT_HOLD_ENABLED": "true"}
        ):
            summary = apply_snapshot_drift_holds(
                [row], _source_sheets(), self.client, self.session_start
            )

        self.assertEqual(summary["manual"], 1)
        self.assertEqual(summary["automation_self_fire_holds"], 0)
        self.assertEqual(row, original)
        self.assertNotIn("__drifted_snapshot_date", row)


class TestTask3UnclassifiedNeverMutated(SnapshotDriftClassifierTestBase):
    def test_unclassified_candidate_never_mutated_even_with_hold_enabled(
        self,
    ) -> None:
        row, baseline = _drift_row_and_baseline()
        original = dict(row)
        self.mock_fetch.return_value = (baseline, "success")
        self.client.Cells.get_cell_history.side_effect = Exception("boom")

        with mock.patch.dict(
            "os.environ", {"SNAPSHOT_DRIFT_HOLD_ENABLED": "true"}
        ):
            summary = apply_snapshot_drift_holds(
                [row], _source_sheets(), self.client, self.session_start
            )

        self.assertEqual(summary["unclassified"], 1)
        self.assertEqual(summary["automation_self_fire_holds"], 0)
        self.assertEqual(row, original)


class TestTask3HoldGateDefaultOffDoesNotMutate(SnapshotDriftClassifierTestBase):
    def test_default_hold_off_flags_but_does_not_mutate(self) -> None:
        row, baseline = _drift_row_and_baseline()
        original = dict(row)
        self.mock_fetch.return_value = (baseline, "success")
        self._wire_history(
            snapshot_entries=[
                _entry("2026-08-12T10:00:00Z",
                       types.SimpleNamespace(email=AUTOMATION_EMAIL)),
            ],
            units_entries=[],
        )

        summary = apply_snapshot_drift_holds(
            [row], _source_sheets(), self.client, self.session_start
        )

        self.assertEqual(summary["automation_self_fire"], 1)
        self.assertEqual(summary["automation_self_fire_holds"], 0)
        self.assertFalse(summary["hold_enabled"])
        self.assertEqual(row, original)


class TestTask3RiskEscalation(unittest.TestCase):
    def _base_summary(self) -> dict:
        return {
            "total_anomalies": 0,
            "total_unauthorized_changes": 0,
            "total_data_issues": 0,
            "total_rate_sanity_mismatches": 0,
            "risk_level": "LOW",
            "recommendations": [],
        }

    def test_four_holds_escalate_to_high(self) -> None:
        from audit_billing_changes import escalate_risk_for_snapshot_drift

        summary = self._base_summary()
        escalate_risk_for_snapshot_drift(summary, 4)

        self.assertEqual(summary["risk_level"], "HIGH")
        self.assertEqual(summary["total_snapshot_drift_holds"], 4)

    def test_zero_holds_leaves_risk_level_unchanged(self) -> None:
        """Manual/unclassified drift never reaches this function as a
        hold count -- it is recorded via Supabase + the run log only,
        never folded into risk_level (RESEARCH caveat 5)."""
        from audit_billing_changes import escalate_risk_for_snapshot_drift

        summary = self._base_summary()
        escalate_risk_for_snapshot_drift(summary, 0)

        self.assertEqual(summary["risk_level"], "LOW")
        self.assertEqual(summary["total_snapshot_drift_holds"], 0)


class TestTask3DriftEventsRecordEveryCandidate(SnapshotDriftClassifierTestBase):
    def test_events_carry_classification_and_held_for_all_outcomes(self) -> None:
        row_self_fire, b1 = _drift_row_and_baseline(row_id=1, wr="90101")
        row_manual, b2 = _drift_row_and_baseline(row_id=2, wr="90102")
        row_unclassified, b3 = _drift_row_and_baseline(row_id=3, wr="90103")
        merged = {**b1, **b2, **b3}
        self.mock_fetch.return_value = (merged, "success")

        def _side_effect(sheet_id, row_id, column_id, include_all=True):
            if row_id == 1:
                if column_id == SNAPSHOT_COL:
                    return _history([_entry(
                        "2026-08-12T10:00:00Z",
                        types.SimpleNamespace(email=AUTOMATION_EMAIL),
                    )])
                return _history([])
            if row_id == 2:
                if column_id == SNAPSHOT_COL:
                    return _history([_entry(
                        "2026-08-12T10:00:00Z",
                        types.SimpleNamespace(email="human@example.com"),
                    )])
                return _history([])
            if row_id == 3:
                raise Exception("api error")
            raise AssertionError(f"unexpected row_id {row_id}")

        self.client.Cells.get_cell_history.side_effect = _side_effect

        with mock.patch.dict(
            "os.environ", {"SNAPSHOT_DRIFT_HOLD_ENABLED": "true"}
        ):
            apply_snapshot_drift_holds(
                [row_self_fire, row_manual, row_unclassified],
                _source_sheets(), self.client, self.session_start,
            )

        self.mock_insert.assert_called_once()
        events = self.mock_insert.call_args[0][0]
        by_row = {e["row_id"]: e for e in events}

        self.assertEqual(by_row[1]["classification"], "automation_self_fire")
        self.assertTrue(by_row[1]["held"])
        self.assertEqual(by_row[2]["classification"], "manual")
        self.assertFalse(by_row[2]["held"])
        self.assertEqual(by_row[3]["classification"], "unclassified")
        self.assertFalse(by_row[3]["held"])


class TestTask3BothSwitchesOffEquivalence(SnapshotDriftClassifierTestBase):
    def test_both_switches_off_reproduces_baseline(self) -> None:
        row, baseline = _drift_row_and_baseline()
        original = dict(row)
        self.mock_fetch.return_value = (baseline, "success")

        with mock.patch.dict(
            "os.environ",
            {
                "SNAPSHOT_DRIFT_AUDIT_ENABLED": "false",
                "SNAPSHOT_DRIFT_HOLD_ENABLED": "false",
            },
        ):
            summary = apply_snapshot_drift_holds(
                [row], _source_sheets(), self.client, self.session_start
            )

        self.assertFalse(summary["enabled"])
        self.assertEqual(row, original)
        self.mock_fetch.assert_not_called()
        self.mock_upsert.assert_not_called()
        self.mock_insert.assert_not_called()


if __name__ == "__main__":
    unittest.main()
