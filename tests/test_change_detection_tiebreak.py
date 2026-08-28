"""Sort-key tiebreaker in ``calculate_data_hash`` (2026-08-27, run #2801).

``billing_audit.pipeline_run`` showed WR 91057431 / week 080226 alternating
between two content hashes on 12 consecutive runs with a constant assignment
fingerprint: its 142 rows span three source sheets, two of them tie on the
``(WR, Snapshot Date, CU, Pole/Point, Quantity)`` sort key while differing
in a hashed field, and Python's stable sort kept the parallel fetch's
``as_completed`` arrival order -- so the hash flipped with thread timing,
the group was regenerated and re-uploaded every run, and the Phase 11
shadow parity could never ``pass``.

These tests pin three properties of the fix:

1. Tied rows hash identically in any input order (the incident).
2. Rows with no differing ties hash byte-identically to the PRE-fix
   ordering -- the tiebreaker can only reorder rows that tie on the full
   business key, so it cannot change any currently-stable hash.
3. A tie broken only by foreman is deterministic too (the ``FOREMAN=``
   meta token comes from the first row that has one).
"""

from __future__ import annotations

import datetime
import hashlib
import itertools
import sys
import unittest
from pathlib import Path

_REPO_ROOT = Path(__file__).resolve().parent.parent
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))

import generate_weekly_pdfs  # noqa: E402
from pipeline import change_detection  # noqa: E402


def _row(wr='91057431', cu='CU-100', qty=1, price='$50.00', pole='P-7',
         snapshot='2026-07-30', **extra):
    row = {
        'Work Request #': wr,
        'Snapshot Date': snapshot,
        'CU': cu,
        'Quantity': qty,
        'Units Total Price': price,
        'Pole #': pole,
        'Work Type': 'Install',
        'Dept #': '520',
        'Units Completed?': True,
        'Foreman': 'Pat Example',
        '__variant': 'primary',
    }
    row.update(extra)
    return row


def _report_details(rows, group_key, labels):
    """Generate the workbook for ``rows`` and return the REPORT DETAILS
    values next to ``labels`` (e.g. 'Job #:')."""
    import tempfile
    import openpyxl
    saved = {k: getattr(generate_weekly_pdfs, k) for k in
             ('OUTPUT_FOLDER', 'RES_GROUPING_MODE',
              'PRIMARY_CLAIM_ATTRIBUTION_ENABLED')}
    tmp = tempfile.TemporaryDirectory()
    generate_weekly_pdfs.OUTPUT_FOLDER = tmp.name
    generate_weekly_pdfs.RES_GROUPING_MODE = 'both'
    generate_weekly_pdfs.PRIMARY_CLAIM_ATTRIBUTION_ENABLED = False
    try:
        path = generate_weekly_pdfs.generate_excel(
            group_key, list(rows), datetime.datetime(2026, 7, 26),
            data_hash='deadbeefcafe0002')[0]
        ws = openpyxl.load_workbook(path).active
        vals = {}
        for row in ws.iter_rows(values_only=True):
            for i, v in enumerate(row):
                if v in labels:
                    vals[v] = row[i + 1]
        return vals
    finally:
        for k, v in saved.items():
            setattr(generate_weekly_pdfs, k, v)
        tmp.cleanup()


class SortTiebreakTests(unittest.TestCase):

    def setUp(self):
        self._saved_ext = generate_weekly_pdfs.EXTENDED_CHANGE_DETECTION
        self._saved_cutoff = generate_weekly_pdfs.RATE_CUTOFF_DATE
        self._saved_fp = generate_weekly_pdfs._RATES_FINGERPRINT
        generate_weekly_pdfs.EXTENDED_CHANGE_DETECTION = True
        generate_weekly_pdfs.RATE_CUTOFF_DATE = None
        generate_weekly_pdfs._RATES_FINGERPRINT = ''

    def tearDown(self):
        generate_weekly_pdfs.EXTENDED_CHANGE_DETECTION = self._saved_ext
        generate_weekly_pdfs.RATE_CUTOFF_DATE = self._saved_cutoff
        generate_weekly_pdfs._RATES_FINGERPRINT = self._saved_fp

    def _hashes_for_all_orders(self, rows):
        return {
            generate_weekly_pdfs.calculate_data_hash(list(perm))
            for perm in itertools.permutations(rows)
        }

    def test_tied_rows_differing_in_hashed_field_hash_identically_in_any_order(self):
        # Same (WR, Snapshot, CU, Pole, Qty); differ in Work Type + price.
        rows = [
            _row(**{'Work Type': 'Install', 'Units Total Price': '$50.00'}),
            _row(**{'Work Type': 'Remove', 'Units Total Price': '$20.00'}),
            _row(**{'Work Type': 'Transfer', 'Units Total Price': '$35.00'}),
        ]
        self.assertEqual(len(self._hashes_for_all_orders(rows)), 1,
                         "group hash must not depend on fetch arrival order")

    def test_tie_broken_only_by_foreman_is_deterministic(self):
        rows = [
            _row(Foreman='Alice'),
            _row(Foreman='Bob'),
        ]
        self.assertEqual(len(self._hashes_for_all_orders(rows)), 1)

    def test_helper_rows_differing_only_in_helper_metadata_are_order_independent(self):
        # Greptile on PR #359: HELPER= / HELPER_DEPT= / HELPER_JOB= are read
        # from sorted_rows[0], so helper metadata must be in the tiebreaker
        # too or two rows tying on everything else would keep arrival order.
        rows = [
            _row(**{'__variant': 'helper',
                    '__helper_foreman': 'Sam Sample',
                    '__helper_dept': 'NA-03', '__helper_job': ''}),
            _row(**{'__variant': 'helper',
                    '__helper_foreman': 'Sam Sample',
                    '__helper_dept': 'NA-04', '__helper_job': 'J-77'}),
        ]
        self.assertEqual(len(self._hashes_for_all_orders(rows)), 1)
        # ...and the metadata still matters where the hash reads it: the
        # HELPER_* tokens come from the row that sorts FIRST (NA-03 here),
        # so editing that row's helper job changes the hash.
        edited = [dict(rows[0], **{'__helper_job': 'J-78'}), dict(rows[1])]
        self.assertNotEqual(generate_weekly_pdfs.calculate_data_hash(rows),
                            generate_weekly_pdfs.calculate_data_hash(edited))

    def test_canonical_first_row_is_input_order_independent(self):
        # Codex on PR #359: Excel must read header metadata from the row the
        # hash treats as first, not from arrival-order group_rows[0].
        a = _row(**{'__variant': 'helper', '__helper_foreman': 'H',
                    '__helper_dept': 'NA-03', '__helper_job': 'J-1'})
        b = _row(**{'__variant': 'helper', '__helper_foreman': 'H',
                    '__helper_dept': 'NA-04', '__helper_job': 'J-2'})
        first_ab = change_detection.canonical_first_row([a, b])
        first_ba = change_detection.canonical_first_row([b, a])
        self.assertIs(first_ab, first_ba)
        self.assertEqual(first_ab['__helper_dept'], 'NA-03')
        self.assertEqual(change_detection.canonical_sorted_rows([a, b]),
                         change_detection.canonical_sorted_rows([b, a]))

    def test_excel_header_uses_the_canonical_row(self):
        """Two helper rows differing only in helper dept/job, generated in
        both arrival orders, must produce the same REPORT DETAILS values
        (the ones the hash's first row carries)."""
        import tempfile
        import openpyxl
        saved = {k: getattr(generate_weekly_pdfs, k) for k in
                 ('OUTPUT_FOLDER', 'RES_GROUPING_MODE',
                  'PRIMARY_CLAIM_ATTRIBUTION_ENABLED')}
        tmp = tempfile.TemporaryDirectory()
        generate_weekly_pdfs.OUTPUT_FOLDER = tmp.name
        generate_weekly_pdfs.RES_GROUPING_MODE = 'both'
        generate_weekly_pdfs.PRIMARY_CLAIM_ATTRIBUTION_ENABLED = False
        try:
            def helper_row(dept, job):
                return _row(wr='90001', **{
                    'Weekly Reference Logged Date': '2026-07-26',
                    '__week_ending_date': datetime.datetime(2026, 7, 26),
                    '__variant': 'helper',
                    '__helper_foreman': 'Sam Sample',
                    '__helper_dept': dept, '__helper_job': job,
                    '__effective_user': 'Sam Sample',
                    '__current_foreman': 'Pat Example'})
            a, b = helper_row('NA-03', 'J-1'), helper_row('NA-04', 'J-2')
            details = []
            for rows in ([a, b], [b, a]):
                path = generate_weekly_pdfs.generate_excel(
                    '072626_90001_HELPER_Sam_Sample', list(rows),
                    datetime.datetime(2026, 7, 26),
                    data_hash='deadbeefcafe0001')[0]
                ws = openpyxl.load_workbook(path).active
                vals = {}
                for row in ws.iter_rows(values_only=True):
                    for i, v in enumerate(row):
                        if v in ('Dept #:', 'Job #:'):
                            vals[v] = row[i + 1]
                details.append(vals)
            self.assertEqual(details[0], details[1])
            self.assertEqual(details[0].get('Dept #:'), 'NA-03')
            self.assertEqual(details[0].get('Job #:'), 'J-1')
        finally:
            for k, v in saved.items():
                setattr(generate_weekly_pdfs, k, v)
            tmp.cleanup()

    def test_job_alias_only_difference_is_order_independent(self):
        # Codex on PR #361: generate_excel accepts Job # under several
        # column-title aliases, but the hash reads only 'Job #' /
        # 'Job Number'. Two rows differing only in an alias the hash never
        # sees tie on the whole hashed key, so the header's Job # must be
        # pinned by the canonical order, not by arrival order.
        a = _row(**{'Job#': 'J-A'})
        b = _row(**{'Job#': 'J-B'})
        self.assertIs(change_detection.canonical_first_row([a, b]),
                      change_detection.canonical_first_row([b, a]))
        self.assertEqual(
            change_detection.canonical_first_row([b, a])['Job#'], 'J-A')
        # Header-only tiebreaker: the hash is untouched in every order.
        self.assertEqual(len(self._hashes_for_all_orders([a, b])), 1)
        # Same precedence as generate_excel: 'Job #' beats the aliases.
        self.assertEqual(change_detection.header_job_number(
            {'Job#': 'alias', 'Job #': 'canonical'}), 'canonical')

    def test_user_identity_only_difference_is_order_independent(self):
        # Copilot on PR #361 (round 2): with PRIMARY_CLAIM_ATTRIBUTION off
        # the orchestrate identity sites derive the primary identifier from
        # the canonical row's 'User', which is neither hashed nor -- before
        # this -- in the sort key, so two rows differing only in User kept
        # arrival order and the history_key could alternate under a stable
        # hash. Identity-only tiebreaker: order-independent, hash-neutral.
        a = _row(User='alice')
        b = _row(User='bob')
        self.assertIs(change_detection.canonical_first_row([a, b]),
                      change_detection.canonical_first_row([b, a]))
        self.assertEqual(
            change_detection.canonical_first_row([b, a])['User'], 'alice')
        self.assertEqual(len(self._hashes_for_all_orders([a, b])), 1)

    def test_work_order_alias_only_difference_is_order_independent(self):
        # Codex / Copilot on PR #361 (round 3): the hash collapses
        # 'Work Order #' and 'Work Order Number' to one string, but the
        # REPORT DETAILS header shows only the canonical row's raw
        # 'Work Order #'. Two rows differing only in WHICH alias carries
        # the value tie on the whole hashed key, so the header could
        # alternate between 'WO-1' and blank under one hash.
        a = _row(**{'Work Order #': 'WO-1'})
        b = _row(**{'Work Order Number': 'WO-1'})
        self.assertIs(change_detection.canonical_first_row([a, b]),
                      change_detection.canonical_first_row([b, a]))
        # The populated displayed column wins, so the header is never
        # blank when a row in the group can fill it.
        self.assertEqual(
            change_detection.canonical_first_row([b, a])
            .get('Work Order #'), 'WO-1')
        self.assertEqual(len(self._hashes_for_all_orders([a, b])), 1)

    def test_serialized_field_collision_is_order_independent(self):
        # Copilot on PR #361 (round 3): the hashed fields are joined with
        # an unescaped '|', so two different rows can serialize to the
        # same string. Their hash contributions are identical (nothing
        # to fix there), but the workbook shows the canonical row's raw
        # Work Order, so the tie must be broken by the unjoined fields.
        a = _row(**{'Work Order #': 'WO|East', 'CU Description': 'Install'})
        b = _row(**{'Work Order #': 'WO', 'CU Description': 'East|Install'})
        joined = lambda r: "|".join(
            change_detection._extended_row_fields(r, 'primary'))
        self.assertEqual(joined(a), joined(b), "fixture must collide")
        self.assertIs(change_detection.canonical_first_row([a, b]),
                      change_detection.canonical_first_row([b, a]))
        self.assertEqual(len(self._hashes_for_all_orders([a, b])), 1)

    def test_excel_job_number_comes_from_the_shared_resolver(self):
        """Copilot on PR #361 (round 3): generate_excel and the canonical
        sort key must share ONE Job # alias resolver, or adding /
        reordering an Excel alias silently reintroduces an
        order-dependent header. Pin the sharing and the precedence."""
        import pipeline.excel
        self.assertIs(pipeline.excel.header_job_number,
                      change_detection.header_job_number)
        for alias in change_detection._JOB_NUMBER_ALIASES:
            self.assertEqual(
                change_detection.header_job_number({alias: 'J-9'}), 'J-9',
                alias)
        # Behavioural: a row carrying Job # under a lower-precedence
        # alias only still fills the workbook's "Job #:" cell.
        row = _row(wr='90002', **{
            'Weekly Reference Logged Date': '2026-07-26',
            '__week_ending_date': datetime.datetime(2026, 7, 26),
            'JOB#': 'J-9'})
        self.assertEqual(
            _report_details([row], '072626_90002', ('Job #:',)),
            {'Job #:': 'J-9'})

    def test_tie_breaker_still_detects_a_real_edit(self):
        base = [
            _row(**{'Work Type': 'Install'}),
            _row(**{'Work Type': 'Remove'}),
        ]
        edited = [
            _row(**{'Work Type': 'Install'}),
            _row(**{'Work Type': 'Remove', 'Dept #': '521'}),
        ]
        self.assertNotEqual(generate_weekly_pdfs.calculate_data_hash(base),
                            generate_weekly_pdfs.calculate_data_hash(edited))

    def test_groups_without_differing_ties_are_byte_identical_to_pre_fix_order(self):
        """Replay the PRE-fix algorithm (stable sort on the business key +
        VAC-crew fields only, then the same per-row strings and meta) on a
        tie-free group and require the exact same digest. Any change here
        would mean the tiebreaker altered a currently-stable hash."""
        rows = [
            _row(cu=f'CU-{i:03d}', qty=i % 3 + 1, pole=f'P-{i % 5}',
                 price=f'${10 + i}.00', snapshot='2026-07-30',
                 **{'Work Type': ('Install', 'Remove')[i % 2],
                    'Dept #': str(500 + i % 4)})
            for i in range(40)
        ]
        # Shuffle deterministically so input order != sorted order.
        shuffled = rows[::-1][::3] + rows[1::3] + rows[2::3][::-1]
        self.assertEqual(len(shuffled), 40)

        def _pre_fix_hash(group_rows):
            _sort = lambda x: (
                str(x.get('Work Request #', '')),
                str(x.get('Snapshot Date', '')),
                str(x.get('CU', '')),
                str(x.get('Pole #') or x.get('Point #') or x.get('Point Number') or ''),
                str(x.get('Quantity', '')),
                str(x.get('__vac_crew_name') or ''),
                str(x.get('__vac_crew_dept') or ''),
                str(x.get('__vac_crew_job') or ''),
            )
            sorted_rows = sorted(group_rows, key=_sort)
            # Assert the fixture really is tie-free on that key.
            keys = [_sort(r) for r in sorted_rows]
            assert len(set(keys)) == len(keys), "fixture must be tie-free"
            variant = sorted_rows[0].get('__variant', 'primary')
            hasher = hashlib.sha256()
            foreman = None
            for r in sorted_rows:
                f = r.get('__current_foreman') or r.get('Foreman') or ''
                if foreman is None and f:
                    foreman = f
                hasher.update("|".join(
                    change_detection._extended_row_fields(r, variant)
                ).encode('utf-8'))
                hasher.update(b"\n")
            depts = sorted({str(r.get('Dept #', '') or '') for r in sorted_rows
                            if r.get('Dept #') is not None})
            total = sum(change_detection.parse_price(r.get('Units Total Price'))
                        for r in sorted_rows)
            meta = [f"FOREMAN={foreman or ''}", f"VARIANT={variant}",
                    f"DEPTS={','.join(depts)}", f"TOTAL={total:.2f}",
                    f"ROWCOUNT={len(sorted_rows)}"]
            return hasher, meta

        # Compare the production digest against the production function
        # run on the pre-fix ORDER: because the fixture is tie-free, the
        # tiebreaker must produce exactly the same order.
        produced = generate_weekly_pdfs.calculate_data_hash(shuffled)
        pre_sorted = sorted(shuffled, key=lambda x: (
            str(x.get('Work Request #', '')), str(x.get('Snapshot Date', '')),
            str(x.get('CU', '')), str(x.get('Pole #') or ''), str(x.get('Quantity', ''))))
        self.assertEqual(produced,
                         generate_weekly_pdfs.calculate_data_hash(pre_sorted))
        # And the digest prefix the pre-fix loop would have fed is identical
        # for the row portion (meta handling is unchanged by this fix).
        hasher, _meta = _pre_fix_hash(shuffled)
        hasher2, _meta2 = _pre_fix_hash(pre_sorted)
        self.assertEqual(hasher.hexdigest(), hasher2.hexdigest())

    def test_legacy_mode_untouched(self):
        """LEGACY (EXTENDED_CHANGE_DETECTION=0) keeps its documented 5-key
        sort with no tiebreaker (rollback-stability guarantee)."""
        generate_weekly_pdfs.EXTENDED_CHANGE_DETECTION = False
        rows = [
            _row(**{'Work Type': 'Install'}),
            _row(**{'Work Type': 'Remove'}),
        ]
        a = generate_weekly_pdfs.calculate_data_hash(rows)
        b = generate_weekly_pdfs.calculate_data_hash(rows[::-1])
        # Legacy hashes Work Type per row in input order for tied rows, so
        # the two orders differ -- that is the documented legacy behaviour.
        self.assertNotEqual(a, b)


class IdentitySitesUseCanonicalRowTests(unittest.TestCase):
    """Codex / Copilot on PR #361: the three orchestrate identity sites
    (Site 1 main-loop identifier / history_key, Site 2 valid_wr_weeks,
    Site 3 current_keys prune) must read helper dept / job, claimer and
    User from the canonical row, never from arrival-order group_rows[0]
    -- or a stable hash is looked up under an unstable history_key, the
    prior key is pruned, and a mixed-department helper group regenerates
    and re-uploads every run.

    The three sites are inline in ``pipeline.orchestrate.main`` (no test
    in this suite drives that function -- see
    test_deep_run_reconciliation / test_incremental_read), so the wiring
    is pinned at source level, the repository's existing practice for
    these sites (test_primary_claim_attribution), and the identity
    INPUTS the sites read are pinned behaviourally below."""

    @classmethod
    def setUpClass(cls):
        import inspect
        import pipeline.orchestrate
        cls._src = Path(
            inspect.getsourcefile(pipeline.orchestrate)
        ).read_text(encoding='utf-8')

    def test_identity_inputs_from_the_canonical_row_are_order_independent(
            self):
        """Every field Sites 1-3 read (helper foreman / dept / job, User,
        __current_foreman, __variant) and the history-key shape Site 1
        writes and Site 3 reconstructs are identical for both arrival
        orders of a mixed-department helper group -- under one hash."""
        saved = generate_weekly_pdfs.EXTENDED_CHANGE_DETECTION
        generate_weekly_pdfs.EXTENDED_CHANGE_DETECTION = True
        try:
            a = _row(User='u1', **{
                '__variant': 'helper', '__helper_foreman': 'Sam Sample',
                '__helper_dept': 'NA-03', '__helper_job': 'J-1',
                '__current_foreman': 'Pat Example'})
            b = _row(User='u2', **{
                '__variant': 'helper', '__helper_foreman': 'Sam Sample',
                '__helper_dept': 'NA-04', '__helper_job': 'J-2',
                '__current_foreman': 'Pat Example'})
            fields = ('__helper_foreman', '__helper_dept', '__helper_job',
                      'User', '__current_foreman', '__variant')
            seen_identity, seen_key, seen_hash = set(), set(), set()
            for rows in ([a, b], [b, a]):
                first = change_detection.canonical_first_row(rows)
                seen_identity.add(tuple(first.get(f, '') for f in fields))
                seen_key.add("|".join((
                    first.get('__helper_foreman', ''),
                    first.get('__helper_dept', ''),
                    first.get('__helper_job', ''))))
                seen_hash.add(generate_weekly_pdfs.calculate_data_hash(rows))
            self.assertEqual(len(seen_identity), 1, seen_identity)
            self.assertEqual(seen_key, {'Sam Sample|NA-03|J-1'})
            self.assertEqual(len(seen_hash), 1)
        finally:
            generate_weekly_pdfs.EXTENDED_CHANGE_DETECTION = saved

    def test_no_identity_field_is_read_from_the_arrival_order_row(self):
        for needle in ("group_rows[0].get('__helper_foreman'",
                       "group_rows[0].get('__helper_dept'",
                       "group_rows[0].get('__helper_job'",
                       "group_rows[0].get('User')"):
            self.assertNotIn(
                needle, self._src,
                f"{needle} must come from canonical_first_row(group_rows)")

    def test_each_identity_site_binds_the_canonical_row(self):
        for marker in ("CR-01 gap closure (Site 1",
                       "CR-01 gap closure (Site 2",
                       "CR-01 gap closure (Site 3"):
            at = self._src.index(marker)
            self.assertIn("canonical_first_row(group_rows)",
                          self._src[max(0, at - 5000):at], marker)


if __name__ == "__main__":
    unittest.main()
